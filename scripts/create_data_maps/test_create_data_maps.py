import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.create_data_maps.run import create_data_maps


def write_workbook(path: Path, sheets: dict[str, list[list[object]]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in sheets.items():
        worksheet = workbook.create_sheet(sheet_name)
        for row in rows:
            worksheet.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


class CreateDataMapsTests(unittest.TestCase):
    def test_creates_platforms_and_class_maps_with_relative_locations_and_descriptions(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study_folder = Path(tmpdir) / "study"
            cleaned_dir = study_folder / "data" / "cleaned"
            write_workbook(
                cleaned_dir / "dictionary.xlsx",
                {
                    "instrument": [
                        ["instrument", "instrument_label"],
                        ["madrs", "MADRS Montgomery Asberg Depression Rating Scale"],
                        ["unmatched_scale", "Unmatched Clinical Scale"],
                    ]
                },
            )
            write_workbook(
                cleaned_dir / "assessments" / "58807-madrs.xlsx",
                {
                    "cleaned": [
                        ["IRB", "subid", "arm", "visit"],
                        ["58807", "s001", "1", "V1"],
                    ]
                },
            )
            write_workbook(
                cleaned_dir / "assessments" / "58807-unmatched_scale.xlsx",
                {
                    "cleaned": [
                        ["IRB", "subid", "arm", "visit"],
                        ["58807", "s001", "1", "V1"],
                    ]
                },
            )
            write_workbook(
                cleaned_dir / "assessments" / "58807-local_notes.xlsx",
                {
                    "notes": [
                        ["note"],
                        ["not a REDCap instrument"],
                    ]
                },
            )
            write_workbook(
                cleaned_dir / "subjects" / "subject_timepoints.xlsx",
                {
                    "subject_timepoints": [
                        ["IRB", "subid", "arm", "visit"],
                        ["58807", "s001", "1", "V1"],
                    ]
                },
            )
            semantics_path = Path(tmpdir) / "semantics.json"
            semantics_path.write_text(
                """
                {
                  "assessments": {
                    "caption": "Clinical assessments.",
                    "examples": [
                      "MADRS Montgomery Asberg Depression Rating Scale clinician-rated depression severity assessment"
                    ]
                  },
                  "subjects": {
                    "caption": "Subject indexing.",
                    "examples": []
                  }
                }
                """,
                encoding="utf-8",
            )

            result = create_data_maps(study_folder, semantics_path=semantics_path)

            self.assertEqual(result.map_count, 2)
            self.assertEqual(result.output_dir, study_folder / "data-map")
            self.assertEqual(result.platforms_path, study_folder / "data-map" / "platforms-data-map.xlsx")
            self.assertEqual(
                sorted(path.name for path in result.class_paths),
                ["assessments-data-map.xlsx", "subjects-data-map.xlsx"],
            )
            platforms = load_workbook(study_folder / "data-map" / "platforms-data-map.xlsx", data_only=True)
            platform_rows = list(platforms.active.iter_rows(values_only=True))
            self.assertEqual(platform_rows[0], ("stage", "privacy", "description", "location"))
            self.assertEqual(len(platform_rows), 11)
            self.assertTrue(all(row == (None, None, None, None) for row in platform_rows[1:]))

            assessments = load_workbook(study_folder / "data-map" / "assessments-data-map.xlsx", data_only=True)
            assessment_rows = list(assessments.active.iter_rows(values_only=True))
            self.assertEqual(assessment_rows[0], ("stage", "description", "location"))
            self.assertEqual(assessment_rows[1], ("raw", None, None))
            assessment_rows_by_location = {row[2]: row for row in assessment_rows[2:]}
            self.assertEqual(
                assessment_rows_by_location["./data/cleaned/assessments/58807-madrs.xlsx"],
                (
                    "cleaned/processed",
                    "MADRS Montgomery Asberg Depression Rating Scale clinician-rated depression severity assessment",
                    "./data/cleaned/assessments/58807-madrs.xlsx",
                ),
            )
            self.assertEqual(
                assessment_rows_by_location["./data/cleaned/assessments/58807-unmatched_scale.xlsx"],
                (
                    "cleaned/processed",
                    "Unmatched Clinical Scale",
                    "./data/cleaned/assessments/58807-unmatched_scale.xlsx",
                ),
            )
            self.assertEqual(
                assessment_rows_by_location["./data/cleaned/assessments/58807-local_notes.xlsx"],
                (
                    "cleaned/processed",
                    "58807-local_notes",
                    "./data/cleaned/assessments/58807-local_notes.xlsx",
                ),
            )

            subjects = load_workbook(study_folder / "data-map" / "subjects-data-map.xlsx", data_only=True)
            subject_rows = list(subjects.active.iter_rows(values_only=True))
            self.assertEqual(subject_rows[1], ("raw", None, None))
            self.assertEqual(
                subject_rows[2],
                (
                    "cleaned/processed",
                    "subject_timepoints",
                    "./data/cleaned/subjects/subject_timepoints.xlsx",
                ),
            )


if __name__ == "__main__":
    unittest.main()
