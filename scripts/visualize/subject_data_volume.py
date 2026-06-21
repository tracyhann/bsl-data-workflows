#!/usr/bin/env python3
"""Visualize cleaned REDCap data-entry volume by subject and visit."""

from __future__ import annotations

import argparse
import os
import re
import sys
import tempfile
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


if __package__ in {None, ""}:
    sys.path.append(str(Path(__file__).resolve().parents[2]))


REQUIRED_CLEANED_PREFIX = ["irb", "subid", "arm", "visit"]
PLOTS_DIRNAME = "plots"


@dataclass(frozen=True)
class SubjectDataVolume:
    entries: pd.DataFrame
    participant_counts: pd.DataFrame
    visit_subid_counts: pd.DataFrame
    visit_totals: pd.DataFrame


@dataclass(frozen=True)
class SubjectDataVolumeResult:
    study_folder: Path
    output_dir: Path
    summary_path: Path
    plot_path: Path
    total_entries: int
    subject_count: int
    visit_count: int
    instrument_count: int


def normalize_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def natural_visit_key(value: object) -> tuple[object, ...]:
    text = str(value or "").strip()
    match = re.search(r"(\d+)", text)
    if match:
        return (text[: match.start()], int(match.group(1)), text)
    return (text, -1, text)


def is_blank(value: object) -> bool:
    return value is None or str(value).strip() == ""


def dictionary_path_for_study(study_folder: Path) -> Path:
    candidates = [
        study_folder / "data" / "cleaned" / "dictionary.xlsx",
        study_folder / "data" / "cleaned" / "redcap" / "dictionary.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def instrument_key_from_workbook(path: Path) -> str:
    stem = path.stem
    stripped = re.sub(r"^(?:\d+[-_])+", "", stem)
    return stripped or stem


def load_dictionary_instruments(study_folder: Path) -> set[str]:
    dictionary_path = dictionary_path_for_study(study_folder)
    if not dictionary_path.exists():
        return set()

    workbook = load_workbook(dictionary_path, read_only=True, data_only=True)
    if "instrument" not in workbook.sheetnames:
        return set()

    rows = list(workbook["instrument"].iter_rows(values_only=True))
    if not rows:
        return set()

    headers = [normalize_header(value) for value in rows[0]]
    if "instrument" not in headers:
        return set()
    instrument_index = headers.index("instrument")

    instruments: set[str] = set()
    for row in rows[1:]:
        if instrument_index >= len(row):
            continue
        instrument = str(row[instrument_index] or "").strip()
        if not instrument:
            continue
        instruments.add(instrument)
        instruments.add(normalize_header(instrument))
    return instruments


def cleaned_workbook_paths(study_folder: Path) -> list[Path]:
    cleaned_dir = study_folder / "data" / "cleaned"
    if not cleaned_dir.exists():
        return []
    return sorted(
        path
        for path in cleaned_dir.rglob("*.xlsx")
        if path.is_file()
        and not path.name.startswith("~$")
        and path.name != "dictionary.xlsx"
        and "histories" not in path.parts
    )


def workbook_matches_dictionary(path: Path, dictionary_instruments: set[str]) -> bool:
    if not dictionary_instruments:
        return False
    instrument_key = instrument_key_from_workbook(path)
    return instrument_key in dictionary_instruments or normalize_header(instrument_key) in dictionary_instruments


def cleaned_sheet_rows(path: Path) -> tuple[list[str], list[tuple[object, ...]]] | None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "cleaned" not in workbook.sheetnames:
        return None
    rows = list(workbook["cleaned"].iter_rows(values_only=True))
    if not rows:
        return None
    headers = [normalize_header(value) for value in rows[0]]
    if headers[: len(REQUIRED_CLEANED_PREFIX)] != REQUIRED_CLEANED_PREFIX:
        return None
    data_rows = [tuple(row) for row in rows[1:] if any(not is_blank(value) for value in row)]
    return headers, data_rows


def source_entries_for_workbook(study_folder: Path, path: Path) -> list[dict[str, object]]:
    cleaned = cleaned_sheet_rows(path)
    if cleaned is None:
        return []
    headers, rows = cleaned
    subid_index = headers.index("subid")
    visit_index = headers.index("visit")
    arm_index = headers.index("arm") if "arm" in headers else None
    irb_index = headers.index("irb") if "irb" in headers else None
    instrument = instrument_key_from_workbook(path)
    relative_path = f"./{path.relative_to(study_folder).as_posix()}"

    entries: list[dict[str, object]] = []
    for row_number, row in enumerate(rows, start=2):
        subid = str(row[subid_index] or "").strip() if subid_index < len(row) else ""
        visit = str(row[visit_index] or "").strip() if visit_index < len(row) else ""
        if not subid or not visit:
            continue
        entries.append(
            {
                "IRB": str(row[irb_index] or "").strip() if irb_index is not None and irb_index < len(row) else "",
                "subid": subid,
                "arm": str(row[arm_index] or "").strip() if arm_index is not None and arm_index < len(row) else "",
                "visit": visit,
                "instrument": instrument,
                "source_workbook": relative_path,
                "cleaned_row": row_number,
            }
        )
    return entries


def collect_subject_data_volume(study_folder: str | Path) -> SubjectDataVolume:
    study_folder = Path(study_folder)
    dictionary_instruments = load_dictionary_instruments(study_folder)
    entries: list[dict[str, object]] = []
    for path in cleaned_workbook_paths(study_folder):
        if not workbook_matches_dictionary(path, dictionary_instruments):
            continue
        entries.extend(source_entries_for_workbook(study_folder, path))

    entries_df = pd.DataFrame(
        entries,
        columns=["IRB", "subid", "arm", "visit", "instrument", "source_workbook", "cleaned_row"],
    )
    if entries_df.empty:
        empty_counts = pd.DataFrame(columns=["subid", "entry_count"])
        empty_visit_counts = pd.DataFrame(columns=["visit", "subid", "entry_count"])
        empty_visit_totals = pd.DataFrame(columns=["visit", "entry_count"])
        return SubjectDataVolume(entries_df, empty_counts, empty_visit_counts, empty_visit_totals)

    participant_counts = (
        entries_df.groupby("subid", dropna=False)
        .size()
        .reset_index(name="entry_count")
        .sort_values(["subid"], kind="stable")
        .reset_index(drop=True)
    )
    visit_subid_counts = (
        entries_df.groupby(["visit", "subid"], dropna=False)
        .size()
        .reset_index(name="entry_count")
        .sort_values(["visit", "subid"], key=lambda column: column.map(natural_visit_key), kind="stable")
        .reset_index(drop=True)
    )
    visit_order = sorted(entries_df["visit"].dropna().unique(), key=natural_visit_key)
    visit_totals = (
        entries_df.groupby("visit", dropna=False)
        .size()
        .reindex(visit_order)
        .reset_index(name="entry_count")
    )
    return SubjectDataVolume(entries_df, participant_counts, visit_subid_counts, visit_totals)


def write_summary_excel(volumes: SubjectDataVolume, output_dir: Path) -> Path:
    output_path = output_dir / "subject_data_volume.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        volumes.participant_counts.to_excel(writer, index=False, sheet_name="participant_counts")
        volumes.visit_subid_counts.to_excel(writer, index=False, sheet_name="visit_subid_counts")
        volumes.visit_totals.to_excel(writer, index=False, sheet_name="visit_totals")
        volumes.entries.to_excel(writer, index=False, sheet_name="source_entries")
    return output_path


def configure_matplotlib_cache() -> None:
    cache_root = Path(tempfile.gettempdir()) / "bsldb_matplotlib_cache"
    matplotlib_cache = cache_root / "matplotlib"
    xdg_cache = cache_root / "xdg"
    matplotlib_cache.mkdir(parents=True, exist_ok=True)
    xdg_cache.mkdir(parents=True, exist_ok=True)
    os.environ.setdefault("MPLCONFIGDIR", str(matplotlib_cache))
    os.environ.setdefault("XDG_CACHE_HOME", str(xdg_cache))


def write_plot(volumes: SubjectDataVolume, output_dir: Path) -> Path:
    configure_matplotlib_cache()

    import matplotlib

    matplotlib.use("Agg", force=True)
    import matplotlib.pyplot as plt

    output_path = output_dir / "subject_data_volume.png"
    fig, axes = plt.subplots(2, 1, figsize=(22, 16))
    fig.subplots_adjust(left=0.07, right=0.82, top=0.94, bottom=0.18, hspace=0.48)

    if volumes.participant_counts.empty:
        for axis in axes:
            axis.text(0.5, 0.5, "No cleaned REDCap entries found", ha="center", va="center")
            axis.set_axis_off()
    else:
        axes[0].bar(volumes.participant_counts["subid"], volumes.participant_counts["entry_count"])
        axes[0].set_title("Raw Data Entry Count Per Participant")
        axes[0].set_xlabel("Participant")
        axes[0].set_ylabel("Data Entry Count")
        axes[0].tick_params(axis="x", rotation=90)
        axes[0].grid(axis="y", alpha=0.25)

        pivot = volumes.visit_subid_counts.pivot(index="visit", columns="subid", values="entry_count").fillna(0)
        pivot = pivot.reindex(sorted(pivot.index, key=natural_visit_key))
        pivot.plot(kind="bar", stacked=True, ax=axes[1], width=0.82, legend=False)
        axes[1].set_title("Data Entry Count By Visit And Participant")
        axes[1].set_xlabel("Visit")
        axes[1].set_ylabel("Data Entry Count")
        axes[1].tick_params(axis="x", rotation=45, labelsize=10)
        for label in axes[1].get_xticklabels():
            label.set_horizontalalignment("right")
        axes[1].grid(axis="y", alpha=0.25)

        max_legend_items = 25
        handles, labels = axes[1].get_legend_handles_labels()
        if handles:
            if len(handles) > max_legend_items:
                handles = handles[:max_legend_items]
                labels = [*labels[:max_legend_items], f"+ {len(pivot.columns) - max_legend_items} more in Excel"]
                handles = [*handles, handles[-1]]
            axes[1].legend(
                handles,
                labels,
                title="Participant",
                bbox_to_anchor=(1.01, 1),
                loc="upper left",
                fontsize=9,
                title_fontsize=10,
            )

    fig.savefig(output_path, dpi=200)
    plt.close(fig)
    return output_path


def latest_history_plots_dir(study_folder: Path) -> Path:
    histories_dir = study_folder / "histories"
    history_dirs = sorted(path for path in histories_dir.iterdir() if path.is_dir()) if histories_dir.exists() else []
    history_dir = history_dirs[-1] if history_dirs else histories_dir / date.today().isoformat()
    return history_dir / PLOTS_DIRNAME


def create_subject_data_volume(
    study_folder: str | Path,
    output_dir: str | Path | None = None,
) -> SubjectDataVolumeResult:
    study_folder = Path(study_folder)
    output_dir = Path(output_dir) if output_dir is not None else latest_history_plots_dir(study_folder)
    output_dir.mkdir(parents=True, exist_ok=True)

    volumes = collect_subject_data_volume(study_folder)
    summary_path = write_summary_excel(volumes, output_dir)
    plot_path = write_plot(volumes, output_dir)

    return SubjectDataVolumeResult(
        study_folder=study_folder,
        output_dir=output_dir,
        summary_path=summary_path,
        plot_path=plot_path,
        total_entries=len(volumes.entries),
        subject_count=volumes.entries["subid"].nunique() if not volumes.entries.empty else 0,
        visit_count=volumes.entries["visit"].nunique() if not volumes.entries.empty else 0,
        instrument_count=volumes.entries["instrument"].nunique() if not volumes.entries.empty else 0,
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--study-folder", required=True, type=Path, help="Cleaned study folder.")
    parser.add_argument(
        "--out-dir",
        type=Path,
        help="Output directory. Defaults to <study-folder>/histories/<latest-date>/plots.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    result = create_subject_data_volume(args.study_folder, output_dir=args.out_dir)
    print(result.output_dir)
    print(f"summary={result.summary_path}")
    print(f"plot={result.plot_path}")
    print(f"entries={result.total_entries}")
    print(f"subjects={result.subject_count}")
    print(f"visits={result.visit_count}")
    print(f"instruments={result.instrument_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
