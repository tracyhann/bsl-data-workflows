import tempfile
import subprocess
import sys
import unittest
from pathlib import Path

from PIL import Image
from openpyxl import Workbook, load_workbook

from scripts.visualize.subject_data_volume import collect_subject_data_volume, create_subject_data_volume


def write_workbook(path: Path, sheets: dict[str, list[list[object]]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in sheets.items():
        worksheet = workbook.create_sheet(sheet_name)
        for row in rows:
            worksheet.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


class SubjectDataVolumeTests(unittest.TestCase):
    def test_counts_cleaned_redcap_rows_by_subid_and_visit_and_writes_plot_and_summary(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study_folder = Path(tmpdir) / "study"
            cleaned_dir = study_folder / "data" / "cleaned"
            write_workbook(
                cleaned_dir / "dictionary.xlsx",
                {
                    "instrument": [
                        ["instrument", "instrument_label"],
                        ["madrs", "MADRS Montgomery Asberg Depression Rating Scale"],
                        ["panas", "PANAS Positive and Negative Affect Schedule"],
                    ]
                },
            )
            cleaned_rows = [
                ["IRB", "subid", "arm", "visit", "date", "record_id", "redcap_event_name", "score"],
                ["58807", "s001", "1", "V1", "2024-01-01", "58807_s001", "baseline_arm_1", 10],
                ["58807", "s001", "1", "V2", "2024-01-08", "58807_s001", "visit_2_arm_1", 8],
                ["58807", "s002", "1", "V1", "2024-01-02", "58807_s002", "baseline_arm_1", 12],
            ]
            write_workbook(cleaned_dir / "assessments" / "58807-madrs.xlsx", {"cleaned": cleaned_rows})
            write_workbook(
                cleaned_dir / "assessments" / "58807-panas.xlsx",
                {
                    "cleaned": [
                        ["IRB", "subid", "arm", "visit", "date", "record_id", "redcap_event_name", "score"],
                        ["58807", "s001", "1", "V1", "2024-01-01", "58807_s001", "baseline_arm_1", 4],
                        ["58807", "s003", "1", "V3", "2024-02-01", "58807_s003", "visit_3_arm_1", 5],
                    ]
                },
            )
            write_workbook(
                cleaned_dir / "assessments" / "58807-not_in_dictionary.xlsx",
                {
                    "cleaned": [
                        ["IRB", "subid", "arm", "visit", "date", "record_id", "redcap_event_name"],
                        ["58807", "s999", "1", "V1", "2024-01-01", "58807_s999", "baseline_arm_1"],
                    ]
                },
            )
            write_workbook(
                cleaned_dir / "subjects" / "subject_timepoints.xlsx",
                {
                    "subject_timepoints": [
                        ["IRB", "subid", "arm", "visit"],
                        ["58807", "s001", "1", "V1"],
                    ]
                },
            )
            (study_folder / "histories" / "2026-06-18").mkdir(parents=True)
            latest_history = study_folder / "histories" / "2026-06-19"
            latest_history.mkdir(parents=True)

            volumes = collect_subject_data_volume(study_folder)

            participant_counts = {
                row["subid"]: row["entry_count"] for row in volumes.participant_counts.to_dict("records")
            }
            self.assertEqual(participant_counts, {"s001": 3, "s002": 1, "s003": 1})
            visit_counts = {
                (row["visit"], row["subid"]): row["entry_count"]
                for row in volumes.visit_subid_counts.to_dict("records")
            }
            self.assertEqual(
                visit_counts,
                {
                    ("V1", "s001"): 2,
                    ("V1", "s002"): 1,
                    ("V2", "s001"): 1,
                    ("V3", "s003"): 1,
                },
            )

            result = create_subject_data_volume(study_folder)

            self.assertEqual(result.output_dir, latest_history / "plots")
            self.assertEqual(result.summary_path, latest_history / "plots" / "subject_data_volume.xlsx")
            self.assertEqual(result.plot_path, latest_history / "plots" / "subject_data_volume.png")
            self.assertEqual(result.total_entries, 5)
            self.assertTrue(result.summary_path.exists())
            self.assertTrue(result.plot_path.exists())
            workbook = load_workbook(result.summary_path, data_only=True)
            self.assertIn("participant_counts", workbook.sheetnames)
            self.assertIn("visit_subid_counts", workbook.sheetnames)
            self.assertIn("source_entries", workbook.sheetnames)

    def test_cli_does_not_emit_matplotlib_cache_warnings(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study_folder = Path(tmpdir) / "study"
            cleaned_dir = study_folder / "data" / "cleaned"
            write_workbook(
                cleaned_dir / "dictionary.xlsx",
                {
                    "instrument": [
                        ["instrument", "instrument_label"],
                        ["madrs", "MADRS Montgomery Asberg Depression Rating Scale"],
                    ]
                },
            )
            write_workbook(
                cleaned_dir / "assessments" / "58807-madrs.xlsx",
                {
                    "cleaned": [
                        ["IRB", "subid", "arm", "visit", "date", "record_id", "redcap_event_name"],
                        ["58807", "s001", "1", "V1", "2024-01-01", "58807_s001", "baseline_arm_1"],
                    ]
                },
            )

            completed = subprocess.run(
                [
                    sys.executable,
                    "scripts/visualize/subject_data_volume.py",
                    "--study-folder",
                    str(study_folder),
                ],
                capture_output=True,
                check=False,
                text=True,
            )

            self.assertEqual(completed.returncode, 0, completed.stderr)
            self.assertNotIn("Matplotlib created a temporary cache directory", completed.stderr)
            self.assertNotIn("Fontconfig error", completed.stderr)

    def test_plot_has_enough_canvas_for_many_subjects_and_visit_labels(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study_folder = Path(tmpdir) / "study"
            cleaned_dir = study_folder / "data" / "cleaned"
            write_workbook(
                cleaned_dir / "dictionary.xlsx",
                {
                    "instrument": [
                        ["instrument", "instrument_label"],
                        ["madrs", "MADRS Montgomery Asberg Depression Rating Scale"],
                    ]
                },
            )
            rows = [["IRB", "subid", "arm", "visit", "date", "record_id", "redcap_event_name"]]
            visits = [
                "1week_followup",
                "1week_post",
                "3week_post",
                "additional_tms_treatment",
                "adverse_events_requiring_review",
                "treatment_day_5",
                "V1",
                "V2",
                "V8",
                "V10",
            ]
            for subject_number in range(1, 61):
                for visit in visits:
                    rows.append(
                        [
                            "58807",
                            f"s{subject_number:03d}",
                            "1",
                            visit,
                            "2024-01-01",
                            f"58807_s{subject_number:03d}",
                            f"{visit}_arm_1",
                        ]
                    )
            write_workbook(cleaned_dir / "assessments" / "58807-madrs.xlsx", {"cleaned": rows})
            (study_folder / "histories" / "2026-06-19").mkdir(parents=True)

            result = create_subject_data_volume(study_folder)

            with Image.open(result.plot_path) as image:
                width, height = image.size
            self.assertGreaterEqual(width, 4400)
            self.assertGreaterEqual(height, 3200)


if __name__ == "__main__":
    unittest.main()
