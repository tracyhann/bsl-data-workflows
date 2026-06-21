import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.workflows.clean_redcap_instruments.steps.discover_and_standardize import run_workflow


class WorkflowTests(unittest.TestCase):
    def test_workflow_writes_discovery_log_and_dictionary(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            input_path = root / "input.csv"
            out_dir = root / "intermediates"
            log_path = root / "log.md"
            with input_path.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(
                    [
                        "record_id",
                        "redcap_event_name",
                        "demo_recordid",
                        "demo_date",
                        "demographics_complete",
                    ]
                )
                writer.writerow(["58807_s001", "baseline_arm_1", "58807s001", "2026-01-01", "2"])
                writer.writerow(["MDD_001", "screening_arm_1", "", "", ""])

            result = run_workflow(input_path=input_path, out_dir=out_dir, log_path=log_path)

            self.assertTrue(result["log"].exists())
            self.assertTrue(result["dictionary"].exists())
            self.assertTrue(result["record_id"]["standardized"].exists())
            self.assertTrue(result["events"]["events_by_arm"].exists())
            self.assertTrue(result["instruments"]["instrument_summary"].exists())

            log_text = log_path.read_text(encoding="utf-8")
            self.assertIn("<summary><h1>Discovery</h1></summary>", log_text)
            self.assertNotIn("\n# Discovery\n", log_text)
            self.assertIn("## Record ID Discovery", log_text)
            self.assertNotIn("===== RECORD ID DISCOVERY =====", log_text)
            self.assertIn("| 1 | 2 | 3 | 4 | 5 |", log_text)
            self.assertIn("<summary><h1>Dictionary</h1></summary>", log_text)
            self.assertNotIn("\n# Dictionary\n", log_text)
            self.assertIn("## Subject ID", log_text)
            self.assertIn("Verification rule:", log_text)
            self.assertNotIn("===== SUBID =====", log_text)

            workbook = load_workbook(result["dictionary"])
            self.assertEqual(workbook.sheetnames, ["subject_id", "event", "instrument"])
            subject_rows = list(workbook["subject_id"].iter_rows(values_only=True))
            self.assertEqual(subject_rows[0], ("IRB", "subid", "standardized", "raw_entries"))
            self.assertEqual(subject_rows[1][2], "58807_s001")

    def test_workflow_logs_optional_codebook_verification_mismatches(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            input_path = root / "input.csv"
            out_dir = root / "intermediates"
            log_path = root / "log.md"
            codebook_path = root / "codebook.xlsx"

            with input_path.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(
                    [
                        "record_id",
                        "redcap_event_name",
                        "demo_date",
                        "demographics_complete",
                    ]
                )
                writer.writerow(["58807_s001", "baseline_arm_1", "2026-01-01", "2"])

            workbook = Workbook()
            instrument_sheet = workbook.active
            instrument_sheet.title = "study-instrument"
            instrument_sheet.append(["Instrument", "Form Name", "Events"])
            instrument_sheet.append(["Demographics", "demographics", "baseline_arm_1"])
            instrument_sheet.append(["Missing Form", "missing_form", "baseline_arm_1"])
            event_sheet = workbook.create_sheet("study-event")
            event_sheet.append(["Event Name", "Unique event name", "Event ID"])
            event_sheet.append(["Baseline", "baseline_arm_1", "1"])
            event_sheet.append(["Missing", "missing_arm_2", "2"])
            workbook.save(codebook_path)

            result = run_workflow(
                input_path=input_path,
                out_dir=out_dir,
                log_path=log_path,
                codebook_path=codebook_path,
                instrument_codebook_sheet="study-instrument",
                event_codebook_sheet="study-event",
            )

            log_text = log_path.read_text(encoding="utf-8")
            instrument_verification_exists = result["instruments"]["instrument_codebook_verification"].exists()
            event_verification_exists = result["events"]["event_codebook_verification"].exists()

        self.assertTrue(instrument_verification_exists)
        self.assertTrue(event_verification_exists)
        self.assertIn("Codebook verification mismatches: **1**", log_text)
        self.assertIn("missing_form", log_text)
        self.assertIn("missing", log_text)
        self.assertIn("Raw discovery remains the source of truth", log_text)


if __name__ == "__main__":
    unittest.main()
