import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from scripts.workflows.clean_redcap_instruments.steps.exclude_instruments import exclude_instruments


def write_dictionary(path: Path, rows: list[tuple[str, ...]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "instrument"
    worksheet.append(["instrument", "instrument_label", "abbreviation", "number_of_events", "events", "number_of_columns"])
    for row in rows:
        instrument, label = row[:2]
        number_of_events = row[2] if len(row) > 2 else "1"
        worksheet.append([instrument, label, "", number_of_events, "Arm 1: baseline", "1"])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_classification(path: Path, rows: list[tuple[str, str, str, float]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "instrument_classification"
    worksheet.append(["instrument_name", "instrument_label", "class", "confidence"])
    for row in rows:
        worksheet.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_instrument_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "cleaned"
    worksheet.append(["IRB", "subid", "arm", "visit", "date", "record_id", "redcap_event_name"])
    worksheet.append(["62822", "s001", "1", "V1", "2026-01-01", "62822_s001", "baseline_arm_1"])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


class ExcludeInstrumentsTests(unittest.TestCase):
    def test_excludes_admin_and_contact_keyword_instrument_workbooks_to_excluded_dir(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study_folder = Path(tmpdir) / "study"
            cleaned_dir = study_folder / "data" / "cleaned" / "redcap"
            history_dir = study_folder / "histories" / "2026-06-18"
            history_dir.mkdir(parents=True)
            (history_dir / "log.md").write_text("# REDCap Workflow Log\n", encoding="utf-8")
            write_dictionary(
                cleaned_dir / "dictionary.xlsx",
                [
                    ("survey_trigger", "Survey Trigger"),
                    ("patient_email", "Patient Email"),
                    ("patient_contact_log", "Patient Contact Log"),
                    ("participant_phone", "Participant Phone"),
                    ("participant_address", "Participant Address"),
                    ("capacity_to_consent", "Capacity To Consent"),
                    ("empty_event_form", "Empty Event Form", "0"),
                    ("madrs", "MADRS"),
                ],
            )
            write_classification(
                history_dir / "instrument_classification.xlsx",
                [
                    ("survey_trigger", "Survey Trigger", "admin", 0.91),
                    ("patient_email", "Patient Email", "subjects", 0.88),
                    ("patient_contact_log", "Patient Contact Log", "subjects", 0.88),
                    ("participant_phone", "Participant Phone", "subjects", 0.88),
                    ("participant_address", "Participant Address", "subjects", 0.88),
                    ("capacity_to_consent", "Capacity To Consent", "subjects", 0.88),
                    ("empty_event_form", "Empty Event Form", "assessments", 0.88),
                    ("madrs", "MADRS", "assessments", 0.88),
                ],
            )
            for instrument in [
                "survey_trigger",
                "patient_email",
                "patient_contact_log",
                "participant_phone",
                "participant_address",
                "capacity_to_consent",
                "empty_event_form",
                "madrs",
            ]:
                write_instrument_workbook(cleaned_dir / f"62822-{instrument}.xlsx")

            result = exclude_instruments(study_folder)
            log_text = (history_dir / "log.md").read_text(encoding="utf-8")

            self.assertEqual(
                [item.instrument for item in result.excluded],
                [
                    "survey_trigger",
                    "patient_email",
                    "patient_contact_log",
                    "participant_phone",
                    "participant_address",
                    "capacity_to_consent",
                    "empty_event_form",
                ],
            )
            self.assertFalse((cleaned_dir / "62822-survey_trigger.xlsx").exists())
            self.assertFalse((cleaned_dir / "62822-patient_email.xlsx").exists())
            self.assertFalse((cleaned_dir / "62822-patient_contact_log.xlsx").exists())
            self.assertFalse((cleaned_dir / "62822-participant_phone.xlsx").exists())
            self.assertFalse((cleaned_dir / "62822-participant_address.xlsx").exists())
            self.assertFalse((cleaned_dir / "62822-capacity_to_consent.xlsx").exists())
            self.assertFalse((cleaned_dir / "62822-empty_event_form.xlsx").exists())
            self.assertTrue((cleaned_dir / "62822-madrs.xlsx").exists())
            self.assertTrue((cleaned_dir / "excluded" / "62822-survey_trigger.xlsx").exists())
            self.assertTrue((cleaned_dir / "excluded" / "62822-patient_email.xlsx").exists())
            self.assertTrue((cleaned_dir / "excluded" / "62822-patient_contact_log.xlsx").exists())
            self.assertTrue((cleaned_dir / "excluded" / "62822-participant_phone.xlsx").exists())
            self.assertTrue((cleaned_dir / "excluded" / "62822-participant_address.xlsx").exists())
            self.assertTrue((cleaned_dir / "excluded" / "62822-capacity_to_consent.xlsx").exists())
            self.assertTrue((cleaned_dir / "excluded" / "62822-empty_event_form.xlsx").exists())
            self.assertIn("<summary><h1>Postprocess</h1></summary>", log_text)
            self.assertIn("## Exclude Instruments", log_text)
            self.assertIn("Excluded instrument workbooks: **7**", log_text)
            self.assertIn("admin category", log_text)
            self.assertIn("keyword:email", log_text)
            self.assertIn("keyword:contact", log_text)
            self.assertIn("keyword:phone", log_text)
            self.assertIn("keyword:address", log_text)
            self.assertIn("keyword:consent", log_text)
            self.assertIn("number_of_events=0", log_text)
            self.assertIn("patient_email", log_text)
            self.assertNotIn("62822-madrs.xlsx", log_text)

    def test_replaces_previous_exclusion_log_block(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study_folder = Path(tmpdir) / "study"
            cleaned_dir = study_folder / "data" / "cleaned" / "redcap"
            history_dir = study_folder / "histories" / "2026-06-18"
            write_dictionary(cleaned_dir / "dictionary.xlsx", [("survey_trigger", "Survey Trigger")])
            write_classification(history_dir / "instrument_classification.xlsx", [("survey_trigger", "Survey Trigger", "admin", 0.91)])
            write_instrument_workbook(cleaned_dir / "62822-survey_trigger.xlsx")
            (history_dir / "log.md").write_text(
                "# REDCap Workflow Log\n\n"
                "<details>\n"
                "<summary><h1>Postprocess</h1></summary>\n\n"
                "<!-- BEGIN EXCLUDE_INSTRUMENTS -->\n\n"
                "old block\n\n"
                "<!-- END EXCLUDE_INSTRUMENTS -->\n\n"
                "</details>\n",
                encoding="utf-8",
            )

            exclude_instruments(study_folder)
            log_text = (history_dir / "log.md").read_text(encoding="utf-8")

            self.assertEqual(log_text.count("<!-- BEGIN EXCLUDE_INSTRUMENTS -->"), 1)
            self.assertEqual(log_text.count("## Exclude Instruments"), 1)
            self.assertNotIn("old block", log_text)

    def test_excludes_do_not_use_deprecated_and_obsolete_variants(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study_folder = Path(tmpdir) / "study"
            cleaned_dir = study_folder / "data" / "cleaned" / "redcap"
            history_dir = study_folder / "histories" / "2026-06-18"
            history_dir.mkdir(parents=True)
            (history_dir / "log.md").write_text("# REDCap Workflow Log\n", encoding="utf-8")
            write_dictionary(
                cleaned_dir / "dictionary.xlsx",
                [
                    ("legacy_mood_do_not_use", "Mood Form DO NOT USE"),
                    ("old_scale_donot_use", "DONOT USE Legacy Scale"),
                    ("pirs_not_used", "PIRS-20 Insomnia Rating Scale (NOT USED)"),
                    ("unused_scale", "Unused Legacy Scale"),
                    ("deprecated_assessment", "Deprecated Assessment"),
                    ("obsolete_survey", "Obsolete Survey"),
                    ("retired_form", "Retired Form"),
                    ("active_madrs", "Active MADRS"),
                ],
            )
            write_classification(
                history_dir / "instrument_classification.xlsx",
                [
                    ("legacy_mood_do_not_use", "Mood Form DO NOT USE", "assessments", 0.88),
                    ("old_scale_donot_use", "DONOT USE Legacy Scale", "assessments", 0.88),
                    ("pirs_not_used", "PIRS-20 Insomnia Rating Scale (NOT USED)", "assessments", 0.88),
                    ("unused_scale", "Unused Legacy Scale", "assessments", 0.88),
                    ("deprecated_assessment", "Deprecated Assessment", "assessments", 0.88),
                    ("obsolete_survey", "Obsolete Survey", "assessments", 0.88),
                    ("retired_form", "Retired Form", "assessments", 0.88),
                    ("active_madrs", "Active MADRS", "assessments", 0.88),
                ],
            )
            for instrument in [
                "legacy_mood_do_not_use",
                "old_scale_donot_use",
                "pirs_not_used",
                "unused_scale",
                "deprecated_assessment",
                "obsolete_survey",
                "retired_form",
                "active_madrs",
            ]:
                write_instrument_workbook(cleaned_dir / f"62822-{instrument}.xlsx")

            result = exclude_instruments(study_folder)
            log_text = (history_dir / "log.md").read_text(encoding="utf-8")

            self.assertEqual(
                [item.instrument for item in result.excluded],
                [
                    "legacy_mood_do_not_use",
                    "old_scale_donot_use",
                    "pirs_not_used",
                    "unused_scale",
                    "deprecated_assessment",
                    "obsolete_survey",
                    "retired_form",
                ],
            )
            self.assertTrue((cleaned_dir / "62822-active_madrs.xlsx").exists())
            self.assertIn("keyword:do not use", log_text)
            self.assertIn("keyword:donot use", log_text)
            self.assertIn("keyword:not used", log_text)
            self.assertIn("keyword:unused", log_text)
            self.assertIn("keyword:deprecated", log_text)
            self.assertIn("keyword:obsolete", log_text)
            self.assertIn("keyword:retired", log_text)
            self.assertNotIn("62822-active_madrs.xlsx", log_text)


if __name__ == "__main__":
    unittest.main()
