import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.workflows.clean_redcap_instruments.steps.match_text_labels import match_event_labels, match_text_labels


class MatchTextLabelsTests(unittest.TestCase):
    def test_fills_event_label_by_row_aligned_raw_and_label_exports(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            raw_csv = root / "raw.csv"
            labels_csv = root / "labels.csv"
            dictionary = root / "dictionary.xlsx"
            audit = root / "audit.csv"

            with raw_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["record_id", "redcap_event_name"])
                writer.writerow(["s001", "baseline_arm_1"])
                writer.writerow(["s002", "baseline_arm_1"])
                writer.writerow(["s001", "study_day_1_arm_2"])

            with labels_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["Record ID", "Event Name"])
                writer.writerow(["s001", "Baseline (Arm 1: Screening)"])
                writer.writerow(["s002", "Baseline (Arm 1: Screening)"])
                writer.writerow(["s001", "Study Day 1 (Arm 2: Active Study)"])

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "event"
            worksheet.append(["arm", "order", "event_name", "event_label", "abbreviation"])
            worksheet.append(["1", "1", "baseline", "", ""])
            worksheet.append(["2", "1", "study_day_1", "", ""])
            workbook.save(dictionary)

            result = match_event_labels(raw_csv, labels_csv, dictionary, audit)

            workbook = load_workbook(dictionary)
            rows = list(workbook["event"].iter_rows(values_only=True))

        self.assertEqual(result.matched_events, 2)
        self.assertEqual(result.unmatched_events, 0)
        self.assertEqual(rows[1][3], "Baseline (Arm 1: Screening)")
        self.assertEqual(rows[2][3], "Study Day 1 (Arm 2: Active Study)")

    def test_suggests_event_abbreviation_from_visit_marker_in_event_label(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            raw_csv = root / "raw.csv"
            labels_csv = root / "labels.csv"
            dictionary = root / "dictionary.xlsx"
            audit = root / "audit.csv"

            with raw_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["record_id", "redcap_event_name"])
                writer.writerow(["s001", "baseline_arm_1"])
                writer.writerow(["s001", "followup_arm_1"])

            with labels_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["Record ID", "Event Name"])
                writer.writerow(["s001", "Baseline (Arm 1: Screening)"])
                writer.writerow(["s001", "Follow Up (Visit 20)"])

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "event"
            worksheet.append(["arm", "order", "event_name", "event_label", "abbreviation"])
            worksheet.append(["1", "1", "baseline", "", ""])
            worksheet.append(["1", "2", "followup", "", ""])
            workbook.save(dictionary)

            match_event_labels(raw_csv, labels_csv, dictionary, audit)

            workbook = load_workbook(dictionary)
            rows = list(workbook["event"].iter_rows(values_only=True))

        self.assertEqual(rows[1][3], "Baseline (Arm 1: Screening)")
        self.assertIsNone(rows[1][4])
        self.assertEqual(rows[2][3], "Follow Up (Visit 20)")
        self.assertEqual(rows[2][4], "V20")

    def test_visit_abbreviation_suggestion_does_not_overwrite_manual_abbreviation(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            raw_csv = root / "raw.csv"
            labels_csv = root / "labels.csv"
            dictionary = root / "dictionary.xlsx"
            audit = root / "audit.csv"

            with raw_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["record_id", "redcap_event_name"])
                writer.writerow(["s001", "followup_arm_1"])

            with labels_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["Record ID", "Event Name"])
                writer.writerow(["s001", "Follow Up (Visit 20)"])

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "event"
            worksheet.append(["arm", "order", "event_name", "event_label", "abbreviation"])
            worksheet.append(["1", "1", "followup", "", "FU20"])
            workbook.save(dictionary)

            match_event_labels(raw_csv, labels_csv, dictionary, audit)

            workbook = load_workbook(dictionary)
            rows = list(workbook["event"].iter_rows(values_only=True))

        self.assertEqual(rows[1][3], "Follow Up (Visit 20)")
        self.assertEqual(rows[1][4], "FU20")

    def test_fills_event_labels_from_labels_csv_and_keeps_longer_auto_label_than_codebook(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            raw_csv = root / "raw.csv"
            labels_csv = root / "labels.csv"
            codebook = root / "codebook.xlsx"
            dictionary = root / "dictionary.xlsx"
            event_audit = root / "event_audit.csv"
            instrument_audit = root / "instrument_audit.csv"

            with raw_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["record_id", "redcap_event_name"])
                writer.writerow(["s001", "baseline_arm_1"])

            with labels_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["Record ID", "Event Name"])
                writer.writerow(["s001", "Baseline (Arm 1: Screening)"])

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "event"
            worksheet.append(["arm", "order", "event_name", "event_label", "abbreviation"])
            worksheet.append(["1", "1", "baseline", "", ""])
            worksheet = workbook.create_sheet("instrument")
            worksheet.append(["instrument", "instrument_label", "abbreviation", "number_of_events", "events", "number_of_columns"])
            workbook.save(dictionary)

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "event-codebook"
            worksheet.append(["Event Name", "Unique event name"])
            worksheet.append(["Baseline", "baseline_arm_1"])
            worksheet = workbook.create_sheet("instrument-codebook")
            worksheet.append(["Instrument", "Form Name", "Events"])
            workbook.save(codebook)

            result = match_text_labels(
                raw_csv=raw_csv,
                labels_csv=labels_csv,
                dictionary_path=dictionary,
                event_audit_path=event_audit,
                instrument_audit_path=instrument_audit,
                codebook_path=codebook,
                event_codebook_sheet="event-codebook",
                instrument_codebook_sheet="instrument-codebook",
            )

            workbook = load_workbook(dictionary)
            event_rows = list(workbook["event"].iter_rows(values_only=True))

        self.assertEqual(result.matched_events, 1)
        self.assertEqual(result.event_codebook_mismatches, 0)
        self.assertEqual(event_rows[1][3], "Baseline (Arm 1: Screening)")

    def test_fills_instrument_labels_from_codebook_form_name_with_empty_form_name_fallback(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            raw_csv = root / "raw.csv"
            labels_csv = root / "labels.csv"
            codebook = root / "codebook.xlsx"
            dictionary = root / "dictionary.xlsx"
            event_audit = root / "event_audit.csv"
            instrument_audit = root / "instrument_audit.csv"

            with raw_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["record_id", "redcap_event_name"])
                writer.writerow(["s001", "baseline_arm_1"])

            with labels_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["Record ID", "Event Name"])
                writer.writerow(["s001", "Baseline (Arm 1: Screening)"])

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "event"
            worksheet.append(["arm", "order", "event_name", "event_label", "abbreviation"])
            worksheet.append(["1", "1", "baseline", "", ""])
            worksheet = workbook.create_sheet("instrument")
            worksheet.append(["instrument", "instrument_label", "abbreviation", "number_of_events", "events", "number_of_columns"])
            worksheet.append(["demographics", "", "", "1", "Arm 1: baseline", "3"])
            worksheet.append(["madrs", "", "", "1", "Arm 1: baseline", "4"])
            worksheet.append(["auto_extra", "", "", "1", "Arm 1: baseline", "2"])
            workbook.save(dictionary)

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "event-codebook"
            worksheet.append(["Event Name", "Unique event name"])
            worksheet.append(["Baseline", "baseline_arm_1"])
            worksheet = workbook.create_sheet("instrument-codebook")
            worksheet.append(["Instrument", "Form Name", "Events"])
            worksheet.append(["Demographics", "demographics", "baseline_arm_1"])
            worksheet.append(["madrs", "", "baseline_arm_1"])
            workbook.save(codebook)

            result = match_text_labels(
                raw_csv=raw_csv,
                labels_csv=labels_csv,
                dictionary_path=dictionary,
                event_audit_path=event_audit,
                instrument_audit_path=instrument_audit,
                codebook_path=codebook,
                event_codebook_sheet="event-codebook",
                instrument_codebook_sheet="instrument-codebook",
            )

            workbook = load_workbook(dictionary)
            instrument_rows = list(workbook["instrument"].iter_rows(values_only=True))

        self.assertEqual(result.matched_instruments, 2)
        self.assertEqual(result.unmatched_instruments, 1)
        self.assertEqual(instrument_rows[1][1], "Demographics")
        self.assertEqual(instrument_rows[2][1], "madrs")
        self.assertIsNone(instrument_rows[3][1])

    def test_updates_dictionary_event_and_instrument_sections_in_log(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            raw_csv = root / "raw.csv"
            labels_csv = root / "labels.csv"
            codebook = root / "codebook.xlsx"
            dictionary = root / "dictionary.xlsx"
            event_audit = root / "event_audit.csv"
            instrument_audit = root / "instrument_audit.csv"
            log_path = root / "log.md"

            with raw_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["record_id", "redcap_event_name"])
                writer.writerow(["s001", "baseline_arm_1"])

            with labels_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["Record ID", "Event Name"])
                writer.writerow(["s001", "Baseline (Arm 1: Screening)"])

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "event"
            worksheet.append(["arm", "order", "event_name", "event_label", "abbreviation"])
            worksheet.append(["1", "1", "baseline", "", ""])
            worksheet = workbook.create_sheet("instrument")
            worksheet.append(["instrument", "instrument_label", "abbreviation", "number_of_events", "events", "number_of_columns"])
            worksheet.append(["demographics", "", "", "1", "Arm 1: baseline", "3"])
            workbook.save(dictionary)

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "event-codebook"
            worksheet.append(["Event Name", "Unique event name"])
            worksheet.append(["Baseline", "baseline_arm_1"])
            worksheet = workbook.create_sheet("instrument-codebook")
            worksheet.append(["Instrument", "Form Name", "Events"])
            worksheet.append(["Demographics", "demographics", "baseline_arm_1"])
            workbook.save(codebook)

            log_path.write_text(
                "# REDCap Workflow Log\n\n"
                "<details>\n<summary><h1>Dictionary</h1></summary>\n\n"
                "## Event\n\n"
                "script: `scripts/workflows/clean_redcap_instruments/steps/create_redcap_index_dictionary.py`\n\n"
                "Events indexed: **1**\n\n"
                "## Instrument\n\n"
                "script: `scripts/workflows/clean_redcap_instruments/steps/create_redcap_index_dictionary.py`\n\n"
                "Instruments indexed: **1**\n\n"
                "</details>\n",
                encoding="utf-8",
            )

            for _ in range(2):
                match_text_labels(
                    raw_csv=raw_csv,
                    labels_csv=labels_csv,
                    dictionary_path=dictionary,
                    event_audit_path=event_audit,
                    instrument_audit_path=instrument_audit,
                    codebook_path=codebook,
                    event_codebook_sheet="event-codebook",
                    instrument_codebook_sheet="instrument-codebook",
                    log_path=log_path,
                )

            log_text = log_path.read_text(encoding="utf-8")

        self.assertIn("### Event Label Matching", log_text)
        self.assertIn("Event labels filled: **1** / 1", log_text)
        self.assertIn(str(event_audit), log_text)
        self.assertIn("### Instrument Label Matching", log_text)
        self.assertIn("Instrument labels filled: **1** / 1", log_text)
        self.assertIn(str(instrument_audit), log_text)
        self.assertEqual(log_text.count("### Event Label Matching"), 1)
        self.assertEqual(log_text.count("### Instrument Label Matching"), 1)


if __name__ == "__main__":
    unittest.main()
