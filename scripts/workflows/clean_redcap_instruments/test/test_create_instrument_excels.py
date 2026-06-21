import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.workflows.clean_redcap_instruments.steps.create_instrument_excels import create_instrument_excels, row_is_sparse_or_incomplete


class CreateInstrumentExcelsTests(unittest.TestCase):
    def test_incomplete_status_does_not_drop_rows_with_abundant_meaningful_entries(self):
        headers = [
            "IRB",
            "subid",
            "arm",
            "visit",
            "date",
            "record_id",
            "redcap_event_name",
            "scan_date",
            "operator___1",
            "operator___2",
            "session_name",
            "setup_complete",
            "pregnancy_test",
            "pregnancy_result",
            "metal_clearance",
            "mri_checklist_complete",
        ]
        row = [
            "53879",
            "s004",
            "2",
            "V15",
            "2021-06-14",
            "53879_s004",
            "xover_baseline_vis_arm_2",
            "2021-06-14 07:20",
            "Checked",
            "Unchecked",
            "s004_V15",
            "Yes",
            "Yes",
            "Negative (not pregnant)",
            "Yes",
            "Incomplete",
        ]

        self.assertFalse(row_is_sparse_or_incomplete(row, headers))

    def test_incomplete_status_still_drops_rows_without_enough_meaningful_entries(self):
        headers = [
            "IRB",
            "subid",
            "arm",
            "visit",
            "date",
            "record_id",
            "redcap_event_name",
            "scan_date",
            "operator___1",
            "operator___2",
            "mri_checklist_complete",
        ]
        row = [
            "53879",
            "s004",
            "2",
            "V15",
            "",
            "53879_s004",
            "xover_baseline_vis_arm_2",
            "",
            "Unchecked",
            "Unchecked",
            "Incomplete",
        ]

        self.assertTrue(row_is_sparse_or_incomplete(row, headers))

    def test_creates_raw_cleaned_dictionary_and_excluded_sheets(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            raw_csv = root / "raw.csv"
            labels_csv = root / "labels.csv"
            dictionary = root / "dictionary.xlsx"
            instrument_columns = root / "instrument_columns.csv"
            out_dir = root / "out"

            with raw_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(
                    [
                        "record_id",
                        "redcap_event_name",
                        "demo_date",
                        "patient_email",
                        "demo_score",
                        "demo_empty",
                        "demographics_complete",
                        "madrs_date",
                        "madrs_total",
                        "madrs_complete",
                    ]
                )
                writer.writerow(["58807_s001", "baseline_arm_1", "2026-01-02 09:30", "person@example.com", "bad\x10text", "", "2", "", "", ""])
                writer.writerow(["MDD_001", "baseline_arm_1", "2026-01-03", "other@example.com", "9", "", "2", "", "", ""])
                writer.writerow(["58807_s001", "followup_arm_1", "", "", "", "", "", "01/04/2026", "12", "2"])

            with labels_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(
                    [
                        "Record ID",
                        "Event Name",
                        "Demo Date",
                        "Patient Email",
                        "Demo Score",
                        "Demo Empty",
                        "Demographics Complete?",
                        "MADRS Date",
                        "MADRS Total",
                        "MADRS Complete?",
                    ]
                )
                writer.writerow(
                    [
                        "58807_s001",
                        "Baseline (Arm 1: Screening)",
                        "Jan 2 2026",
                        "person@example.com",
                        "Bad Text Label",
                        "",
                        "Complete",
                        "",
                        "",
                        "",
                    ]
                )
                writer.writerow(
                    [
                        "MDD_001",
                        "Baseline (Arm 1: Screening)",
                        "Jan 3 2026",
                        "other@example.com",
                        "Nine Label",
                        "",
                        "Complete",
                        "",
                        "",
                        "",
                    ]
                )
                writer.writerow(
                    [
                        "58807_s001",
                        "Followup (Arm 1: Screening)",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "Jan 4 2026",
                        "Twelve Label",
                        "Complete",
                    ]
                )

            with instrument_columns.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["instrument_order", "instrument_key", "column_index", "column_name", "is_stop_signal"])
                writer.writerow(["1", "demographics", "2", "demo_date", "False"])
                writer.writerow(["1", "demographics", "3", "patient_email", "False"])
                writer.writerow(["1", "demographics", "4", "demo_score", "False"])
                writer.writerow(["1", "demographics", "5", "demo_empty", "False"])
                writer.writerow(["1", "demographics", "6", "demographics_complete", "True"])
                writer.writerow(["2", "madrs", "7", "madrs_date", "False"])
                writer.writerow(["2", "madrs", "8", "madrs_total", "False"])
                writer.writerow(["2", "madrs", "9", "madrs_complete", "True"])

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "subject_id"
            worksheet.append(["IRB", "subid", "standardized", "raw_entries"])
            worksheet.append(["58807", "s001", "58807_s001", "58807_s001"])
            worksheet = workbook.create_sheet("event")
            worksheet.append(["arm", "order", "event_name", "event_label", "abbreviation"])
            worksheet.append(["1", "1", "baseline", "Baseline (Arm 1: Screening)", "bl"])
            worksheet.append(["1", "2", "followup", "Followup (Arm 1: Screening)", ""])
            worksheet = workbook.create_sheet("instrument")
            worksheet.append(["instrument", "instrument_label", "abbreviation", "number_of_events", "events", "number_of_columns"])
            worksheet.append(["demographics", "Demographics", "", "1", "Arm 1: baseline", "3"])
            worksheet.append(["madrs", "MADRS", "", "1", "Arm 1: followup", "3"])
            workbook.save(dictionary)

            result = create_instrument_excels(
                input_path=raw_csv,
                labels_path=labels_csv,
                dictionary_path=dictionary,
                instrument_columns_path=instrument_columns,
                irb="58807",
                out_dir=out_dir,
            )

            demo_path = result.output_paths["demographics"]
            madrs_path = result.output_paths["madrs"]
            demo_wb = load_workbook(demo_path, data_only=True)
            madrs_wb = load_workbook(madrs_path, data_only=True)
            demo_cleaned = list(demo_wb["cleaned"].iter_rows(values_only=True))
            demo_raw = list(demo_wb["raw"].iter_rows(values_only=True))
            demo_raw_labels = list(demo_wb["raw_labels"].iter_rows(values_only=True))
            demo_excluded = list(demo_wb["excluded_rows"].iter_rows(values_only=True))
            demo_dictionary = list(demo_wb["column_variable_dictionary"].iter_rows(values_only=True))
            madrs_cleaned = list(madrs_wb["cleaned"].iter_rows(values_only=True))

        self.assertEqual(set(result.output_paths), {"demographics", "madrs"})
        self.assertTrue(demo_path.parent.match("*/cleaned/redcap"))
        self.assertEqual(
            demo_wb.sheetnames,
            ["raw", "raw_labels", "cleaned", "timepoint_dictionary", "column_variable_dictionary", "excluded_rows"],
        )
        self.assertEqual(
            demo_raw[0],
                ("record_id", "redcap_event_name", "demo_date", "demo_score", "demo_empty", "demographics_complete"),
        )
        self.assertNotIn("patient_email", demo_raw[0])
        self.assertEqual(demo_raw[1][3], "badtext")
        self.assertEqual(
            demo_raw_labels[0],
                ("Record ID", "Event Name", "Demo Date", "Demo Score", "Demo Empty", "Demographics Complete?"),
        )
        self.assertNotIn("Patient Email", demo_raw_labels[0])
        self.assertEqual(demo_raw_labels[1][1], "Baseline (Arm 1: Screening)")
        self.assertEqual(demo_raw_labels[1][3], "Bad Text Label")
        self.assertEqual(demo_raw_labels[2][0], "MDD_001")
        self.assertEqual(
            demo_cleaned[0],
            (
                "IRB",
                "subid",
                "arm",
                "visit",
                "date",
                "record_id",
                "redcap_event_name",
                "demo_date",
                "demo_score",
                "demographics_complete",
            ),
        )
        self.assertNotIn("patient_email", demo_cleaned[0])
        self.assertEqual(
            demo_cleaned[1][:7],
            ("58807", "s001", "1", "bl", "2026-01-02", "58807_s001", "Baseline (Arm 1: Screening)"),
        )
        self.assertEqual(demo_cleaned[1][7], "Jan 2 2026")
        self.assertEqual(demo_cleaned[1][8], "Bad Text Label")
        self.assertEqual(demo_cleaned[1][9], "Complete")
        self.assertEqual(len(demo_cleaned), 2)
        self.assertEqual(demo_excluded[0], demo_raw[0])
        self.assertEqual(demo_excluded[1][0], "MDD_001")
        self.assertEqual(demo_excluded[2][0], "58807_s001")
        self.assertEqual(demo_excluded[2][1], "followup_arm_1")
        self.assertEqual(demo_dictionary[1], ("record_id", "Record ID", None, "True"))
        self.assertEqual(demo_dictionary[2], ("redcap_event_name", "Event Name", None, "True"))
        self.assertEqual(demo_dictionary[3], ("demo_date", "Demo Date", None, "True"))
        self.assertEqual(demo_dictionary[4], ("patient_email", "Patient Email", None, "False"))
        self.assertEqual(demo_dictionary[6], ("demo_empty", "Demo Empty", None, "False"))
        self.assertEqual(
            madrs_cleaned[1][:7],
            ("58807", "s001", "1", "followup", "2026-01-04", "58807_s001", "Followup (Arm 1: Screening)"),
        )
        self.assertEqual(madrs_cleaned[1][7], "Jan 4 2026")
        self.assertEqual(madrs_cleaned[1][8], "Twelve Label")
        self.assertEqual(madrs_cleaned[1][9], "Complete")
        self.assertEqual(len(madrs_cleaned), 2)

    def test_skips_excluded_instruments_when_creating_workbooks(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            raw_csv = root / "raw.csv"
            labels_csv = root / "labels.csv"
            dictionary = root / "dictionary.xlsx"
            instrument_columns = root / "instrument_columns.csv"
            out_dir = root / "out"

            with raw_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["record_id", "redcap_event_name", "demo_value", "demo_complete", "admin_value", "admin_complete"])
                writer.writerow(["58807_s001", "baseline_arm_1", "A", "2", "B", "2"])

            with labels_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["Record ID", "Event Name", "Demo Value", "Demo Complete?", "Admin Value", "Admin Complete?"])
                writer.writerow(["58807_s001", "Baseline (Arm 1)", "A Label", "Complete", "B Label", "Complete"])

            with instrument_columns.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["instrument_order", "instrument_key", "column_index", "column_name", "is_stop_signal"])
                writer.writerow(["1", "demographics", "2", "demo_value", "False"])
                writer.writerow(["1", "demographics", "3", "demo_complete", "True"])
                writer.writerow(["2", "survey_trigger", "4", "admin_value", "False"])
                writer.writerow(["2", "survey_trigger", "5", "admin_complete", "True"])

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "subject_id"
            worksheet.append(["IRB", "subid", "standardized", "raw_entries"])
            worksheet.append(["58807", "s001", "58807_s001", "58807_s001"])
            worksheet = workbook.create_sheet("event")
            worksheet.append(["arm", "order", "event_name", "event_label", "abbreviation"])
            worksheet.append(["1", "1", "baseline", "Baseline (Arm 1)", "bl"])
            worksheet = workbook.create_sheet("instrument")
            worksheet.append(["instrument", "instrument_label", "abbreviation", "number_of_events", "events", "number_of_columns"])
            worksheet.append(["demographics", "Demographics", "", "1", "Arm 1: baseline", "2"])
            worksheet.append(["survey_trigger", "Survey Trigger", "", "1", "Arm 1: baseline", "2"])
            workbook.save(dictionary)

            result = create_instrument_excels(
                input_path=raw_csv,
                labels_path=labels_csv,
                dictionary_path=dictionary,
                instrument_columns_path=instrument_columns,
                irb="58807",
                out_dir=out_dir,
                excluded_instruments={"survey_trigger"},
            )
            demographics_exists = (out_dir / "cleaned" / "redcap" / "58807-demographics.xlsx").exists()
            survey_trigger_exists = (out_dir / "cleaned" / "redcap" / "58807-survey_trigger.xlsx").exists()

        self.assertEqual(set(result.output_paths), {"demographics"})
        self.assertTrue(demographics_exists)
        self.assertFalse(survey_trigger_exists)


if __name__ == "__main__":
    unittest.main()
