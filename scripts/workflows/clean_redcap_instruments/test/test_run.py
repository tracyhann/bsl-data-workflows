import csv
import tempfile
import unittest
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook

from scripts.workflows.clean_redcap_instruments.run import discover_column_name_if_default_index_misses, run_study_workflow


def add_malformed_shared_strings(path: Path) -> None:
    temp_path = path.with_name(f"{path.stem}_bad{path.suffix}")
    with ZipFile(path, "r") as source, ZipFile(temp_path, "w", compression=ZIP_DEFLATED) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "xl/workbook.xml":
                data = data.replace(b">", b'><workbookPr codeName="&quot" />', 1)
            target.writestr(item, data)
    temp_path.replace(path)


class StudyWorkflowRunnerTests(unittest.TestCase):
    def test_default_column_resolution_falls_back_to_known_redcap_names(self):
        headers = ["site", "misc", "Record ID", "Event Name"]

        record_name = discover_column_name_if_default_index_misses(
            headers,
            column_name=None,
            column_index=0,
            default_index=0,
            known_names={"record_id", "record id", "recordid"},
        )
        event_name = discover_column_name_if_default_index_misses(
            headers,
            column_name=None,
            column_index=1,
            default_index=1,
            known_names={"redcap_event_name", "event_name", "event name"},
        )
        explicit_index = discover_column_name_if_default_index_misses(
            headers,
            column_name=None,
            column_index=2,
            default_index=0,
            known_names={"record id"},
        )

        self.assertEqual(record_name, "Record ID")
        self.assertEqual(event_name, "Event Name")
        self.assertIsNone(explicit_index)

    def test_runner_creates_study_folder_layout_and_final_cleaned_outputs(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            source_dir = root / "source"
            source_dir.mkdir()
            study_folder = root / "study"
            data_csv = source_dir / "Study_DATA.csv"
            labels_csv = source_dir / "Study_DATA_LABELS.csv"
            codebook = source_dir / "Study_Codebook.xlsx"

            with data_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(
                    [
                        "record_id",
                        "redcap_event_name",
                        "demo_date",
                        "demographics_complete",
                        "patient_email",
                        "patient_email_complete",
                    ]
                )
                writer.writerow(["58807_s001", "baseline_arm_1", "2026-01-01", "2", "person@example.com", "2"])
                writer.writerow(["MDD_001", "baseline_arm_1", "2026-01-02", "2", "other@example.com", "2"])

            with labels_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["Record ID", "Event Name", "Demo Date", "Complete?", "Patient Email", "Complete?"])
                writer.writerow(["58807_s001", "Baseline (Arm 1: Screening)", "", "", "person@example.com", "Complete"])
                writer.writerow(["MDD_001", "Baseline (Arm 1: Screening)", "", "", "other@example.com", "Complete"])

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "58807-instrument"
            worksheet.append(["Instrument", "Form Name", "Events"])
            worksheet.append(["Demographics", "demographics", "baseline_arm_1"])
            worksheet.append(["Patient Email", "patient_email", "baseline_arm_1"])
            worksheet = workbook.create_sheet("58807-event")
            worksheet.append(["Event Name", "Unique event name", "Event ID"])
            worksheet.append(["Baseline", "baseline_arm_1", "1"])
            workbook.save(codebook)

            result = run_study_workflow(
                study_folder=study_folder,
                data_csv=data_csv,
                labels_csv=labels_csv,
                codebook=codebook,
                irb="58807",
                instrument_codebook_sheet="58807-instrument",
                event_codebook_sheet="58807-event",
                run_date="2026-06-17",
            )

            raw_dir = study_folder / "data" / "raw_exports" / "redcap" / "all"
            cleaned_root = study_folder / "data" / "cleaned"
            cleaned_dir = cleaned_root / "redcap"
            history_dir = study_folder / "histories" / "2026-06-17"

            self.assertEqual(result.study_folder, study_folder)
            self.assertTrue((raw_dir / data_csv.name).exists())
            self.assertTrue((raw_dir / labels_csv.name).exists())
            self.assertTrue((raw_dir / codebook.name).exists())
            self.assertTrue((history_dir / "log.md").exists())
            self.assertTrue((history_dir / "Study_DATA_instrument_columns.csv").exists())
            self.assertTrue((history_dir / "instrument_classification.xlsx").exists())
            self.assertEqual(result.dictionary_path, cleaned_root / "dictionary.xlsx")
            self.assertTrue(result.dictionary_path.exists())
            self.assertFalse((cleaned_dir / "dictionary.xlsx").exists())
            self.assertIn("demographics", result.instrument_workbooks)
            self.assertNotIn("patient_email", result.instrument_workbooks)
            self.assertEqual(result.instrument_workbooks["demographics"], cleaned_root / "subjects" / "58807-demographics.xlsx")
            self.assertTrue((cleaned_root / "subjects" / "58807-demographics.xlsx").exists())
            self.assertFalse((cleaned_dir / "58807-demographics.xlsx").exists())
            self.assertFalse((cleaned_dir / "58807-patient_email.xlsx").exists())
            self.assertFalse(cleaned_dir.exists())

            log_text = (history_dir / "log.md").read_text(encoding="utf-8")
            self.assertIn("## Workflow Arguments", log_text)
            self.assertIn("| stage | all |", log_text)
            self.assertIn("| data_csv |", log_text)
            self.assertIn("| labels_csv |", log_text)
            self.assertIn("| codebook |", log_text)
            self.assertIn("| irb | 58807 |", log_text)
            self.assertIn("<summary><h1>Clean</h1></summary>", log_text)
            self.assertIn("## Instrument Cleaning", log_text)
            self.assertIn("| drop_fully_empty_rows | True |", log_text)
            self.assertIn("| drop_fully_empty_columns | True |", log_text)
            self.assertIn("## Exclude Instruments", log_text)
            self.assertIn("instrument excluded:", log_text)
            self.assertIn("patient_email", log_text)
            self.assertIn("keyword:email", log_text)
            self.assertIn("## Drop Stale Instruments", log_text)
            self.assertIn("## Sort Instuments", log_text)
            self.assertLess(log_text.index("<summary><h1>Clean</h1></summary>"), log_text.index("<summary><h1>Postprocess</h1></summary>"))

            dictionary = load_workbook(result.dictionary_path, data_only=True)
            self.assertEqual(dictionary["event"].cell(2, 4).value, "Baseline (Arm 1: Screening)")
            self.assertEqual(dictionary["instrument"].cell(2, 2).value, "Demographics")

            instrument_wb = load_workbook(result.instrument_workbooks["demographics"], data_only=True)
            cleaned_headers = [instrument_wb["cleaned"].cell(1, column).value for column in range(1, 9)]
            self.assertEqual(
                cleaned_headers,
                ["IRB", "subid", "arm", "visit", "date", "record_id", "redcap_event_name", "demo_date"],
            )
            self.assertEqual(instrument_wb["cleaned"].cell(2, 1).value, "58807")
            self.assertEqual(instrument_wb["cleaned"].cell(2, 3).value, "1")
            self.assertEqual(instrument_wb["excluded_rows"].cell(2, 1).value, "MDD_001")

    def test_runner_supports_discovery_then_clean_stages_with_manual_dictionary_edits(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            source_dir = root / "source"
            source_dir.mkdir()
            study_folder = root / "study"
            data_csv = source_dir / "Study_DATA.csv"
            labels_csv = source_dir / "Study_DATA_LABELS.csv"
            codebook = source_dir / "Study_Codebook.xlsx"

            with data_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["record_id", "redcap_event_name", "demo_date", "demographics_complete"])
                writer.writerow(["58807_s001", "baseline_arm_1", "2026-01-01", "2"])

            with labels_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["Record ID", "Event Name", "Demo Date", "Complete?"])
                writer.writerow(["58807_s001", "Baseline (Arm 1: Screening)", "January 1, 2026", "Complete"])

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "58807-instrument"
            worksheet.append(["Instrument", "Form Name", "Events"])
            worksheet.append(["Demographics", "demographics", "baseline_arm_1"])
            worksheet = workbook.create_sheet("58807-event")
            worksheet.append(["Event Name", "Unique event name", "Event ID"])
            worksheet.append(["Baseline", "baseline_arm_1", "1"])
            workbook.save(codebook)

            discovery_result = run_study_workflow(
                study_folder=study_folder,
                data_csv=data_csv,
                labels_csv=labels_csv,
                codebook=codebook,
                irb="58807",
                instrument_codebook_sheet="58807-instrument",
                event_codebook_sheet="58807-event",
                run_date="2026-06-17",
                stage="discovery",
            )

            cleaned_dir = study_folder / "data" / "cleaned" / "redcap"
            self.assertEqual(discovery_result.instrument_workbooks, {})
            self.assertTrue(discovery_result.dictionary_path.exists())
            self.assertFalse((cleaned_dir / "58807-demographics.xlsx").exists())
            discovery_log = discovery_result.log_path.read_text(encoding="utf-8")
            self.assertIn("| stage | discovery |", discovery_log)
            self.assertNotIn("<summary><h1>Clean</h1></summary>", discovery_log)

            dictionary = load_workbook(discovery_result.dictionary_path)
            dictionary["event"].cell(2, 5).value = "bl"
            dictionary.save(discovery_result.dictionary_path)

            clean_result = run_study_workflow(
                study_folder=study_folder,
                irb="58807",
                run_date="2026-06-17",
                stage="clean",
            )

            workbook = load_workbook(clean_result.instrument_workbooks["demographics"], data_only=True)
            cleaned = workbook["cleaned"]
            clean_log = clean_result.log_path.read_text(encoding="utf-8")

        self.assertEqual(clean_result.dictionary_path, cleaned_dir / "dictionary.xlsx")
        self.assertEqual(cleaned.cell(2, 3).value, "1")
        self.assertEqual(cleaned.cell(2, 4).value, "bl")
        self.assertEqual(cleaned.cell(2, 8).value, "January 1, 2026")
        self.assertIn("| stage | clean |", clean_log)
        self.assertIn("## Instrument Cleaning", clean_log)
        self.assertIn("58807-demographics.xlsx", clean_log)

    def test_runner_repairs_malformed_codebook_xml_before_discovery(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            source_dir = root / "source"
            source_dir.mkdir()
            study_folder = root / "study"
            data_csv = source_dir / "Study_DATA.csv"
            labels_csv = source_dir / "Study_DATA_LABELS.csv"
            codebook = source_dir / "Study_Codebook.xlsx"

            with data_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["record_id", "redcap_event_name", "demo_date", "demographics_complete"])
                writer.writerow(["58807_s001", "baseline_arm_1", "2026-01-01", "2"])

            with labels_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["Record ID", "Event Name", "Demo Date", "Complete?"])
                writer.writerow(["58807_s001", "Baseline (Arm 1: Screening)", "January 1, 2026", "Complete"])

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "58807-instrument"
            worksheet.append(["Instrument", "Form Name", "Events"])
            worksheet.append(["Demographics", "demographics", "baseline_arm_1"])
            worksheet = workbook.create_sheet("58807-event")
            worksheet.append(["Event Name", "Unique event name", "Event ID"])
            worksheet.append(["Baseline", "baseline_arm_1", "1"])
            workbook.save(codebook)
            add_malformed_shared_strings(codebook)

            result = run_study_workflow(
                study_folder=study_folder,
                data_csv=data_csv,
                labels_csv=labels_csv,
                codebook=codebook,
                irb="58807",
                run_date="2026-06-17",
                stage="discovery",
            )

            self.assertEqual(result.codebook.name, "Study_Codebook_xml_repaired.xlsx")
            self.assertTrue(result.codebook.exists())
            self.assertTrue(result.dictionary_path.exists())
            log_text = result.log_path.read_text(encoding="utf-8")
            self.assertIn("Study_Codebook_xml_repaired.xlsx", log_text)

    def test_runner_supports_postprocess_stage_after_clean(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            source_dir = root / "source"
            source_dir.mkdir()
            study_folder = root / "study"
            data_csv = source_dir / "Study_DATA.csv"
            labels_csv = source_dir / "Study_DATA_LABELS.csv"
            codebook = source_dir / "Study_Codebook.xlsx"

            with data_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["record_id", "redcap_event_name", "demo_date", "demographics_complete"])
                writer.writerow(["58807_s001", "baseline_arm_1", "2026-01-01", "2"])

            with labels_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["Record ID", "Event Name", "Demo Date", "Complete?"])
                writer.writerow(["58807_s001", "Baseline (Arm 1: Screening)", "January 1, 2026", "Complete"])

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "58807-instrument"
            worksheet.append(["Instrument", "Form Name", "Events"])
            worksheet.append(["Demographics", "demographics", "baseline_arm_1"])
            worksheet = workbook.create_sheet("58807-event")
            worksheet.append(["Event Name", "Unique event name", "Event ID"])
            worksheet.append(["Baseline", "baseline_arm_1", "1"])
            workbook.save(codebook)

            run_study_workflow(
                study_folder=study_folder,
                data_csv=data_csv,
                labels_csv=labels_csv,
                codebook=codebook,
                irb="58807",
                instrument_codebook_sheet="58807-instrument",
                event_codebook_sheet="58807-event",
                run_date="2026-06-17",
                stage="discovery",
            )
            clean_result = run_study_workflow(
                study_folder=study_folder,
                irb="58807",
                run_date="2026-06-17",
                stage="clean",
            )
            self.assertTrue((study_folder / "data" / "cleaned" / "redcap" / "58807-demographics.xlsx").exists())

            postprocess_result = run_study_workflow(
                study_folder=study_folder,
                irb="58807",
                run_date="2026-06-17",
                stage="postprocess",
            )
            log_text = postprocess_result.log_path.read_text(encoding="utf-8")

            self.assertEqual(
                clean_result.instrument_workbooks["demographics"],
                study_folder / "data" / "cleaned" / "redcap" / "58807-demographics.xlsx",
            )
            self.assertEqual(
                postprocess_result.instrument_workbooks["demographics"],
                study_folder / "data" / "cleaned" / "subjects" / "58807-demographics.xlsx",
            )
            self.assertFalse((study_folder / "data" / "cleaned" / "redcap" / "58807-demographics.xlsx").exists())
            self.assertTrue((study_folder / "data" / "cleaned" / "subjects" / "58807-demographics.xlsx").exists())
            self.assertEqual(postprocess_result.dictionary_path, study_folder / "data" / "cleaned" / "dictionary.xlsx")
            self.assertTrue(postprocess_result.dictionary_path.exists())
            self.assertFalse((study_folder / "data" / "cleaned" / "redcap").exists())
            self.assertIn("## Drop Stale Instruments", log_text)
            self.assertIn("## Sort Instuments", log_text)


if __name__ == "__main__":
    unittest.main()
