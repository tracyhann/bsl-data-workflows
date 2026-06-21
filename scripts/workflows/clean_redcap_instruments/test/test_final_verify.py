import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.workflows.clean_redcap_instruments.steps.final_verify import FinalVerificationError, final_verify


def write_dictionary(path: Path, rows: list[tuple[str, str, str, str]]) -> None:
    workbook = Workbook()
    subject = workbook.active
    subject.title = "subject_id"
    subject.append(["IRB", "subid", "standardized", "raw_entries"])
    subject.append(["58807", "s001", "58807_s001", "58807_s001"])

    event = workbook.create_sheet("event")
    event.append(["arm", "order", "event_name", "event_label", "abbreviation"])
    event.append(["1", "1", "baseline", "Baseline (Arm 1: Screening)", ""])

    instrument = workbook.create_sheet("instrument")
    instrument.append(["instrument", "instrument_label", "abbreviation", "number_of_events", "events", "number_of_columns"])
    for row in rows:
        name, label, abbreviation, number_of_events = row
        instrument.append([name, label, abbreviation, number_of_events, "baseline", "3"])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_classification(path: Path, rows: list[tuple[str, str, str, float]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "instrument_classification"
    worksheet.append(["instrument_name", "instrument_label", "class", "confidence"])
    for row in rows:
        worksheet.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_instrument_workbook(path: Path, instrument_column: str = "madrs_total", keep_sensitive: bool = False) -> None:
    workbook = Workbook()
    raw = workbook.active
    raw.title = "raw"
    raw.append(["record_id", "redcap_event_name", instrument_column])
    raw.append(["58807_s001", "baseline_arm_1", "10"])

    raw_labels = workbook.create_sheet("raw_labels")
    raw_labels.append(["Record ID", "Event Name", instrument_column])
    raw_labels.append(["58807_s001", "Baseline (Arm 1: Screening)", "10"])

    cleaned = workbook.create_sheet("cleaned")
    cleaned.append(["IRB", "subid", "arm", "visit", "date", "record_id", "redcap_event_name", instrument_column])
    cleaned.append(["58807", "s001", "1", "baseline", "2026-01-01", "58807_s001", "baseline_arm_1", "10"])

    timepoint = workbook.create_sheet("timepoint_dictionary")
    timepoint.append(["arm", "order", "event_name", "event_label", "abbreviation"])
    timepoint.append(["1", "1", "baseline", "Baseline (Arm 1: Screening)", ""])

    columns = workbook.create_sheet("column_variable_dictionary")
    columns.append(["column_name", "column_labels", "clean_column_name", "keep"])
    columns.append(["record_id", "Record ID", "", "True"])
    columns.append(["redcap_event_name", "Event Name", "", "True"])
    columns.append([instrument_column, instrument_column, "", "False" if keep_sensitive else "True"])

    excluded = workbook.create_sheet("excluded_rows")
    excluded.append(["record_id", "redcap_event_name", instrument_column])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_workbook_with_dropped_non_sensitive_column(path: Path) -> None:
    workbook = Workbook()
    raw = workbook.active
    raw.title = "raw"
    raw.append(["record_id", "redcap_event_name", "survey_timestamp", "kept_score"])
    raw.append(["58807_s001", "baseline_arm_1", "2026-01-01 10:00", "10"])

    raw_labels = workbook.create_sheet("raw_labels")
    raw_labels.append(["Record ID", "Event Name", "Survey Timestamp", "Kept Score"])
    raw_labels.append(["58807_s001", "Baseline (Arm 1: Screening)", "January 1, 2026 10:00", "10"])

    cleaned = workbook.create_sheet("cleaned")
    cleaned.append(["IRB", "subid", "arm", "visit", "date", "record_id", "redcap_event_name", "kept_score"])
    cleaned.append(["58807", "s001", "1", "baseline", "2026-01-01", "58807_s001", "baseline_arm_1", "10"])

    timepoint = workbook.create_sheet("timepoint_dictionary")
    timepoint.append(["arm", "order", "event_name", "event_label", "abbreviation"])
    timepoint.append(["1", "1", "baseline", "Baseline (Arm 1: Screening)", ""])

    columns = workbook.create_sheet("column_variable_dictionary")
    columns.append(["column_name", "column_labels", "clean_column_name", "keep"])
    columns.append(["record_id", "Record ID", "", "True"])
    columns.append(["redcap_event_name", "Event Name", "", "True"])
    columns.append(["survey_timestamp", "Survey Timestamp", "", "False"])
    columns.append(["kept_score", "Kept Score", "", "True"])

    excluded = workbook.create_sheet("excluded_rows")
    excluded.append(["record_id", "redcap_event_name", "survey_timestamp", "kept_score"])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_workbook_with_dropped_raw_visit_field(path: Path) -> None:
    workbook = Workbook()
    raw = workbook.active
    raw.title = "raw"
    raw.append(["record_id", "redcap_event_name", "raw_visit", "kept_score"])
    raw.append(["58807_s001", "baseline_arm_1", "1", "10"])

    raw_labels = workbook.create_sheet("raw_labels")
    raw_labels.append(["Record ID", "Event Name", "Visit #:", "Kept Score"])
    raw_labels.append(["58807_s001", "Baseline (Arm 1: Screening)", "1", "10"])

    cleaned = workbook.create_sheet("cleaned")
    cleaned.append(["IRB", "subid", "arm", "visit", "date", "record_id", "redcap_event_name", "kept_score"])
    cleaned.append(["58807", "s001", "1", "baseline", "2026-01-01", "58807_s001", "baseline_arm_1", "10"])

    timepoint = workbook.create_sheet("timepoint_dictionary")
    timepoint.append(["arm", "order", "event_name", "event_label", "abbreviation"])
    timepoint.append(["1", "1", "baseline", "Baseline (Arm 1: Screening)", ""])

    columns = workbook.create_sheet("column_variable_dictionary")
    columns.append(["column_name", "column_labels", "clean_column_name", "keep"])
    columns.append(["record_id", "Record ID", "", "True"])
    columns.append(["redcap_event_name", "Event Name", "", "True"])
    columns.append(["raw_visit", "Visit #:", "", "False"])
    columns.append(["kept_score", "Kept Score", "", "True"])

    excluded = workbook.create_sheet("excluded_rows")
    excluded.append(["record_id", "redcap_event_name", "raw_visit", "kept_score"])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_workbook_with_kept_column_and_dropped_duplicate_label(path: Path) -> None:
    workbook = Workbook()
    raw = workbook.active
    raw.title = "raw"
    raw.append(["record_id", "redcap_event_name", "eeg_data_folder_name", "eeg_data_folder_name_2"])
    raw.append(["62822_s001", "baseline_arm_1", "folder-a", ""])

    raw_labels = workbook.create_sheet("raw_labels")
    raw_labels.append(["Record ID", "Event Name", "EEG Data Folder Name:", "EEG Data Folder Name:"])
    raw_labels.append(["62822_s001", "Baseline (Arm 1: Screening)", "folder-a", ""])

    cleaned = workbook.create_sheet("cleaned")
    cleaned.append(["IRB", "subid", "arm", "visit", "date", "record_id", "redcap_event_name", "eeg_data_folder_name"])
    cleaned.append(["62822", "s001", "1", "baseline", "2026-01-01", "62822_s001", "baseline_arm_1", "folder-a"])

    timepoint = workbook.create_sheet("timepoint_dictionary")
    timepoint.append(["arm", "order", "event_name", "event_label", "abbreviation"])
    timepoint.append(["1", "1", "baseline", "Baseline (Arm 1: Screening)", ""])

    columns = workbook.create_sheet("column_variable_dictionary")
    columns.append(["column_name", "column_labels", "clean_column_name", "keep"])
    columns.append(["record_id", "Record ID", "", "True"])
    columns.append(["redcap_event_name", "Event Name", "", "True"])
    columns.append(["eeg_data_folder_name", "EEG Data Folder Name:", "", "True"])
    columns.append(["eeg_data_folder_name_2", "EEG Data Folder Name:", "", "False"])

    excluded = workbook.create_sheet("excluded_rows")
    excluded.append(["record_id", "redcap_event_name", "eeg_data_folder_name", "eeg_data_folder_name_2"])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


class FinalVerifyTests(unittest.TestCase):
    def test_final_verify_script_runs_directly_from_repo_root(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            repo_root = Path(__file__).resolve().parents[4]
            study_folder = Path(tmpdir) / "study"
            cleaned_dir = study_folder / "data" / "cleaned"
            history_dir = study_folder / "histories" / "2026-06-18"
            history_dir.mkdir(parents=True)
            (history_dir / "log.md").write_text("# REDCap Workflow Log\n", encoding="utf-8")

            write_dictionary(cleaned_dir / "dictionary.xlsx", [("madrs", "MADRS", "", "1")])
            write_classification(history_dir / "instrument_classification.xlsx", [("madrs", "MADRS", "assessments", 0.91)])
            write_instrument_workbook(cleaned_dir / "assessments" / "58807-madrs.xlsx")

            result = subprocess.run(
                [
                    sys.executable,
                    "scripts/workflows/clean_redcap_instruments/steps/final_verify.py",
                    "--study-folder",
                    str(study_folder),
                ],
                cwd=repo_root,
                text=True,
                capture_output=True,
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertTrue((history_dir / "final_verification.xlsx").exists())

    def test_final_verify_writes_audit_workbook_and_warns_without_failing_on_missing_abbreviations(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study_folder = Path(tmpdir) / "study"
            cleaned_dir = study_folder / "data" / "cleaned"
            history_dir = study_folder / "histories" / "2026-06-18"
            history_dir.mkdir(parents=True)
            (history_dir / "log.md").write_text("# REDCap Workflow Log\n", encoding="utf-8")

            write_dictionary(
                cleaned_dir / "dictionary.xlsx",
                [
                    ("madrs", "MADRS", "", "1"),
                    ("survey_trigger", "Survey Trigger", "", "1"),
                    ("patient_email", "Patient Email", "", "1"),
                ],
            )
            write_classification(
                history_dir / "instrument_classification.xlsx",
                [
                    ("madrs", "MADRS", "assessments", 0.91),
                    ("survey_trigger", "Survey Trigger", "admin", 0.99),
                    ("patient_email", "Patient Email", "subjects", 0.8),
                ],
            )
            write_instrument_workbook(cleaned_dir / "assessments" / "58807-madrs.xlsx")

            result = final_verify(study_folder)
            log_text = (history_dir / "log.md").read_text(encoding="utf-8")

            self.assertTrue(result.output_path.exists())
            self.assertEqual(result.failure_count, 0)
            self.assertGreater(result.warning_count, 0)
            self.assertIn("## Final Verification", log_text)
            self.assertIn("PASS", log_text)
            self.assertIn("WARN", log_text)
            self.assertNotIn("| FAIL |", log_text)

            workbook = load_workbook(result.output_path, data_only=True)
            self.assertIn("summary", workbook.sheetnames)
            self.assertIn("instrument_workbooks", workbook.sheetnames)
            self.assertIn("workbook_checks", workbook.sheetnames)

    def test_final_verify_allows_non_sensitive_keep_false_columns_to_remain_in_raw_audit_sheets(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study_folder = Path(tmpdir) / "study"
            cleaned_dir = study_folder / "data" / "cleaned"
            history_dir = study_folder / "histories" / "2026-06-18"
            history_dir.mkdir(parents=True)
            (history_dir / "log.md").write_text("# REDCap Workflow Log\n", encoding="utf-8")

            write_dictionary(cleaned_dir / "dictionary.xlsx", [("madrs", "MADRS", "", "1")])
            write_classification(history_dir / "instrument_classification.xlsx", [("madrs", "MADRS", "assessments", 0.91)])
            write_workbook_with_dropped_non_sensitive_column(cleaned_dir / "assessments" / "58807-madrs.xlsx")

            result = final_verify(study_folder)
            log_text = (history_dir / "log.md").read_text(encoding="utf-8")

            self.assertEqual(result.failure_count, 0)
            self.assertNotIn("excluded column still present", log_text)

    def test_final_verify_does_not_treat_protected_visit_index_as_excluded_raw_visit_field(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study_folder = Path(tmpdir) / "study"
            cleaned_dir = study_folder / "data" / "cleaned"
            history_dir = study_folder / "histories" / "2026-06-18"
            history_dir.mkdir(parents=True)
            (history_dir / "log.md").write_text("# REDCap Workflow Log\n", encoding="utf-8")

            write_dictionary(cleaned_dir / "dictionary.xlsx", [("ae_log", "AE Log", "", "1")])
            write_classification(history_dir / "instrument_classification.xlsx", [("ae_log", "AE Log", "safety_regulatory", 0.91)])
            write_workbook_with_dropped_raw_visit_field(cleaned_dir / "safety_regulatory" / "58807-ae_log.xlsx")

            result = final_verify(study_folder)
            log_text = (history_dir / "log.md").read_text(encoding="utf-8")

            self.assertEqual(result.failure_count, 0)
            self.assertNotIn("excluded column still present in cleaned: visit", log_text)

    def test_final_verify_uses_variable_names_not_repeated_labels_for_cleaned_exclusion_check(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study_folder = Path(tmpdir) / "study"
            cleaned_dir = study_folder / "data" / "cleaned"
            history_dir = study_folder / "histories" / "2026-06-18"
            history_dir.mkdir(parents=True)
            (history_dir / "log.md").write_text("# REDCap Workflow Log\n", encoding="utf-8")

            write_dictionary(cleaned_dir / "dictionary.xlsx", [("eeg", "EEG", "", "1")])
            write_classification(history_dir / "instrument_classification.xlsx", [("eeg", "EEG", "neuroimaging", 0.91)])
            write_workbook_with_kept_column_and_dropped_duplicate_label(cleaned_dir / "neuroimaging" / "62822-eeg.xlsx")

            result = final_verify(study_folder)
            log_text = (history_dir / "log.md").read_text(encoding="utf-8")

            self.assertEqual(result.failure_count, 0)
            self.assertNotIn("excluded column still present in cleaned: eeg data folder name", log_text)

    def test_final_verify_treats_postprocess_stale_deleted_instruments_as_expected(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study_folder = Path(tmpdir) / "study"
            cleaned_dir = study_folder / "data" / "cleaned"
            history_dir = study_folder / "histories" / "2026-06-18"
            history_dir.mkdir(parents=True)
            (history_dir / "log.md").write_text(
                "# REDCap Workflow Log\n\n"
                "<details>\n"
                "<summary><h1>Postprocess</h1></summary>\n\n"
                "<!-- BEGIN DROP_STALE_INSTRUMENTS -->\n\n"
                "## Drop Stale Instruments\n\n"
                "| deleted_workbook | cleaned_data_rows | reason |\n"
                "| --- | --- | --- |\n"
                "| /tmp/study/data/cleaned/redcap/58807-cage.xlsx | 0 | cleaned sheet has no data rows |\n\n"
                "<!-- END DROP_STALE_INSTRUMENTS -->\n\n"
                "</details>\n",
                encoding="utf-8",
            )

            write_dictionary(cleaned_dir / "dictionary.xlsx", [("cage", "CAGE", "", "1")])
            write_classification(history_dir / "instrument_classification.xlsx", [("cage", "CAGE", "assessments", 0.91)])

            result = final_verify(study_folder)
            log_text = (history_dir / "log.md").read_text(encoding="utf-8")

            self.assertEqual(result.failure_count, 0)
            self.assertIn("stale workbook deleted in postprocess", log_text)
            self.assertNotIn("missing workbook", log_text)

    def test_final_verify_raises_after_writing_outputs_for_structural_failures(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study_folder = Path(tmpdir) / "study"
            cleaned_dir = study_folder / "data" / "cleaned"
            history_dir = study_folder / "histories" / "2026-06-18"
            history_dir.mkdir(parents=True)
            (history_dir / "log.md").write_text("# REDCap Workflow Log\n", encoding="utf-8")

            write_dictionary(
                cleaned_dir / "dictionary.xlsx",
                [
                    ("madrs", "MADRS", "madrs", "1"),
                    ("hamd", "HAMD", "hamd", "1"),
                ],
            )
            write_classification(
                history_dir / "instrument_classification.xlsx",
                [
                    ("madrs", "MADRS", "assessments", 0.91),
                    ("hamd", "HAMD", "assessments", 0.91),
                ],
            )
            write_instrument_workbook(cleaned_dir / "subjects" / "58807-madrs.xlsx", instrument_column="patient_phone", keep_sensitive=True)

            with self.assertRaises(FinalVerificationError) as context:
                final_verify(study_folder)

            result = context.exception.result
            log_text = (history_dir / "log.md").read_text(encoding="utf-8")
            self.assertTrue(result.output_path.exists())
            self.assertGreaterEqual(result.failure_count, 3)
            self.assertIn("missing workbook", log_text)
            self.assertIn("folder mismatch", log_text)
            self.assertIn("sensitive excluded column still present", log_text)
            self.assertIn("| FAIL |", log_text)


if __name__ == "__main__":
    unittest.main()
