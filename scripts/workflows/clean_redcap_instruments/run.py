#!/usr/bin/env python3
"""End-to-end REDCap cleaning workflow for a study folder."""

from __future__ import annotations

import argparse
import csv
import re
import shutil
import sys
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from zipfile import ZIP_DEFLATED, BadZipFile, ZipFile

from openpyxl import load_workbook


if __package__ in {None, ""}:
    sys.path.append(str(Path(__file__).resolve().parents[3]))

from scripts.classify_instruments.run import classify_instruments
from scripts.discover_events.run import process_file as discover_events
from scripts.discover_instruments.run import process_file as discover_instruments
from scripts.standardize_record_id.run import process_file as standardize_record_id
from scripts.workflows.clean_redcap_instruments.steps.create_instrument_excels import InstrumentExcelResult, create_instrument_excels
from scripts.workflows.clean_redcap_instruments.steps.create_redcap_index_dictionary import DictionaryResult, create_dictionary
from scripts.workflows.clean_redcap_instruments.steps.discover_and_standardize import build_log, markdown_table, path_list, timestamp
from scripts.workflows.clean_redcap_instruments.steps.drop_stale_instruments import DropStaleResult, drop_stale_instruments
from scripts.workflows.clean_redcap_instruments.steps.exclude_instruments import exclude_instruments, log_exclusion_result
from scripts.workflows.clean_redcap_instruments.steps.match_text_labels import MatchResult, match_text_labels
from scripts.workflows.clean_redcap_instruments.steps.organize_instruments import OrganizeInstrumentsResult, organize_instruments


RECORD_ID_NAMES = {"record_id", "record id", "recordid"}
EVENT_NAMES = {"redcap_event_name", "event_name", "event name", "unique_event_name"}
MISSING_XML_ENTITY_RE = re.compile(rb"&(amp|lt|gt|quot|apos)(?=([^A-Za-z0-9#;]|$))")
BARE_AMPERSAND_RE = re.compile(rb"&(?!amp;|lt;|gt;|quot;|apos;|#\d+;|#x[0-9A-Fa-f]+;)")


@dataclass(frozen=True)
class StudyWorkflowResult:
    study_folder: Path
    raw_dir: Path
    history_dir: Path
    cleaned_redcap_dir: Path
    data_csv: Path | None
    labels_csv: Path | None
    codebook: Path | None
    dictionary_path: Path
    log_path: Path
    record_paths: dict[str, Path]
    event_paths: dict[str, Path]
    instrument_paths: dict[str, Path]
    label_match: MatchResult | None
    instrument_workbooks: dict[str, Path]
    drop_stale_result: DropStaleResult | None = None
    organize_result: OrganizeInstrumentsResult | None = None


def study_paths(study_folder: str | Path, run_date: str | None = None) -> tuple[Path, Path, Path, Path]:
    study_folder = Path(study_folder)
    raw_dir = study_folder / "data" / "raw_exports" / "redcap" / "all"
    cleaned_redcap_dir = study_folder / "data" / "cleaned" / "redcap"
    history_dir = study_folder / "histories" / (run_date or date.today().isoformat())
    return study_folder, raw_dir, cleaned_redcap_dir, history_dir


def latest_history_dir(study_folder: Path) -> Path | None:
    histories_root = study_folder / "histories"
    if not histories_root.exists():
        return None
    dated_dirs = sorted(path for path in histories_root.iterdir() if path.is_dir())
    return dated_dirs[-1] if dated_dirs else None


def copy_into_raw_dir(source: str | Path, raw_dir: Path) -> Path:
    source = Path(source)
    raw_dir.mkdir(parents=True, exist_ok=True)
    destination = raw_dir / source.name
    if destination.resolve() == source.resolve():
        return destination
    if not destination.exists():
        shutil.copy2(source, destination)
    return destination


def sanitize_xml_entities(data: bytes) -> bytes:
    data = MISSING_XML_ENTITY_RE.sub(rb"&\1;", data)
    return BARE_AMPERSAND_RE.sub(b"&amp;", data)


def repair_xlsx_xml_entities(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    with ZipFile(source, "r") as input_archive, ZipFile(destination, "w", compression=ZIP_DEFLATED) as output_archive:
        for item in input_archive.infolist():
            data = input_archive.read(item.filename)
            if item.filename.endswith((".xml", ".rels")):
                data = sanitize_xml_entities(data)
            output_archive.writestr(item, data)


def assert_workbook_readable(path: Path) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    workbook.close()


def repair_codebook_if_needed(codebook: Path, raw_dir: Path) -> Path:
    try:
        assert_workbook_readable(codebook)
        return codebook
    except (BadZipFile, OSError):
        raise
    except Exception as original_error:
        repaired = raw_dir / f"{codebook.stem}_xml_repaired{codebook.suffix}"
        repair_xlsx_xml_entities(codebook, repaired)
        try:
            assert_workbook_readable(repaired)
        except Exception as repaired_error:
            raise RuntimeError(
                f"Could not read codebook {codebook}; attempted XML repair at {repaired}, "
                f"but the repaired workbook still failed: {repaired_error}"
            ) from original_error
        return repaired


def require_source_path(source: str | Path | None, description: str, stage: str) -> Path:
    if source is None:
        raise ValueError(f"{description} is required for stage={stage!r}.")
    return Path(source)


def find_one_file(directory: Path, patterns: list[str], description: str, exclude_names: set[str] | None = None) -> Path:
    exclude_names = exclude_names or set()
    candidates: list[Path] = []
    for pattern in patterns:
        candidates.extend(path for path in directory.glob(pattern) if path.is_file())

    unique_candidates = sorted({path.resolve(): path for path in candidates}.values())
    unique_candidates = [path for path in unique_candidates if path.name not in exclude_names]
    if not unique_candidates:
        raise FileNotFoundError(f"Could not find {description} in {directory}.")
    if len(unique_candidates) > 1:
        names = ", ".join(path.name for path in unique_candidates)
        raise ValueError(f"Found multiple {description} files in {directory}: {names}")
    return unique_candidates[0]


def find_data_csv(raw_dir: Path) -> Path:
    candidates = sorted(
        path
        for path in raw_dir.glob("*.csv")
        if path.is_file() and "DATA" in path.name.upper() and "LABELS" not in path.name.upper()
    )
    if not candidates:
        candidates = sorted(path for path in raw_dir.glob("*.csv") if path.is_file() and "LABELS" not in path.name.upper())
    if not candidates:
        raise FileNotFoundError(f"Could not find raw DATA CSV in {raw_dir}.")
    if len(candidates) > 1:
        names = ", ".join(path.name for path in candidates)
        raise ValueError(f"Found multiple possible raw DATA CSV files in {raw_dir}: {names}; pass --data-csv.")
    return candidates[0]


def resolve_raw_inputs(
    raw_dir: Path,
    stage: str,
    data_csv: str | Path | None,
    labels_csv: str | Path | None,
    codebook: str | Path | None,
) -> tuple[Path, Path, Path | None]:
    if stage in {"all", "discovery"}:
        resolved_data = copy_into_raw_dir(require_source_path(data_csv, "--data-csv", stage), raw_dir)
        resolved_labels = copy_into_raw_dir(require_source_path(labels_csv, "--labels-csv", stage), raw_dir)
        resolved_codebook = copy_into_raw_dir(require_source_path(codebook, "--codebook", stage), raw_dir)
        return resolved_data, resolved_labels, resolved_codebook

    if data_csv is not None:
        resolved_data = copy_into_raw_dir(data_csv, raw_dir)
    else:
        resolved_data = find_data_csv(raw_dir)

    if labels_csv is not None:
        resolved_labels = copy_into_raw_dir(labels_csv, raw_dir)
    else:
        resolved_labels = find_one_file(raw_dir, ["*DATA_LABELS*.csv", "*LABELS*.csv"], "DATA_LABELS CSV")

    if codebook is not None:
        resolved_codebook: Path | None = copy_into_raw_dir(codebook, raw_dir)
    else:
        codebook_candidates = sorted(
            path
            for pattern in ["*Codebook*.xlsx", "*codebook*.xlsx", "*Codebook*.xls", "*codebook*.xls"]
            for path in raw_dir.glob(pattern)
            if path.is_file()
        )
        resolved_codebook = codebook_candidates[0] if len(codebook_candidates) == 1 else None

    return resolved_data, resolved_labels, resolved_codebook


def resolve_instrument_columns(history_dir: Path, data_stem: str) -> Path:
    expected = history_dir / f"{data_stem}_instrument_columns.csv"
    if expected.exists():
        return expected
    return find_one_file(history_dir, ["*_instrument_columns.csv"], "instrument columns CSV")


def normalize_stage(stage: str) -> str:
    normalized = stage.strip().lower()
    aliases = {
        "discover": "discovery",
        "dictionary": "discovery",
        "discover_and_dictionary": "discovery",
        "discover-dictionary": "discovery",
        "clean_instruments": "clean",
        "clean-instruments": "clean",
        "post-process": "postprocess",
        "organize": "postprocess",
        "sort": "postprocess",
        "full": "all",
    }
    normalized = aliases.get(normalized, normalized)
    if normalized not in {"all", "discovery", "clean", "postprocess"}:
        raise ValueError("stage must be one of: all, discovery, clean, postprocess")
    return normalized


def normalized_header(value: object) -> str:
    return str(value or "").strip().lower()


def read_csv_headers(path: Path) -> list[str]:
    with path.open(newline="", encoding="utf-8-sig") as file:
        sample = file.read(8192)
        file.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(file, dialect)
        return next(reader)


def discover_column_name_if_default_index_misses(
    headers: list[str],
    column_name: str | None,
    column_index: int,
    default_index: int,
    known_names: set[str],
) -> str | None:
    if column_name or column_index != default_index:
        return column_name
    if 0 <= column_index < len(headers) and normalized_header(headers[column_index]) in known_names:
        return column_name
    for header in headers:
        if normalized_header(header) in known_names:
            return header
    return column_name


def infer_codebook_sheet(codebook: Path, kind: str, irb: str, explicit: str | None = None) -> str | None:
    if explicit:
        return explicit

    workbook = load_workbook(codebook, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    kind_matches = [name for name in sheet_names if kind.lower() in name.lower()]
    if not kind_matches:
        return None

    irb_matches = [name for name in kind_matches if irb and irb in name]
    if len(irb_matches) == 1:
        return irb_matches[0]
    if len(kind_matches) == 1:
        return kind_matches[0]
    return None


def workflow_args(
    stage: str,
    study_folder: Path,
    raw_dir: Path,
    cleaned_redcap_dir: Path,
    history_dir: Path,
    data_csv: Path | None,
    labels_csv: Path | None,
    codebook: Path | None,
    irb: str,
    instrument_codebook_sheet: str | None,
    event_codebook_sheet: str | None,
    record_id_column_name: str | None,
    record_id_column_index: int,
    event_column_name: str | None,
    event_column_index: int,
    drop_fully_empty_rows: bool,
    drop_fully_empty_columns: bool,
    run_date: str | None,
) -> dict[str, object]:
    return {
        "stage": stage,
        "study_folder": study_folder,
        "raw_dir": raw_dir,
        "cleaned_redcap_dir": cleaned_redcap_dir,
        "history_dir": history_dir,
        "data_csv": data_csv,
        "labels_csv": labels_csv,
        "codebook": codebook,
        "irb": irb,
        "instrument_codebook_sheet": instrument_codebook_sheet,
        "event_codebook_sheet": event_codebook_sheet,
        "record_id_column_name": record_id_column_name,
        "record_id_column_index": record_id_column_index,
        "event_column_name": event_column_name,
        "event_column_index": event_column_index,
        "drop_fully_empty_rows": drop_fully_empty_rows,
        "drop_fully_empty_columns": drop_fully_empty_columns,
        "run_date": run_date,
    }


def clean_log(
    clean_time: str,
    args: dict[str, object],
    dictionary_path: Path,
    instrument_columns_path: Path,
    instrument_workbooks: dict[str, Path],
) -> str:
    clean_args = dict(args)
    clean_args["dictionary_path"] = dictionary_path
    clean_args["instrument_columns_path"] = instrument_columns_path

    return (
        "<details>\n"
        "<summary><h1>Clean</h1></summary>\n\n"
        "## Instrument Cleaning\n\n"
        "script: `scripts/workflows/clean_redcap_instruments/steps/create_instrument_excels.py`\n\n"
        f"time: {clean_time}\n\n"
        "args:\n\n"
        f"{markdown_table(['arg', 'value'], [[key, value] for key, value in clean_args.items()])}\n\n"
        "outputs:\n"
        f"{path_list(list(instrument_workbooks.values()))}\n\n"
        f"Instruments cleaned: **{len(instrument_workbooks)}**\n\n"
        "</details>\n"
    )


def run_study_workflow(
    study_folder: str | Path,
    data_csv: str | Path | None = None,
    labels_csv: str | Path | None = None,
    codebook: str | Path | None = None,
    irb: str = "",
    stage: str = "all",
    instrument_codebook_sheet: str | None = None,
    event_codebook_sheet: str | None = None,
    record_id_column_name: str | None = None,
    record_id_column_index: int = 0,
    event_column_name: str | None = None,
    event_column_index: int = 1,
    drop_fully_empty_rows: bool = True,
    drop_fully_empty_columns: bool = True,
    run_date: str | None = None,
) -> StudyWorkflowResult:
    if not irb:
        raise ValueError("IRB is required.")
    stage = normalize_stage(stage)
    study_folder = Path(study_folder)
    if stage in {"clean", "postprocess"} and run_date is None:
        existing_history_dir = latest_history_dir(study_folder)
        run_date = existing_history_dir.name if existing_history_dir else None

    study_folder, raw_dir, cleaned_redcap_dir, history_dir = study_paths(study_folder, run_date=run_date)
    if stage in {"all", "discovery", "clean"}:
        cleaned_redcap_dir.mkdir(parents=True, exist_ok=True)
    history_dir.mkdir(parents=True, exist_ok=True)

    if stage in {"all", "discovery", "clean"}:
        data_csv, labels_csv, codebook = resolve_raw_inputs(raw_dir, stage, data_csv, labels_csv, codebook)
        if codebook is not None:
            codebook = repair_codebook_if_needed(codebook, raw_dir)
        stem = data_csv.stem
        headers = read_csv_headers(data_csv)
        record_id_column_name = discover_column_name_if_default_index_misses(
            headers,
            record_id_column_name,
            record_id_column_index,
            default_index=0,
            known_names=RECORD_ID_NAMES,
        )
        event_column_name = discover_column_name_if_default_index_misses(
            headers,
            event_column_name,
            event_column_index,
            default_index=1,
            known_names=EVENT_NAMES,
        )
    else:
        data_csv = Path(data_csv) if data_csv is not None else None
        labels_csv = Path(labels_csv) if labels_csv is not None else None
        codebook = Path(codebook) if codebook is not None else None
        stem = ""

    dictionary_path = cleaned_redcap_dir / "dictionary.xlsx"
    log_path = history_dir / "log.md"
    record_paths: dict[str, Path] = {}
    event_paths: dict[str, Path] = {}
    instrument_paths: dict[str, Path] = {}
    label_match: MatchResult | None = None
    instrument_workbooks: dict[str, Path] = {}
    drop_stale_result: DropStaleResult | None = None
    organize_result: OrganizeInstrumentsResult | None = None
    resolved_workflow_args = workflow_args(
        stage=stage,
        study_folder=study_folder,
        raw_dir=raw_dir,
        cleaned_redcap_dir=cleaned_redcap_dir,
        history_dir=history_dir,
        data_csv=data_csv,
        labels_csv=labels_csv,
        codebook=codebook,
        irb=irb,
        instrument_codebook_sheet=instrument_codebook_sheet,
        event_codebook_sheet=event_codebook_sheet,
        record_id_column_name=record_id_column_name,
        record_id_column_index=record_id_column_index,
        event_column_name=event_column_name,
        event_column_index=event_column_index,
        drop_fully_empty_rows=drop_fully_empty_rows,
        drop_fully_empty_columns=drop_fully_empty_columns,
        run_date=history_dir.name,
    )

    if stage in {"all", "discovery"}:
        if codebook is None:
            raise ValueError("--codebook is required for discovery stage.")
        instrument_codebook_sheet = infer_codebook_sheet(codebook, "instrument", irb, instrument_codebook_sheet)
        event_codebook_sheet = infer_codebook_sheet(codebook, "event", irb, event_codebook_sheet)
        resolved_workflow_args["instrument_codebook_sheet"] = instrument_codebook_sheet
        resolved_workflow_args["event_codebook_sheet"] = event_codebook_sheet

        record_time = timestamp()
        record_paths = standardize_record_id(
            input_path=data_csv,
            output_path=history_dir / f"{stem}_redcap_subid_standardized.csv",
            column_name=record_id_column_name,
            column_index=record_id_column_index,
            summary_output_path=history_dir / f"{stem}_record_id_format_summary.csv",
        )

        instrument_time = timestamp()
        instrument_paths = discover_instruments(
            input_path=data_csv,
            out_dir=history_dir,
            codebook_path=codebook if instrument_codebook_sheet else None,
            codebook_sheet=instrument_codebook_sheet,
        )

        event_time = timestamp()
        event_paths = discover_events(
            input_path=data_csv,
            out_dir=history_dir,
            column_name=event_column_name,
            column_index=event_column_index,
            codebook_path=codebook if event_codebook_sheet else None,
            codebook_sheet=event_codebook_sheet,
        )

        dictionary_time = timestamp()
        dictionary_result = create_dictionary(
            input_path=data_csv,
            standardized_output_path=record_paths["standardized"],
            events_by_arm_path=event_paths["events_by_arm"],
            instrument_summary_path=instrument_paths["instrument_summary"],
            output_path=dictionary_path,
            event_column_name=event_column_name,
            event_column_index=event_column_index,
        )

        log_path.write_text(
            build_log(
                input_path=data_csv,
                record_time=record_time,
                instrument_time=instrument_time,
                event_time=event_time,
                dictionary_time=dictionary_time,
                record_paths=record_paths,
                instrument_paths=instrument_paths,
                event_paths=event_paths,
                dictionary_result=dictionary_result,
                workflow_args=resolved_workflow_args,
            ),
            encoding="utf-8",
        )

        label_match = match_text_labels(
            raw_csv=data_csv,
            labels_csv=labels_csv,
            dictionary_path=dictionary_path,
            event_audit_path=history_dir / "event_label_match_audit.csv",
            instrument_audit_path=history_dir / "instrument_label_match_audit.csv",
            codebook_path=codebook,
            event_codebook_sheet=event_codebook_sheet,
            instrument_codebook_sheet=instrument_codebook_sheet,
            log_path=log_path,
        )

    if stage in {"all", "clean"}:
        if not dictionary_path.exists():
            raise FileNotFoundError(f"Missing dictionary for clean stage: {dictionary_path}")
        instrument_columns_path = (
            instrument_paths["instrument_columns"]
            if instrument_paths.get("instrument_columns")
            else resolve_instrument_columns(history_dir, stem)
        )
        classification_result = classify_instruments(
            study_folder=study_folder,
            output_path=history_dir / "instrument_classification.xlsx",
        )
        exclusion_result = exclude_instruments(study_folder, history_dir=history_dir, write_log=False)
        excluded_instruments = {
            item.instrument for item in [*exclusion_result.excluded, *exclusion_result.missing_workbooks]
        }
        instrument_result: InstrumentExcelResult = create_instrument_excels(
            input_path=data_csv,
            labels_path=labels_csv,
            dictionary_path=dictionary_path,
            instrument_columns_path=instrument_columns_path,
            irb=irb,
            out_dir=study_folder / "data",
            record_id_column_name=record_id_column_name,
            record_id_column_index=record_id_column_index,
            event_column_name=event_column_name,
            event_column_index=event_column_index,
            drop_fully_empty_rows=drop_fully_empty_rows,
            drop_fully_empty_columns=drop_fully_empty_columns,
            excluded_instruments=excluded_instruments,
        )
        instrument_workbooks = instrument_result.output_paths
        if not log_path.exists():
            log_path.write_text("# REDCap Workflow Log\n\n", encoding="utf-8")
        with log_path.open("a", encoding="utf-8") as file:
            file.write(
                "\n\n"
                + clean_log(
                    clean_time=timestamp(),
                    args=resolved_workflow_args,
                    dictionary_path=dictionary_path,
                    instrument_columns_path=instrument_columns_path,
                    instrument_workbooks=instrument_workbooks,
                )
            )
        log_exclusion_result(exclusion_result)

    if stage in {"all", "postprocess"}:
        drop_stale_result = drop_stale_instruments(study_folder, history_dir=history_dir)
        organize_result = organize_instruments(study_folder, history_dir=history_dir)
        if organize_result.dictionary_path is not None:
            dictionary_path = organize_result.dictionary_path
        moved_workbooks = {item.instrument: item.destination_path for item in organize_result.moved}
        if instrument_workbooks:
            instrument_workbooks = {
                instrument: moved_workbooks.get(instrument, workbook_path)
                for instrument, workbook_path in instrument_workbooks.items()
                if workbook_path.exists() or instrument in moved_workbooks
            }
        else:
            instrument_workbooks = moved_workbooks

    return StudyWorkflowResult(
        study_folder=study_folder,
        raw_dir=raw_dir,
        history_dir=history_dir,
        cleaned_redcap_dir=cleaned_redcap_dir,
        data_csv=data_csv,
        labels_csv=labels_csv,
        codebook=codebook,
        dictionary_path=dictionary_path,
        log_path=log_path,
        record_paths=record_paths,
        event_paths=event_paths,
        instrument_paths=instrument_paths,
        label_match=label_match,
        instrument_workbooks=instrument_workbooks,
        drop_stale_result=drop_stale_result,
        organize_result=organize_result,
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run the end-to-end REDCap study-folder cleaning workflow.")
    parser.add_argument("--study-folder", type=Path, required=True, help="Study folder output root")
    parser.add_argument(
        "--stage",
        choices=[
            "all",
            "discovery",
            "clean",
            "postprocess",
            "discover",
            "dictionary",
            "clean-instruments",
            "post-process",
            "organize",
            "sort",
        ],
        default="all",
        help=(
            "Workflow stage: discovery creates raw exports, intermediates, log, and dictionary; "
            "clean creates per-instrument workbooks; postprocess drops stale workbooks and sorts instruments"
        ),
    )
    parser.add_argument("--data-csv", type=Path, help="Raw REDCap DATA CSV; required for all/discovery, inferred for clean")
    parser.add_argument("--labels-csv", type=Path, help="Matching REDCap DATA_LABELS CSV; required for all/discovery, inferred for clean")
    parser.add_argument("--codebook", type=Path, help="REDCap codebook workbook; required for all/discovery")
    parser.add_argument("--irb", required=True, help="IRB/study number for filenames and codebook sheet inference")
    parser.add_argument("--instrument-codebook-sheet", help="Instrument sheet in the codebook workbook")
    parser.add_argument("--event-codebook-sheet", help="Event sheet in the codebook workbook")
    parser.add_argument("--record-id-column-name", help="Record ID column name override")
    parser.add_argument("--record-id-column-index", type=int, default=0, help="Record ID zero-based column index")
    parser.add_argument("--event-column-name", help="Event column name override")
    parser.add_argument("--event-column-index", type=int, default=1, help="Event zero-based column index")
    parser.add_argument(
        "--drop-fully-empty-rows",
        "--drop_fully_empty_rows",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Drop sparse/incomplete rows from per-instrument cleaned sheets",
    )
    parser.add_argument(
        "--drop-fully-empty-columns",
        "--drop_fully_empty_columns",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Drop fully empty data columns from per-instrument cleaned sheets",
    )
    parser.add_argument("--run-date", help="History folder date override in YYYY-MM-DD format")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    result = run_study_workflow(
        study_folder=args.study_folder,
        data_csv=args.data_csv,
        labels_csv=args.labels_csv,
        codebook=args.codebook,
        irb=args.irb,
        stage=args.stage,
        instrument_codebook_sheet=args.instrument_codebook_sheet,
        event_codebook_sheet=args.event_codebook_sheet,
        record_id_column_name=args.record_id_column_name,
        record_id_column_index=args.record_id_column_index,
        event_column_name=args.event_column_name,
        event_column_index=args.event_column_index,
        drop_fully_empty_rows=args.drop_fully_empty_rows,
        drop_fully_empty_columns=args.drop_fully_empty_columns,
        run_date=args.run_date,
    )
    print(result.log_path)
    print(result.dictionary_path)
    print(f"instruments={len(result.instrument_workbooks)}")
    print(result.cleaned_redcap_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
