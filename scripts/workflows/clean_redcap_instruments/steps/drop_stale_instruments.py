#!/usr/bin/env python3
"""Remove per-instrument REDCap workbooks whose cleaned sheet has no data rows."""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


SCRIPT_PATH = Path("scripts/workflows/clean_redcap_instruments/steps/drop_stale_instruments.py")
POSTPROCESS_SUMMARY = "<summary><h1>Postprocess</h1></summary>"
DROP_STALE_BEGIN = "<!-- BEGIN DROP_STALE_INSTRUMENTS -->"
DROP_STALE_END = "<!-- END DROP_STALE_INSTRUMENTS -->"


@dataclass(frozen=True)
class DropStaleResult:
    study_folder: Path
    cleaned_redcap_dir: Path
    log_path: Path
    deleted_paths: list[Path]
    kept_paths: list[Path]
    skipped_paths: list[Path]


def timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d-%H%M%S")


def latest_history_dir(study_folder: Path) -> Path:
    histories_root = study_folder / "histories"
    histories_root.mkdir(parents=True, exist_ok=True)
    dated_dirs = sorted(path for path in histories_root.iterdir() if path.is_dir())
    if dated_dirs:
        return dated_dirs[-1]

    history_dir = histories_root / date.today().isoformat()
    history_dir.mkdir(parents=True, exist_ok=True)
    return history_dir


def nonblank(value: object) -> bool:
    return value is not None and str(value).strip() != ""


def cleaned_data_row_count(path: Path) -> int | None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "cleaned" not in workbook.sheetnames:
        return None

    worksheet = workbook["cleaned"]
    count = 0
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if any(nonblank(value) for value in row):
            count += 1
    return count


def instrument_workbooks(cleaned_redcap_dir: Path) -> list[Path]:
    return sorted(
        path
        for path in cleaned_redcap_dir.glob("*.xlsx")
        if path.is_file() and path.name != "dictionary.xlsx" and not path.name.startswith("~$")
    )


def markdown_table(headers: list[str], rows: list[list[object]]) -> str:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(str(cell) for cell in row) + " |")
    return "\n".join(lines)


def _replace_marker_block(text: str, begin: str, end: str, block: str) -> tuple[str, bool]:
    start = text.find(begin)
    finish = text.find(end)
    if start == -1 or finish == -1 or finish < start:
        return text, False
    finish += len(end)
    return text[:start].rstrip() + "\n\n" + block + "\n\n" + text[finish:].lstrip(), True


def _insert_into_postprocess(text: str, block: str) -> str:
    summary_index = text.find(POSTPROCESS_SUMMARY)
    if summary_index == -1:
        return (
            text.rstrip()
            + "\n\n"
            + "<details>\n"
            + f"{POSTPROCESS_SUMMARY}\n\n"
            + block
            + "\n\n"
            + "</details>\n"
        )

    details_end = text.find("</details>", summary_index)
    if details_end == -1:
        raise ValueError("Found Postprocess section without closing </details> marker.")
    return text[:details_end].rstrip() + "\n\n" + block + "\n\n" + text[details_end:].lstrip()


def append_log(
    log_path: Path,
    study_folder: Path,
    deleted_paths: list[Path],
    kept_paths: list[Path],
    skipped_paths: list[Path],
    data_row_counts: dict[Path, int | None],
) -> None:
    rows = []
    for path in deleted_paths:
        rows.append([path, data_row_counts.get(path, ""), "cleaned sheet has no data rows"])

    block = (
        f"{DROP_STALE_BEGIN}\n\n"
        "## Drop Stale Instruments\n\n"
        f"script: `{SCRIPT_PATH}`\n\n"
        f"time: {timestamp()}\n\n"
        "args:\n\n"
        f"{markdown_table(['arg', 'value'], [['study_folder', study_folder]])}\n\n"
        f"Deleted stale instrument workbooks: **{len(deleted_paths)}**\n\n"
        f"{markdown_table(['deleted_workbook', 'cleaned_data_rows', 'reason'], rows)}\n\n"
        f"{DROP_STALE_END}"
    )
    log_path.parent.mkdir(parents=True, exist_ok=True)
    if not log_path.exists():
        log_path.write_text("# REDCap Workflow Log\n", encoding="utf-8")

    text = log_path.read_text(encoding="utf-8")
    text, replaced = _replace_marker_block(text, DROP_STALE_BEGIN, DROP_STALE_END, block)
    if not replaced:
        text = _insert_into_postprocess(text, block)
    log_path.write_text(text.rstrip() + "\n", encoding="utf-8")


def drop_stale_instruments(study_folder: str | Path, history_dir: str | Path | None = None) -> DropStaleResult:
    study_folder = Path(study_folder)
    cleaned_redcap_dir = study_folder / "data" / "cleaned" / "redcap"
    if not cleaned_redcap_dir.exists():
        raise FileNotFoundError(f"Missing cleaned REDCap directory: {cleaned_redcap_dir}")

    history_dir = Path(history_dir) if history_dir is not None else latest_history_dir(study_folder)
    history_dir.mkdir(parents=True, exist_ok=True)
    log_path = history_dir / "log.md"

    deleted_paths: list[Path] = []
    kept_paths: list[Path] = []
    skipped_paths: list[Path] = []
    data_row_counts: dict[Path, int | None] = {}

    for path in instrument_workbooks(cleaned_redcap_dir):
        data_rows = cleaned_data_row_count(path)
        data_row_counts[path] = data_rows
        if data_rows is None:
            skipped_paths.append(path)
            continue
        if data_rows == 0:
            path.unlink()
            deleted_paths.append(path)
        else:
            kept_paths.append(path)

    append_log(log_path, study_folder, deleted_paths, kept_paths, skipped_paths, data_row_counts)
    return DropStaleResult(
        study_folder=study_folder,
        cleaned_redcap_dir=cleaned_redcap_dir,
        log_path=log_path,
        deleted_paths=deleted_paths,
        kept_paths=kept_paths,
        skipped_paths=skipped_paths,
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Delete stale REDCap instrument workbooks with empty cleaned sheets.")
    parser.add_argument("--study-folder", type=Path, required=True, help="Study folder root")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    result = drop_stale_instruments(args.study_folder)
    print(result.log_path)
    print(f"deleted={len(result.deleted_paths)}")
    print(f"kept={len(result.kept_paths)}")
    print(f"skipped={len(result.skipped_paths)}")
    for path in result.deleted_paths:
        print(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
