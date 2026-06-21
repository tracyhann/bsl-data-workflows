#!/usr/bin/env python3
"""Final verification for sorted REDCap cleaned outputs."""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook


if __package__ in {None, ""}:
    sys.path.append(str(Path(__file__).resolve().parents[4]))

from scripts.workflows.clean_redcap_instruments.steps.exclude_instruments import (
    DEFAULT_EXCLUDE_KEYWORDS,
    ClassificationRecord,
    InstrumentRecord,
    exclusion_reasons,
    read_classifications,
    read_dictionary_instruments,
)
from scripts.workflows.clean_redcap_instruments.steps.organize_instruments import safe_category_name, safe_instrument_name
from scripts.workflows.clean_redcap_instruments.steps.exclude_column import DEFAULT_EXCLUDE_KEYWORDS as SENSITIVE_COLUMN_KEYWORDS


SCRIPT_PATH = Path("scripts/workflows/clean_redcap_instruments/steps/final_verify.py")
FINAL_VERIFY_BEGIN = "<!-- BEGIN FINAL_VERIFICATION -->"
FINAL_VERIFY_END = "<!-- END FINAL_VERIFICATION -->"
FINAL_VERIFY_SUMMARY = "<summary><h1>Final Verification</h1></summary>"
DROP_STALE_BEGIN = "<!-- BEGIN DROP_STALE_INSTRUMENTS -->"
DROP_STALE_END = "<!-- END DROP_STALE_INSTRUMENTS -->"
REQUIRED_SHEETS = ["raw", "raw_labels", "cleaned", "timepoint_dictionary", "column_variable_dictionary", "excluded_rows"]
REQUIRED_CLEANED_PREFIX = ["IRB", "subid", "arm", "visit", "date"]
PROTECTED_CLEANED_COLUMN_COUNT = 7
DEFAULT_RECORD_ID_NAMES = {"record_id", "record id", "recordid", "Record ID"}
DEFAULT_EVENT_NAMES = {"redcap_event_name", "event_name", "event name", "Event Name", "unique_event_name"}


@dataclass(frozen=True)
class VerificationItem:
    check: str
    status: str
    detail: str
    path: Path | None = None


@dataclass(frozen=True)
class InstrumentVerification:
    instrument: str
    instrument_label: str
    expected_category: str
    workbook_count: int
    status: str
    detail: str
    paths: list[Path]


@dataclass(frozen=True)
class WorkbookVerification:
    workbook_path: Path
    instrument: str
    category_folder: str
    status: str
    detail: str


@dataclass(frozen=True)
class FinalVerificationResult:
    study_folder: Path
    dictionary_path: Path
    classification_path: Path
    output_path: Path
    log_path: Path
    checks: list[VerificationItem]
    instrument_checks: list[InstrumentVerification]
    workbook_checks: list[WorkbookVerification]

    @property
    def failure_count(self) -> int:
        return sum(1 for item in self.checks if item.status == "FAIL")

    @property
    def warning_count(self) -> int:
        return sum(1 for item in self.checks if item.status == "WARN")


class FinalVerificationError(RuntimeError):
    def __init__(self, result: FinalVerificationResult):
        super().__init__(f"Final verification failed with {result.failure_count} failure(s).")
        self.result = result


def timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d-%H%M%S")


def latest_history_dir(study_folder: Path) -> Path:
    histories_root = study_folder / "histories"
    histories_root.mkdir(parents=True, exist_ok=True)
    dated_dirs = sorted(path for path in histories_root.iterdir() if path.is_dir())
    if dated_dirs:
        return dated_dirs[-1]

    history_dir = histories_root / date.today().isoformat()
    history_dir.mkdir(parents=True, exist_ok=True)
    return history_dir


def markdown_table(headers: list[str], rows: list[list[object]]) -> str:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(str(cell) for cell in row) + " |")
    return "\n".join(lines)


def normalized(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()


def status_from_failures(failures: list[str], warnings: list[str] | None = None) -> tuple[str, str]:
    warnings = warnings or []
    if failures:
        return "FAIL", "; ".join(failures)
    if warnings:
        return "WARN", "; ".join(warnings)
    return "PASS", "ok"


def workbook_paths_by_instrument(cleaned_dir: Path, instruments: list[InstrumentRecord]) -> tuple[dict[str, list[Path]], list[Path]]:
    instrument_by_suffix = {
        f"-{safe_instrument_name(instrument.instrument)}.xlsx": instrument.instrument
        for instrument in instruments
    }
    paths_by_instrument = {instrument.instrument: [] for instrument in instruments}
    unmatched: list[Path] = []

    for child in sorted(cleaned_dir.iterdir() if cleaned_dir.exists() else []):
        if not child.is_dir() or child.name == "redcap":
            continue
        for path in sorted(child.rglob("*.xlsx")):
            if path.name.startswith("~$"):
                continue
            matched_instrument = None
            for suffix, instrument in instrument_by_suffix.items():
                if path.name.endswith(suffix):
                    matched_instrument = instrument
                    break
            if matched_instrument is None:
                unmatched.append(path)
            else:
                paths_by_instrument[matched_instrument].append(path)
    return paths_by_instrument, unmatched


def stale_deleted_instruments(log_path: Path, instruments: list[InstrumentRecord]) -> set[str]:
    if not log_path.exists():
        return set()
    text = log_path.read_text(encoding="utf-8")
    start = text.find(DROP_STALE_BEGIN)
    finish = text.find(DROP_STALE_END)
    if start == -1 or finish == -1 or finish < start:
        return set()
    block = text[start:finish]
    deleted: set[str] = set()
    for instrument in instruments:
        suffix = f"-{safe_instrument_name(instrument.instrument)}.xlsx"
        if suffix in block:
            deleted.add(instrument.instrument)
    return deleted


def non_excluded_instruments(
    instruments: list[InstrumentRecord],
    classifications: dict[str, ClassificationRecord],
    exclude_keywords: tuple[str, ...],
) -> tuple[list[InstrumentRecord], dict[str, list[str]]]:
    included: list[InstrumentRecord] = []
    excluded_reasons: dict[str, list[str]] = {}
    for instrument in instruments:
        reasons = exclusion_reasons(instrument, classifications.get(instrument.instrument), exclude_keywords)
        if reasons:
            excluded_reasons[instrument.instrument] = reasons
        else:
            included.append(instrument)
    return included, excluded_reasons


def header_row(workbook, sheet_name: str) -> list[str]:
    if sheet_name not in workbook.sheetnames:
        return []
    worksheet = workbook[sheet_name]
    if worksheet.max_row < 1:
        return []
    return [str(value or "").strip() for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]


def cleaned_prefix_failures(
    headers: list[str],
    record_id_column_name: str | None,
    event_column_name: str | None,
) -> list[str]:
    failures: list[str] = []
    if len(headers) < 7:
        return [f"cleaned sheet has {len(headers)} columns; expected at least 7"]

    expected_prefix = REQUIRED_CLEANED_PREFIX
    if headers[:5] != expected_prefix:
        failures.append(f"first five cleaned columns are {headers[:5]}, expected {expected_prefix}")

    record_header = headers[5]
    event_header = headers[6]
    if record_id_column_name:
        if record_header != record_id_column_name:
            failures.append(f"record id column is {record_header!r}, expected {record_id_column_name!r}")
    elif normalized(record_header) not in {normalized(name) for name in DEFAULT_RECORD_ID_NAMES}:
        failures.append(f"record id column is {record_header!r}, expected a known REDCap record id header")

    if event_column_name:
        if event_header != event_column_name:
            failures.append(f"event column is {event_header!r}, expected {event_column_name!r}")
    elif normalized(event_header) not in {normalized(name) for name in DEFAULT_EVENT_NAMES}:
        failures.append(f"event column is {event_header!r}, expected a known REDCap event header")
    return failures


def matches_sensitive_keyword(value: object) -> bool:
    text = normalized(value)
    compact_text = text.replace(" ", "")
    return any(keyword in text.split() or keyword in compact_text for keyword in SENSITIVE_COLUMN_KEYWORDS)


def excluded_column_names(workbook) -> tuple[set[str], set[str], set[str]]:
    if "column_variable_dictionary" not in workbook.sheetnames:
        return set(), set(), set()
    worksheet = workbook["column_variable_dictionary"]
    headers = header_row(workbook, "column_variable_dictionary")
    normalized_headers = {normalized(header): index for index, header in enumerate(headers)}
    name_index = normalized_headers.get("column name")
    clean_name_index = normalized_headers.get("clean column name")
    label_index = normalized_headers.get("column labels")
    keep_index = normalized_headers.get("keep")
    if name_index is None or keep_index is None:
        return set(), set(), set()

    excluded: set[str] = set()
    sensitive_data_excluded: set[str] = set()
    sensitive_label_excluded: set[str] = set()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        keep_value = normalized(row[keep_index] if keep_index < len(row) else "")
        if keep_value not in {"false", "0", "no"}:
            continue

        data_values = []
        label_values = []
        for index in [name_index, clean_name_index]:
            if index is None or index >= len(row):
                continue
            raw_value = row[index]
            data_values.append(raw_value)
            value = normalized(raw_value)
            if value:
                excluded.add(value)

        if label_index is not None and label_index < len(row):
            label_values.append(row[label_index])

        if any(matches_sensitive_keyword(value) for value in data_values + label_values):
            for value in data_values:
                normalized_value = normalized(value)
                if normalized_value:
                    sensitive_data_excluded.add(normalized_value)
            for value in label_values:
                normalized_value = normalized(value)
                if normalized_value:
                    sensitive_label_excluded.add(normalized_value)
    return excluded, sensitive_data_excluded, sensitive_label_excluded


def verify_workbook_structure(
    path: Path,
    instrument: str,
    category: str,
    record_id_column_name: str | None,
    event_column_name: str | None,
) -> tuple[list[VerificationItem], WorkbookVerification]:
    checks: list[VerificationItem] = []
    failures: list[str] = []
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as error:
        detail = f"could not open workbook: {error}"
        item = VerificationItem("workbook can be opened", "FAIL", detail, path)
        return [item], WorkbookVerification(path, instrument, category, "FAIL", detail)

    missing_sheets = [sheet for sheet in REQUIRED_SHEETS if sheet not in workbook.sheetnames]
    status, detail = status_from_failures([f"missing sheets: {', '.join(missing_sheets)}"] if missing_sheets else [])
    checks.append(VerificationItem("required sheets exist", status, detail, path))
    failures.extend(missing_sheets)

    cleaned_headers = header_row(workbook, "cleaned")
    prefix_failures = cleaned_prefix_failures(cleaned_headers, record_id_column_name, event_column_name)
    status, detail = status_from_failures(prefix_failures)
    checks.append(VerificationItem("cleaned sheet required first columns", status, detail, path))
    failures.extend(prefix_failures)

    excluded_names, sensitive_data_excluded_names, sensitive_label_excluded_names = excluded_column_names(workbook)
    present_excluded_cleaned: list[str] = []
    cleaned_headers_set = {
        normalized(header)
        for header in header_row(workbook, "cleaned")[PROTECTED_CLEANED_COLUMN_COUNT:]
    }
    present_excluded_cleaned.extend(sorted(excluded_names & cleaned_headers_set))
    cleaned_failures = (
        [f"excluded column still present in cleaned: {', '.join(sorted(set(present_excluded_cleaned)))}"]
        if present_excluded_cleaned
        else []
    )
    status, detail = status_from_failures(cleaned_failures)
    checks.append(VerificationItem("excluded columns absent from cleaned sheet", status, detail, path))
    failures.extend(cleaned_failures)

    present_sensitive_excluded: list[str] = []
    for sheet_name in ["raw", "cleaned", "excluded_rows"]:
        headers = {normalized(header) for header in header_row(workbook, sheet_name)}
        present_sensitive_excluded.extend(sorted(sensitive_data_excluded_names & headers))
    raw_label_headers = {normalized(header) for header in header_row(workbook, "raw_labels")}
    present_sensitive_excluded.extend(sorted((sensitive_data_excluded_names | sensitive_label_excluded_names) & raw_label_headers))
    present_sensitive_excluded = sorted(set(present_sensitive_excluded))
    excluded_failures = (
        [f"sensitive excluded column still present: {', '.join(present_sensitive_excluded)}"]
        if present_sensitive_excluded
        else []
    )
    status, detail = status_from_failures(excluded_failures)
    checks.append(VerificationItem("sensitive excluded columns absent from workbook data sheets", status, detail, path))
    failures.extend(excluded_failures)

    workbook_status, workbook_detail = status_from_failures(failures)
    return checks, WorkbookVerification(path, instrument, category, workbook_status, workbook_detail)


def abbreviation_warnings(dictionary_path: Path) -> list[VerificationItem]:
    workbook = load_workbook(dictionary_path, read_only=True, data_only=True)
    warnings: list[VerificationItem] = []
    for sheet_name in ["event", "instrument"]:
        if sheet_name not in workbook.sheetnames:
            warnings.append(VerificationItem(f"{sheet_name} abbreviation column", "WARN", f"dictionary has no {sheet_name} sheet", dictionary_path))
            continue
        worksheet = workbook[sheet_name]
        headers = header_row(workbook, sheet_name)
        normalized_headers = {normalized(header): index for index, header in enumerate(headers)}
        abbreviation_index = normalized_headers.get("abbreviation")
        if abbreviation_index is None:
            warnings.append(VerificationItem(f"{sheet_name} abbreviations", "WARN", "missing abbreviation column", dictionary_path))
            continue
        blank_count = 0
        row_count = 0
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_count += 1
            value = row[abbreviation_index] if abbreviation_index < len(row) else ""
            if str(value or "").strip() == "":
                blank_count += 1
        if blank_count:
            warnings.append(
                VerificationItem(
                    f"{sheet_name} abbreviations",
                    "WARN",
                    f"{blank_count}/{row_count} abbreviation values are blank",
                    dictionary_path,
                )
            )
        else:
            warnings.append(VerificationItem(f"{sheet_name} abbreviations", "PASS", "ok", dictionary_path))
    return warnings


def set_readable_widths(workbook: Workbook) -> None:
    for worksheet in workbook.worksheets:
        for column_cells in worksheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 90)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width


def write_report_workbook(result: FinalVerificationResult) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "summary"
    summary.append(["check", "status", "detail", "path"])
    for item in result.checks:
        summary.append([item.check, item.status, item.detail, str(item.path or "")])

    instruments = workbook.create_sheet("instrument_workbooks")
    instruments.append(["instrument", "instrument_label", "expected_category", "workbook_count", "status", "detail", "paths"])
    for item in result.instrument_checks:
        instruments.append(
            [
                item.instrument,
                item.instrument_label,
                item.expected_category,
                item.workbook_count,
                item.status,
                item.detail,
                ", ".join(str(path) for path in item.paths),
            ]
        )

    workbooks = workbook.create_sheet("workbook_checks")
    workbooks.append(["workbook_path", "instrument", "category_folder", "status", "detail"])
    for item in result.workbook_checks:
        workbooks.append([str(item.workbook_path), item.instrument, item.category_folder, item.status, item.detail])

    set_readable_widths(workbook)
    result.output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(result.output_path)


def _replace_marker_block(text: str, begin: str, end: str, block: str) -> tuple[str, bool]:
    start = text.find(begin)
    finish = text.find(end)
    if start == -1 or finish == -1 or finish < start:
        return text, False
    finish += len(end)
    return text[:start].rstrip() + "\n\n" + block + "\n\n" + text[finish:].lstrip(), True


def append_log(result: FinalVerificationResult) -> None:
    rows = [[item.check, item.status, item.detail, item.path or ""] for item in result.checks]
    block = (
        f"{FINAL_VERIFY_BEGIN}\n\n"
        "## Final Verification\n\n"
        f"script: `{SCRIPT_PATH}`\n\n"
        f"time: {timestamp()}\n\n"
        "outputs:\n"
        f"- `{result.output_path}`\n\n"
        f"Failures: **{result.failure_count}**\n\n"
        f"Warnings: **{result.warning_count}**\n\n"
        f"{markdown_table(['check', 'status', 'detail', 'path'], rows)}\n\n"
        f"{FINAL_VERIFY_END}"
    )
    details = (
        "<details>\n"
        f"{FINAL_VERIFY_SUMMARY}\n\n"
        f"{block}\n\n"
        "</details>"
    )
    result.log_path.parent.mkdir(parents=True, exist_ok=True)
    if not result.log_path.exists():
        result.log_path.write_text("# REDCap Workflow Log\n", encoding="utf-8")

    text = result.log_path.read_text(encoding="utf-8")
    text, replaced = _replace_marker_block(text, FINAL_VERIFY_BEGIN, FINAL_VERIFY_END, block)
    if not replaced:
        text = text.rstrip() + "\n\n" + details + "\n"
    result.log_path.write_text(text.rstrip() + "\n", encoding="utf-8")


def build_result(
    study_folder: Path,
    history_dir: Path,
    dictionary_path: Path,
    classification_path: Path,
    output_path: Path,
    record_id_column_name: str | None,
    event_column_name: str | None,
    exclude_keywords: tuple[str, ...],
) -> FinalVerificationResult:
    cleaned_dir = study_folder / "data" / "cleaned"
    log_path = history_dir / "log.md"
    checks: list[VerificationItem] = []
    instrument_checks: list[InstrumentVerification] = []
    workbook_checks: list[WorkbookVerification] = []

    if dictionary_path.exists():
        checks.append(VerificationItem("final dictionary exists", "PASS", "ok", dictionary_path))
    else:
        checks.append(VerificationItem("final dictionary exists", "FAIL", "missing final dictionary", dictionary_path))

    legacy_dictionary = cleaned_dir / "redcap" / "dictionary.xlsx"
    if legacy_dictionary.exists():
        checks.append(VerificationItem("legacy redcap dictionary removed", "FAIL", "dictionary remains under cleaned/redcap", legacy_dictionary))
    else:
        checks.append(VerificationItem("legacy redcap dictionary removed", "PASS", "ok", legacy_dictionary))

    if not dictionary_path.exists():
        return FinalVerificationResult(study_folder, dictionary_path, classification_path, output_path, log_path, checks, [], [])
    if classification_path.exists():
        checks.append(VerificationItem("instrument classification exists", "PASS", "ok", classification_path))
    else:
        checks.append(VerificationItem("instrument classification exists", "FAIL", "missing instrument classification workbook", classification_path))
        return FinalVerificationResult(study_folder, dictionary_path, classification_path, output_path, log_path, checks, [], [])

    instruments = read_dictionary_instruments(dictionary_path)
    classifications = read_classifications(classification_path)
    included_instruments, excluded_reasons = non_excluded_instruments(instruments, classifications, exclude_keywords)
    paths_by_instrument, unmatched_workbooks = workbook_paths_by_instrument(cleaned_dir, instruments)
    stale_deleted = stale_deleted_instruments(log_path, included_instruments)

    for path in unmatched_workbooks:
        checks.append(VerificationItem("workbook maps to dictionary instrument", "FAIL", "workbook under category folder does not match any dictionary instrument", path))

    for instrument in included_instruments:
        classification = classifications.get(instrument.instrument)
        expected_category = safe_category_name(classification.category) if classification else ""
        paths = paths_by_instrument.get(instrument.instrument, [])
        failures: list[str] = []
        if classification is None:
            failures.append("missing classification")
        if len(paths) == 0 and instrument.instrument in stale_deleted:
            pass
        elif len(paths) == 0:
            failures.append("missing workbook")
        elif len(paths) > 1:
            failures.append(f"expected exactly one workbook, found {len(paths)}")
        if len(paths) == 1 and expected_category and paths[0].parent.name != expected_category:
            failures.append(f"folder mismatch: expected {expected_category}, found {paths[0].parent.name}")

        if len(paths) == 0 and instrument.instrument in stale_deleted and not failures:
            status, detail = "PASS", "stale workbook deleted in postprocess"
        else:
            status, detail = status_from_failures(failures)
        instrument_checks.append(
            InstrumentVerification(
                instrument.instrument,
                instrument.instrument_label,
                expected_category,
                len(paths),
                status,
                detail,
                paths,
            )
        )
        checks.append(VerificationItem(f"instrument workbook: {instrument.instrument}", status, detail, paths[0] if paths else None))

    for instrument in instruments:
        for path in paths_by_instrument.get(instrument.instrument, []):
            classification = classifications.get(instrument.instrument)
            category_folder = path.parent.name
            failures: list[str] = []
            if instrument.instrument in excluded_reasons:
                failures.append(f"excluded instrument delivered: {', '.join(excluded_reasons[instrument.instrument])}")
            if classification is None:
                failures.append("missing classification")
            else:
                expected_category = safe_category_name(classification.category)
                if expected_category != category_folder:
                    failures.append(f"folder mismatch: expected {expected_category}, found {category_folder}")
            workbook_items, workbook_check = verify_workbook_structure(
                path,
                instrument.instrument,
                category_folder,
                record_id_column_name,
                event_column_name,
            )
            failures.extend(item.detail for item in workbook_items if item.status == "FAIL")
            workbook_status, workbook_detail = status_from_failures(failures)
            workbook_checks.append(
                WorkbookVerification(path, instrument.instrument, category_folder, workbook_status, workbook_detail)
            )
            checks.extend(workbook_items)
            checks.append(VerificationItem(f"workbook folder/category: {path.name}", workbook_status, workbook_detail, path))

    checks.extend(abbreviation_warnings(dictionary_path))
    return FinalVerificationResult(
        study_folder=study_folder,
        dictionary_path=dictionary_path,
        classification_path=classification_path,
        output_path=output_path,
        log_path=log_path,
        checks=checks,
        instrument_checks=instrument_checks,
        workbook_checks=workbook_checks,
    )


def final_verify(
    study_folder: str | Path,
    history_dir: str | Path | None = None,
    dictionary_path: str | Path | None = None,
    classification_path: str | Path | None = None,
    output_path: str | Path | None = None,
    record_id_column_name: str | None = None,
    event_column_name: str | None = None,
    exclude_keywords: tuple[str, ...] = DEFAULT_EXCLUDE_KEYWORDS,
) -> FinalVerificationResult:
    study_folder = Path(study_folder)
    history_dir = Path(history_dir) if history_dir is not None else latest_history_dir(study_folder)
    history_dir.mkdir(parents=True, exist_ok=True)
    dictionary_path = Path(dictionary_path) if dictionary_path is not None else study_folder / "data" / "cleaned" / "dictionary.xlsx"
    classification_path = (
        Path(classification_path)
        if classification_path is not None
        else history_dir / "instrument_classification.xlsx"
    )
    output_path = Path(output_path) if output_path is not None else history_dir / "final_verification.xlsx"

    result = build_result(
        study_folder=study_folder,
        history_dir=history_dir,
        dictionary_path=dictionary_path,
        classification_path=classification_path,
        output_path=output_path,
        record_id_column_name=record_id_column_name,
        event_column_name=event_column_name,
        exclude_keywords=exclude_keywords,
    )
    write_report_workbook(result)
    append_log(result)
    if result.failure_count:
        raise FinalVerificationError(result)
    return result


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run final verification on sorted REDCap cleaned outputs.")
    parser.add_argument("--study-folder", type=Path, required=True, help="Study folder root")
    parser.add_argument("--history-dir", type=Path, help="History folder containing instrument_classification.xlsx")
    parser.add_argument("--dictionary", type=Path, help="Final dictionary workbook path")
    parser.add_argument("--classification", type=Path, help="Instrument classification workbook path")
    parser.add_argument("--out", type=Path, help="Output final_verification.xlsx path")
    parser.add_argument("--record-id-column-name", help="Expected cleaned record id column header")
    parser.add_argument("--event-column-name", help="Expected cleaned event column header")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        result = final_verify(
            study_folder=args.study_folder,
            history_dir=args.history_dir,
            dictionary_path=args.dictionary,
            classification_path=args.classification,
            output_path=args.out,
            record_id_column_name=args.record_id_column_name,
            event_column_name=args.event_column_name,
        )
    except FinalVerificationError as error:
        result = error.result
        print(result.log_path)
        print(result.output_path)
        print(f"failures={result.failure_count}")
        print(f"warnings={result.warning_count}")
        return 1

    print(result.log_path)
    print(result.output_path)
    print(f"failures={result.failure_count}")
    print(f"warnings={result.warning_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
