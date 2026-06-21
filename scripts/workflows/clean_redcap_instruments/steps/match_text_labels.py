#!/usr/bin/env python3
"""Fill dictionary event and instrument labels from REDCap labels/codebook exports."""

from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


if __package__ in {None, ""}:
    sys.path.append(str(Path(__file__).resolve().parents[4]))

from scripts.discover_events.event_discovery import parse_event_value, resolve_event_column_index


DEFAULT_RAW_CSV = Path(
    "58807-54909-BRAINS-wiki/data/FREEZE/raw_exports/redcap/all/"
    "58807SchatzbergRapid_DATA_2026-06-15_0333.csv"
)
DEFAULT_LABELS_CSV = Path(
    "58807-54909-BRAINS-wiki/data/FREEZE/raw_exports/redcap/all/"
    "58807SchatzbergRapid_DATA_LABELS_2026-06-11_1539.csv"
)
DEFAULT_DICTIONARY = Path("scripts/workflows/dev/intermediates/dictionary.xlsx")
DEFAULT_AUDIT = Path("scripts/workflows/dev/intermediates/event_label_match_audit.csv")
DEFAULT_INSTRUMENT_AUDIT = Path("scripts/workflows/dev/intermediates/instrument_label_match_audit.csv")
DEFAULT_LOG = Path("scripts/workflows/log.md")
SCRIPT_PATH = Path("scripts/workflows/clean_redcap_instruments/steps/match_text_labels.py")
EVENT_LOG_BEGIN = "<!-- BEGIN MATCH_TEXT_LABELS EVENT -->"
EVENT_LOG_END = "<!-- END MATCH_TEXT_LABELS EVENT -->"
INSTRUMENT_LOG_BEGIN = "<!-- BEGIN MATCH_TEXT_LABELS INSTRUMENT -->"
INSTRUMENT_LOG_END = "<!-- END MATCH_TEXT_LABELS INSTRUMENT -->"
VISIT_LABEL_RE = re.compile(r"\(\s*Visit\s+([^)]+?)\s*\)", re.IGNORECASE)


@dataclass(frozen=True)
class MatchResult:
    dictionary_path: Path
    audit_path: Path
    raw_unique_events: int
    label_unique_events: int
    matched_events: int
    unmatched_events: int
    conflict_events: int
    instrument_audit_path: Path | None = None
    matched_instruments: int = 0
    unmatched_instruments: int = 0
    event_codebook_mismatches: int = 0
    instrument_codebook_mismatches: int = 0
    log_path: Path | None = None


def read_csv_table(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open(newline="", encoding="utf-8-sig") as file:
        reader = csv.reader(file)
        headers = next(reader)
        rows = list(reader)
    return headers, rows


def timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d-%H%M%S")


def read_codebook_table(path: Path, sheet_name: str | int | None = None) -> tuple[list[str], list[list[str]]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with path.open(newline="", encoding="utf-8-sig") as file:
            reader = csv.reader(file, delimiter=delimiter)
            headers = next(reader)
            rows = list(reader)
        return headers, rows

    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, data_only=True)
        selected_sheet = sheet_name or workbook.sheetnames[0]
        if selected_sheet not in workbook.sheetnames:
            raise ValueError(f"Sheet {selected_sheet!r} was not found in {path}.")
        worksheet = workbook[selected_sheet]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError(f"Sheet {selected_sheet!r} in {path} is empty.")
        headers = ["" if value is None else str(value).strip() for value in rows[0]]
        data_rows = [["" if value is None else str(value).strip() for value in row] for row in rows[1:]]
        return headers, data_rows

    raise ValueError(f"Unsupported codebook extension {path.suffix!r}; expected CSV, TSV, XLSX, or XLSM.")


def _normalized_header(header: str) -> str:
    return str(header).strip().lower().replace("_", " ")


def _find_header(headers: list[str], header_name: str, aliases: set[str] | None = None) -> int:
    expected = {_normalized_header(header_name)}
    if aliases:
        expected.update(_normalized_header(alias) for alias in aliases)
    for index, header in enumerate(headers):
        if _normalized_header(header) in expected:
            return index
    raise ValueError(f"Expected header {header_name!r} in {headers!r}.")


def _cell(row: list[str], index: int) -> str:
    return row[index].strip() if index < len(row) else ""


def suggest_visit_abbreviation(event_label: str) -> str:
    match = VISIT_LABEL_RE.search(event_label or "")
    if not match:
        return ""
    token = re.sub(r"[^A-Za-z0-9]+", "", match.group(1)).upper()
    return f"V{token}" if token else ""


def build_codebook_event_keys(
    codebook_path: str | Path | None,
    event_codebook_sheet: str | int | None,
    unique_event_column: str = "Unique event name",
) -> set[tuple[str, str]]:
    if not codebook_path or not event_codebook_sheet:
        return set()

    headers, rows = read_codebook_table(Path(codebook_path), event_codebook_sheet)
    unique_event_index = _find_header(headers, unique_event_column, {"unique_event_name"})
    keys: set[tuple[str, str]] = set()
    for row in rows:
        unique_event = _cell(row, unique_event_index)
        if not unique_event:
            continue
        parsed = parse_event_value(unique_event)
        keys.add((parsed.arm, parsed.event_name))
    return keys


def build_event_label_map(
    raw_csv: str | Path,
    labels_csv: str | Path,
    codebook_event_keys: set[tuple[str, str]] | None = None,
) -> tuple[dict[tuple[str, str], str], list[dict[str, str]]]:
    raw_csv = Path(raw_csv)
    labels_csv = Path(labels_csv)
    raw_headers, raw_rows = read_csv_table(raw_csv)
    label_headers, label_rows = read_csv_table(labels_csv)
    if len(raw_rows) != len(label_rows):
        raise ValueError(
            f"Raw and labels exports must have the same number of data rows for row-index matching: "
            f"{raw_csv} has {len(raw_rows)}, {labels_csv} has {len(label_rows)}."
        )
    raw_event_index = resolve_event_column_index(raw_headers)
    label_event_index = resolve_event_column_index(label_headers)

    label_counts: dict[tuple[str, str], Counter[str]] = defaultdict(Counter)
    raw_examples: dict[tuple[str, str], str] = {}
    for row_number, (raw_row, label_row) in enumerate(zip(raw_rows, label_rows), start=1):
        raw_event = raw_row[raw_event_index] if raw_event_index < len(raw_row) else ""
        label_event = label_row[label_event_index] if label_event_index < len(label_row) else ""
        parsed = parse_event_value(raw_event)
        key = (parsed.arm, parsed.event_name)
        raw_examples.setdefault(key, parsed.raw_event)
        if label_event:
            label_counts[key][label_event] += 1

    codebook_event_keys = codebook_event_keys or set()
    label_map: dict[tuple[str, str], str] = {}
    audit_rows: list[dict[str, str]] = []
    for key in sorted(label_counts, key=lambda item: (int(item[0]) if item[0].isdigit() else 10**9, item[1])):
        counter = label_counts[key]
        label, count = counter.most_common(1)[0]
        label_map[key] = label
        codebook_status = "not_checked"
        if codebook_event_keys:
            codebook_status = "matched" if key in codebook_event_keys else "auto_only"
        audit_rows.append(
            {
                "arm": key[0],
                "event_name": key[1],
                "event_label": label,
                "raw_event_example": raw_examples.get(key, ""),
                "label_row_count": str(count),
                "label_variants": "; ".join(f"{value}={variant_count}" for value, variant_count in counter.most_common()),
                "conflict": str(len(counter) > 1),
                "codebook_status": codebook_status,
            }
        )

    for key in sorted(codebook_event_keys - set(label_map), key=lambda item: (int(item[0]) if item[0].isdigit() else 10**9, item[1])):
        audit_rows.append(
            {
                "arm": key[0],
                "event_name": key[1],
                "event_label": "",
                "raw_event_example": "",
                "label_row_count": "0",
                "label_variants": "",
                "conflict": "False",
                "codebook_status": "codebook_only",
            }
        )

    return label_map, audit_rows


def _header_index(headers: list[str], header: str) -> int:
    try:
        return headers.index(header)
    except ValueError as exc:
        raise ValueError(f"Expected header {header!r} in event sheet; found {headers!r}") from exc


def write_audit(path: Path, audit_rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "arm",
        "event_name",
        "event_label",
        "raw_event_example",
        "label_row_count",
        "label_variants",
        "conflict",
        "codebook_status",
    ]
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(audit_rows)


def fill_event_sheet(dictionary_path: Path, label_map: dict[tuple[str, str], str]) -> tuple[int, int]:
    workbook = load_workbook(dictionary_path)
    if "event" not in workbook.sheetnames:
        raise ValueError(f"Workbook {dictionary_path} has no 'event' sheet.")
    worksheet = workbook["event"]

    headers = [str(cell.value or "") for cell in worksheet[1]]
    arm_col = _header_index(headers, "arm") + 1
    event_name_col = _header_index(headers, "event_name") + 1
    event_label_col = _header_index(headers, "event_label") + 1
    abbreviation_col = _header_index(headers, "abbreviation") + 1

    matched = 0
    unmatched = 0
    for row_index in range(2, worksheet.max_row + 1):
        arm = str(worksheet.cell(row=row_index, column=arm_col).value or "")
        event_name = str(worksheet.cell(row=row_index, column=event_name_col).value or "")
        label = label_map.get((arm, event_name))
        if label:
            worksheet.cell(row=row_index, column=event_label_col).value = label
            abbreviation = suggest_visit_abbreviation(label)
            existing_abbreviation = str(worksheet.cell(row=row_index, column=abbreviation_col).value or "").strip()
            if abbreviation and not existing_abbreviation:
                worksheet.cell(row=row_index, column=abbreviation_col).value = abbreviation
            matched += 1
        else:
            unmatched += 1

    workbook.save(dictionary_path)
    return matched, unmatched


def dictionary_instrument_keys(dictionary_path: Path) -> set[str]:
    workbook = load_workbook(dictionary_path, read_only=True, data_only=True)
    if "instrument" not in workbook.sheetnames:
        return set()
    worksheet = workbook["instrument"]
    headers = [str(cell.value or "") for cell in worksheet[1]]
    instrument_col = _header_index(headers, "instrument") + 1
    return {
        str(worksheet.cell(row=row_index, column=instrument_col).value or "").strip()
        for row_index in range(2, worksheet.max_row + 1)
        if str(worksheet.cell(row=row_index, column=instrument_col).value or "").strip()
    }


def build_instrument_label_map(
    codebook_path: str | Path | None,
    instrument_codebook_sheet: str | int | None,
    discovered_instruments: set[str],
    form_name_column: str = "Form Name",
    instrument_name_column: str = "Instrument",
) -> dict[str, str]:
    if not codebook_path or not instrument_codebook_sheet:
        return {}

    headers, rows = read_codebook_table(Path(codebook_path), instrument_codebook_sheet)
    form_name_index = _find_header(headers, form_name_column, {"form_name"})
    instrument_index = _find_header(headers, instrument_name_column, {"instrument"})
    label_map: dict[str, str] = {}

    for row in rows:
        form_name = _cell(row, form_name_index)
        instrument_name = _cell(row, instrument_index)
        key = form_name or (instrument_name if instrument_name in discovered_instruments else "")
        if key and key not in label_map:
            label_map[key] = instrument_name or key

    return label_map


def write_instrument_audit(
    path: Path,
    discovered_instruments: set[str],
    instrument_label_map: dict[str, str],
) -> list[dict[str, str]]:
    path.parent.mkdir(parents=True, exist_ok=True)
    rows: list[dict[str, str]] = []
    for instrument in sorted(discovered_instruments):
        label = instrument_label_map.get(instrument, "")
        rows.append(
            {
                "instrument": instrument,
                "instrument_label": label,
                "codebook_status": "matched" if label else "auto_only",
            }
        )
    for instrument, label in sorted(instrument_label_map.items()):
        if instrument not in discovered_instruments:
            rows.append(
                {
                    "instrument": instrument,
                    "instrument_label": label,
                    "codebook_status": "codebook_only",
                }
            )

    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=["instrument", "instrument_label", "codebook_status"])
        writer.writeheader()
        writer.writerows(rows)
    return rows


def fill_instrument_sheet(dictionary_path: Path, instrument_label_map: dict[str, str]) -> tuple[int, int]:
    workbook = load_workbook(dictionary_path)
    if "instrument" not in workbook.sheetnames:
        return 0, 0
    worksheet = workbook["instrument"]
    headers = [str(cell.value or "") for cell in worksheet[1]]
    instrument_col = _header_index(headers, "instrument") + 1
    instrument_label_col = _header_index(headers, "instrument_label") + 1

    matched = 0
    unmatched = 0
    for row_index in range(2, worksheet.max_row + 1):
        instrument = str(worksheet.cell(row=row_index, column=instrument_col).value or "").strip()
        label = instrument_label_map.get(instrument, "")
        if label:
            worksheet.cell(row=row_index, column=instrument_label_col).value = label
            matched += 1
        else:
            unmatched += 1

    workbook.save(dictionary_path)
    return matched, unmatched


def _path_list(paths: list[Path | None]) -> str:
    return "\n".join(f"- `{path}`" for path in paths if path is not None)


def _event_log_block(result: MatchResult, match_time: str) -> str:
    total_events = result.matched_events + result.unmatched_events
    return (
        f"{EVENT_LOG_BEGIN}\n\n"
        "### Event Label Matching\n\n"
        f"script: `{SCRIPT_PATH}`\n\n"
        f"time: {match_time}\n\n"
        "outputs:\n"
        f"{_path_list([result.dictionary_path, result.audit_path])}\n\n"
        f"Event labels filled: **{result.matched_events}** / {total_events}\n\n"
        f"Unique event labels: **{result.label_unique_events}**\n\n"
        f"Label conflicts: **{result.conflict_events}**\n\n"
        f"Codebook verification mismatches: **{result.event_codebook_mismatches}**\n\n"
        "Automatic label discovery remains the source of truth; codebook verification is advisory.\n\n"
        f"{EVENT_LOG_END}"
    )


def _instrument_log_block(result: MatchResult, match_time: str) -> str:
    total_instruments = result.matched_instruments + result.unmatched_instruments
    return (
        f"{INSTRUMENT_LOG_BEGIN}\n\n"
        "### Instrument Label Matching\n\n"
        f"script: `{SCRIPT_PATH}`\n\n"
        f"time: {match_time}\n\n"
        "outputs:\n"
        f"{_path_list([result.dictionary_path, result.instrument_audit_path])}\n\n"
        f"Instrument labels filled: **{result.matched_instruments}** / {total_instruments}\n\n"
        f"Codebook verification mismatches: **{result.instrument_codebook_mismatches}**\n\n"
        "Automatic label discovery remains the source of truth; codebook verification is advisory.\n\n"
        f"{INSTRUMENT_LOG_END}"
    )


def _replace_marker_block(text: str, begin: str, end: str, block: str) -> tuple[str, bool]:
    start = text.find(begin)
    finish = text.find(end)
    if start == -1 or finish == -1 or finish < start:
        return text, False
    finish += len(end)
    return text[:start].rstrip() + "\n\n" + block + "\n\n" + text[finish:].lstrip(), True


def _insert_before(text: str, marker: str, block: str, start_at: int = 0) -> str:
    index = text.find(marker, start_at)
    if index == -1:
        raise ValueError(f"Could not find log marker {marker!r}.")
    return text[:index].rstrip() + "\n\n" + block + "\n\n" + text[index:].lstrip()


def _minimal_log() -> str:
    return (
        "# REDCap Workflow Log\n\n"
        "<details>\n"
        "<summary><h1>Dictionary</h1></summary>\n\n"
        "## Event\n\n"
        "## Instrument\n\n"
        "</details>\n"
    )


def update_log(log_path: str | Path, result: MatchResult, match_time: str) -> None:
    log_path = Path(log_path)
    text = log_path.read_text(encoding="utf-8") if log_path.exists() else _minimal_log()

    event_block = _event_log_block(result, match_time)
    text, replaced = _replace_marker_block(text, EVENT_LOG_BEGIN, EVENT_LOG_END, event_block)
    if not replaced:
        event_heading = text.find("\n## Event\n")
        if event_heading == -1:
            raise ValueError("Could not find dictionary Event section in log.md.")
        text = _insert_before(text, "\n## Instrument\n", event_block, start_at=event_heading)

    instrument_block = _instrument_log_block(result, match_time)
    text, replaced = _replace_marker_block(text, INSTRUMENT_LOG_BEGIN, INSTRUMENT_LOG_END, instrument_block)
    if not replaced:
        instrument_heading = text.find("\n## Instrument\n")
        if instrument_heading == -1:
            raise ValueError("Could not find dictionary Instrument section in log.md.")
        text = _insert_before(text, "\n</details>", instrument_block, start_at=instrument_heading)

    log_path.parent.mkdir(parents=True, exist_ok=True)
    log_path.write_text(text.rstrip() + "\n", encoding="utf-8")


def match_text_labels(
    raw_csv: str | Path = DEFAULT_RAW_CSV,
    labels_csv: str | Path = DEFAULT_LABELS_CSV,
    dictionary_path: str | Path = DEFAULT_DICTIONARY,
    event_audit_path: str | Path = DEFAULT_AUDIT,
    instrument_audit_path: str | Path | None = DEFAULT_INSTRUMENT_AUDIT,
    codebook_path: str | Path | None = None,
    event_codebook_sheet: str | int | None = None,
    instrument_codebook_sheet: str | int | None = None,
    unique_event_column: str = "Unique event name",
    form_name_column: str = "Form Name",
    instrument_name_column: str = "Instrument",
    log_path: str | Path | None = None,
) -> MatchResult:
    match_time = timestamp()
    raw_csv = Path(raw_csv)
    labels_csv = Path(labels_csv)
    dictionary_path = Path(dictionary_path)
    event_audit_path = Path(event_audit_path)
    instrument_audit = Path(instrument_audit_path) if instrument_audit_path is not None else None
    codebook = Path(codebook_path) if codebook_path else None

    codebook_event_keys = build_codebook_event_keys(codebook, event_codebook_sheet, unique_event_column)
    label_map, audit_rows = build_event_label_map(raw_csv, labels_csv, codebook_event_keys)
    matched_events, unmatched_events = fill_event_sheet(dictionary_path, label_map)
    write_audit(event_audit_path, audit_rows)

    discovered_instruments = dictionary_instrument_keys(dictionary_path)
    instrument_label_map = build_instrument_label_map(
        codebook,
        instrument_codebook_sheet,
        discovered_instruments,
        form_name_column=form_name_column,
        instrument_name_column=instrument_name_column,
    )
    matched_instruments, unmatched_instruments = fill_instrument_sheet(dictionary_path, instrument_label_map)
    instrument_audit_rows = []
    if instrument_audit is not None:
        instrument_audit_rows = write_instrument_audit(instrument_audit, discovered_instruments, instrument_label_map)

    result = MatchResult(
        dictionary_path=dictionary_path,
        audit_path=event_audit_path,
        instrument_audit_path=instrument_audit,
        log_path=Path(log_path) if log_path else None,
        raw_unique_events=len(label_map),
        label_unique_events=len({row["event_label"] for row in audit_rows if row["event_label"]}),
        matched_events=matched_events,
        unmatched_events=unmatched_events,
        conflict_events=sum(1 for row in audit_rows if row["conflict"] == "True"),
        matched_instruments=matched_instruments,
        unmatched_instruments=unmatched_instruments,
        event_codebook_mismatches=sum(1 for row in audit_rows if row["codebook_status"] in {"auto_only", "codebook_only"}),
        instrument_codebook_mismatches=sum(
            1 for row in instrument_audit_rows if row["codebook_status"] in {"auto_only", "codebook_only"}
        ),
    )
    if log_path:
        update_log(log_path, result, match_time)
    return result


def match_event_labels(
    raw_csv: str | Path = DEFAULT_RAW_CSV,
    labels_csv: str | Path = DEFAULT_LABELS_CSV,
    dictionary_path: str | Path = DEFAULT_DICTIONARY,
    audit_path: str | Path = DEFAULT_AUDIT,
) -> MatchResult:
    return match_text_labels(
        raw_csv=raw_csv,
        labels_csv=labels_csv,
        dictionary_path=dictionary_path,
        event_audit_path=audit_path,
        instrument_audit_path=None,
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fill dictionary event and instrument labels from REDCap exports.")
    parser.add_argument("--raw-csv", type=Path, default=DEFAULT_RAW_CSV, help="Raw REDCap data CSV")
    parser.add_argument("--labels-csv", type=Path, default=DEFAULT_LABELS_CSV, help="Matching labels REDCap CSV")
    parser.add_argument("--dictionary", type=Path, default=DEFAULT_DICTIONARY, help="Dictionary workbook to update")
    parser.add_argument("--audit", type=Path, default=DEFAULT_AUDIT, help="Event audit CSV output path")
    parser.add_argument("--instrument-audit", type=Path, default=DEFAULT_INSTRUMENT_AUDIT, help="Instrument audit CSV output path")
    parser.add_argument("--log", type=Path, default=DEFAULT_LOG, help="Workflow log.md path to update")
    parser.add_argument("--codebook", type=Path, help="Optional REDCap codebook workbook/CSV")
    parser.add_argument("--event-codebook-sheet", help="Event sheet name in the codebook")
    parser.add_argument("--instrument-codebook-sheet", help="Instrument sheet name in the codebook")
    parser.add_argument("--unique-event-column", default="Unique event name", help="Codebook unique event name column")
    parser.add_argument("--form-name-column", default="Form Name", help="Codebook instrument form-name column")
    parser.add_argument("--instrument-name-column", default="Instrument", help="Codebook instrument label column")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    result = match_text_labels(
        raw_csv=args.raw_csv,
        labels_csv=args.labels_csv,
        dictionary_path=args.dictionary,
        event_audit_path=args.audit,
        instrument_audit_path=args.instrument_audit,
        codebook_path=args.codebook,
        event_codebook_sheet=args.event_codebook_sheet,
        instrument_codebook_sheet=args.instrument_codebook_sheet,
        unique_event_column=args.unique_event_column,
        form_name_column=args.form_name_column,
        instrument_name_column=args.instrument_name_column,
        log_path=args.log,
    )
    print(result.dictionary_path)
    print(result.audit_path)
    if result.instrument_audit_path:
        print(result.instrument_audit_path)
    if result.log_path:
        print(result.log_path)
    print(f"matched_events={result.matched_events}")
    print(f"unmatched_events={result.unmatched_events}")
    print(f"label_unique_events={result.label_unique_events}")
    print(f"conflict_events={result.conflict_events}")
    print(f"event_codebook_mismatches={result.event_codebook_mismatches}")
    print(f"matched_instruments={result.matched_instruments}")
    print(f"unmatched_instruments={result.unmatched_instruments}")
    print(f"instrument_codebook_mismatches={result.instrument_codebook_mismatches}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
