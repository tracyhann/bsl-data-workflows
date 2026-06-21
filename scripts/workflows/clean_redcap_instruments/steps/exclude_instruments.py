#!/usr/bin/env python3
"""Exclude REDCap instrument workbooks that should not remain in cleaned outputs."""

from __future__ import annotations

import argparse
import re
import shutil
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


SCRIPT_PATH = Path("scripts/workflows/clean_redcap_instruments/steps/exclude_instruments.py")
POSTPROCESS_SUMMARY = "<summary><h1>Postprocess</h1></summary>"
EXCLUDE_BEGIN = "<!-- BEGIN EXCLUDE_INSTRUMENTS -->"
EXCLUDE_END = "<!-- END EXCLUDE_INSTRUMENTS -->"
DEFAULT_EXCLUDE_KEYWORDS = (
    "contact",
    "email",
    "consent",
    "phone",
    "address",
    "do not use",
    "donot use",
    "donotuse",
    "not used",
    "notused",
    "unused",
    "deprecated",
    "deprecate",
    "obsolete",
    "retired",
    "inactive",
    "superseded",
    "decommissioned",
    "disabled",
)


@dataclass(frozen=True)
class InstrumentRecord:
    instrument: str
    instrument_label: str
    number_of_events: object = ""


@dataclass(frozen=True)
class ClassificationRecord:
    instrument: str
    instrument_label: str
    category: str
    confidence: object


@dataclass(frozen=True)
class ExcludedInstrument:
    instrument: str
    instrument_label: str
    category: str
    reasons: list[str]
    source_path: Path | None
    excluded_path: Path | None


@dataclass(frozen=True)
class ExcludeInstrumentsResult:
    study_folder: Path
    dictionary_path: Path
    classification_path: Path
    cleaned_redcap_dir: Path
    excluded_dir: Path
    log_path: Path
    excluded: list[ExcludedInstrument]
    kept: list[InstrumentRecord]
    missing_workbooks: list[ExcludedInstrument]
    exclude_keywords: tuple[str, ...] = DEFAULT_EXCLUDE_KEYWORDS


def timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d-%H%M%S")


def latest_history_dir(study_folder: Path) -> Path:
    histories_root = study_folder / "histories"
    histories_root.mkdir(parents=True, exist_ok=True)
    dated_dirs = sorted(path for path in histories_root.iterdir() if path.is_dir())
    if dated_dirs:
        return dated_dirs[-1]

    history_dir = histories_root / date.today().isoformat()
    history_dir.mkdir(parents=True, exist_ok=True)
    return history_dir


def markdown_table(headers: list[str], rows: list[list[object]]) -> str:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(str(cell) for cell in row) + " |")
    return "\n".join(lines)


def header_index(headers: list[str], *candidates: str) -> int:
    normalized = {header.strip().lower(): index for index, header in enumerate(headers)}
    for candidate in candidates:
        index = normalized.get(candidate.strip().lower())
        if index is not None:
            return index
    raise ValueError(f"Missing required column. Expected one of: {', '.join(candidates)}")


def read_dictionary_instruments(dictionary_path: Path) -> list[InstrumentRecord]:
    workbook = load_workbook(dictionary_path, read_only=True, data_only=True)
    if "instrument" not in workbook.sheetnames:
        raise ValueError(f"Workbook {dictionary_path} has no 'instrument' sheet.")

    worksheet = workbook["instrument"]
    headers = [str(value or "") for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    instrument_index = header_index(headers, "instrument", "instrument_name")
    label_index = header_index(headers, "instrument_label", "form name", "form_name")
    number_of_events_index = header_index(headers, "number_of_events")

    records: list[InstrumentRecord] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        instrument = str(row[instrument_index] or "").strip()
        label = str(row[label_index] or "").strip()
        if instrument:
            records.append(InstrumentRecord(instrument, label, row[number_of_events_index]))
    return records


def read_classifications(classification_path: Path) -> dict[str, ClassificationRecord]:
    workbook = load_workbook(classification_path, read_only=True, data_only=True)
    if "instrument_classification" not in workbook.sheetnames:
        raise ValueError(f"Workbook {classification_path} has no 'instrument_classification' sheet.")

    worksheet = workbook["instrument_classification"]
    headers = [str(value or "") for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    instrument_index = header_index(headers, "instrument_name", "instrument")
    label_index = header_index(headers, "instrument_label", "form name", "form_name")
    category_index = header_index(headers, "class", "category")
    confidence_index = header_index(headers, "confidence")

    records: dict[str, ClassificationRecord] = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        instrument = str(row[instrument_index] or "").strip()
        if not instrument:
            continue
        records[instrument] = ClassificationRecord(
            instrument=instrument,
            instrument_label=str(row[label_index] or "").strip(),
            category=str(row[category_index] or "").strip(),
            confidence=row[confidence_index],
        )
    return records


def normalized_contains_keyword(value: str, keyword: str) -> bool:
    normalized_value = re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()
    normalized_keyword = re.sub(r"[^a-z0-9]+", " ", keyword.lower()).strip()
    if not normalized_value or not normalized_keyword:
        return False
    if " " not in normalized_keyword:
        return normalized_keyword in normalized_value.split()
    return f" {normalized_keyword} " in f" {normalized_value} "


def has_zero_events(value: object) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    if text == "":
        return False
    try:
        return float(text) == 0.0
    except ValueError:
        return False


def exclusion_reasons(
    instrument: InstrumentRecord,
    classification: ClassificationRecord | None,
    keywords: tuple[str, ...],
) -> list[str]:
    reasons: list[str] = []
    if classification and classification.category.lower() == "admin":
        reasons.append("admin category")

    text = f"{instrument.instrument} {instrument.instrument_label}"
    for keyword in keywords:
        if normalized_contains_keyword(text, keyword):
            reasons.append(f"keyword:{keyword}")
    if has_zero_events(instrument.number_of_events):
        reasons.append("number_of_events=0")
    return reasons


def safe_instrument_name(instrument: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", instrument).strip("_")


def matching_workbook(cleaned_redcap_dir: Path, instrument: str) -> Path | None:
    suffix = f"-{safe_instrument_name(instrument)}.xlsx"
    matches = sorted(
        path
        for path in cleaned_redcap_dir.glob(f"*{suffix}")
        if path.is_file() and not path.name.startswith("~$") and path.name != "dictionary.xlsx"
    )
    return matches[0] if matches else None


def unique_destination(path: Path) -> Path:
    if not path.exists():
        return path

    for index in range(1, 1000):
        candidate = path.with_name(f"{path.stem}_{index}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise FileExistsError(f"Could not choose unique destination for {path}")


def _replace_marker_block(text: str, begin: str, end: str, block: str) -> tuple[str, bool]:
    start = text.find(begin)
    finish = text.find(end)
    if start == -1 or finish == -1 or finish < start:
        return text, False
    finish += len(end)
    return text[:start].rstrip() + "\n\n" + block + "\n\n" + text[finish:].lstrip(), True


def _insert_into_postprocess(text: str, block: str) -> str:
    summary_index = text.find(POSTPROCESS_SUMMARY)
    if summary_index == -1:
        return (
            text.rstrip()
            + "\n\n"
            + "<details>\n"
            + f"{POSTPROCESS_SUMMARY}\n\n"
            + block
            + "\n\n"
            + "</details>\n"
        )

    details_end = text.find("</details>", summary_index)
    if details_end == -1:
        raise ValueError("Found Postprocess section without closing </details> marker.")
    return text[:details_end].rstrip() + "\n\n" + block + "\n\n" + text[details_end:].lstrip()


def append_log(
    log_path: Path,
    study_folder: Path,
    dictionary_path: Path,
    classification_path: Path,
    excluded: list[ExcludedInstrument],
    missing_workbooks: list[ExcludedInstrument],
    keywords: tuple[str, ...],
) -> None:
    rows = [
        [
            item.instrument,
            item.instrument_label,
            item.category,
            ", ".join(item.reasons),
            item.source_path or "",
            item.excluded_path or "",
        ]
        for item in excluded
    ]
    missing_rows = [
        [item.instrument, item.instrument_label, item.category, ", ".join(item.reasons)]
        for item in missing_workbooks
    ]
    missing_section = ""
    if missing_rows:
        missing_section = (
            "\n\nExcluded instruments missing cleaned workbook: "
            f"**{len(missing_workbooks)}**\n\n"
            f"{markdown_table(['instrument', 'instrument_label', 'class', 'reason'], missing_rows)}"
        )

    block = (
        f"{EXCLUDE_BEGIN}\n\n"
        "## Exclude Instruments\n\n"
        f"script: `{SCRIPT_PATH}`\n\n"
        f"time: {timestamp()}\n\n"
        "args:\n\n"
        f"{markdown_table(['arg', 'value'], [['study_folder', study_folder], ['dictionary_path', dictionary_path], ['classification_path', classification_path], ['exclude_keywords', ', '.join(keywords)]])}\n\n"
        f"instrument excluded: **{len(excluded) + len(missing_workbooks)}**\n\n"
        f"Excluded instrument workbooks: **{len(excluded)}**\n\n"
        f"{markdown_table(['instrument', 'instrument_label', 'class', 'reason', 'source_workbook', 'excluded_workbook'], rows)}"
        f"{missing_section}\n\n"
        f"{EXCLUDE_END}"
    )

    log_path.parent.mkdir(parents=True, exist_ok=True)
    if not log_path.exists():
        log_path.write_text("# REDCap Workflow Log\n", encoding="utf-8")

    text = log_path.read_text(encoding="utf-8")
    text, replaced = _replace_marker_block(text, EXCLUDE_BEGIN, EXCLUDE_END, block)
    if not replaced:
        text = _insert_into_postprocess(text, block)
    log_path.write_text(text.rstrip() + "\n", encoding="utf-8")


def log_exclusion_result(result: ExcludeInstrumentsResult) -> None:
    append_log(
        result.log_path,
        result.study_folder,
        result.dictionary_path,
        result.classification_path,
        result.excluded,
        result.missing_workbooks,
        result.exclude_keywords,
    )


def exclude_instruments(
    study_folder: str | Path,
    exclude_keywords: tuple[str, ...] = DEFAULT_EXCLUDE_KEYWORDS,
    history_dir: str | Path | None = None,
    write_log: bool = True,
) -> ExcludeInstrumentsResult:
    study_folder = Path(study_folder)
    cleaned_redcap_dir = study_folder / "data" / "cleaned" / "redcap"
    dictionary_path = cleaned_redcap_dir / "dictionary.xlsx"
    if not dictionary_path.exists():
        raise FileNotFoundError(f"Missing dictionary workbook: {dictionary_path}")

    history_dir = Path(history_dir) if history_dir is not None else latest_history_dir(study_folder)
    classification_path = history_dir / "instrument_classification.xlsx"
    if not classification_path.exists():
        raise FileNotFoundError(f"Missing instrument classification workbook: {classification_path}")

    log_path = history_dir / "log.md"
    excluded_dir = cleaned_redcap_dir / "excluded"
    excluded_dir.mkdir(parents=True, exist_ok=True)

    instruments = read_dictionary_instruments(dictionary_path)
    classifications = read_classifications(classification_path)
    excluded: list[ExcludedInstrument] = []
    missing_workbooks: list[ExcludedInstrument] = []
    kept: list[InstrumentRecord] = []

    for instrument in instruments:
        classification = classifications.get(instrument.instrument)
        reasons = exclusion_reasons(instrument, classification, exclude_keywords)
        if not reasons:
            kept.append(instrument)
            continue

        category = classification.category if classification else ""
        source_path = matching_workbook(cleaned_redcap_dir, instrument.instrument)
        if source_path is None:
            missing_workbooks.append(
                ExcludedInstrument(instrument.instrument, instrument.instrument_label, category, reasons, None, None)
            )
            continue

        destination = unique_destination(excluded_dir / source_path.name)
        shutil.move(str(source_path), str(destination))
        excluded.append(
            ExcludedInstrument(
                instrument.instrument,
                instrument.instrument_label,
                category,
                reasons,
                source_path,
                destination,
            )
        )

    result = ExcludeInstrumentsResult(
        study_folder=study_folder,
        dictionary_path=dictionary_path,
        classification_path=classification_path,
        cleaned_redcap_dir=cleaned_redcap_dir,
        excluded_dir=excluded_dir,
        log_path=log_path,
        excluded=excluded,
        kept=kept,
        missing_workbooks=missing_workbooks,
        exclude_keywords=exclude_keywords,
    )
    if write_log:
        log_exclusion_result(result)
    return result


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Exclude REDCap instrument workbooks by category and sensitive keywords.")
    parser.add_argument("--study-folder", type=Path, required=True, help="Study folder root")
    parser.add_argument(
        "--exclude-keyword",
        action="append",
        default=[],
        help="Keyword that excludes an instrument when present in its name or label. Repeatable.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    keywords = tuple(args.exclude_keyword) if args.exclude_keyword else DEFAULT_EXCLUDE_KEYWORDS
    result = exclude_instruments(args.study_folder, exclude_keywords=keywords)
    print(result.log_path)
    print(f"excluded={len(result.excluded)}")
    print(f"kept={len(result.kept)}")
    print(f"missing_workbooks={len(result.missing_workbooks)}")
    for item in result.excluded:
        print(item.excluded_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
