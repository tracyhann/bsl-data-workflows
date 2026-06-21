#!/usr/bin/env python3
"""Move cleaned REDCap instrument workbooks into class-specific folders."""

from __future__ import annotations

import argparse
import re
import shutil
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


SCRIPT_PATH = Path("scripts/workflows/clean_redcap_instruments/steps/organize_instruments.py")
POSTPROCESS_SUMMARY = "<summary><h1>Postprocess</h1></summary>"
SORT_BEGIN = "<!-- BEGIN SORT_INSTRUMENTS -->"
SORT_END = "<!-- END SORT_INSTRUMENTS -->"


@dataclass(frozen=True)
class ClassificationRecord:
    instrument: str
    instrument_label: str
    category: str
    confidence: object


@dataclass(frozen=True)
class MovedInstrument:
    instrument: str
    category: str
    source_path: Path
    destination_path: Path


@dataclass(frozen=True)
class SkippedInstrument:
    instrument: str
    category: str
    reason: str


@dataclass(frozen=True)
class OrganizeInstrumentsResult:
    study_folder: Path
    cleaned_dir: Path
    redcap_dir: Path
    classification_path: Path
    log_path: Path
    created_class_dirs: list[Path]
    moved: list[MovedInstrument]
    skipped: list[SkippedInstrument]
    redcap_dir_removed: bool
    dictionary_path: Path | None = None


def timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d-%H%M%S")


def latest_history_dir(study_folder: Path) -> Path:
    histories_root = study_folder / "histories"
    histories_root.mkdir(parents=True, exist_ok=True)
    dated_dirs = sorted(path for path in histories_root.iterdir() if path.is_dir())
    if dated_dirs:
        return dated_dirs[-1]

    history_dir = histories_root / date.today().isoformat()
    history_dir.mkdir(parents=True, exist_ok=True)
    return history_dir


def markdown_table(headers: list[str], rows: list[list[object]]) -> str:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(str(cell) for cell in row) + " |")
    return "\n".join(lines)


def header_index(headers: list[str], *candidates: str) -> int:
    normalized = {header.strip().lower(): index for index, header in enumerate(headers)}
    for candidate in candidates:
        index = normalized.get(candidate.strip().lower())
        if index is not None:
            return index
    raise ValueError(f"Missing required column. Expected one of: {', '.join(candidates)}")


def read_classifications(classification_path: Path) -> list[ClassificationRecord]:
    workbook = load_workbook(classification_path, read_only=True, data_only=True)
    if "instrument_classification" not in workbook.sheetnames:
        raise ValueError(f"Workbook {classification_path} has no 'instrument_classification' sheet.")

    worksheet = workbook["instrument_classification"]
    headers = [str(value or "") for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    instrument_index = header_index(headers, "instrument_name", "instrument")
    label_index = header_index(headers, "instrument_label", "form name", "form_name")
    category_index = header_index(headers, "class", "category")
    confidence_index = header_index(headers, "confidence")

    records: list[ClassificationRecord] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        instrument = str(row[instrument_index] or "").strip()
        if not instrument:
            continue
        records.append(
            ClassificationRecord(
                instrument=instrument,
                instrument_label=str(row[label_index] or "").strip(),
                category=str(row[category_index] or "").strip(),
                confidence=row[confidence_index],
            )
        )
    return records


def safe_instrument_name(instrument: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", instrument).strip("_")


def safe_category_name(category: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", category.strip().lower()).strip("_")


def matching_workbook(redcap_dir: Path, instrument: str) -> Path | None:
    suffix = f"-{safe_instrument_name(instrument)}.xlsx"
    matches = sorted(
        path
        for path in redcap_dir.glob(f"*{suffix}")
        if path.is_file() and not path.name.startswith("~$") and path.name != "dictionary.xlsx"
    )
    return matches[0] if matches else None


def matching_cleaned_workbook(cleaned_dir: Path, instrument: str) -> Path | None:
    suffix = f"-{safe_instrument_name(instrument)}.xlsx"
    matches = sorted(
        path
        for path in cleaned_dir.rglob(f"*{suffix}")
        if path.is_file()
        and not path.name.startswith("~$")
        and path.name != "dictionary.xlsx"
    )
    return matches[0] if matches else None


def unique_destination(path: Path) -> Path:
    if not path.exists():
        return path
    for index in range(1, 1000):
        candidate = path.with_name(f"{path.stem}_{index}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise FileExistsError(f"Could not choose unique destination for {path}")


def promote_dictionary(redcap_dir: Path, cleaned_dir: Path) -> Path | None:
    source = redcap_dir / "dictionary.xlsx"
    destination = cleaned_dir / "dictionary.xlsx"
    if source.exists():
        destination.parent.mkdir(parents=True, exist_ok=True)
        source.replace(destination)
        return destination
    if destination.exists():
        return destination
    return None


def remove_empty_child_dirs(directory: Path) -> None:
    if not directory.exists():
        return
    subdirs = sorted((path for path in directory.rglob("*") if path.is_dir()), key=lambda path: len(path.parts), reverse=True)
    for subdir in subdirs:
        try:
            subdir.rmdir()
        except OSError:
            continue


def _replace_marker_block(text: str, begin: str, end: str, block: str) -> tuple[str, bool]:
    start = text.find(begin)
    finish = text.find(end)
    if start == -1 or finish == -1 or finish < start:
        return text, False
    finish += len(end)
    return text[:start].rstrip() + "\n\n" + block + "\n\n" + text[finish:].lstrip(), True


def _insert_into_postprocess(text: str, block: str) -> str:
    summary_index = text.find(POSTPROCESS_SUMMARY)
    if summary_index == -1:
        return (
            text.rstrip()
            + "\n\n"
            + "<details>\n"
            + f"{POSTPROCESS_SUMMARY}\n\n"
            + block
            + "\n\n"
            + "</details>\n"
        )

    details_end = text.find("</details>", summary_index)
    if details_end == -1:
        raise ValueError("Found Postprocess section without closing </details> marker.")
    return text[:details_end].rstrip() + "\n\n" + block + "\n\n" + text[details_end:].lstrip()


def append_log(
    log_path: Path,
    study_folder: Path,
    classification_path: Path,
    dictionary_path: Path | None,
    moved: list[MovedInstrument],
    skipped: list[SkippedInstrument],
    redcap_dir_removed: bool,
) -> None:
    moved_rows = [
        [item.instrument, item.category, item.source_path, item.destination_path]
        for item in moved
    ]
    skipped_rows = [
        [item.instrument, item.category, item.reason]
        for item in skipped
    ]
    skipped_section = ""
    if skipped_rows:
        skipped_section = (
            "\n\nSkipped instruments: "
            f"**{len(skipped)}**\n\n"
            f"{markdown_table(['instrument', 'class', 'reason'], skipped_rows)}"
        )

    block = (
        f"{SORT_BEGIN}\n\n"
        "## Sort Instuments\n\n"
        f"script: `{SCRIPT_PATH}`\n\n"
        f"time: {timestamp()}\n\n"
        "args:\n\n"
        f"{markdown_table(['arg', 'value'], [['study_folder', study_folder], ['classification_path', classification_path], ['dictionary_path', dictionary_path or '']])}\n\n"
        f"Moved instrument workbooks: **{len(moved)}**\n\n"
        f"{markdown_table(['instrument', 'class', 'source_workbook', 'destination_workbook'], moved_rows)}"
        f"{skipped_section}\n\n"
        f"cleaned/redcap removed: **{redcap_dir_removed}**\n\n"
        f"{SORT_END}"
    )
    log_path.parent.mkdir(parents=True, exist_ok=True)
    if not log_path.exists():
        log_path.write_text("# REDCap Workflow Log\n", encoding="utf-8")

    text = log_path.read_text(encoding="utf-8")
    text, replaced = _replace_marker_block(text, SORT_BEGIN, SORT_END, block)
    if not replaced:
        text = _insert_into_postprocess(text, block)
    log_path.write_text(text.rstrip() + "\n", encoding="utf-8")


def organize_instruments(study_folder: str | Path, history_dir: str | Path | None = None) -> OrganizeInstrumentsResult:
    study_folder = Path(study_folder)
    cleaned_dir = study_folder / "data" / "cleaned"
    redcap_dir = cleaned_dir / "redcap"

    history_dir = Path(history_dir) if history_dir is not None else latest_history_dir(study_folder)
    history_dir.mkdir(parents=True, exist_ok=True)
    classification_path = history_dir / "instrument_classification.xlsx"
    if not classification_path.exists():
        raise FileNotFoundError(f"Missing instrument classification workbook: {classification_path}")
    log_path = history_dir / "log.md"

    classifications = read_classifications(classification_path)
    class_dirs: dict[str, Path] = {}
    for record in classifications:
        category = safe_category_name(record.category)
        if category and category != "admin":
            class_dir = cleaned_dir / category
            class_dir.mkdir(parents=True, exist_ok=True)
            class_dirs[category] = class_dir

    moved: list[MovedInstrument] = []
    skipped: list[SkippedInstrument] = []
    for record in classifications:
        category = safe_category_name(record.category)
        if category == "admin":
            skipped.append(SkippedInstrument(record.instrument, category, "admin category excluded"))
            continue
        if not category:
            skipped.append(SkippedInstrument(record.instrument, record.category, "empty class"))
            continue

        source_path = matching_workbook(redcap_dir, record.instrument) if redcap_dir.exists() else None
        if source_path is None:
            source_path = matching_cleaned_workbook(cleaned_dir, record.instrument)
        if source_path is None:
            skipped.append(SkippedInstrument(record.instrument, category, "workbook not found"))
            continue

        destination = class_dirs[category] / source_path.name
        if source_path.resolve() == destination.resolve():
            skipped.append(SkippedInstrument(record.instrument, category, "already in target class"))
            continue
        destination = unique_destination(destination)
        shutil.move(str(source_path), str(destination))
        moved.append(MovedInstrument(record.instrument, category, source_path, destination))

    dictionary_path = promote_dictionary(redcap_dir, cleaned_dir)
    if redcap_dir.exists():
        remove_empty_child_dirs(redcap_dir)
    redcap_dir_removed = False
    if redcap_dir.exists():
        try:
            redcap_dir.rmdir()
            redcap_dir_removed = True
        except OSError:
            redcap_dir_removed = False

    append_log(log_path, study_folder, classification_path, dictionary_path, moved, skipped, redcap_dir_removed)
    return OrganizeInstrumentsResult(
        study_folder=study_folder,
        cleaned_dir=cleaned_dir,
        redcap_dir=redcap_dir,
        classification_path=classification_path,
        log_path=log_path,
        created_class_dirs=sorted(class_dirs.values()),
        moved=moved,
        skipped=skipped,
        redcap_dir_removed=redcap_dir_removed,
        dictionary_path=dictionary_path,
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Sort cleaned REDCap instrument workbooks into class folders.")
    parser.add_argument("--study-folder", type=Path, required=True, help="Study folder root")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    result = organize_instruments(args.study_folder)
    print(result.log_path)
    print(f"moved={len(result.moved)}")
    print(f"skipped={len(result.skipped)}")
    print(f"redcap_dir_removed={result.redcap_dir_removed}")
    for item in result.moved:
        print(item.destination_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
