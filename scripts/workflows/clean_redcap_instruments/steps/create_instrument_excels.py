#!/usr/bin/env python3
"""Create one cleaned REDCap workbook per discovered instrument."""

from __future__ import annotations

import argparse
import csv
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl import Workbook, load_workbook


if __package__ in {None, ""}:
    sys.path.append(str(Path(__file__).resolve().parents[4]))

from scripts.discover_events.event_discovery import parse_event_value
from scripts.standardize_date.date_standardizer import standardize_date
from scripts.standardize_record_id.presidio.redcap_subid_presidio import find_redcap_subids
from scripts.workflows.clean_redcap_instruments.steps.exclude_column import flag_sensitive_columns


DEFAULT_DICTIONARY = Path("scripts/workflows/dev/intermediates/dictionary.xlsx")
DEFAULT_INSTRUMENT_COLUMNS = Path(
    "scripts/workflows/dev/intermediates/58807SchatzbergRapid_DATA_2026-06-15_0333_instrument_columns.csv"
)
DEFAULT_OUT_DIR = Path("scripts/workflows/dev")
REDCAP_ID_NAMES = {"record_id", "record id", "recordid", "Record ID"}
EVENT_NAMES = {"redcap_event_name", "event_name", "event name", "Event Name", "unique_event_name"}
DATE_COLUMN_RE = re.compile(r"date|time|timestamp", re.IGNORECASE)
PROTECTED_CLEANED_COLUMN_COUNT = 7
MIN_MEANINGFUL_VALUES_TO_KEEP_INCOMPLETE_ROW = 5
NON_MEANINGFUL_VALUES = {
    "",
    ".",
    "0",
    "unchecked",
    "incomplete",
    "unverified",
    "na",
    "n/a",
    "nan",
    "none",
    "null",
}


@dataclass(frozen=True)
class InstrumentExcelResult:
    output_dir: Path
    output_paths: dict[str, Path]


def read_csv_table(path: str | Path) -> tuple[list[str], list[list[str]]]:
    path = Path(path)
    with path.open(newline="", encoding="utf-8-sig") as file:
        reader = csv.reader(file)
        headers = next(reader)
        rows = list(reader)
    return headers, rows


def read_csv_dicts(path: str | Path) -> list[dict[str, str]]:
    path = Path(path)
    with path.open(newline="", encoding="utf-8-sig") as file:
        return list(csv.DictReader(file))


def _normalized(value: object) -> str:
    return str(value or "").strip().lower()


def resolve_column_index(
    headers: list[str],
    column_name: str | None = None,
    column_index: int = 0,
    known_names: set[str] | None = None,
) -> int:
    if column_name:
        lowered = column_name.lower()
        for index, header in enumerate(headers):
            if str(header).lower() == lowered:
                return index
        raise ValueError(f"Column name {column_name!r} was not found.")

    if 0 <= column_index < len(headers):
        return column_index

    known = {_normalized(name) for name in (known_names or set())}
    for index, header in enumerate(headers):
        if _normalized(header) in known:
            return index
    raise ValueError(f"Column index {column_index} is out of range for {len(headers)} columns.")


def read_subject_mapping(dictionary_path: Path) -> tuple[dict[str, str], set[str]]:
    workbook = load_workbook(dictionary_path, read_only=True, data_only=True)
    worksheet = workbook["subject_id"]
    headers = [str(cell.value or "") for cell in worksheet[1]]
    standardized_index = headers.index("standardized") + 1
    raw_entries_index = headers.index("raw_entries") + 1
    mapping: dict[str, str] = {}
    standardized_values: set[str] = set()

    for row_index in range(2, worksheet.max_row + 1):
        standardized = str(worksheet.cell(row=row_index, column=standardized_index).value or "").strip()
        raw_entries = str(worksheet.cell(row=row_index, column=raw_entries_index).value or "").strip()
        if not standardized:
            continue
        standardized_values.add(standardized)
        mapping[standardized] = standardized
        mapping[standardized.lower()] = standardized
        for raw_entry in raw_entries.split(","):
            raw_entry = raw_entry.strip()
            if raw_entry:
                mapping[raw_entry] = standardized
                mapping[raw_entry.lower()] = standardized

    return mapping, standardized_values


def standardize_subject_id(raw_value: object, raw_mapping: dict[str, str], standardized_values: set[str]) -> str | None:
    text = str(raw_value or "").strip()
    if not text:
        return None
    if text in raw_mapping:
        return raw_mapping[text]
    if text.lower() in raw_mapping:
        return raw_mapping[text.lower()]

    matches = find_redcap_subids(text)
    for match in matches:
        if match.canonicalized in standardized_values:
            return match.canonicalized
    return None


def parse_standardized_subject(standardized: str) -> tuple[str, str]:
    parts = standardized.split("_", 1)
    if len(parts) != 2:
        return "", ""
    return parts[0], parts[1]


def read_timepoint_rows(dictionary_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(dictionary_path, read_only=True, data_only=True)
    worksheet = workbook["event"]
    headers = [str(cell.value or "") for cell in worksheet[1]]
    rows: list[dict[str, str]] = []
    for row_index in range(2, worksheet.max_row + 1):
        row = {
            header: str(worksheet.cell(row=row_index, column=column_index).value or "").strip()
            for column_index, header in enumerate(headers, start=1)
        }
        rows.append(row)
    return rows


def visit_lookup(timepoint_rows: Iterable[dict[str, str]]) -> dict[tuple[str, str], str]:
    lookup: dict[tuple[str, str], str] = {}
    for row in timepoint_rows:
        arm = row.get("arm", "")
        event_name = row.get("event_name", "")
        abbreviation = row.get("abbreviation", "")
        lookup[(arm, event_name)] = abbreviation or event_name
    return lookup


def arm_lookup(timepoint_rows: Iterable[dict[str, str]]) -> dict[tuple[str, str], str]:
    lookup: dict[tuple[str, str], str] = {}
    for row in timepoint_rows:
        arm = row.get("arm", "")
        event_name = row.get("event_name", "")
        lookup[(arm, event_name)] = arm
    return lookup


def read_dictionary_instruments(dictionary_path: Path) -> list[str]:
    workbook = load_workbook(dictionary_path, read_only=True, data_only=True)
    worksheet = workbook["instrument"]
    headers = [str(cell.value or "") for cell in worksheet[1]]
    instrument_index = headers.index("instrument") + 1
    instruments: list[str] = []
    for row_index in range(2, worksheet.max_row + 1):
        instrument = str(worksheet.cell(row=row_index, column=instrument_index).value or "").strip()
        if instrument:
            instruments.append(instrument)
    return instruments


def read_instrument_columns(instrument_columns_path: str | Path) -> dict[str, list[int]]:
    rows = read_csv_dicts(instrument_columns_path)
    columns: dict[str, list[int]] = {}
    for row in rows:
        instrument = row.get("instrument_key", "").strip()
        column_index = row.get("column_index", "").strip()
        if not instrument or not column_index:
            continue
        columns.setdefault(instrument, []).append(int(column_index))
    for instrument in columns:
        columns[instrument] = sorted(dict.fromkeys(columns[instrument]))
    return columns


def output_path_for_instrument(cleaned_dir: Path, irb: str, instrument: str) -> Path:
    safe_instrument = re.sub(r"[^A-Za-z0-9_.-]+", "_", instrument).strip("_")
    return cleaned_dir / f"{irb}-{safe_instrument}.xlsx"


def _row_values(row: list[str], indices: list[int]) -> list[str]:
    return [row[index] if index < len(row) else "" for index in indices]


def _label_or_raw_value(raw_row: list[str], label_row: list[str], index: int) -> str:
    label_value = label_row[index] if index < len(label_row) else ""
    if str(label_value or "").strip() != "":
        return label_value
    return raw_row[index] if index < len(raw_row) else ""


def _label_or_raw_values(raw_row: list[str], label_row: list[str], indices: list[int]) -> list[str]:
    return [_label_or_raw_value(raw_row, label_row, index) for index in indices]


def excel_safe_value(value: object) -> object:
    if not isinstance(value, str):
        return value
    return ILLEGAL_CHARACTERS_RE.sub("", value)[:32767]


def append_excel_row(worksheet, row: Iterable[object]) -> None:
    worksheet.append([excel_safe_value(value) for value in row])


def copy_timepoint_dictionary(workbook: Workbook, timepoint_rows: list[dict[str, str]]) -> None:
    worksheet = workbook.create_sheet("timepoint_dictionary")
    headers = ["arm", "order", "event_name", "event_label", "abbreviation"]
    append_excel_row(worksheet, headers)
    for row in timepoint_rows:
        append_excel_row(worksheet, [row.get(header, "") for header in headers])


def write_rows(worksheet, headers: list[str], rows: Iterable[list[object]]) -> None:
    append_excel_row(worksheet, headers)
    for row in rows:
        append_excel_row(worksheet, row)


def candidate_date_indices(
    raw_headers: list[str],
    label_headers: list[str],
    instrument_indices: list[int],
    valid_rows: list[list[str]],
) -> list[int]:
    candidates: list[int] = []
    for index in instrument_indices:
        raw_name = raw_headers[index] if index < len(raw_headers) else ""
        label_name = label_headers[index] if index < len(label_headers) else ""
        name_hint = DATE_COLUMN_RE.search(raw_name) or DATE_COLUMN_RE.search(label_name)
        has_parseable_value = any(standardize_date(row[index] if index < len(row) else "") for row in valid_rows)
        if name_hint or has_parseable_value:
            candidates.append(index)
    return candidates


def row_date(row: list[str], date_indices: list[int]) -> str:
    for index in date_indices:
        standardized = standardize_date(row[index] if index < len(row) else "")
        if standardized:
            return standardized
    return ""


def is_empty(value: object) -> bool:
    return str(value or "").strip() == ""


def is_meaningful_value(value: object) -> bool:
    normalized = str(value or "").strip().lower()
    return normalized not in NON_MEANINGFUL_VALUES


def meaningful_value_count(
    row: list[object],
    protected_column_count: int = PROTECTED_CLEANED_COLUMN_COUNT,
) -> int:
    return sum(1 for value in row[protected_column_count:] if is_meaningful_value(value))


def is_complete_column_name(column_name: str) -> bool:
    normalized = column_name.strip().lower()
    return normalized.endswith("_complete") or normalized == "complete?"


def drop_empty_data_columns(
    headers: list[str],
    row_pairs: list[tuple[list[object], list[str]]],
    cleaned_to_raw_indices: list[int | None],
    keep_by_raw_index: dict[int, bool],
    protected_column_count: int = PROTECTED_CLEANED_COLUMN_COUNT,
) -> tuple[list[str], list[tuple[list[object], list[str]]], list[int | None]]:
    keep_column = [True] * len(headers)
    for column_index in range(protected_column_count, len(headers)):
        if all(is_empty(row[column_index]) for row, _ in row_pairs):
            keep_column[column_index] = False
            raw_index = cleaned_to_raw_indices[column_index]
            if raw_index is not None:
                keep_by_raw_index[raw_index] = False

    headers = [header for header, keep in zip(headers, keep_column) if keep]
    cleaned_to_raw_indices = [index for index, keep in zip(cleaned_to_raw_indices, keep_column) if keep]
    row_pairs = [
        ([value for value, keep in zip(row, keep_column) if keep], raw_view_row)
        for row, raw_view_row in row_pairs
    ]
    return headers, row_pairs, cleaned_to_raw_indices


def row_is_sparse_or_incomplete(
    row: list[object],
    headers: list[str],
    protected_column_count: int = PROTECTED_CLEANED_COLUMN_COUNT,
    empty_threshold: float = 0.80,
    min_meaningful_values: int = MIN_MEANINGFUL_VALUES_TO_KEEP_INCOMPLETE_ROW,
) -> bool:
    data_values = row[protected_column_count:]
    if not data_values:
        return True

    meaningful_count = meaningful_value_count(row, protected_column_count)
    empty_ratio = sum(1 for value in data_values if is_empty(value)) / len(data_values)
    if empty_ratio > empty_threshold and meaningful_count < min_meaningful_values:
        return True

    complete_indices = [
        index
        for index in range(protected_column_count, len(headers))
        if is_complete_column_name(headers[index])
    ]
    for index in complete_indices:
        value = str(row[index] or "").strip()
        normalized = value.lower()
        if (
            (value == "" or value == "0" or normalized in {"incomplete", "unverified"})
            and meaningful_count < min_meaningful_values
        ):
            return True
    return False


def drop_sparse_or_incomplete_rows(
    headers: list[str],
    row_pairs: list[tuple[list[object], list[str]]],
    excluded_rows: list[list[str]],
) -> list[tuple[list[object], list[str]]]:
    kept: list[tuple[list[object], list[str]]] = []
    for row, raw_view_row in row_pairs:
        if row_is_sparse_or_incomplete(row, headers):
            excluded_rows.append(raw_view_row)
        else:
            kept.append((row, raw_view_row))
    return kept


def write_column_dictionary(
    worksheet,
    raw_headers: list[str],
    label_headers: list[str],
    selected_indices: list[int],
    keep_by_raw_index: dict[int, bool],
) -> None:
    append_excel_row(worksheet, ["column_name", "column_labels", "clean_column_name", "keep"])
    for index in selected_indices:
        column_name = raw_headers[index] if index < len(raw_headers) else ""
        column_label = label_headers[index] if index < len(label_headers) else ""
        keep = str(keep_by_raw_index.get(index, True))
        append_excel_row(worksheet, [column_name, column_label, "", keep])


def set_readable_widths(workbook: Workbook) -> None:
    for worksheet in workbook.worksheets:
        for column_cells in worksheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 80)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width


def create_instrument_workbook(
    output_path: Path,
    raw_headers: list[str],
    raw_rows: list[list[str]],
    label_headers: list[str],
    label_rows: list[list[str]],
    dictionary_path: Path,
    selected_indices: list[int],
    instrument_indices: list[int],
    record_id_index: int,
    event_index: int,
    subject_mapping: dict[str, str],
    standardized_subjects: set[str],
    timepoint_rows: list[dict[str, str]],
    visit_by_event: dict[tuple[str, str], str],
    drop_fully_empty_rows: bool = True,
    drop_fully_empty_columns: bool = True,
) -> None:
    flagged_columns = flag_sensitive_columns(raw_headers, label_headers)
    protected_indices = {record_id_index, event_index}
    sensitive_indices = set(flagged_columns) - protected_indices
    visible_selected_indices = [index for index in selected_indices if index not in sensitive_indices]
    visible_instrument_indices = [index for index in instrument_indices if index not in sensitive_indices]

    raw_view_headers = [raw_headers[index] if index < len(raw_headers) else "" for index in visible_selected_indices]
    raw_view_rows = [_row_values(row, visible_selected_indices) for row in raw_rows]
    label_view_headers = [label_headers[index] if index < len(label_headers) else "" for index in visible_selected_indices]
    label_view_rows = [_row_values(row, visible_selected_indices) for row in label_rows]

    valid_rows: list[tuple[list[str], list[str], list[str], str]] = []
    excluded_rows: list[list[str]] = []
    for raw_row, raw_view_row, label_row in zip(raw_rows, raw_view_rows, label_rows):
        standardized = standardize_subject_id(
            raw_row[record_id_index] if record_id_index < len(raw_row) else "",
            subject_mapping,
            standardized_subjects,
        )
        if standardized:
            valid_rows.append((raw_row, raw_view_row, label_row, standardized))
        else:
            excluded_rows.append(raw_view_row)

    date_indices = candidate_date_indices(
        raw_headers,
        label_headers,
        visible_instrument_indices,
        [row for row, _, _, _ in valid_rows],
    )
    keep_by_raw_index = {index: True for index in selected_indices}
    for index in sensitive_indices:
        if index in keep_by_raw_index:
            keep_by_raw_index[index] = False

    workbook = Workbook()
    raw_sheet = workbook.active
    raw_sheet.title = "raw"
    write_rows(raw_sheet, raw_view_headers, raw_view_rows)

    raw_labels_sheet = workbook.create_sheet("raw_labels")
    write_rows(raw_labels_sheet, label_view_headers, label_view_rows)

    cleaned_sheet = workbook.create_sheet("cleaned")
    cleaned_headers = [
        "IRB",
        "subid",
        "arm",
        "visit",
        "date",
        raw_headers[record_id_index] if record_id_index < len(raw_headers) else "record_id",
        raw_headers[event_index] if event_index < len(raw_headers) else "redcap_event_name",
    ] + [
        raw_headers[index] if index < len(raw_headers) else "" for index in visible_instrument_indices
    ]
    cleaned_to_raw_indices: list[int | None] = [
        None,
        None,
        None,
        None,
        None,
        record_id_index,
        event_index,
    ] + visible_instrument_indices
    cleaned_row_pairs: list[tuple[list[object], list[str]]] = []
    arm_by_event = arm_lookup(timepoint_rows)
    for raw_row, raw_view_row, label_row, standardized in valid_rows:
        parsed_irb, subid = parse_standardized_subject(standardized)
        parsed_event = parse_event_value(raw_row[event_index] if event_index < len(raw_row) else "")
        arm = arm_by_event.get((parsed_event.arm, parsed_event.event_name), parsed_event.arm)
        visit = visit_by_event.get((parsed_event.arm, parsed_event.event_name), parsed_event.event_name)
        cleaned_row_pairs.append(
            (
                [
                    parsed_irb,
                    subid,
                    arm,
                    visit,
                    row_date(raw_row, date_indices),
                    _label_or_raw_value(raw_row, label_row, record_id_index),
                    _label_or_raw_value(raw_row, label_row, event_index),
                ]
                + _label_or_raw_values(raw_row, label_row, visible_instrument_indices),
                raw_view_row,
            )
        )

    if drop_fully_empty_columns:
        cleaned_headers, cleaned_row_pairs, cleaned_to_raw_indices = drop_empty_data_columns(
            cleaned_headers,
            cleaned_row_pairs,
            cleaned_to_raw_indices,
            keep_by_raw_index,
        )

    if drop_fully_empty_rows:
        cleaned_row_pairs = drop_sparse_or_incomplete_rows(cleaned_headers, cleaned_row_pairs, excluded_rows)

    append_excel_row(cleaned_sheet, cleaned_headers)
    for cleaned_row, _ in cleaned_row_pairs:
        append_excel_row(cleaned_sheet, cleaned_row)

    copy_timepoint_dictionary(workbook, timepoint_rows)
    column_dictionary_sheet = workbook.create_sheet("column_variable_dictionary")
    write_column_dictionary(
        column_dictionary_sheet,
        raw_headers,
        label_headers,
        selected_indices,
        keep_by_raw_index=keep_by_raw_index,
    )
    excluded_sheet = workbook.create_sheet("excluded_rows")
    write_rows(excluded_sheet, raw_view_headers, excluded_rows)

    set_readable_widths(workbook)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def create_instrument_excels(
    input_path: str | Path,
    labels_path: str | Path,
    dictionary_path: str | Path = DEFAULT_DICTIONARY,
    instrument_columns_path: str | Path = DEFAULT_INSTRUMENT_COLUMNS,
    irb: str = "",
    out_dir: str | Path = DEFAULT_OUT_DIR,
    record_id_column_name: str | None = None,
    record_id_column_index: int = 0,
    event_column_name: str | None = None,
    event_column_index: int = 1,
    drop_fully_empty_rows: bool = True,
    drop_fully_empty_columns: bool = True,
    excluded_instruments: Iterable[str] | None = None,
) -> InstrumentExcelResult:
    if not irb:
        raise ValueError("IRB is required for instrument workbook filenames.")

    input_path = Path(input_path)
    labels_path = Path(labels_path)
    dictionary_path = Path(dictionary_path)
    out_dir = Path(out_dir)
    cleaned_dir = out_dir / "cleaned" / "redcap"

    raw_headers, raw_rows = read_csv_table(input_path)
    label_headers, label_rows = read_csv_table(labels_path)
    if len(raw_headers) != len(label_headers):
        raise ValueError(
            f"Raw and labels exports must have the same number of columns: "
            f"{input_path} has {len(raw_headers)}, {labels_path} has {len(label_headers)}."
        )
    if len(raw_rows) != len(label_rows):
        raise ValueError(
            f"Raw and labels exports must have the same number of data rows: "
            f"{input_path} has {len(raw_rows)}, {labels_path} has {len(label_rows)}."
        )

    record_id_index = resolve_column_index(raw_headers, record_id_column_name, record_id_column_index, REDCAP_ID_NAMES)
    event_index = resolve_column_index(raw_headers, event_column_name, event_column_index, EVENT_NAMES)
    subject_mapping, standardized_subjects = read_subject_mapping(dictionary_path)
    timepoint_rows = read_timepoint_rows(dictionary_path)
    visit_by_event = visit_lookup(timepoint_rows)
    dictionary_instruments = read_dictionary_instruments(dictionary_path)
    instrument_columns = read_instrument_columns(instrument_columns_path)
    excluded_instrument_set = {str(instrument).strip() for instrument in excluded_instruments or set()}

    output_paths: dict[str, Path] = {}
    for instrument in dictionary_instruments:
        if instrument in excluded_instrument_set:
            continue
        columns = instrument_columns.get(instrument)
        if not columns:
            continue
        selected_indices = [record_id_index, event_index] + [
            index for index in columns if index not in {record_id_index, event_index}
        ]
        output_path = output_path_for_instrument(cleaned_dir, irb, instrument)
        create_instrument_workbook(
            output_path=output_path,
            raw_headers=raw_headers,
            raw_rows=raw_rows,
            label_headers=label_headers,
            label_rows=label_rows,
            dictionary_path=dictionary_path,
            selected_indices=selected_indices,
            instrument_indices=columns,
            record_id_index=record_id_index,
            event_index=event_index,
            subject_mapping=subject_mapping,
            standardized_subjects=standardized_subjects,
            timepoint_rows=timepoint_rows,
            visit_by_event=visit_by_event,
            drop_fully_empty_rows=drop_fully_empty_rows,
            drop_fully_empty_columns=drop_fully_empty_columns,
        )
        output_paths[instrument] = output_path

    return InstrumentExcelResult(output_dir=cleaned_dir, output_paths=output_paths)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create one cleaned REDCap workbook per discovered instrument.")
    parser.add_argument("--input", type=Path, required=True, help="Raw REDCap CSV export")
    parser.add_argument("--labels-csv", type=Path, required=True, help="Matching REDCap LABELS CSV export")
    parser.add_argument("--dictionary", type=Path, default=DEFAULT_DICTIONARY, help="dictionary.xlsx path")
    parser.add_argument("--instrument-columns", type=Path, default=DEFAULT_INSTRUMENT_COLUMNS, help="Instrument column mapping CSV")
    parser.add_argument("--irb", required=True, help="IRB/study number used in output filenames")
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR, help="Output directory; cleaned/redcap is created inside it")
    parser.add_argument("--record-id-column-name", help="Record ID column name override")
    parser.add_argument("--record-id-column-index", type=int, default=0, help="Record ID zero-based column index")
    parser.add_argument("--event-column-name", help="Event column name override")
    parser.add_argument("--event-column-index", type=int, default=1, help="Event zero-based column index")
    parser.add_argument(
        "--drop-fully-empty-rows",
        "--drop_fully_empty_rows",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Drop sparse or incomplete cleaned rows after population",
    )
    parser.add_argument(
        "--drop-fully-empty-columns",
        "--drop_fully_empty_columns",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Drop fully empty cleaned data columns after population",
    )
    parser.add_argument(
        "--exclude-instrument",
        action="append",
        default=[],
        help="Instrument key to skip while creating per-instrument workbooks. Repeatable.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    result = create_instrument_excels(
        input_path=args.input,
        labels_path=args.labels_csv,
        dictionary_path=args.dictionary,
        instrument_columns_path=args.instrument_columns,
        irb=args.irb,
        out_dir=args.out_dir,
        record_id_column_name=args.record_id_column_name,
        record_id_column_index=args.record_id_column_index,
        event_column_name=args.event_column_name,
        event_column_index=args.event_column_index,
        drop_fully_empty_rows=args.drop_fully_empty_rows,
        drop_fully_empty_columns=args.drop_fully_empty_columns,
        excluded_instruments=set(args.exclude_instrument),
    )
    print(result.output_dir)
    print(f"instruments={len(result.output_paths)}")
    for path in result.output_paths.values():
        print(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
