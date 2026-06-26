import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from scripts.workflows.create_study_folder_gdrive.run import (
    DriveFile,
    UploadedFile,
    GoogleDriveClient,
    RefreshingSheetsHttpClient,
    add_protected_sheet_editors,
    copy_template_tree,
    copy_template_permissions,
    find_template_by_name,
    find_drive_path,
    is_redcap_instrument_workbook,
    plan_cleaned_uploads,
    replace_placeholders,
    rewrite_data_map_locations,
    rewrite_subject_timepoint_source_locations,
    update_subject_timepoints_source_links,
    upload_cleaned_data,
)


class FakeDriveClient:
    def __init__(self):
        self.files = {
            "template": DriveFile(
                id="template",
                name="STUDY IRB Template",
                mime_type="application/vnd.google-apps.folder",
                web_url="https://drive.google.com/drive/folders/template",
            ),
            "meta": DriveFile(
                id="meta",
                name="IRB-meta",
                mime_type="application/vnd.google-apps.spreadsheet",
                web_url="https://docs.google.com/spreadsheets/d/meta/edit",
            ),
            "overview_folder": DriveFile(
                id="overview_folder",
                name="Overview",
                mime_type="application/vnd.google-apps.folder",
                web_url="https://drive.google.com/drive/folders/overview_folder",
            ),
            "overview_file": DriveFile(
                id="overview_file",
                name="STUDY_IRB",
                mime_type="application/vnd.google-apps.spreadsheet",
                web_url="https://docs.google.com/spreadsheets/d/overview_file/edit",
            ),
            "data_folder": DriveFile(
                id="data_folder",
                name="Data (internal/approved-access)",
                mime_type="application/vnd.google-apps.folder",
                web_url="https://drive.google.com/drive/folders/data_folder",
            ),
            "nophi_folder": DriveFile(
                id="nophi_folder",
                name="No-PHI Data (internal/approved-access)",
                mime_type="application/vnd.google-apps.folder",
                web_url="https://drive.google.com/drive/folders/nophi_folder",
            ),
            "templates_folder": DriveFile(
                id="templates_folder",
                name="blank_templates",
                mime_type="application/vnd.google-apps.folder",
                web_url="https://drive.google.com/drive/folders/templates_folder",
            ),
            "blank_template": DriveFile(
                id="blank_template",
                name="BLANK",
                mime_type="application/vnd.google-apps.spreadsheet",
                web_url="https://docs.google.com/spreadsheets/d/blank_template/edit",
            ),
            "redcap_template": DriveFile(
                id="redcap_template",
                name="REDCap_INSTRUMENT",
                mime_type="application/vnd.google-apps.spreadsheet",
                web_url="https://docs.google.com/spreadsheets/d/redcap_template/edit",
            ),
            "data_map_folder": DriveFile(
                id="data_map_folder",
                name="Data Map (internal/approved-access)",
                mime_type="application/vnd.google-apps.folder",
                web_url="https://drive.google.com/drive/folders/data_map_folder",
            ),
            "data_map_file": DriveFile(
                id="data_map_file",
                name="IRB-data-map",
                mime_type="application/vnd.google-apps.spreadsheet",
                web_url="https://docs.google.com/spreadsheets/d/data_map_file/edit",
            ),
        }
        self.children = {
            "template": ["meta", "overview_folder", "data_folder", "data_map_folder"],
            "overview_folder": ["overview_file"],
            "data_folder": ["nophi_folder"],
            "nophi_folder": ["templates_folder"],
            "templates_folder": ["blank_template", "redcap_template"],
            "data_map_folder": ["data_map_file"],
            "dest": [],
        }
        self.created_folders = []
        self.copied_files = []
        self.uploaded_files = []
        self.updated_files = []
        self.trashed_files = []
        self.permissions = {
            "template": [
                {"id": "owner", "type": "user", "emailAddress": "owner@example.com", "role": "owner"},
                {"id": "writer", "type": "user", "emailAddress": "editor@example.com", "role": "writer"},
                {"id": "reader", "type": "group", "emailAddress": "readers@example.com", "role": "reader"},
                {
                    "id": "inherited",
                    "type": "user",
                    "emailAddress": "inherited@example.com",
                    "role": "writer",
                    "permissionDetails": [{"inherited": True}],
                },
            ],
            "dest": [],
        }
        self.created_permissions = []
        self.rejected_permission_emails = set()

    def list_children(self, folder_id):
        return [self.files[file_id] for file_id in self.children.get(folder_id, [])]

    def get_file(self, file_id):
        return self.files[file_id]

    def create_folder(self, name, parent_id):
        file_id = f"folder_{len(self.created_folders) + 1}"
        drive_file = DriveFile(
            id=file_id,
            name=name,
            mime_type="application/vnd.google-apps.folder",
            web_url=f"https://drive.google.com/drive/folders/{file_id}",
        )
        self.files[file_id] = drive_file
        self.children[file_id] = []
        self.children.setdefault(parent_id, []).append(file_id)
        self.created_folders.append((name, parent_id))
        self.permissions[file_id] = []
        return drive_file

    def copy_file(self, file_id, name, parent_id):
        copy_id = f"copy_{len(self.copied_files) + 1}"
        source = self.files[file_id]
        drive_file = DriveFile(
            id=copy_id,
            name=name,
            mime_type=source.mime_type,
            web_url=f"https://docs.google.com/spreadsheets/d/{copy_id}/edit",
        )
        self.files[copy_id] = drive_file
        self.children.setdefault(parent_id, []).append(copy_id)
        self.copied_files.append((file_id, name, parent_id))
        self.permissions[copy_id] = []
        return drive_file

    def upload_file(self, local_path, name, parent_id):
        upload_id = f"upload_{len(self.uploaded_files) + 1}"
        drive_file = DriveFile(
            id=upload_id,
            name=name,
            mime_type="application/octet-stream",
            web_url=f"https://drive.google.com/file/d/{upload_id}/view",
        )
        self.files[upload_id] = drive_file
        self.children.setdefault(parent_id, []).append(upload_id)
        self.uploaded_files.append((str(local_path), name, parent_id))
        return drive_file

    def update_file(self, local_path, file_id, name=None):
        existing = self.files[file_id]
        updated = DriveFile(
            id=existing.id,
            name=name or existing.name,
            mime_type=existing.mime_type,
            web_url=existing.web_url,
        )
        self.files[file_id] = updated
        self.updated_files.append((str(local_path), file_id, name))
        return updated

    def trash_file(self, file_id):
        self.trashed_files.append(file_id)
        for children in self.children.values():
            if file_id in children:
                children.remove(file_id)

    def list_permissions(self, file_id):
        return list(self.permissions.get(file_id, []))

    def create_permission(self, file_id, permission):
        if permission.get("emailAddress") in self.rejected_permission_emails:
            raise RuntimeError("teamDriveDomainUsersOnlyRestriction")
        created = {"id": f"permission_{len(self.created_permissions) + 1}", **permission}
        self.permissions.setdefault(file_id, []).append(created)
        self.created_permissions.append((file_id, permission))
        return created


class FakeDriveHttpClient:
    def __init__(self):
        self.get_headers = []

    def get(self, url, headers, timeout):
        self.get_headers.append(dict(headers))
        return type(
            "Response",
            (),
            {"payload": {"id": "file", "name": "File", "mimeType": "application/vnd.google-apps.folder"}},
        )()

    def post(self, url, body, headers, timeout):
        return self.get(url, headers, timeout)

    def patch(self, url, body, headers, timeout):
        return self.get(url, headers, timeout)


class FakeSheetsHttpClient:
    def __init__(self, payload=None):
        self.payload = payload or {}
        self.get_headers = []
        self.post_bodies = []

    def get(self, url, headers, timeout):
        self.get_headers.append(dict(headers))
        return type("Response", (), {"payload": self.payload})()

    def post(self, url, body, headers, timeout):
        self.get_headers.append(dict(headers))
        self.post_bodies.append(body)
        return type("Response", (), {"payload": {}})()


def create_redcap_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "raw"
    for sheet in [
        "raw_labels",
        "cleaned",
        "timepoint_dictionary",
        "column_variable_dictionary",
        "excluded_rows",
    ]:
        workbook.create_sheet(sheet)
    workbook.save(path)


def create_plain_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "cleaned"
    workbook.save(path)


class CreateStudyFolderGDriveTests(unittest.TestCase):
    def test_drive_client_refreshes_authorization_header_for_each_request(self):
        tokens = iter(["fresh_1", "fresh_2"])
        fake_http = FakeDriveHttpClient()
        drive = GoogleDriveClient(
            "stale",
            http_client=fake_http,
            token_provider=lambda: next(tokens),
        )

        drive.get_file("one")
        drive.get_file("two")

        self.assertEqual(fake_http.get_headers[0]["Authorization"], "Bearer fresh_1")
        self.assertEqual(fake_http.get_headers[1]["Authorization"], "Bearer fresh_2")

    def test_refreshing_sheets_client_refreshes_authorization_header_for_each_request(self):
        tokens = iter(["sheet_1", "sheet_2"])
        fake_http = FakeSheetsHttpClient()
        client = RefreshingSheetsHttpClient(fake_http, lambda: next(tokens))

        client.get("https://example.test/get", {"Authorization": "Bearer stale"}, 1)
        client.post("https://example.test/post", b"{}", {"Authorization": "Bearer stale"}, 1)

        self.assertEqual(fake_http.get_headers[0]["Authorization"], "Bearer sheet_1")
        self.assertEqual(fake_http.get_headers[1]["Authorization"], "Bearer sheet_2")

    def test_add_protected_sheet_editors_updates_existing_protections(self):
        fake_http = FakeSheetsHttpClient(
            {
                "sheets": [
                    {
                        "properties": {"sheetId": 1, "title": "cleaned"},
                        "protectedRanges": [
                            {
                                "protectedRangeId": 99,
                                "editors": {"users": ["owner@example.com"], "groups": []},
                            }
                        ],
                    }
                ]
            }
        )

        result = add_protected_sheet_editors(
            sheets_client=fake_http,
            spreadsheet_id="sheet123",
            access_token="token",
            emails=("new@example.com", "owner@example.com"),
        )

        self.assertEqual(result.copied_count, 1)
        body = json.loads(fake_http.post_bodies[0].decode("utf-8"))
        editors = body["requests"][0]["updateProtectedRange"]["protectedRange"]["editors"]
        self.assertEqual(editors["users"], ["owner@example.com", "new@example.com"])

    def test_add_protected_sheet_editors_skips_warning_only_protections(self):
        fake_http = FakeSheetsHttpClient(
            {
                "sheets": [
                    {
                        "protectedRanges": [
                            {
                                "protectedRangeId": 99,
                                "warningOnly": True,
                                "editors": {"users": ["owner@example.com"]},
                            }
                        ]
                    }
                ]
            }
        )

        result = add_protected_sheet_editors(
            sheets_client=fake_http,
            spreadsheet_id="sheet123",
            access_token="token",
            emails=("new@example.com",),
        )

        self.assertEqual(result.copied_count, 0)
        self.assertEqual(fake_http.post_bodies, [])

    def test_replaces_study_and_irb_placeholders_case_sensitively(self):
        self.assertEqual(
            replace_placeholders("STUDY_IRB overview", study_name="OCD-TMS", irb="53879"),
            "OCD-TMS_53879 overview",
        )
        self.assertEqual(
            replace_placeholders("IRB-meta", study_name="OCD-TMS", irb="53879"),
            "53879-meta",
        )

    def test_copies_template_tree_with_placeholder_renames(self):
        fake_drive = FakeDriveClient()

        result = copy_template_tree(
            drive=fake_drive,
            template_folder_id="template",
            destination_parent_id="dest",
            study_name="OCD-TMS",
            irb="53879",
        )

        self.assertEqual(result.root.name, "OCD-TMS 53879 Template")
        self.assertGreater(result.copied_permission_count, 3)
        copied_names = [copy_call[1] for copy_call in fake_drive.copied_files]
        self.assertIn("53879-meta", copied_names)
        self.assertIn("OCD-TMS_53879", copied_names)
        self.assertEqual(result.files_by_relative_path["Overview/OCD-TMS_53879"].name, "OCD-TMS_53879")
        self.assertEqual(result.files_by_relative_path["53879-meta"].name, "53879-meta")

    def test_copies_template_root_permissions_by_default(self):
        fake_drive = FakeDriveClient()

        result = copy_template_tree(
            drive=fake_drive,
            template_folder_id="template",
            destination_parent_id="dest",
            study_name="OCD-TMS",
            irb="53879",
        )

        self.assertIn(
            (result.root.id, {"type": "user", "role": "writer", "emailAddress": "editor@example.com"}),
            fake_drive.created_permissions,
        )
        self.assertIn(
            (result.files_by_relative_path["Overview/OCD-TMS_53879"].id, {"type": "user", "role": "writer", "emailAddress": "editor@example.com"}),
            fake_drive.created_permissions,
        )

    def test_can_skip_template_root_permission_copy(self):
        fake_drive = FakeDriveClient()

        copy_template_tree(
            drive=fake_drive,
            template_folder_id="template",
            destination_parent_id="dest",
            study_name="OCD-TMS",
            irb="53879",
            copy_template_permissions_to_root=False,
        )

        self.assertEqual(fake_drive.created_permissions, [])

    def test_does_not_duplicate_existing_template_root_permission(self):
        fake_drive = FakeDriveClient()
        fake_drive.permissions["target"] = [
            {"id": "existing", "type": "user", "emailAddress": "editor@example.com", "role": "writer"}
        ]

        copied = copy_template_permissions(fake_drive, "template", "target")

        self.assertEqual(copied.copied_count, 2)
        self.assertEqual(copied.error_count, 0)
        self.assertEqual(
            fake_drive.created_permissions,
            [
                ("target", {"type": "group", "role": "reader", "emailAddress": "readers@example.com"}),
                ("target", {"type": "user", "role": "writer", "emailAddress": "inherited@example.com"}),
            ],
        )

    def test_permission_copy_continues_when_drive_rejects_permission(self):
        fake_drive = FakeDriveClient()
        fake_drive.rejected_permission_emails.add("inherited@example.com")

        result = copy_template_tree(
            drive=fake_drive,
            template_folder_id="template",
            destination_parent_id="dest",
            study_name="OCD-TMS",
            irb="53879",
        )

        self.assertGreater(result.copied_permission_count, 2)
        self.assertGreater(result.permission_error_count, 1)
        self.assertIn("teamDriveDomainUsersOnlyRestriction", result.permission_errors[0])

    def test_explicit_sheet_editors_apply_only_to_template_google_sheets(self):
        fake_drive = FakeDriveClient()

        result = copy_template_tree(
            drive=fake_drive,
            template_folder_id="template",
            destination_parent_id="dest",
            study_name="OCD-TMS",
            irb="53879",
            copy_template_permissions_to_root=False,
            sheet_editor_emails=("analyst@gmail.com",),
        )

        explicit_permission = {"type": "user", "role": "writer", "emailAddress": "analyst@gmail.com"}
        explicit_targets = [
            file_id for file_id, permission in fake_drive.created_permissions if permission == explicit_permission
        ]
        self.assertNotIn(result.root.id, explicit_targets)
        self.assertTrue(explicit_targets)
        self.assertTrue(all(fake_drive.files[file_id].is_google_sheet for file_id in explicit_targets))

    def test_cleaned_upload_stamps_permissions_onto_generated_workbook(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            fake_drive = FakeDriveClient()
            study = Path(tmpdir) / "study"
            subjects = study / "data" / "cleaned" / "subjects"
            subjects.mkdir(parents=True)
            dictionary = subjects / "dictionary.xlsx"
            create_plain_workbook(dictionary)

            with patch("scripts.workflows.create_study_folder_gdrive.run.fill_in_overview"):
                results = upload_cleaned_data(
                    drive=fake_drive,
                    target_data_folder_id="nophi_folder",
                    template_folder_id="templates_folder",
                    study_folder=study,
                    access_token="token",
                    permission_source_file_id="template",
                )

            self.assertEqual(len(results), 1)
            self.assertEqual(results[0].drive_file.id, "copy_1")
            self.assertIn(
                ("copy_1", {"type": "user", "role": "writer", "emailAddress": "editor@example.com"}),
                fake_drive.created_permissions,
            )

    def test_cleaned_upload_adds_explicit_editors_only_to_google_sheets(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            fake_drive = FakeDriveClient()
            study = Path(tmpdir) / "study"
            subjects = study / "data" / "cleaned" / "subjects"
            subjects.mkdir(parents=True)
            dictionary = subjects / "dictionary.xlsx"
            create_plain_workbook(dictionary)

            with patch("scripts.workflows.create_study_folder_gdrive.run.fill_in_overview"):
                results = upload_cleaned_data(
                    drive=fake_drive,
                    target_data_folder_id="nophi_folder",
                    template_folder_id="templates_folder",
                    study_folder=study,
                    access_token="token",
                    sheet_editor_emails=("analyst@gmail.com",),
                )

            self.assertEqual(len(results), 1)
            self.assertEqual(results[0].drive_file.id, "copy_1")
            self.assertIn(
                ("copy_1", {"type": "user", "role": "writer", "emailAddress": "analyst@gmail.com"}),
                fake_drive.created_permissions,
            )

    def test_update_or_create_template_tree_reuses_existing_study_root(self):
        fake_drive = FakeDriveClient()
        existing_root = DriveFile(
            id="existing_root",
            name="OCD-TMS 53879 Template",
            mime_type="application/vnd.google-apps.folder",
            web_url="https://drive.google.com/drive/folders/existing_root",
        )
        fake_drive.files[existing_root.id] = existing_root
        fake_drive.children["dest"].append(existing_root.id)
        fake_drive.children[existing_root.id] = []

        result = copy_template_tree(
            drive=fake_drive,
            template_folder_id="template",
            destination_parent_id="dest",
            study_name="OCD-TMS",
            irb="53879",
            existing_file_policy="update-or-create",
        )

        self.assertEqual(result.root.id, "existing_root")
        self.assertNotIn(("OCD-TMS 53879 Template", "dest"), fake_drive.created_folders)
        self.assertIn("53879-meta", result.files_by_relative_path)

    def test_find_drive_path_matches_nested_template_locations(self):
        fake_drive = FakeDriveClient()

        found = find_drive_path(
            fake_drive,
            "template",
            "Data (internal/approved-access)/No-PHI Data (internal/approved-access)/blank_templates",
        )

        self.assertIsNotNone(found)
        self.assertEqual(found.id, "templates_folder")

    def test_plans_cleaned_uploads_with_template_selection_and_class_folder(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study = Path(tmpdir) / "study"
            assessments = study / "data" / "cleaned" / "assessments"
            subjects = study / "data" / "cleaned" / "subjects"
            assessments.mkdir(parents=True)
            subjects.mkdir(parents=True)
            redcap = assessments / "53879-madrs.xlsx"
            plain = subjects / "subject_timepoints.xlsx"
            create_redcap_workbook(redcap)
            create_plain_workbook(plain)

            planned = plan_cleaned_uploads(study)

            by_name = {item.local_path.name: item for item in planned}
            self.assertEqual(by_name["53879-madrs.xlsx"].template_name, "REDCap_INSTRUMENT")
            self.assertEqual(by_name["53879-madrs.xlsx"].relative_parent, Path("assessments"))
            self.assertEqual(by_name["subject_timepoints.xlsx"].template_name, "BLANK")
            self.assertEqual(by_name["subject_timepoints.xlsx"].relative_parent, Path("subjects"))

    def test_update_or_create_cleaned_upload_reuses_existing_template_sheet(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            fake_drive = FakeDriveClient()
            study = Path(tmpdir) / "study"
            subjects = study / "data" / "cleaned" / "subjects"
            subjects.mkdir(parents=True)
            dictionary = subjects / "dictionary.xlsx"
            create_plain_workbook(dictionary)
            subjects_folder = fake_drive.create_folder("subjects", "nophi_folder")
            existing = DriveFile(
                id="existing_dictionary",
                name="dictionary",
                mime_type="application/vnd.google-apps.spreadsheet",
                web_url="https://docs.google.com/spreadsheets/d/existing_dictionary/edit",
            )
            fake_drive.files[existing.id] = existing
            fake_drive.children[subjects_folder.id].append(existing.id)

            with patch("scripts.workflows.create_study_folder_gdrive.run.fill_in_overview") as fill:
                results = upload_cleaned_data(
                    drive=fake_drive,
                    target_data_folder_id="nophi_folder",
                    template_folder_id="templates_folder",
                    study_folder=study,
                    access_token="token",
                    existing_file_policy="update-or-create",
                )

            self.assertEqual(len(results), 1)
            self.assertEqual(results[0].drive_file.id, "existing_dictionary")
            self.assertEqual(fake_drive.copied_files, [])
            fill.assert_called_once()
            self.assertEqual(fill.call_args.kwargs["target"], "existing_dictionary")

    def test_replace_cleaned_upload_trashes_existing_template_sheet_then_copies(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            fake_drive = FakeDriveClient()
            study = Path(tmpdir) / "study"
            subjects = study / "data" / "cleaned" / "subjects"
            subjects.mkdir(parents=True)
            dictionary = subjects / "dictionary.xlsx"
            create_plain_workbook(dictionary)
            subjects_folder = fake_drive.create_folder("subjects", "nophi_folder")
            existing = DriveFile(
                id="existing_dictionary",
                name="dictionary",
                mime_type="application/vnd.google-apps.spreadsheet",
                web_url="https://docs.google.com/spreadsheets/d/existing_dictionary/edit",
            )
            fake_drive.files[existing.id] = existing
            fake_drive.children[subjects_folder.id].append(existing.id)

            with patch("scripts.workflows.create_study_folder_gdrive.run.fill_in_overview") as fill:
                results = upload_cleaned_data(
                    drive=fake_drive,
                    target_data_folder_id="nophi_folder",
                    template_folder_id="templates_folder",
                    study_folder=study,
                    access_token="token",
                    existing_file_policy="replace",
                )

            self.assertEqual(fake_drive.trashed_files, ["existing_dictionary"])
            self.assertEqual(results[0].drive_file.id, "copy_1")
            self.assertEqual(fake_drive.copied_files, [("blank_template", "dictionary", subjects_folder.id)])
            fill.assert_called_once()
            self.assertEqual(fill.call_args.kwargs["target"], "copy_1")

    def test_redcap_template_lookup_uses_actual_template_name_and_legacy_aliases(self):
        template = DriveFile(
            id="redcap_template",
            name="REDCap_INSTRUMENT",
            mime_type="application/vnd.google-apps.spreadsheet",
        )

        self.assertEqual(find_template_by_name([template], "REDCap_INSTRUMENT"), template)
        self.assertEqual(find_template_by_name([template], "REDCap_INSTRUMENTS"), template)
        self.assertEqual(find_template_by_name([template], "REDCap Instrument"), template)

    def test_redcap_instrument_detection_requires_self_contained_sheet_set(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            redcap = Path(tmpdir) / "instrument.xlsx"
            plain = Path(tmpdir) / "plain.xlsx"
            create_redcap_workbook(redcap)
            create_plain_workbook(plain)

            self.assertTrue(is_redcap_instrument_workbook(redcap))
            self.assertFalse(is_redcap_instrument_workbook(plain))

    def test_rewrites_data_map_locations_to_uploaded_drive_paths(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            source_dir = Path(tmpdir) / "data-map"
            output_dir = Path(tmpdir) / "rewritten"
            source_dir.mkdir()
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "data_map"
            worksheet.append(["stage", "description", "location"])
            worksheet.append(["raw", "", ""])
            worksheet.append(
                [
                    "cleaned/processed",
                    "MADRS",
                    "./data/cleaned/assessments/53879-madrs.xlsx",
                ]
            )
            worksheet.append(
                [
                    "cleaned/processed",
                    "Missing upload",
                    "./data/cleaned/assessments/missing.xlsx",
                ]
            )
            workbook.save(source_dir / "assessments-data-map.xlsx")

            rewritten = rewrite_data_map_locations(
                source_dir,
                output_dir,
                {
                    "data/cleaned/assessments/53879-madrs.xlsx": "https://docs.google.com/spreadsheets/d/madrs/edit"
                },
            )

            self.assertEqual(len(rewritten), 1)
            result = load_workbook(rewritten[0], data_only=True)
            rows = list(result.active.iter_rows(values_only=True))
            self.assertEqual(rows[2][2], "https://docs.google.com/spreadsheets/d/madrs/edit")
            self.assertIsNone(rows[3][2])

    def test_rewrites_data_map_locations_from_cleaned_upload_relative_paths(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            source_dir = Path(tmpdir) / "data-map"
            output_dir = Path(tmpdir) / "rewritten"
            source_dir.mkdir()
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "data_map"
            worksheet.append(["stage", "description", "location"])
            worksheet.append(
                [
                    "cleaned/processed",
                    "MADRS",
                    "./data/cleaned/assessments/53879-madrs.xlsx",
                ]
            )
            workbook.save(source_dir / "assessments-data-map.xlsx")

            rewritten = rewrite_data_map_locations(
                source_dir,
                output_dir,
                {
                    "assessments/53879-madrs.xlsx": "https://docs.google.com/spreadsheets/d/madrs/edit"
                },
            )

            result = load_workbook(rewritten[0], data_only=True)
            rows = list(result.active.iter_rows(values_only=True))
            self.assertEqual(rows[1][2], "https://docs.google.com/spreadsheets/d/madrs/edit")

    def test_rewrites_subject_timepoint_source_paths_to_uploaded_drive_links(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            source_path = Path(tmpdir) / "subject_timepoints.xlsx"
            output_path = Path(tmpdir) / "rewritten" / "subject_timepoints.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "subject_timepoints"
            worksheet.append(
                [
                    "IRB",
                    "subid",
                    "arm",
                    "visit",
                    "earliest_entry_date",
                    "earliest_date_source",
                    "latest_entry_date",
                    "latest_date_source",
                    "span",
                    "values",
                ]
            )
            worksheet.append(
                [
                    "53879",
                    "s001",
                    "1",
                    "V2",
                    "2021-01-01",
                    "studies/example/data/cleaned/assessments/53879-madrs.xlsx; /tmp/example/data/cleaned/treatments/53879-medication.xlsx",
                    "2021-01-03",
                    "/tmp/example/data/cleaned/neuroimaging/53879-eeg.xlsx",
                    2,
                    "2021-01-01; 2021-01-03",
                ]
            )
            worksheet.append(
                [
                    "53879",
                    "s002",
                    "1",
                    "V2",
                    "2021-01-01",
                    "studies/example/data/cleaned/assessments/missing.xlsx",
                    "2021-01-03",
                    "",
                    2,
                    "2021-01-01; 2021-01-03",
                ]
            )
            workbook.save(source_path)

            rewritten = rewrite_subject_timepoint_source_locations(
                source_path,
                output_path,
                {
                    "assessments/53879-madrs.xlsx": "https://docs.google.com/spreadsheets/d/madrs/edit",
                    "treatments/53879-medication.xlsx": "https://docs.google.com/spreadsheets/d/medication/edit",
                    "neuroimaging/53879-eeg.xlsx": "https://docs.google.com/spreadsheets/d/eeg/edit",
                },
            )

            result = load_workbook(rewritten, data_only=True)
            rows = list(result["subject_timepoints"].iter_rows(values_only=True))
            self.assertEqual(
                rows[1][5],
                "https://docs.google.com/spreadsheets/d/madrs/edit; https://docs.google.com/spreadsheets/d/medication/edit",
            )
            self.assertEqual(rows[1][7], "https://docs.google.com/spreadsheets/d/eeg/edit")
            self.assertIsNone(rows[2][5])

    def test_updates_uploaded_subject_timepoints_google_sheet_with_drive_source_links(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study = Path(tmpdir) / "study"
            subject_timepoints = study / "data" / "cleaned" / "subjects" / "subject_timepoints.xlsx"
            subject_timepoints.parent.mkdir(parents=True)
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "subject_timepoints"
            worksheet.append(
                [
                    "IRB",
                    "subid",
                    "arm",
                    "visit",
                    "earliest_entry_date",
                    "earliest_date_source",
                    "latest_entry_date",
                    "latest_date_source",
                    "span",
                    "values",
                ]
            )
            worksheet.append(
                [
                    "53879",
                    "s001",
                    "1",
                    "V2",
                    "2021-01-01",
                    str(study / "data" / "cleaned" / "assessments" / "53879-madrs.xlsx"),
                    "2021-01-03",
                    "",
                    2,
                    "2021-01-01; 2021-01-03",
                ]
            )
            workbook.save(subject_timepoints)
            upload_results = [
                UploadedFile(
                    local_path=study / "data" / "cleaned" / "assessments" / "53879-madrs.xlsx",
                    relative_path=Path("assessments/53879-madrs.xlsx"),
                    drive_file=DriveFile(
                        id="madrs_sheet",
                        name="53879-madrs",
                        mime_type="application/vnd.google-apps.spreadsheet",
                        web_url="https://docs.google.com/spreadsheets/d/madrs/edit",
                    ),
                ),
                UploadedFile(
                    local_path=subject_timepoints,
                    relative_path=Path("subjects/subject_timepoints.xlsx"),
                    drive_file=DriveFile(
                        id="subject_timepoints_sheet",
                        name="subject_timepoints",
                        mime_type="application/vnd.google-apps.spreadsheet",
                        web_url="https://docs.google.com/spreadsheets/d/subject_timepoints/edit",
                    ),
                ),
            ]

            with patch("scripts.workflows.create_study_folder_gdrive.run.fill_in_overview") as fill:
                rewritten = update_subject_timepoints_source_links(
                    study_folder=study,
                    upload_results=upload_results,
                    access_token="token",
                    sheets_client=None,
                    timeout=1,
                )

            self.assertIsNotNone(rewritten)
            result = load_workbook(rewritten, data_only=True)
            rows = list(result["subject_timepoints"].iter_rows(values_only=True))
            self.assertEqual(rows[1][5], "https://docs.google.com/spreadsheets/d/madrs/edit")
            fill.assert_called_once()
            self.assertEqual(fill.call_args.kwargs["target"], "subject_timepoints_sheet")


if __name__ == "__main__":
    unittest.main()
