#!/usr/bin/env python3
"""Create and populate a Google Drive study folder from local cleaned outputs."""

from __future__ import annotations

import argparse
import csv
import json
import mimetypes
import re
import sys
import tempfile
import urllib.error
import urllib.parse
import urllib.request
import uuid
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any, Callable, Mapping, Protocol

from openpyxl import Workbook, load_workbook


if __package__ in {None, ""}:
    sys.path.append(str(Path(__file__).resolve().parents[3]))

from scripts.push_to_gdrive.fill_in_overview import (  # noqa: E402
    OverviewSheet,
    SheetsHttpClient,
    UrllibSheetsHttpClient,
    build_value_updates,
    json_headers,
    load_sheet_properties,
    load_target_headers,
    sheets_values_batch_update_url,
)
from scripts.push_to_gdrive.fill_in_overview import fill_in_overview  # noqa: E402
from scripts.push_to_gdrive.push_data_map import push_data_map  # noqa: E402
from scripts.push_to_gdrive.push_instrument_workbook import (  # noqa: E402
    resolve_access_token,
)


DRIVE_API_ROOT = "https://www.googleapis.com/drive/v3/files"
DRIVE_UPLOAD_ROOT = "https://www.googleapis.com/upload/drive/v3/files"
FOLDER_MIME_TYPE = "application/vnd.google-apps.folder"
SHEET_MIME_TYPE = "application/vnd.google-apps.spreadsheet"
REDCAP_TEMPLATE_NAME = "REDCap_INSTRUMENT"
REDCAP_TEMPLATE_ALIASES = (
    REDCAP_TEMPLATE_NAME,
    "REDCap_INSTRUMENTS",
    "REDCap_Instrument",
    "REDCap Instruments",
    "REDCap Instrument",
)
BLANK_TEMPLATE_NAME = "BLANK"
REDCAP_INSTRUMENT_SHEETS = {
    "raw",
    "raw_labels",
    "cleaned",
    "timepoint_dictionary",
    "column_variable_dictionary",
    "excluded_rows",
}
LOCAL_TOP_LEVEL_EXCLUDE = {"data", "data-map", "overview", "histories", "studies"}
DEFAULT_CLEANED_GDRIVE_PATH = "Data (internal/approved-access)/No-PHI Data (internal/approved-access)"
DEFAULT_TEMPLATE_GDRIVE_PATH = f"{DEFAULT_CLEANED_GDRIVE_PATH}/blank_templates"
DEFAULT_DATA_MAP_FOLDER = "Data Map (internal/approved-access)"
EXISTING_FILE_POLICIES = ("update-or-create", "skip", "duplicate", "replace", "fail")
DEFAULT_EXISTING_FILE_POLICY = "update-or-create"


@dataclass(frozen=True)
class DriveFile:
    id: str
    name: str
    mime_type: str
    web_url: str | None = None

    @property
    def is_folder(self) -> bool:
        return self.mime_type == FOLDER_MIME_TYPE

    @property
    def is_google_sheet(self) -> bool:
        return self.mime_type == SHEET_MIME_TYPE


@dataclass(frozen=True)
class DriveResponse:
    status_code: int
    payload: dict[str, Any]
    raw_text: str


class DriveHttpClient(Protocol):
    def get(self, url: str, headers: Mapping[str, str], timeout: float) -> DriveResponse:
        ...

    def post(
        self,
        url: str,
        body: bytes,
        headers: Mapping[str, str],
        timeout: float,
    ) -> DriveResponse:
        ...

    def patch(
        self,
        url: str,
        body: bytes,
        headers: Mapping[str, str],
        timeout: float,
    ) -> DriveResponse:
        ...


class UrllibDriveHttpClient:
    def get(self, url: str, headers: Mapping[str, str], timeout: float) -> DriveResponse:
        request = urllib.request.Request(url, headers=dict(headers), method="GET")
        return _send_drive_request(request, timeout)

    def post(
        self,
        url: str,
        body: bytes,
        headers: Mapping[str, str],
        timeout: float,
    ) -> DriveResponse:
        request = urllib.request.Request(
            url,
            data=body,
            headers=dict(headers),
            method="POST",
        )
        return _send_drive_request(request, timeout)

    def patch(
        self,
        url: str,
        body: bytes,
        headers: Mapping[str, str],
        timeout: float,
    ) -> DriveResponse:
        request = urllib.request.Request(
            url,
            data=body,
            headers=dict(headers),
            method="PATCH",
        )
        return _send_drive_request(request, timeout)


class RefreshingSheetsHttpClient:
    def __init__(self, wrapped: SheetsHttpClient, token_provider: Callable[[], str]) -> None:
        self.wrapped = wrapped
        self.token_provider = token_provider

    def _headers(self, headers: Mapping[str, str]) -> dict[str, str]:
        refreshed = dict(headers)
        refreshed["Authorization"] = f"Bearer {self.token_provider()}"
        return refreshed

    def get(self, url: str, headers: Mapping[str, str], timeout: float) -> Any:
        return self.wrapped.get(url, self._headers(headers), timeout)

    def post(self, url: str, body: bytes, headers: Mapping[str, str], timeout: float) -> Any:
        return self.wrapped.post(url, body, self._headers(headers), timeout)


def _send_drive_request(request: urllib.request.Request, timeout: float) -> DriveResponse:
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            raw = response.read().decode("utf-8")
            payload = json.loads(raw) if raw.strip() else {}
            return DriveResponse(response.status, payload, raw)
    except urllib.error.HTTPError as exc:
        raw = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Google Drive request failed with HTTP {exc.code}: {raw}") from exc


class GoogleDriveClient:
    def __init__(
        self,
        access_token: str,
        *,
        http_client: DriveHttpClient | None = None,
        timeout: float = 120.0,
        token_provider: Callable[[], str] | None = None,
    ) -> None:
        self.token_provider = token_provider
        self.headers = {
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json; charset=UTF-8",
        }
        self.upload_headers = {"Authorization": f"Bearer {access_token}"}
        self.http_client = http_client or UrllibDriveHttpClient()
        self.timeout = timeout

    def _refresh_authorization_headers(self) -> None:
        if self.token_provider is None:
            return
        token = self.token_provider()
        self.headers["Authorization"] = f"Bearer {token}"
        self.upload_headers["Authorization"] = f"Bearer {token}"

    def get_file(self, file_id: str) -> DriveFile:
        self._refresh_authorization_headers()
        url = (
            f"{DRIVE_API_ROOT}/{file_id}"
            "?supportsAllDrives=true&fields=id,name,mimeType,webViewLink"
        )
        return drive_file_from_payload(self.http_client.get(url, self.headers, self.timeout).payload)

    def list_children(self, folder_id: str) -> list[DriveFile]:
        self._refresh_authorization_headers()
        files: list[DriveFile] = []
        page_token: str | None = None
        while True:
            params = {
                "q": f"'{folder_id}' in parents and trashed=false",
                "fields": "nextPageToken,files(id,name,mimeType,webViewLink)",
                "supportsAllDrives": "true",
                "includeItemsFromAllDrives": "true",
                "pageSize": "1000",
            }
            if page_token:
                params["pageToken"] = page_token
            url = f"{DRIVE_API_ROOT}?{urllib.parse.urlencode(params)}"
            payload = self.http_client.get(url, self.headers, self.timeout).payload
            files.extend(drive_file_from_payload(file) for file in payload.get("files", []))
            page_token = payload.get("nextPageToken")
            if not page_token:
                break
        return files

    def create_folder(self, name: str, parent_id: str) -> DriveFile:
        self._refresh_authorization_headers()
        payload = {
            "name": name,
            "mimeType": FOLDER_MIME_TYPE,
            "parents": [parent_id],
        }
        url = f"{DRIVE_API_ROOT}?supportsAllDrives=true&fields=id,name,mimeType,webViewLink"
        return drive_file_from_payload(
            self.http_client.post(
                url,
                json.dumps(payload).encode("utf-8"),
                self.headers,
                self.timeout,
            ).payload
        )

    def copy_file(self, file_id: str, name: str, parent_id: str) -> DriveFile:
        self._refresh_authorization_headers()
        payload = {"name": name, "parents": [parent_id]}
        url = (
            f"{DRIVE_API_ROOT}/{file_id}/copy"
            "?supportsAllDrives=true&fields=id,name,mimeType,webViewLink"
        )
        return drive_file_from_payload(
            self.http_client.post(
                url,
                json.dumps(payload).encode("utf-8"),
                self.headers,
                self.timeout,
            ).payload
        )

    def upload_file(self, local_path: Path, name: str, parent_id: str) -> DriveFile:
        self._refresh_authorization_headers()
        mime_type = mimetypes.guess_type(local_path.name)[0] or "application/octet-stream"
        boundary = f"codex_{uuid.uuid4().hex}"
        metadata = {"name": name, "parents": [parent_id]}
        body = build_multipart_upload_body(
            metadata=metadata,
            file_bytes=local_path.read_bytes(),
            file_mime_type=mime_type,
            boundary=boundary,
        )
        headers = {
            **self.upload_headers,
            "Content-Type": f"multipart/related; boundary={boundary}",
            "Content-Length": str(len(body)),
        }
        url = (
            f"{DRIVE_UPLOAD_ROOT}"
            "?uploadType=multipart&supportsAllDrives=true&fields=id,name,mimeType,webViewLink"
        )
        return drive_file_from_payload(self.http_client.post(url, body, headers, self.timeout).payload)

    def update_file(self, local_path: Path, file_id: str, name: str | None = None) -> DriveFile:
        self._refresh_authorization_headers()
        mime_type = mimetypes.guess_type(local_path.name)[0] or "application/octet-stream"
        boundary = f"codex_{uuid.uuid4().hex}"
        metadata = {"name": name} if name else {}
        body = build_multipart_upload_body(
            metadata=metadata,
            file_bytes=local_path.read_bytes(),
            file_mime_type=mime_type,
            boundary=boundary,
        )
        headers = {
            **self.upload_headers,
            "Content-Type": f"multipart/related; boundary={boundary}",
            "Content-Length": str(len(body)),
        }
        url = (
            f"{DRIVE_UPLOAD_ROOT}/{file_id}"
            "?uploadType=multipart&supportsAllDrives=true&fields=id,name,mimeType,webViewLink"
        )
        return drive_file_from_payload(self.http_client.patch(url, body, headers, self.timeout).payload)

    def trash_file(self, file_id: str) -> DriveFile:
        self._refresh_authorization_headers()
        url = (
            f"{DRIVE_API_ROOT}/{file_id}"
            "?supportsAllDrives=true&fields=id,name,mimeType,webViewLink"
        )
        payload = {"trashed": True}
        return drive_file_from_payload(
            self.http_client.patch(
                url,
                json.dumps(payload).encode("utf-8"),
                self.headers,
                self.timeout,
            ).payload
        )

    def list_permissions(self, file_id: str) -> list[dict[str, Any]]:
        self._refresh_authorization_headers()
        params = {
            "supportsAllDrives": "true",
            "fields": (
                "permissions("
                "id,type,emailAddress,domain,role,allowFileDiscovery,deleted,"
                "permissionDetails(inherited)"
                ")"
            ),
        }
        url = f"{DRIVE_API_ROOT}/{file_id}/permissions?{urllib.parse.urlencode(params)}"
        payload = self.http_client.get(url, self.headers, self.timeout).payload
        return [dict(permission) for permission in payload.get("permissions", [])]

    def create_permission(self, file_id: str, permission: Mapping[str, Any]) -> dict[str, Any]:
        self._refresh_authorization_headers()
        params = {
            "supportsAllDrives": "true",
            "sendNotificationEmail": "false",
            "fields": "id,type,emailAddress,domain,role,allowFileDiscovery",
        }
        url = f"{DRIVE_API_ROOT}/{file_id}/permissions?{urllib.parse.urlencode(params)}"
        return dict(
            self.http_client.post(
                url,
                json.dumps(dict(permission)).encode("utf-8"),
                self.headers,
                self.timeout,
            ).payload
        )


@dataclass(frozen=True)
class PermissionCopyResult:
    copied_count: int = 0
    errors: tuple[str, ...] = ()

    @property
    def error_count(self) -> int:
        return len(self.errors)


@dataclass(frozen=True)
class TemplateCopyResult:
    root: DriveFile
    files_by_relative_path: dict[str, DriveFile]
    copied_permission_count: int = 0
    permission_errors: tuple[str, ...] = ()

    @property
    def permission_error_count(self) -> int:
        return len(self.permission_errors)


@dataclass(frozen=True)
class CleanedUploadPlan:
    local_path: Path
    relative_path: Path
    relative_parent: Path
    template_name: str | None
    is_redcap_instrument: bool


@dataclass
class UploadedFile:
    local_path: Path
    relative_path: Path
    drive_file: DriveFile | None = None
    error: str | None = None

    @property
    def ok(self) -> bool:
        return self.drive_file is not None and self.error is None


@dataclass
class WorkflowState:
    study_folder: Path
    study_name: str
    irb: str
    initialized_root: DriveFile | None = None
    template_files: dict[str, DriveFile] = field(default_factory=dict)
    upload_results: list[UploadedFile] = field(default_factory=list)
    overview_result: Any | None = None
    data_map_result: Any | None = None


def drive_file_from_payload(payload: Mapping[str, Any]) -> DriveFile:
    return DriveFile(
        id=str(payload.get("id", "")),
        name=str(payload.get("name", "")),
        mime_type=str(payload.get("mimeType", "")),
        web_url=payload.get("webViewLink"),
    )


def build_multipart_upload_body(
    *,
    metadata: Mapping[str, Any],
    file_bytes: bytes,
    file_mime_type: str,
    boundary: str,
) -> bytes:
    metadata_json = json.dumps(dict(metadata), separators=(",", ":")).encode("utf-8")
    chunks = [
        f"--{boundary}\r\n".encode("utf-8"),
        b"Content-Type: application/json; charset=UTF-8\r\n\r\n",
        metadata_json,
        b"\r\n",
        f"--{boundary}\r\n".encode("utf-8"),
        f"Content-Type: {file_mime_type}\r\n\r\n".encode("utf-8"),
        file_bytes,
        b"\r\n",
        f"--{boundary}--\r\n".encode("utf-8"),
    ]
    return b"".join(chunks)


def resolve_drive_id(value: str | Path) -> str:
    text = str(value).strip()
    if text == "root":
        return text
    path = Path(text).expanduser()
    if path.exists() and path.is_file():
        text = path.read_text(encoding="utf-8").strip()
    patterns = [
        r"drive\.google\.com/drive/folders/([A-Za-z0-9_-]+)",
        r"drive\.google\.com/file/d/([A-Za-z0-9_-]+)",
        r"docs\.google\.com/[a-z]+/d/([A-Za-z0-9_-]+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1)
    if re.fullmatch(r"[A-Za-z0-9_-]{10,}", text):
        return text
    raise ValueError(f"Could not resolve Google Drive id from: {value}")


def replace_placeholders(name: str, *, study_name: str, irb: str) -> str:
    return name.replace("STUDY", study_name).replace("IRB", irb)


def validate_existing_file_policy(policy: str) -> str:
    if policy not in EXISTING_FILE_POLICIES:
        raise ValueError(
            f"Invalid existing file policy {policy!r}; expected one of {', '.join(EXISTING_FILE_POLICIES)}"
        )
    return policy


def parse_email_values(values: list[str] | tuple[str, ...] | None) -> tuple[str, ...]:
    emails: list[str] = []
    seen: set[str] = set()
    for value in values or []:
        for email in re.split(r"[\s,;]+", str(value).strip()):
            email = email.strip()
            if not email:
                continue
            key = email.lower()
            if key in seen:
                continue
            seen.add(key)
            emails.append(email)
    return tuple(emails)


def google_auth_default_token_provider() -> Callable[[], str] | None:
    try:
        import google.auth  # type: ignore
        from google.auth.transport.requests import Request  # type: ignore
    except Exception:
        return None

    try:
        credentials, _ = google.auth.default(
            scopes=[
                "https://www.googleapis.com/auth/drive",
                "https://www.googleapis.com/auth/spreadsheets",
            ]
        )
    except Exception:
        return None

    request = Request()

    def token_provider() -> str:
        if not getattr(credentials, "valid", False):
            credentials.refresh(request)
        token = getattr(credentials, "token", None)
        if not token:
            credentials.refresh(request)
            token = getattr(credentials, "token", None)
        if not token:
            raise RuntimeError("Could not refresh Google OAuth access token from google.auth.default().")
        return str(token)

    return token_provider


def build_token_provider(initial_token: str) -> Callable[[], str]:
    google_provider = google_auth_default_token_provider()
    if google_provider is not None:
        return google_provider
    return lambda: initial_token


def resolve_existing_drive_file(
    *,
    drive: Any,
    parent_id: str,
    name: str,
    existing_file_policy: str,
) -> DriveFile | None:
    policy = validate_existing_file_policy(existing_file_policy)
    if policy == "duplicate":
        return None
    existing = find_child_by_name(drive, parent_id, name)
    if existing is None:
        return None
    if policy == "fail":
        raise FileExistsError(f"Google Drive file already exists: {name}")
    if policy == "replace":
        drive.trash_file(existing.id)
        return None
    return existing


def find_or_create_drive_folder(
    *,
    drive: Any,
    parent_id: str,
    name: str,
    existing_file_policy: str,
) -> DriveFile:
    existing = resolve_existing_drive_file(
        drive=drive,
        parent_id=parent_id,
        name=name,
        existing_file_policy=existing_file_policy,
    )
    if existing is not None:
        if not existing.is_folder:
            raise FileExistsError(f"Google Drive item exists but is not a folder: {name}")
        return existing
    return drive.create_folder(name, parent_id)


def permission_key(permission: Mapping[str, Any]) -> tuple[str, str]:
    permission_type = str(permission.get("type", "")).lower()
    if permission_type in {"user", "group"}:
        return permission_type, str(permission.get("emailAddress", "")).lower()
    if permission_type == "domain":
        return permission_type, str(permission.get("domain", "")).lower()
    return permission_type, ""


def permission_is_inherited(permission: Mapping[str, Any]) -> bool:
    details = permission.get("permissionDetails")
    if not isinstance(details, list) or not details:
        return False
    return all(bool(detail.get("inherited")) for detail in details if isinstance(detail, Mapping))


def copyable_permission_body(permission: Mapping[str, Any]) -> dict[str, Any] | None:
    if permission.get("deleted"):
        return None
    role = str(permission.get("role", "")).lower()
    permission_type = str(permission.get("type", "")).lower()
    if not role or role == "owner" or not permission_type:
        return None

    body: dict[str, Any] = {"type": permission_type, "role": role}
    if permission_type in {"user", "group"}:
        email = str(permission.get("emailAddress", "")).strip()
        if not email:
            return None
        body["emailAddress"] = email
    elif permission_type == "domain":
        domain = str(permission.get("domain", "")).strip()
        if not domain:
            return None
        body["domain"] = domain
    elif permission_type != "anyone":
        return None

    if "allowFileDiscovery" in permission:
        body["allowFileDiscovery"] = bool(permission.get("allowFileDiscovery"))
    return body


def copy_template_permissions(drive: Any, template_file_id: str, target_file_id: str) -> PermissionCopyResult:
    existing_keys = {
        permission_key(permission)
        for permission in drive.list_permissions(target_file_id)
        if permission_key(permission) != ("", "")
    }
    copied_count = 0
    errors: list[str] = []
    for source_permission in drive.list_permissions(template_file_id):
        body = copyable_permission_body(source_permission)
        if body is None:
            continue
        key = permission_key(body)
        if not key[0] or key in existing_keys:
            continue
        try:
            drive.create_permission(target_file_id, body)
        except Exception as exc:
            errors.append(f"{key[0]}:{key[1]} role={body.get('role')}: {exc}")
            continue
        else:
            existing_keys.add(key)
            copied_count += 1
    return PermissionCopyResult(copied_count=copied_count, errors=tuple(errors))


def add_sheet_editor_permissions(
    drive: Any,
    drive_file: DriveFile,
    emails: tuple[str, ...] | list[str] | None,
) -> PermissionCopyResult:
    parsed_emails = parse_email_values(tuple(emails or ()))
    if not parsed_emails or not drive_file.is_google_sheet:
        return PermissionCopyResult()

    existing_keys = {
        permission_key(permission)
        for permission in drive.list_permissions(drive_file.id)
        if permission_key(permission) != ("", "")
    }
    copied_count = 0
    errors: list[str] = []
    for email in parsed_emails:
        body = {"type": "user", "role": "writer", "emailAddress": email}
        key = permission_key(body)
        if key in existing_keys:
            continue
        try:
            drive.create_permission(drive_file.id, body)
        except Exception as exc:
            errors.append(f"user:{email.lower()} role=writer: {exc}")
            continue
        existing_keys.add(key)
        copied_count += 1
    return PermissionCopyResult(copied_count=copied_count, errors=tuple(errors))


def irb_meta_path_candidates(irb_meta_path: str, *, study_name: str, irb: str) -> list[str]:
    resolved = replace_placeholders(irb_meta_path, study_name=study_name, irb=irb)
    candidates = [resolved]
    if irb_meta_path != resolved:
        candidates.append(irb_meta_path)
    return candidates


def copy_template_tree(
    *,
    drive: Any,
    template_folder_id: str,
    destination_parent_id: str,
    study_name: str,
    irb: str,
    existing_file_policy: str = DEFAULT_EXISTING_FILE_POLICY,
    copy_template_permissions_to_root: bool = True,
    sheet_editor_emails: tuple[str, ...] = (),
) -> TemplateCopyResult:
    template_root = drive.get_file(template_folder_id)
    root_name = replace_placeholders(template_root.name, study_name=study_name, irb=irb)
    root_copy = find_or_create_drive_folder(
        drive=drive,
        parent_id=destination_parent_id,
        name=root_name,
        existing_file_policy=existing_file_policy,
    )
    copied_permission_count = 0
    permission_errors: list[str] = []

    def stamp_template_permissions(target_file_id: str) -> None:
        nonlocal copied_permission_count, permission_errors
        if not copy_template_permissions_to_root:
            return
        result = copy_template_permissions(drive, template_folder_id, target_file_id)
        copied_permission_count += result.copied_count
        permission_errors.extend(result.errors)

    stamp_template_permissions(root_copy.id)

    def stamp_sheet_editors(drive_file: DriveFile) -> None:
        nonlocal copied_permission_count, permission_errors
        result = add_sheet_editor_permissions(drive, drive_file, sheet_editor_emails)
        copied_permission_count += result.copied_count
        permission_errors.extend(result.errors)

    files_by_relative_path: dict[str, DriveFile] = {}

    def copy_children(source_folder_id: str, dest_folder: DriveFile, relative_parent: Path) -> None:
        for child in drive.list_children(source_folder_id):
            child_name = replace_placeholders(child.name, study_name=study_name, irb=irb)
            relative_path = relative_parent / child_name
            if child.is_folder:
                copied_folder = find_or_create_drive_folder(
                    drive=drive,
                    parent_id=dest_folder.id,
                    name=child_name,
                    existing_file_policy=existing_file_policy,
                )
                stamp_template_permissions(copied_folder.id)
                stamp_sheet_editors(copied_folder)
                files_by_relative_path[relative_path.as_posix()] = copied_folder
                copy_children(child.id, copied_folder, relative_path)
            else:
                copied_file = resolve_existing_drive_file(
                    drive=drive,
                    parent_id=dest_folder.id,
                    name=child_name,
                    existing_file_policy=existing_file_policy,
                )
                if copied_file is not None and copied_file.is_folder:
                    raise FileExistsError(f"Google Drive item exists but is a folder: {child_name}")
                if copied_file is None:
                    copied_file = drive.copy_file(child.id, child_name, dest_folder.id)
                stamp_template_permissions(copied_file.id)
                stamp_sheet_editors(copied_file)
                files_by_relative_path[relative_path.as_posix()] = copied_file

    copy_children(template_folder_id, root_copy, Path(""))
    return TemplateCopyResult(
        root=root_copy,
        files_by_relative_path=files_by_relative_path,
        copied_permission_count=copied_permission_count,
        permission_errors=tuple(permission_errors),
    )


def latest_history_dir(study_folder: Path) -> Path:
    histories = study_folder / "histories"
    candidates = sorted(path for path in histories.glob("*") if path.is_dir()) if histories.exists() else []
    if candidates:
        return candidates[-1]
    history = histories / date.today().isoformat()
    history.mkdir(parents=True, exist_ok=True)
    return history


def append_log(study_folder: Path, section: str, lines: list[str]) -> Path:
    history = latest_history_dir(study_folder)
    log_path = history / "log.md"
    with log_path.open("a", encoding="utf-8") as file:
        file.write("\n\n<details open>\n")
        file.write(f"<summary>{section}</summary>\n\n")
        for line in lines:
            file.write(f"{line}\n")
        file.write("\n</details>\n")
    return log_path


def find_by_relative_path(files_by_relative_path: Mapping[str, DriveFile], relative_path: str) -> DriveFile | None:
    target_key = normalize_path_key(relative_path)
    for key, drive_file in files_by_relative_path.items():
        if normalize_path_key(key) == target_key:
            return drive_file
    return None


def normalize_path_key(value: str | Path) -> str:
    text = str(value).strip().replace("\\", "/")
    text = re.sub(r"/+", "/", text)
    return text.strip("/").lower()


def find_child_by_name(drive: Any, parent_id: str, name: str) -> DriveFile | None:
    target = normalize_name(name)
    for child in drive.list_children(parent_id):
        if normalize_name(child.name) == target:
            return child
    return None


def find_drive_path(drive: Any, root_folder_id: str, relative_path: str | Path) -> DriveFile | None:
    current = DriveFile(id=root_folder_id, name="", mime_type=FOLDER_MIME_TYPE)
    for part in split_drive_path(relative_path):
        if part in {"", "."}:
            continue
        child = find_child_by_name(drive, current.id, part)
        if child is None:
            return None
        current = child
    return current


def split_drive_path(relative_path: str | Path) -> list[str]:
    text = str(relative_path).replace("\\", "/").strip().strip("/")
    parts: list[str] = []
    current: list[str] = []
    paren_depth = 0
    for character in text:
        if character == "(":
            paren_depth += 1
        elif character == ")" and paren_depth:
            paren_depth -= 1
        if character == "/" and paren_depth == 0:
            part = "".join(current).strip()
            if part:
                parts.append(part)
            current = []
            continue
        current.append(character)
    part = "".join(current).strip()
    if part:
        parts.append(part)
    return parts


def normalize_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", name.lower())


def fill_irb_meta(
    *,
    sheets_client: SheetsHttpClient,
    spreadsheet_id: str,
    access_token: str,
    study_name: str,
    irb: str,
    timeout: float = 120.0,
) -> None:
    headers = json_headers(access_token)
    sheets = load_sheet_properties(
        spreadsheet_id,
        http_client=sheets_client,
        headers=headers,
        timeout=timeout,
    )
    if not sheets:
        return
    sheet = next(iter(sheets.values()))
    target_headers = load_target_headers(
        spreadsheet_id,
        sheet.title,
        http_client=sheets_client,
        headers=headers,
        timeout=timeout,
    )
    updates, _, _, _ = build_value_updates(
        OverviewSheet(title=sheet.title, headers=["STUDY", "IRB"], rows=[[study_name, irb]]),
        sheet.title,
        target_headers,
    )
    if not updates:
        return
    sheets_client.post(
        sheets_values_batch_update_url(spreadsheet_id),
        json.dumps({"valueInputOption": "RAW", "data": updates}).encode("utf-8"),
        headers,
        timeout,
    )


def is_redcap_instrument_workbook(path: str | Path) -> bool:
    path = Path(path)
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return False
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False
    sheet_names = {sheet.lower() for sheet in workbook.sheetnames}
    return REDCAP_INSTRUMENT_SHEETS.issubset(sheet_names)


def cleaned_dir_for_study(study_folder: Path) -> Path:
    return study_folder / "data" / "cleaned"


def plan_cleaned_uploads(study_folder: str | Path, cleaned_data_dir: str | Path | None = None) -> list[CleanedUploadPlan]:
    study_folder = Path(study_folder)
    cleaned_dir = Path(cleaned_data_dir) if cleaned_data_dir else cleaned_dir_for_study(study_folder)
    if not cleaned_dir.exists():
        return []
    plans: list[CleanedUploadPlan] = []
    for path in sorted(item for item in cleaned_dir.rglob("*") if item.is_file() and not item.name.startswith("~$")):
        relative_path = path.relative_to(cleaned_dir)
        relative_parent = relative_path.parent
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            is_redcap = is_redcap_instrument_workbook(path)
            template_name = REDCAP_TEMPLATE_NAME if is_redcap else BLANK_TEMPLATE_NAME
        elif path.suffix.lower() == ".csv":
            is_redcap = False
            template_name = BLANK_TEMPLATE_NAME
        else:
            is_redcap = False
            template_name = None
        plans.append(
            CleanedUploadPlan(
                local_path=path,
                relative_path=relative_path,
                relative_parent=relative_parent,
                template_name=template_name,
                is_redcap_instrument=is_redcap,
            )
        )
    return plans


def template_name_candidates(template_name: str) -> tuple[str, ...]:
    if normalize_name(template_name) in {normalize_name(name) for name in REDCAP_TEMPLATE_ALIASES}:
        return REDCAP_TEMPLATE_ALIASES
    return (template_name,)


def find_template_by_name(template_folder_children: list[DriveFile], template_name: str) -> DriveFile | None:
    targets = {normalize_name(name) for name in template_name_candidates(template_name)}
    for child in template_folder_children:
        if normalize_name(child.name) in targets:
            return child
    return None


def find_or_create_drive_folder_path(
    drive: Any,
    root_folder_id: str,
    relative_path: Path,
    existing_file_policy: str = DEFAULT_EXISTING_FILE_POLICY,
) -> DriveFile:
    current = DriveFile(id=root_folder_id, name="", mime_type=FOLDER_MIME_TYPE)
    for part in relative_path.parts:
        if part in {"", "."}:
            continue
        current = find_or_create_drive_folder(
            drive=drive,
            parent_id=current.id,
            name=part,
            existing_file_policy=existing_file_policy,
        )
    return current


def workbook_from_csv(csv_path: Path, output_dir: Path) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "data"
    with csv_path.open(newline="", encoding="utf-8-sig") as file:
        for row in csv.reader(file):
            worksheet.append(row)
    output_path = output_dir / f"{csv_path.stem}.xlsx"
    workbook.save(output_path)
    return output_path


def upload_cleaned_data(
    *,
    drive: Any,
    target_data_folder_id: str,
    template_folder_id: str,
    study_folder: Path,
    access_token: str,
    cleaned_data_dir: Path | None = None,
    sheets_client: SheetsHttpClient | None = None,
    timeout: float = 120.0,
    existing_file_policy: str = DEFAULT_EXISTING_FILE_POLICY,
    permission_source_file_id: str | None = None,
    sheet_editor_emails: tuple[str, ...] = (),
) -> list[UploadedFile]:
    validate_existing_file_policy(existing_file_policy)
    plans = plan_cleaned_uploads(study_folder, cleaned_data_dir)
    templates = drive.list_children(template_folder_id) if template_folder_id else []
    sheets_client = sheets_client or UrllibSheetsHttpClient()
    results: list[UploadedFile] = []
    with tempfile.TemporaryDirectory() as tmpdir:
        temp_dir = Path(tmpdir)
        def stamp_uploaded_permissions(drive_file: DriveFile) -> None:
            if permission_source_file_id:
                copy_template_permissions(drive, permission_source_file_id, drive_file.id)
            add_sheet_editor_permissions(drive, drive_file, sheet_editor_emails)

        for plan in plans:
            parent = find_or_create_drive_folder_path(
                drive,
                target_data_folder_id,
                plan.relative_parent,
                existing_file_policy=DEFAULT_EXISTING_FILE_POLICY,
            )
            try:
                if plan.template_name:
                    template = find_template_by_name(templates, plan.template_name)
                    if template is None:
                        raise FileNotFoundError(f"Missing Google template: {plan.template_name}")
                    copied = resolve_existing_drive_file(
                        drive=drive,
                        parent_id=parent.id,
                        name=plan.local_path.stem,
                        existing_file_policy=existing_file_policy,
                    )
                    if copied is not None and not copied.is_google_sheet:
                        raise FileExistsError(
                            f"Existing Google Drive item is not a Google Sheet: {plan.local_path.stem}"
                        )
                    reused_existing = copied is not None
                    if copied is None:
                        copied = drive.copy_file(template.id, plan.local_path.stem, parent.id)
                    stamp_uploaded_permissions(copied)
                    workbook_path = (
                        workbook_from_csv(plan.local_path, temp_dir)
                        if plan.local_path.suffix.lower() == ".csv"
                        else plan.local_path
                    )
                    if existing_file_policy != "skip" or not reused_existing:
                        fill_in_overview(
                            target=copied.id,
                            overview_file=workbook_path,
                            access_token=access_token,
                            http_client=sheets_client,
                            timeout=timeout,
                            write_full_sheet_when_no_headers=True,
                            sync_template_tab_to_source_sheets=plan.template_name == BLANK_TEMPLATE_NAME,
                        )
                    results.append(UploadedFile(plan.local_path, plan.relative_path, copied))
                else:
                    existing = resolve_existing_drive_file(
                        drive=drive,
                        parent_id=parent.id,
                        name=plan.local_path.name,
                        existing_file_policy=existing_file_policy,
                    )
                    if existing is None:
                        uploaded = drive.upload_file(plan.local_path, plan.local_path.name, parent.id)
                    elif existing_file_policy == "skip":
                        uploaded = existing
                    else:
                        uploaded = drive.update_file(plan.local_path, existing.id, plan.local_path.name)
                    stamp_uploaded_permissions(uploaded)
                    results.append(UploadedFile(plan.local_path, plan.relative_path, uploaded))
            except Exception as exc:  # continue batch and log failures
                results.append(UploadedFile(plan.local_path, plan.relative_path, None, str(exc)))
    return results


def clean_relative_location(value: object) -> str:
    text = str(value or "").strip()
    if text.startswith("./"):
        text = text[2:]
    normalized = normalize_path_key(text)
    cleaned_prefix = "data/cleaned/"
    if cleaned_prefix in normalized:
        normalized = normalized.split(cleaned_prefix, 1)[1]
    return normalized


def rewrite_data_map_locations(
    source_data_map_dir: str | Path,
    output_dir: str | Path,
    uploaded_locations: Mapping[str, str],
) -> list[Path]:
    source_dir = Path(source_data_map_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    rewritten_paths: list[Path] = []
    upload_lookup = {clean_relative_location(key): value for key, value in uploaded_locations.items()}
    for source_path in sorted(source_dir.glob("*-data-map.xlsx")):
        workbook = load_workbook(source_path)
        for worksheet in workbook.worksheets:
            headers = [str(cell.value or "").strip() for cell in worksheet[1]]
            location_index = next(
                (index + 1 for index, header in enumerate(headers) if normalize_name(header) == "location"),
                None,
            )
            if location_index is None:
                continue
            for row_index in range(2, worksheet.max_row + 1):
                original = worksheet.cell(row=row_index, column=location_index).value
                normalized = clean_relative_location(original)
                if not normalized:
                    continue
                worksheet.cell(row=row_index, column=location_index).value = upload_lookup.get(normalized, "")
        output_path = output_dir / source_path.name
        workbook.save(output_path)
        rewritten_paths.append(output_path)
    return rewritten_paths


SUBJECT_TIMEPOINT_SOURCE_COLUMNS = {
    normalize_name("earliest_date_source"),
    normalize_name("latest_date_source"),
}


def rewrite_source_location_list(value: object, upload_lookup: Mapping[str, str]) -> str:
    links: list[str] = []
    for part in str(value or "").split(";"):
        normalized = clean_relative_location(part)
        if not normalized:
            continue
        link = upload_lookup.get(normalized)
        if link:
            links.append(link)
    return "; ".join(links)


def rewrite_subject_timepoint_source_locations(
    source_path: str | Path,
    output_path: str | Path,
    uploaded_locations: Mapping[str, str],
) -> Path:
    source_path = Path(source_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(source_path)
    upload_lookup = {clean_relative_location(key): value for key, value in uploaded_locations.items()}

    for worksheet in workbook.worksheets:
        headers = [str(cell.value or "").strip() for cell in worksheet[1]]
        source_indexes = [
            index + 1
            for index, header in enumerate(headers)
            if normalize_name(header) in SUBJECT_TIMEPOINT_SOURCE_COLUMNS
        ]
        if not source_indexes:
            continue
        for row_index in range(2, worksheet.max_row + 1):
            for column_index in source_indexes:
                original = worksheet.cell(row=row_index, column=column_index).value
                if not str(original or "").strip():
                    continue
                rewritten = rewrite_source_location_list(original, upload_lookup)
                worksheet.cell(row=row_index, column=column_index).value = rewritten

    workbook.save(output_path)
    return output_path


def uploaded_locations_from_results(upload_results: list[UploadedFile]) -> dict[str, str]:
    return {
        result.relative_path.as_posix(): result.drive_file.web_url or result.drive_file.id
        for result in upload_results
        if result.ok and result.drive_file is not None
    }


def update_subject_timepoints_source_links(
    *,
    study_folder: Path,
    upload_results: list[UploadedFile],
    access_token: str,
    sheets_client: SheetsHttpClient | None,
    timeout: float,
) -> Path | None:
    uploaded_locations = uploaded_locations_from_results(upload_results)
    if not uploaded_locations:
        return None

    subject_result = next(
        (
            result
            for result in upload_results
            if result.ok
            and result.drive_file is not None
            and result.relative_path.as_posix() == "subjects/subject_timepoints.xlsx"
        ),
        None,
    )
    if subject_result is None:
        return None

    rewritten_path = latest_history_dir(study_folder) / "gdrive_subject_timepoints" / "subject_timepoints.xlsx"
    rewrite_subject_timepoint_source_locations(
        subject_result.local_path,
        rewritten_path,
        uploaded_locations,
    )
    fill_in_overview(
        target=subject_result.drive_file.id,
        overview_file=rewritten_path,
        access_token=access_token,
        http_client=sheets_client,
        timeout=timeout,
        write_full_sheet_when_no_headers=True,
        sync_template_tab_to_source_sheets=True,
    )
    return rewritten_path


def upload_extra_local_folders(
    *,
    drive: Any,
    study_folder: Path,
    gdrive_study_folder_id: str,
    existing_file_policy: str = DEFAULT_EXISTING_FILE_POLICY,
) -> list[UploadedFile]:
    validate_existing_file_policy(existing_file_policy)
    results: list[UploadedFile] = []
    for item in sorted(study_folder.iterdir()):
        if not item.is_dir() or item.name in LOCAL_TOP_LEVEL_EXCLUDE or item.name.startswith("."):
            continue
        parent = find_or_create_drive_folder(
            drive=drive,
            parent_id=gdrive_study_folder_id,
            name=item.name,
            existing_file_policy=DEFAULT_EXISTING_FILE_POLICY,
        )
        for path in sorted(child for child in item.rglob("*") if child.is_file()):
            relative_path = path.relative_to(study_folder)
            target_parent = find_or_create_drive_folder_path(
                drive,
                parent.id,
                path.parent.relative_to(item),
                existing_file_policy=DEFAULT_EXISTING_FILE_POLICY,
            )
            try:
                existing = resolve_existing_drive_file(
                    drive=drive,
                    parent_id=target_parent.id,
                    name=path.name,
                    existing_file_policy=existing_file_policy,
                )
                if existing is None:
                    uploaded = drive.upload_file(path, path.name, target_parent.id)
                elif existing_file_policy == "skip":
                    uploaded = existing
                else:
                    uploaded = drive.update_file(path, existing.id, path.name)
                results.append(UploadedFile(path, relative_path, uploaded))
            except Exception as exc:
                results.append(UploadedFile(path, relative_path, None, str(exc)))
    return results


def default_overview_path(study_folder: Path) -> Path | None:
    overview_dir = study_folder / "overview"
    if not overview_dir.exists():
        return None
    candidates = sorted(overview_dir.glob("*.xlsx"))
    return candidates[0] if candidates else None


def default_data_map_dir(study_folder: Path) -> Path:
    return study_folder / "data-map"


def initialize_drive_folder(
    *,
    drive: Any,
    sheets_client: SheetsHttpClient,
    access_token: str,
    template: str,
    destination: str,
    study_folder: Path,
    study_name: str,
    irb: str,
    irb_meta_path: str = "IRB-meta",
    timeout: float = 120.0,
    existing_file_policy: str = DEFAULT_EXISTING_FILE_POLICY,
    copy_template_permissions_to_root: bool = True,
    sheet_editor_emails: tuple[str, ...] = (),
) -> TemplateCopyResult:
    result = copy_template_tree(
        drive=drive,
        template_folder_id=resolve_drive_id(template),
        destination_parent_id=resolve_drive_id(destination),
        study_name=study_name,
        irb=irb,
        existing_file_policy=existing_file_policy,
        copy_template_permissions_to_root=copy_template_permissions_to_root,
        sheet_editor_emails=sheet_editor_emails,
    )
    irb_meta = None
    for candidate in irb_meta_path_candidates(irb_meta_path, study_name=study_name, irb=irb):
        irb_meta = find_by_relative_path(result.files_by_relative_path, candidate)
        if irb_meta is not None:
            break
    if irb_meta and irb_meta.is_google_sheet:
        fill_irb_meta(
            sheets_client=sheets_client,
            spreadsheet_id=irb_meta.id,
            access_token=access_token,
            study_name=study_name,
            irb=irb,
            timeout=timeout,
        )
    append_log(
        study_folder,
        "Google Drive Study Folder",
        [
            f"## Initialize",
            f"google drive folder initialized at: {result.root.web_url or result.root.id}",
            f"renamed by Study name: {study_name}",
            f"IRB: {irb}",
            f"copied template permissions: {copy_template_permissions_to_root}",
            f"copied permission count: {result.copied_permission_count}",
            f"permission copy errors: {result.permission_error_count}",
            f"explicit sheet editors: {', '.join(sheet_editor_emails) if sheet_editor_emails else ''}",
            *[f"- {error}" for error in result.permission_errors],
        ],
    )
    return result


def run_workflow(
    *,
    study_folder: str | Path,
    study_name: str,
    irb: str,
    template: str | None = None,
    destination: str = "root",
    stage: str = "all",
    initialized_folder_id: str | None = None,
    overview_destination: str | None = None,
    overview_path: str | Path | None = None,
    cleaned_data_folder_id: str | None = None,
    cleaned_data_dir: str | Path | None = None,
    templates_folder_id: str | None = None,
    data_map_destination: str | None = None,
    data_map_dir: str | Path | None = None,
    irb_meta_path: str = "IRB-meta",
    access_token: str | None = None,
    timeout: float = 120.0,
    existing_file_policy: str = DEFAULT_EXISTING_FILE_POLICY,
    copy_template_permissions_to_root: bool = True,
    sheet_editor_emails: tuple[str, ...] = (),
) -> WorkflowState:
    study_folder = Path(study_folder)
    validate_existing_file_policy(existing_file_policy)
    token = resolve_access_token(access_token)
    token_provider = build_token_provider(token)
    drive = GoogleDriveClient(token, timeout=timeout, token_provider=token_provider)
    sheets_client = RefreshingSheetsHttpClient(UrllibSheetsHttpClient(), token_provider)
    state = WorkflowState(study_folder=study_folder, study_name=study_name, irb=irb)

    copy_result: TemplateCopyResult | None = None
    if stage in {"all", "initialize"}:
        if template is None:
            raise ValueError("--template is required for initialize/all")
        copy_result = initialize_drive_folder(
            drive=drive,
            sheets_client=sheets_client,
            access_token=token,
            template=template,
            destination=destination,
            study_folder=study_folder,
            study_name=study_name,
            irb=irb,
            irb_meta_path=irb_meta_path,
            timeout=timeout,
            existing_file_policy=existing_file_policy,
            copy_template_permissions_to_root=copy_template_permissions_to_root,
            sheet_editor_emails=sheet_editor_emails,
        )
        state.initialized_root = copy_result.root
    elif initialized_folder_id:
        state.initialized_root = drive.get_file(resolve_drive_id(initialized_folder_id))

    files_by_relative_path = copy_result.files_by_relative_path if copy_result else {}

    if stage in {"all", "upload"}:
        if overview_destination is None:
            overview_file = find_by_relative_path(files_by_relative_path, f"Overview/{study_name}_{irb}")
            if overview_file is None and state.initialized_root:
                overview_file = find_drive_path(drive, state.initialized_root.id, f"Overview/{study_name}_{irb}")
            overview_destination = overview_file.id if overview_file else None
        local_overview = Path(overview_path) if overview_path else default_overview_path(study_folder)
        if overview_destination and local_overview and local_overview.exists():
            state.overview_result = fill_in_overview(
                target=overview_destination,
                overview_file=local_overview,
                access_token=token,
                http_client=sheets_client,
                timeout=timeout,
            )

        if state.initialized_root:
            if cleaned_data_folder_id is None:
                cleaned_folder = find_drive_path(drive, state.initialized_root.id, DEFAULT_CLEANED_GDRIVE_PATH)
                cleaned_data_folder_id = cleaned_folder.id if cleaned_folder else None
            if templates_folder_id is None:
                template_folder = find_drive_path(drive, state.initialized_root.id, DEFAULT_TEMPLATE_GDRIVE_PATH)
                templates_folder_id = template_folder.id if template_folder else None

        if cleaned_data_folder_id and templates_folder_id:
            state.upload_results = upload_cleaned_data(
                drive=drive,
                target_data_folder_id=resolve_drive_id(cleaned_data_folder_id),
                template_folder_id=resolve_drive_id(templates_folder_id),
                study_folder=study_folder,
                access_token=token,
                cleaned_data_dir=Path(cleaned_data_dir) if cleaned_data_dir else None,
                sheets_client=sheets_client,
                timeout=timeout,
                existing_file_policy=existing_file_policy,
                permission_source_file_id=state.initialized_root.id if state.initialized_root else None,
                sheet_editor_emails=sheet_editor_emails,
            )
            successes = [result for result in state.upload_results if result.ok]
            failures = [result for result in state.upload_results if not result.ok]
            append_log(
                study_folder,
                "Google Drive Upload",
                [
                    "## Cleaned Data Upload",
                    f"successfully uploaded {len(successes)} files",
                    "failed files:",
                    *[f"- {failure.relative_path}: {failure.error}" for failure in failures],
                ],
            )
        if state.initialized_root:
            extra_uploads = upload_extra_local_folders(
                drive=drive,
                study_folder=study_folder,
                gdrive_study_folder_id=state.initialized_root.id,
                existing_file_policy=existing_file_policy,
            )
            if extra_uploads:
                state.upload_results.extend(extra_uploads)
        if existing_file_policy != "skip":
            rewritten_subject_timepoints = update_subject_timepoints_source_links(
                study_folder=study_folder,
                upload_results=state.upload_results,
                access_token=token,
                sheets_client=sheets_client,
                timeout=timeout,
            )
            if rewritten_subject_timepoints is not None:
                append_log(
                    study_folder,
                    "Google Drive Upload",
                    [
                        "## Subject Timepoints Source Links",
                        f"rewritten source links: {rewritten_subject_timepoints}",
                    ],
                )

    if stage in {"all", "data-map"}:
        local_data_map_dir = Path(data_map_dir) if data_map_dir else default_data_map_dir(study_folder)
        uploaded_locations = uploaded_locations_from_results(state.upload_results)
        if uploaded_locations:
            rewritten_dir = latest_history_dir(study_folder) / "gdrive_data_map"
            rewrite_data_map_locations(local_data_map_dir, rewritten_dir, uploaded_locations)
            local_data_map_dir = rewritten_dir
        if data_map_destination is None and state.initialized_root:
            data_map_file = find_drive_path(drive, state.initialized_root.id, f"{DEFAULT_DATA_MAP_FOLDER}/{irb}-data-map")
            if data_map_file is None:
                data_map_file = find_drive_path(drive, state.initialized_root.id, f"{DEFAULT_DATA_MAP_FOLDER}/{study_name}_{irb}-data-map")
            data_map_destination = data_map_file.id if data_map_file else None

        if data_map_destination and local_data_map_dir.exists():
            state.data_map_result = push_data_map(
                target=data_map_destination,
                data_map_dir=local_data_map_dir,
                access_token=token,
                http_client=sheets_client,
                timeout=timeout,
            )
            append_log(
                study_folder,
                "Google Drive Data Map",
                [
                    "## Data Map",
                    "data map successfully created",
                    f"updated tabs: {', '.join(state.data_map_result.updated_tabs)}",
                ],
            )
    return state


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--study-folder", required=True, type=Path, help="Local cleaned study folder.")
    parser.add_argument("--study-name", required=True, help="Study name used to replace STUDY placeholders.")
    parser.add_argument("--irb", required=True, help="IRB value used to replace IRB placeholders.")
    parser.add_argument(
        "--stage",
        choices=["all", "initialize", "upload", "data-map"],
        default="all",
        help="Workflow stage to run.",
    )
    parser.add_argument("--template", help="Google Drive template folder URL/id/path.")
    parser.add_argument("--destination", default="root", help="Destination parent folder URL/id/path. Defaults to Drive root.")
    parser.add_argument("--initialized-folder-id", help="Existing initialized Google Drive study folder id/link.")
    parser.add_argument("--irb-meta-path", default="IRB-meta", help="Relative path to IRB-meta in initialized template.")
    parser.add_argument("--overview-destination", help="Google Sheet URL/id/path for the overview destination.")
    parser.add_argument("--overview-path", type=Path, help="Local overview workbook path.")
    parser.add_argument("--cleaned-data-folder-id", help="Google Drive folder id/link for cleaned data uploads.")
    parser.add_argument("--cleaned-data-dir", type=Path, help="Local cleaned data directory.")
    parser.add_argument("--templates-folder-id", help="Google Drive folder id/link containing REDCap_INSTRUMENT and BLANK templates.")
    parser.add_argument("--data-map-destination", help="Google Sheet URL/id/path for data map destination.")
    parser.add_argument("--data-map-dir", type=Path, help="Local data-map directory.")
    parser.add_argument(
        "--existing-file-policy",
        choices=EXISTING_FILE_POLICIES,
        default=DEFAULT_EXISTING_FILE_POLICY,
        help=(
            "How to handle Drive files that already exist at the target path. "
            "Default updates existing files or creates missing files."
        ),
    )
    parser.add_argument(
        "--no-copy-template-permissions",
        action="store_true",
        help="Do not copy direct non-owner permissions from the template folder to the initialized study folder.",
    )
    parser.add_argument(
        "--share-sheet-editor",
        action="append",
        default=[],
        help=(
            "Email address to grant writer access on Google Sheets files only. "
            "May be repeated, or contain comma/semicolon/space-separated emails."
        ),
    )
    parser.add_argument("--access-token", help="Google OAuth token. Defaults to env/gcloud lookup.")
    parser.add_argument("--timeout", type=float, default=120.0, help="Google API timeout in seconds.")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    state = run_workflow(
        study_folder=args.study_folder,
        study_name=args.study_name,
        irb=args.irb,
        template=args.template,
        destination=args.destination,
        stage=args.stage,
        initialized_folder_id=args.initialized_folder_id,
        overview_destination=args.overview_destination,
        overview_path=args.overview_path,
        cleaned_data_folder_id=args.cleaned_data_folder_id,
        cleaned_data_dir=args.cleaned_data_dir,
        templates_folder_id=args.templates_folder_id,
        data_map_destination=args.data_map_destination,
        data_map_dir=args.data_map_dir,
        irb_meta_path=args.irb_meta_path,
        access_token=args.access_token,
        timeout=args.timeout,
        existing_file_policy=args.existing_file_policy,
        copy_template_permissions_to_root=not args.no_copy_template_permissions,
        sheet_editor_emails=parse_email_values(args.share_sheet_editor),
    )
    print(
        json.dumps(
            {
                "study_folder": str(state.study_folder),
                "study_name": state.study_name,
                "irb": state.irb,
                "initialized_folder": state.initialized_root.web_url if state.initialized_root else None,
                "uploaded": len([result for result in state.upload_results if result.ok]),
                "failed": len([result for result in state.upload_results if not result.ok]),
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
