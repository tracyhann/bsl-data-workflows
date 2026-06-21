import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.workflows.clean_mri_session.run import clean_mri_session


def write_dictionary(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "event"
    worksheet.append(["arm", "order", "event_name", "event_label", "abbreviation"])
    worksheet.append(["2", "1", "baseline_visit_2", "Baseline (Visit 2)", "V2"])
    worksheet.append(["2", "2", "treatment_visit_1", "Treatment Visit 1 (Visit 3)", "V3"])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_subject_timepoints(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "subject_timepoints"
    worksheet.append(
        [
            "IRB",
            "subid",
            "arm",
            "visit",
            "earliest_entry_date",
            "earliest_date_source",
            "latest_entry_date",
            "latest_date_source",
            "span",
            "values",
        ]
    )
    worksheet.append(["12345", "s001", "2", "V2", "2026-01-10", None, "2026-01-10", None, 0, "2026-01-10"])
    worksheet.append(["12345", "s001", "2", "V3", "2026-01-11", None, "2026-01-15", None, 4, "2026-01-11; 2026-01-15"])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


class CleanMriSessionTests(unittest.TestCase):
    def test_creates_raw_timepoint_dictionary_and_cleaned_mri_session(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            study = root / "study"
            input_csv = root / "mri_session.csv"
            write_dictionary(study / "data" / "cleaned" / "dictionary.xlsx")
            write_subject_timepoints(study / "data" / "cleaned" / "subjects" / "subject_timepoints.xlsx")
            with input_csv.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(
                    [
                        "session.id",
                        "session.label",
                        "session.operator",
                        "session.timestamp",
                        "session.timezone",
                        "session.age",
                        "session.weight",
                        "session.url",
                        "subject.label",
                        "errors",
                    ]
                )
                writer.writerow(
                    [
                        "session-a",
                        "V2",
                        "operator",
                        "2026-01-10 15:55:02+00:00",
                        "America/Los_Angeles",
                        "",
                        "70",
                        "https://example.test/a",
                        "12345s001",
                        "",
                    ]
                )
                writer.writerow(
                    [
                        "session-b",
                        "V3fmap",
                        "operator",
                        "2026-01-12 14:00:00+00:00",
                        "America/Los_Angeles",
                        "",
                        "70",
                        "https://example.test/b",
                        "12345_s001",
                        "minor warning",
                    ]
                )
                writer.writerow(
                    [
                        "session-c",
                        "V4",
                        "operator",
                        "2026-01-12 14:00:00+00:00",
                        "America/Los_Angeles",
                        "",
                        "70",
                        "https://example.test/c",
                        "12345_s001",
                        "",
                    ]
                )

            result = clean_mri_session(input_csv, study)

            workbook = load_workbook(result.output_path, data_only=True)
            raw_rows = list(workbook["raw"].iter_rows(values_only=True))
            dictionary_rows = list(workbook["timepoint_dictionary"].iter_rows(values_only=True))
            cleaned_rows = list(workbook["cleaned"].iter_rows(values_only=True))

        self.assertEqual(result.output_path, study / "data" / "cleaned" / "neuroimaging" / "mri_session.xlsx")
        self.assertEqual(result.raw_rows, 3)
        self.assertEqual(result.cleaned_rows, 3)
        self.assertEqual(workbook.sheetnames, ["raw", "timepoint_dictionary", "cleaned"])
        self.assertEqual(raw_rows[0][0:3], ("session.id", "session.label", "session.operator"))
        self.assertEqual(raw_rows[1][0], "session-a")
        self.assertEqual(dictionary_rows[0], ("arm", "order", "event_name", "event_label", "abbreviation"))
        self.assertEqual(dictionary_rows[1], ("2", "1", "baseline_visit_2", "Baseline (Visit 2)", "V2"))
        self.assertEqual(
            cleaned_rows[0],
            (
                "IRB",
                "subid",
                "arm",
                "visit",
                "date",
                "subject.label",
                "session.label",
                "session.url",
                "errors",
                "session.timestamp",
                "session.timezone",
            ),
        )
        self.assertEqual(
            cleaned_rows[1],
            (
                "12345",
                "s001",
                "2",
                "V2",
                "2026-01-10",
                "12345s001",
                "V2",
                "https://example.test/a",
                None,
                "2026-01-10 15:55:02+00:00",
                "America/Los_Angeles",
            ),
        )
        self.assertEqual(cleaned_rows[2][0:5], ("12345", "s001", "2", "V3", "2026-01-12"))
        self.assertEqual(cleaned_rows[2][5:9], ("12345_s001", "V3fmap", "https://example.test/b", "minor warning"))
        self.assertEqual(cleaned_rows[3][0:5], ("12345", "s001", None, None, "2026-01-12"))


if __name__ == "__main__":
    unittest.main()
