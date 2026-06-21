#!/usr/bin/env python3
"""Clean Flywheel MRI session exports into the study cleaned folder."""

from __future__ import annotations

import argparse
import csv
import re
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook


if __package__ in {None, ""}:
    sys.path.append(str(Path(__file__).resolve().parents[3]))

from scripts.standardize_date.date_standardizer import standardize_date
from scripts.standardize_record_id.presidio.redcap_subid_presidio import standardize_redcap_subid
from scripts.subject_timepoints.run import build_subject_timepoints


CLEANED_COLUMNS = [
    "IRB",
    "subid",
    "arm",
    "visit",
    "date",
    "subject.label",
    "session.label",
    "session.url",
    "errors",
    "session.timestamp",
    "session.timezone",
]
REQUIRED_INPUT_COLUMNS = [
    "subject.label",
    "session.label",
    "session.timestamp",
    "session.timezone",
    "session.url",
    "errors",
]
VISIT_TOKEN_RE = re.compile(r"(?i)(?:\bv[\s._-]*|visit[\s._-]*)(0*\d+)")


@dataclass(frozen=True)
class MriTimepoint:
    irb: str
    subid: str
    arm: str
    visit: str
    earliest: str | None
    latest: str | None


@dataclass(frozen=True)
class CleanMriSessionResult:
    input_path: Path
    study_folder: Path
    output_path: Path
    raw_rows: int
    cleaned_rows: int


def split_irb_values(value: object) -> set[str]:
    return {item.strip() for item in re.split(r"[;,]", str(value or "")) if item.strip()}


def split_standardized_id(value: str | None) -> tuple[str | None, str | None]:
    if not value or "_s" not in value:
        return None, None
    irb, subject_number = value.split("_", 1)
    return irb, subject_number


def visit_tokens(value: object) -> set[str]:
    tokens: set[str] = set()
    for match in VISIT_TOKEN_RE.finditer(str(value or "")):
        try:
            tokens.add(f"V{int(match.group(1))}")
        except ValueError:
            continue
    return tokens


def normalize_text(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def visit_overlaps(session_label: object, visit: object) -> bool:
    session_tokens = visit_tokens(session_label)
    visit_token_set = visit_tokens(visit)
    if session_tokens and visit_token_set:
        return bool(session_tokens & visit_token_set)
    return normalize_text(session_label) == normalize_text(visit)


def date_in_timepoint_range(value: str | None, timepoint: MriTimepoint) -> bool:
    if not value:
        return False
    earliest = timepoint.earliest or timepoint.latest
    latest = timepoint.latest or timepoint.earliest
    if not earliest or not latest:
        return False
    return earliest <= value <= latest


def read_csv_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8-sig") as file:
        sample = file.read(8192)
        file.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(file, dialect=dialect)
        headers = list(reader.fieldnames or [])
        rows = [dict(row) for row in reader]
    missing = [column for column in REQUIRED_INPUT_COLUMNS if column not in headers]
    if missing:
        raise ValueError(f"MRI session CSV missing required columns: {', '.join(missing)}")
    return headers, rows


def dictionary_path_for_study(study_folder: Path) -> Path:
    candidates = [
        study_folder / "data" / "cleaned" / "dictionary.xlsx",
        study_folder / "data" / "cleaned" / "redcap" / "dictionary.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    raise FileNotFoundError(f"Missing dictionary.xlsx under {study_folder / 'data' / 'cleaned'}")


def subject_timepoints_path_for_study(study_folder: Path) -> Path:
    path = study_folder / "data" / "cleaned" / "subjects" / "subject_timepoints.xlsx"
    if not path.exists():
        build_subject_timepoints(study_folder, output_path=path)
    return path


def read_sheet_rows(workbook_path: Path, sheet_name: str) -> list[list[object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"{workbook_path} does not contain sheet {sheet_name!r}")
    return [list(row) for row in workbook[sheet_name].iter_rows(values_only=True)]


def load_timepoints(study_folder: Path) -> list[MriTimepoint]:
    path = subject_timepoints_path_for_study(study_folder)
    rows = read_sheet_rows(path, "subject_timepoints")
    if not rows:
        return []
    headers = [str(header or "") for header in rows[0]]
    lookup = {header: index for index, header in enumerate(headers)}
    required = ["IRB", "subid", "arm", "visit", "earliest_entry_date", "latest_entry_date"]
    missing = [column for column in required if column not in lookup]
    if missing:
        raise ValueError(f"{path} missing required columns: {', '.join(missing)}")

    timepoints: list[MriTimepoint] = []
    for row in rows[1:]:
        if not any(value is not None and str(value).strip() for value in row):
            continue
        timepoints.append(
            MriTimepoint(
                irb=str(row[lookup["IRB"]] or "").strip(),
                subid=str(row[lookup["subid"]] or "").strip(),
                arm=str(row[lookup["arm"]] or "").strip(),
                visit=str(row[lookup["visit"]] or "").strip(),
                earliest=standardize_date(row[lookup["earliest_entry_date"]]),
                latest=standardize_date(row[lookup["latest_entry_date"]]),
            )
        )
    return timepoints


def match_mri_visit(
    timepoints: list[MriTimepoint],
    irb: str | None,
    subid: str | None,
    mri_date: str | None,
    session_label: object,
) -> MriTimepoint | None:
    if not irb or not subid or not mri_date:
        return None

    candidates: list[MriTimepoint] = []
    for timepoint in timepoints:
        if timepoint.subid != subid:
            continue
        if timepoint.irb and irb not in split_irb_values(timepoint.irb):
            continue
        if not date_in_timepoint_range(mri_date, timepoint):
            continue
        if not visit_overlaps(session_label, timepoint.visit):
            continue
        candidates.append(timepoint)

    if not candidates:
        return None
    return sorted(
        candidates,
        key=lambda candidate: (
            candidate.earliest or "",
            candidate.latest or "",
            candidate.arm,
            candidate.visit,
        ),
    )[0]


def cleaned_row(raw_row: dict[str, str], timepoints: list[MriTimepoint]) -> list[object]:
    standardized_id = standardize_redcap_subid(raw_row.get("subject.label"))
    irb, subid = split_standardized_id(standardized_id)
    mri_date = standardize_date(raw_row.get("session.timestamp"))
    matched = match_mri_visit(timepoints, irb, subid, mri_date, raw_row.get("session.label"))
    return [
        irb,
        subid,
        matched.arm if matched else None,
        matched.visit if matched else None,
        mri_date,
        raw_row.get("subject.label") or None,
        raw_row.get("session.label") or None,
        raw_row.get("session.url") or None,
        raw_row.get("errors") or None,
        raw_row.get("session.timestamp") or None,
        raw_row.get("session.timezone") or None,
    ]


def append_table(worksheet, rows: list[list[object]]) -> None:
    for row in rows:
        worksheet.append(row)


def autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 80)
        worksheet.column_dimensions[column_cells[0].column_letter].width = width


def clean_mri_session(
    input_csv: str | Path,
    study_folder: str | Path,
    output_path: str | Path | None = None,
) -> CleanMriSessionResult:
    input_csv = Path(input_csv)
    study_folder = Path(study_folder)
    output_path = (
        Path(output_path)
        if output_path is not None
        else study_folder / "data" / "cleaned" / "neuroimaging" / "mri_session.xlsx"
    )

    raw_headers, raw_rows = read_csv_rows(input_csv)
    dictionary_rows = read_sheet_rows(dictionary_path_for_study(study_folder), "event")
    timepoints = load_timepoints(study_folder)

    workbook = Workbook()
    raw_sheet = workbook.active
    raw_sheet.title = "raw"
    raw_sheet.append(raw_headers)
    for row in raw_rows:
        raw_sheet.append([row.get(header) for header in raw_headers])

    dictionary_sheet = workbook.create_sheet("timepoint_dictionary")
    append_table(dictionary_sheet, dictionary_rows)

    cleaned_sheet = workbook.create_sheet("cleaned")
    cleaned_sheet.append(CLEANED_COLUMNS)
    for row in raw_rows:
        cleaned_sheet.append(cleaned_row(row, timepoints))

    for worksheet in workbook.worksheets:
        autosize_columns(worksheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return CleanMriSessionResult(
        input_path=input_csv,
        study_folder=study_folder,
        output_path=output_path,
        raw_rows=len(raw_rows),
        cleaned_rows=len(raw_rows),
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--csv", "--input", dest="input_csv", required=True, type=Path, help="MRI session CSV.")
    parser.add_argument("--study-folder", required=True, type=Path, help="Study folder with data/cleaned/dictionary.xlsx.")
    parser.add_argument("--out", type=Path, help="Optional output workbook path.")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    result = clean_mri_session(args.input_csv, args.study_folder, output_path=args.out)
    print(result.output_path)
    print(f"raw_rows={result.raw_rows}")
    print(f"cleaned_rows={result.cleaned_rows}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
