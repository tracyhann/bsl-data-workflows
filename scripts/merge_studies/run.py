#!/usr/bin/env python3
"""Merge two cleaned study folders into one combined cleaned study folder."""

from __future__ import annotations

import argparse
import re
import shutil
import sys
from dataclasses import dataclass
from datetime import datetime
from itertools import zip_longest
from pathlib import Path

from openpyxl import Workbook, load_workbook


if __package__ in {None, ""}:
    sys.path.append(str(Path(__file__).resolve().parents[2]))

from scripts.subject_timepoints.run import build_subject_timepoints


WORKBOOK_SHEETS = [
    "raw",
    "raw_labels",
    "cleaned",
    "timepoint_dictionary",
    "column_variable_dictionary",
    "excluded_rows",
]
INDEX_COLUMNS = ["IRB", "subid", "arm", "visit", "date", "record_id", "redcap_event_name"]
DICTIONARY_SHEETS = ["subject_id", "event", "instrument"]


@dataclass(frozen=True)
class StudyInfo:
    folder: Path
    irb: str
    study_name: str
    cleaned_dir: Path
    dictionary_path: Path


@dataclass(frozen=True)
class InstrumentWorkbook:
    instrument: str
    instrument_label: str
    category: str
    path: Path


@dataclass(frozen=True)
class MergeStudiesResult:
    study1: Path
    study2: Path
    output_folder: Path
    dictionary_path: Path
    merged_workbooks: list[Path]
    copied_workbooks: list[Path]
    archived_study_folders: list[Path]


def normalize_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()


def dedupe_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    used: set[str] = set()
    output: list[str] = []
    for header in headers:
        if not header:
            output.append(header)
            continue
        if header not in seen:
            seen[header] = 0
            candidate = header
        else:
            seen[header] += 1
            candidate = f"{header}.{seen[header]}"
            while candidate in used:
                seen[header] += 1
                candidate = f"{header}.{seen[header]}"
        used.add(candidate)
        output.append(candidate)
    return output


def safe_name(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", value.strip()).strip("_")


def parse_study_folder_name(folder: Path) -> tuple[str, str]:
    match = re.match(r"(?P<irb>\d+)[-_]?(?P<name>.*)$", folder.name)
    if not match:
        raise ValueError(f"Could not infer IRB from study folder name: {folder}")
    irb = match.group("irb")
    study_name = match.group("name").strip("-_") or "study"
    return irb, study_name


def study_info(folder: str | Path) -> StudyInfo:
    folder = Path(folder)
    irb, study_name = parse_study_folder_name(folder)
    cleaned_dir = folder / "data" / "cleaned"
    dictionary_path = cleaned_dir / "dictionary.xlsx"
    if not dictionary_path.exists():
        raise FileNotFoundError(f"Missing cleaned dictionary: {dictionary_path}")
    return StudyInfo(folder=folder, irb=irb, study_name=study_name, cleaned_dir=cleaned_dir, dictionary_path=dictionary_path)


def inferred_output_folder(study1: StudyInfo, study2: StudyInfo) -> Path:
    if normalize_key(study1.study_name) == normalize_key(study2.study_name):
        name = study1.study_name
    else:
        name = f"{study1.study_name}-{study2.study_name}"
    return study1.folder.parent / f"{study1.irb}-{study2.irb}-{name}"


def read_sheet(path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, object]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return [], []
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = dedupe_headers([str(value or "").strip() for value in rows[0]])
    data: list[dict[str, object]] = []
    for values in rows[1:]:
        row = {header: values[index] if index < len(values) else None for index, header in enumerate(headers) if header}
        if any(not is_blank(value) for value in row.values()):
            data.append(row)
    return [header for header in headers if header], data


def write_sheet(workbook: Workbook, sheet_name: str, headers: list[str], rows: list[dict[str, object]]) -> None:
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.append(headers)
    for row in rows:
        worksheet.append([blank_to_none(row.get(header)) for header in headers])


def is_blank(value: object) -> bool:
    return value is None or str(value).strip() == ""


def blank_to_none(value: object) -> object:
    return None if is_blank(value) else value


def ordered_union(left: list[str], right: list[str]) -> list[str]:
    output = [header for header in left if header]
    seen = set(output)
    for index, header in enumerate(right):
        if not header or header in seen:
            continue

        next_existing = next((candidate for candidate in right[index + 1 :] if candidate in seen), None)
        if next_existing is not None:
            output.insert(output.index(next_existing), header)
        else:
            output.append(header)
        seen.add(header)
    return output


def cleaned_headers(left: list[str], right: list[str]) -> list[str]:
    merged = ordered_union(left, right)
    prefix = [header for header in INDEX_COLUMNS if header in merged or header in left or header in right]
    suffix = [header for header in merged if header not in prefix]
    return prefix + suffix


def numeric_value(value: object) -> float | None:
    if isinstance(value, bool) or is_blank(value):
        return None
    try:
        return float(str(value).strip())
    except ValueError:
        return None


def date_sort_value(value: object) -> tuple[int, str]:
    text = str(value or "").strip()
    if not text:
        return (1, "")
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d"):
        try:
            return (0, datetime.strptime(text, fmt).strftime("%Y-%m-%d"))
        except ValueError:
            continue
    return (0, text)


def earliest_date(*values: object) -> object:
    present = [value for value in values if not is_blank(value)]
    if not present:
        return None
    return sorted(present, key=date_sort_value)[0]


def cell_values_equal(left: object, right: object) -> bool:
    if is_blank(left) and is_blank(right):
        return True
    left_number = numeric_value(left)
    right_number = numeric_value(right)
    if left_number is not None and right_number is not None:
        return left_number == right_number
    return str(left or "").strip() == str(right or "").strip()


def combine_unique_values(values: list[object]) -> object:
    output: list[object] = []
    normalized_seen: set[str] = set()
    for value in values:
        if is_blank(value):
            continue
        key = str(value).strip()
        if key in normalized_seen:
            continue
        normalized_seen.add(key)
        output.append(value)
    if not output:
        return None
    if len(output) == 1:
        return output[0]
    return "; ".join(str(value).strip() for value in output)


def merge_cell(left: object, right: object) -> object:
    if is_blank(left):
        return blank_to_none(right)
    if is_blank(right):
        return blank_to_none(left)
    if cell_values_equal(left, right):
        return left
    return combine_unique_values([left, right])


def cleaned_row_key(row: dict[str, object]) -> tuple[str, str, str]:
    return (
        str(row.get("subid") or "").strip(),
        str(row.get("arm") or "").strip(),
        str(row.get("visit") or "").strip(),
    )


def natural_subid(value: object) -> tuple[str, int, str]:
    text = str(value or "").strip()
    match = re.search(r"(\d+)", text)
    if match:
        return (re.sub(r"\d+", "", text), int(match.group(1)), text)
    return (text, -1, text)


def natural_label(value: object) -> tuple[str, int, str]:
    text = str(value or "").strip()
    match = re.search(r"(\d+)", text)
    if match:
        return (text[: match.start()], int(match.group(1)), text)
    return (text, -1, text)


def date_then_visit_sort_value(row: dict[str, object]) -> tuple[object, ...]:
    date_value = date_sort_value(row.get("date"))
    if date_value[0] == 0:
        return (0, date_value[1], natural_label(row.get("visit")))
    return (1, natural_label(row.get("visit")))


def cleaned_sort_key(row: dict[str, object]) -> tuple[object, ...]:
    return (
        natural_subid(row.get("subid")),
        str(row.get("arm") or ""),
        date_then_visit_sort_value(row),
        str(row.get("record_id") or ""),
    )


def merge_cleaned_pair(left: dict[str, object], right: dict[str, object], headers: list[str]) -> dict[str, object]:
    merged: dict[str, object] = {}
    for header in headers:
        if header == "IRB":
            merged[header] = combine_unique_values([left.get(header), right.get(header)])
        elif header == "date":
            merged[header] = earliest_date(left.get(header), right.get(header))
        else:
            merged[header] = merge_cell(left.get(header), right.get(header))
    return merged


def merge_cleaned_sheet(
    left_headers: list[str],
    left_rows: list[dict[str, object]],
    right_headers: list[str],
    right_rows: list[dict[str, object]],
) -> tuple[list[str], list[dict[str, object]]]:
    headers = cleaned_headers(left_headers, right_headers)
    by_key: dict[tuple[str, str, str], dict[int, list[dict[str, object]]]] = {}
    for row in left_rows:
        by_key.setdefault(cleaned_row_key(row), {}).setdefault(1, []).append(row)
    for row in right_rows:
        by_key.setdefault(cleaned_row_key(row), {}).setdefault(2, []).append(row)

    output_rows: list[dict[str, object]] = []
    for key in sorted(by_key, key=lambda item: (natural_subid(item[0]), item[1], item[2])):
        group = by_key[key]
        left_group = group.get(1, [])
        right_group = group.get(2, [])
        if left_group and right_group:
            for left_row, right_row in zip_longest(left_group, right_group):
                if left_row is not None and right_row is not None:
                    output_rows.append(merge_cleaned_pair(left_row, right_row, headers))
                elif left_row is not None:
                    output_rows.append(left_row)
                elif right_row is not None:
                    output_rows.append(right_row)
        else:
            output_rows.extend(left_group)
            output_rows.extend(right_group)

    return headers, sorted(output_rows, key=cleaned_sort_key)


def merge_concat_sheet(
    left_headers: list[str],
    left_rows: list[dict[str, object]],
    right_headers: list[str],
    right_rows: list[dict[str, object]],
) -> tuple[list[str], list[dict[str, object]]]:
    return ordered_union(left_headers, right_headers), [*left_rows, *right_rows]


def row_identity(sheet_name: str, row: dict[str, object]) -> tuple[str, str]:
    lowered = sheet_name.lower()
    if lowered in {"event", "timepoint_dictionary"}:
        for field in ("event_name", "event_label"):
            value = normalize_key(row.get(field))
            if value:
                return (field, value)
    elif lowered == "instrument":
        for field in ("instrument", "instrument_label"):
            value = normalize_key(row.get(field))
            if value:
                return (field, value)
    elif lowered == "column_variable_dictionary":
        for field in ("column_name", "column_labels"):
            value = normalize_key(row.get(field))
            if value:
                return (field, value)
    elif lowered == "subject_id":
        standardized = normalize_key(row.get("standardized"))
        if standardized:
            return ("standardized", standardized)
        return ("subject", f"{normalize_key(row.get('IRB'))}:{normalize_key(row.get('subid'))}")

    non_irb = tuple((key, normalize_key(value)) for key, value in sorted(row.items()) if key != "IRB")
    return ("row", repr(non_irb))


def add_irb_column(headers: list[str]) -> list[str]:
    headers = [header for header in headers if header]
    if "IRB" in headers:
        return ["IRB", *[header for header in headers if header != "IRB"]]
    return ["IRB", *headers]


def source_rows_with_irb(headers: list[str], rows: list[dict[str, object]], irb: str) -> list[dict[str, object]]:
    output = []
    for row in rows:
        copied = dict(row)
        copied["IRB"] = copied.get("IRB") or irb
        output.append(copied)
    return output


def merge_dictionary_rows(
    sheet_name: str,
    left_irb: str,
    left_headers: list[str],
    left_rows: list[dict[str, object]],
    right_irb: str,
    right_headers: list[str],
    right_rows: list[dict[str, object]],
) -> tuple[list[str], list[dict[str, object]]]:
    headers = add_irb_column(ordered_union(left_headers, right_headers))
    merged_by_identity: dict[tuple[str, str], dict[str, object]] = {}
    ordered_keys: list[tuple[str, str]] = []

    for row in [*source_rows_with_irb(left_headers, left_rows, left_irb), *source_rows_with_irb(right_headers, right_rows, right_irb)]:
        key = row_identity(sheet_name, row)
        if key not in merged_by_identity:
            merged_by_identity[key] = dict(row)
            ordered_keys.append(key)
            continue
        current = merged_by_identity[key]
        for header in headers:
            if header == "IRB":
                current[header] = combine_unique_values([current.get(header), row.get(header)])
            else:
                current[header] = merge_cell(current.get(header), row.get(header))

    rows_out = [merged_by_identity[key] for key in ordered_keys]
    if sheet_name.lower() in {"event", "timepoint_dictionary"}:
        rows_out.sort(key=lambda row: (str(row.get("arm") or ""), numeric_value(row.get("order")) or 999999, str(row.get("event_name") or "")))
    elif sheet_name.lower() == "instrument":
        rows_out.sort(key=lambda row: str(row.get("instrument") or row.get("instrument_label") or ""))
    elif sheet_name.lower() == "subject_id":
        rows_out.sort(key=lambda row: (str(row.get("IRB") or ""), natural_subid(row.get("subid"))))
    return headers, rows_out


def instrument_label_map(dictionary_path: Path) -> dict[str, str]:
    if not dictionary_path.exists():
        return {}
    workbook = load_workbook(dictionary_path, read_only=True, data_only=True)
    if "instrument" not in workbook.sheetnames:
        return {}
    worksheet = workbook["instrument"]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = [str(value or "").strip() for value in rows[0]]
    header_lookup = {normalize_key(header): index for index, header in enumerate(headers)}
    instrument_index = header_lookup.get("instrument")
    label_index = header_lookup.get("instrument label")
    if instrument_index is None:
        return {}
    labels: dict[str, str] = {}
    for row in rows[1:]:
        instrument = str(row[instrument_index] or "").strip()
        if not instrument:
            continue
        label = str(row[label_index] or "").strip() if label_index is not None and label_index < len(row) else ""
        labels[instrument] = label
    return labels


def collect_instrument_workbooks(study: StudyInfo) -> dict[str, InstrumentWorkbook]:
    labels = instrument_label_map(study.dictionary_path)
    workbooks: dict[str, InstrumentWorkbook] = {}
    for path in sorted(study.cleaned_dir.glob("*/*.xlsx")):
        if path.name.startswith("~$") or path.name == "dictionary.xlsx":
            continue
        instrument = path.stem.split("-", 1)[1] if "-" in path.stem else path.stem
        workbooks[instrument] = InstrumentWorkbook(
            instrument=instrument,
            instrument_label=labels.get(instrument, ""),
            category=path.parent.name,
            path=path,
        )
    return workbooks


def match_instruments(
    left: dict[str, InstrumentWorkbook],
    right: dict[str, InstrumentWorkbook],
) -> tuple[list[tuple[InstrumentWorkbook, InstrumentWorkbook]], list[InstrumentWorkbook], list[InstrumentWorkbook]]:
    pairs: list[tuple[InstrumentWorkbook, InstrumentWorkbook]] = []
    unmatched_left = dict(left)
    unmatched_right = dict(right)

    for instrument in sorted(set(unmatched_left) & set(unmatched_right)):
        pairs.append((unmatched_left.pop(instrument), unmatched_right.pop(instrument)))

    right_by_label: dict[str, str] = {}
    for instrument, workbook in unmatched_right.items():
        label_key = normalize_key(workbook.instrument_label)
        if label_key and label_key not in right_by_label:
            right_by_label[label_key] = instrument

    for instrument, workbook in list(unmatched_left.items()):
        label_key = normalize_key(workbook.instrument_label)
        right_instrument = right_by_label.get(label_key)
        if right_instrument and right_instrument in unmatched_right:
            pairs.append((unmatched_left.pop(instrument), unmatched_right.pop(right_instrument)))

    return pairs, list(unmatched_left.values()), list(unmatched_right.values())


def merge_workbook(
    left: InstrumentWorkbook,
    right: InstrumentWorkbook,
    study1: StudyInfo,
    study2: StudyInfo,
    output_dir: Path,
) -> Path:
    category = left.category or right.category or "unknown"
    destination = output_dir / category / f"{study1.irb}-{study2.irb}-{safe_name(left.instrument)}.xlsx"
    destination.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in WORKBOOK_SHEETS:
        left_headers, left_rows = read_sheet(left.path, sheet_name)
        right_headers, right_rows = read_sheet(right.path, sheet_name)
        if sheet_name == "cleaned":
            headers, rows = merge_cleaned_sheet(left_headers, left_rows, right_headers, right_rows)
        elif sheet_name in {"timepoint_dictionary", "column_variable_dictionary"}:
            headers, rows = merge_dictionary_rows(sheet_name, study1.irb, left_headers, left_rows, study2.irb, right_headers, right_rows)
        else:
            headers, rows = merge_concat_sheet(left_headers, left_rows, right_headers, right_rows)
        write_sheet(workbook, sheet_name, headers, rows)

    workbook.save(destination)
    return destination


def copy_unique_workbook(workbook: InstrumentWorkbook, output_dir: Path) -> Path:
    destination = output_dir / workbook.category / workbook.path.name
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.exists():
        destination = destination.with_name(f"{destination.stem}_copy{destination.suffix}")
    shutil.copy2(workbook.path, destination)
    return destination


def merge_dictionary(study1: StudyInfo, study2: StudyInfo, output_cleaned_dir: Path) -> Path:
    destination = output_cleaned_dir / "dictionary.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in DICTIONARY_SHEETS:
        left_headers, left_rows = read_sheet(study1.dictionary_path, sheet_name)
        right_headers, right_rows = read_sheet(study2.dictionary_path, sheet_name)
        headers, rows = merge_dictionary_rows(sheet_name, study1.irb, left_headers, left_rows, study2.irb, right_headers, right_rows)
        write_sheet(workbook, sheet_name, headers, rows)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def copy_source_study_folders(study1: StudyInfo, study2: StudyInfo, output_folder: Path) -> list[Path]:
    archive_root = output_folder / "studies"
    archive_root.mkdir(parents=True, exist_ok=True)
    archived: list[Path] = []
    for study in (study1, study2):
        destination = archive_root / study.folder.name
        if destination.exists():
            raise FileExistsError(f"Archived study folder already exists: {destination}")
        shutil.copytree(study.folder, destination)
        archived.append(destination)
    return archived


def merge_studies(
    study1_folder: str | Path,
    study2_folder: str | Path,
    output_folder: str | Path | None = None,
    overwrite: bool = False,
) -> MergeStudiesResult:
    study1 = study_info(study1_folder)
    study2 = study_info(study2_folder)
    output = Path(output_folder) if output_folder is not None else inferred_output_folder(study1, study2)
    if output.exists() and overwrite:
        shutil.rmtree(output)
    if output.exists() and any(output.iterdir()):
        raise FileExistsError(f"Output folder already exists and is not empty: {output}")

    output_cleaned_dir = output / "data" / "cleaned"
    output_cleaned_dir.mkdir(parents=True, exist_ok=True)
    archived_study_folders = copy_source_study_folders(study1, study2, output)
    dictionary_path = merge_dictionary(study1, study2, output_cleaned_dir)

    left_workbooks = collect_instrument_workbooks(study1)
    right_workbooks = collect_instrument_workbooks(study2)
    pairs, left_only, right_only = match_instruments(left_workbooks, right_workbooks)

    merged_paths = [
        merge_workbook(left, right, study1, study2, output_cleaned_dir)
        for left, right in pairs
    ]
    copied_paths = [copy_unique_workbook(workbook, output_cleaned_dir) for workbook in [*left_only, *right_only]]
    build_subject_timepoints(output)

    return MergeStudiesResult(
        study1=study1.folder,
        study2=study2.folder,
        output_folder=output,
        dictionary_path=dictionary_path,
        merged_workbooks=merged_paths,
        copied_workbooks=copied_paths,
        archived_study_folders=archived_study_folders,
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--study1", required=True, help="First cleaned study folder.")
    parser.add_argument("--study2", required=True, help="Second cleaned study folder.")
    parser.add_argument("--out", help="Output merged study folder. Defaults to IRB1-IRB2-study name beside study1.")
    parser.add_argument("--overwrite", action="store_true", help="Remove existing output folder before writing.")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    result = merge_studies(args.study1, args.study2, output_folder=args.out, overwrite=args.overwrite)
    print(result.output_folder)
    print(f"dictionary={result.dictionary_path}")
    print(f"merged_workbooks={len(result.merged_workbooks)}")
    print(f"copied_workbooks={len(result.copied_workbooks)}")
    print(f"archived_study_folders={len(result.archived_study_folders)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
