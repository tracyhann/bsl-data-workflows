import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.merge_studies.run import merge_cleaned_sheet, merge_studies


REQUIRED_SHEETS = [
    "raw",
    "raw_labels",
    "cleaned",
    "timepoint_dictionary",
    "column_variable_dictionary",
    "excluded_rows",
]


def write_dictionary(path: Path, irb: str, instruments: list[tuple[str, str]]) -> None:
    workbook = Workbook()
    subject = workbook.active
    subject.title = "subject_id"
    subject.append(["IRB", "subid", "standardized", "raw_entries"])
    subject.append([irb, "s001", f"{irb}_s001", f"{irb}_s001"])

    event = workbook.create_sheet("event")
    event.append(["arm", "order", "event_name", "event_label", "abbreviation"])
    event.append(["1", 1, "baseline_arm_1", "Baseline (Visit 1)", "V1"])

    instrument = workbook.create_sheet("instrument")
    instrument.append(["instrument", "instrument_label", "abbreviation", "number_of_events", "events", "number_of_columns"])
    for name, label in instruments:
        instrument.append([name, label, "", 1, "baseline_arm_1", 3])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_instrument_workbook(
    path: Path,
    irb: str,
    instrument_name: str,
    cleaned_extra_headers: list[str],
    cleaned_rows: list[list[object]],
    raw_extra_headers: list[str] | None = None,
) -> None:
    raw_extra_headers = raw_extra_headers or cleaned_extra_headers
    workbook = Workbook()
    raw = workbook.active
    raw.title = "raw"
    raw_headers = ["record_id", "redcap_event_name", *raw_extra_headers]
    raw.append(raw_headers)
    for row in cleaned_rows:
        values_by_header = dict(zip(["IRB", "subid", "arm", "visit", "date", "record_id", "redcap_event_name", *cleaned_extra_headers], row))
        raw.append([values_by_header.get("record_id"), values_by_header.get("redcap_event_name"), *[values_by_header.get(h) for h in raw_extra_headers]])

    raw_labels = workbook.create_sheet("raw_labels")
    raw_labels.append(raw_headers)
    raw_labels.append(["Record ID", "Event Name", *[f"Label {h}" for h in raw_extra_headers]])

    cleaned = workbook.create_sheet("cleaned")
    cleaned.append(["IRB", "subid", "arm", "visit", "date", "record_id", "redcap_event_name", *cleaned_extra_headers])
    for row in cleaned_rows:
        cleaned.append(row)

    timepoints = workbook.create_sheet("timepoint_dictionary")
    timepoints.append(["arm", "order", "event_name", "event_label", "abbreviation"])
    timepoints.append(["1", 1, "baseline_arm_1", "Baseline (Visit 1)", "V1"])

    columns = workbook.create_sheet("column_variable_dictionary")
    columns.append(["column_name", "column_labels", "clean_column_name", "keep"])
    for header in ["record_id", "redcap_event_name", *cleaned_extra_headers]:
        columns.append([header, f"Label {header}", "", True])

    excluded = workbook.create_sheet("excluded_rows")
    excluded.append(raw_headers)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_duplicate_visit_workbook(path: Path, irb: str) -> None:
    workbook = Workbook()
    raw = workbook.active
    raw.title = "raw"
    raw.append(["record_id", "redcap_event_name", "visit", "score"])
    raw.append([f"{irb}_s001", "baseline_arm_2", "Screening (Visit 1)", 10])

    raw_labels = workbook.create_sheet("raw_labels")
    raw_labels.append(["Record ID", "Event Name", "Visit", "Score"])
    raw_labels.append([f"{irb}_s001", "Baseline (Visit 2)", "Screening (Visit 1)", 10])

    cleaned = workbook.create_sheet("cleaned")
    cleaned.append(
        [
            "IRB",
            "subid",
            "arm",
            "visit",
            "date",
            "record_id",
            "redcap_event_name",
            "visit",
            "score",
        ]
    )
    cleaned.append(
        [
            irb,
            "s001",
            "2",
            "V2",
            "2026-01-01",
            f"{irb}_s001",
            "Baseline (Visit 2) (Arm 2: Baseline/Active Study/Follow-ups)",
            "Screening (Visit 1)",
            10,
        ]
    )

    timepoints = workbook.create_sheet("timepoint_dictionary")
    timepoints.append(["arm", "order", "event_name", "event_label", "abbreviation"])
    timepoints.append(["2", 1, "baseline_arm_2", "Baseline (Visit 2)", "V2"])

    columns = workbook.create_sheet("column_variable_dictionary")
    columns.append(["column_name", "column_labels", "clean_column_name", "keep"])
    for header in ["record_id", "redcap_event_name", "visit", "score"]:
        columns.append([header, f"Label {header}", "", True])

    excluded = workbook.create_sheet("excluded_rows")
    excluded.append(["record_id", "redcap_event_name", "visit", "score"])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


class MergeStudiesTests(unittest.TestCase):
    def test_cleaned_rows_sort_by_subid_arm_date_then_visit_label(self):
        headers = ["IRB", "subid", "arm", "visit", "date", "record_id", "redcap_event_name", "score"]
        rows = [
            {"IRB": "58807", "subid": "s031", "arm": "2", "visit": "V10", "date": "2023-06-02", "record_id": "58807_s031", "redcap_event_name": "v10_arm_2", "score": 10},
            {"IRB": "54909; 58807", "subid": "s031", "arm": "2", "visit": "V2", "date": "2023-04-25", "record_id": "54909_s031; 58807_s031", "redcap_event_name": "v2_arm_2", "score": 2},
            {"IRB": "58807", "subid": "s031", "arm": "2", "visit": "V1", "date": "", "record_id": "58807_s031", "redcap_event_name": "v1_arm_2", "score": 1},
            {"IRB": "58807", "subid": "s030", "arm": "2", "visit": "V99", "date": "2023-12-01", "record_id": "58807_s030", "redcap_event_name": "v99_arm_2", "score": 99},
        ]

        _, sorted_rows = merge_cleaned_sheet(headers, rows, headers, [])

        self.assertEqual(
            [(row["subid"], row["arm"], row["visit"], row["date"]) for row in sorted_rows],
            [
                ("s030", "2", "V99", "2023-12-01"),
                ("s031", "2", "V2", "2023-04-25"),
                ("s031", "2", "V10", "2023-06-02"),
                ("s031", "2", "V1", ""),
            ],
        )

    def test_merges_overlapping_instrument_rows_and_column_dictionaries(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            study1 = root / "111-BRAINS"
            study2 = root / "222-BRAINS"
            write_dictionary(study1 / "data" / "cleaned" / "dictionary.xlsx", "111", [("shared_form", "Shared Form")])
            write_dictionary(study2 / "data" / "cleaned" / "dictionary.xlsx", "222", [("shared_form", "Shared Form")])

            write_instrument_workbook(
                study1 / "data" / "cleaned" / "assessments" / "111-shared_form.xlsx",
                "111",
                "shared_form",
                ["score", "note", "left_only"],
                [["111", "s001", "1", "V1", "2026-01-02", "111_s001", "baseline_arm_1", "5", "alpha", "left"]],
            )
            write_instrument_workbook(
                study2 / "data" / "cleaned" / "assessments" / "222-shared_form.xlsx",
                "222",
                "shared_form",
                ["score", "note", "right_only"],
                [
                    ["222", "s001", "1", "V1", "2026-01-01", "222_s001", "baseline_arm_1", 5.0, "beta", "right"],
                    ["222", "s002", "1", "V1", "2026-01-03", "222_s002", "baseline_arm_1", 9, "gamma", "right2"],
                ],
            )

            result = merge_studies(study1, study2)
            output = result.output_folder
            self.assertEqual(output, root / "111-222-BRAINS")

            merged_path = output / "data" / "cleaned" / "assessments" / "111-222-shared_form.xlsx"
            self.assertTrue(merged_path.exists())
            workbook = load_workbook(merged_path, data_only=True)
            self.assertEqual(workbook.sheetnames, REQUIRED_SHEETS)

            cleaned_rows = list(workbook["cleaned"].iter_rows(values_only=True))
            self.assertEqual(
                cleaned_rows[0],
                ("IRB", "subid", "arm", "visit", "date", "record_id", "redcap_event_name", "score", "note", "left_only", "right_only"),
            )
            self.assertEqual(cleaned_rows[1], ("111; 222", "s001", "1", "V1", "2026-01-01", "111_s001; 222_s001", "baseline_arm_1", "5", "alpha; beta", "left", "right"))
            self.assertEqual(cleaned_rows[2], ("222", "s002", "1", "V1", "2026-01-03", "222_s002", "baseline_arm_1", 9, "gamma", None, "right2"))

            raw_rows = list(workbook["raw"].iter_rows(values_only=True))
            self.assertEqual(raw_rows[0], ("record_id", "redcap_event_name", "score", "note", "left_only", "right_only"))
            self.assertEqual(len(raw_rows), 4)

            column_rows = list(workbook["column_variable_dictionary"].iter_rows(values_only=True))
            self.assertEqual(column_rows[0], ("IRB", "column_name", "column_labels", "clean_column_name", "keep"))
            score_row = next(row for row in column_rows if row[1] == "score")
            self.assertEqual(score_row[0], "111; 222")

            dictionary = load_workbook(output / "data" / "cleaned" / "dictionary.xlsx", data_only=True)
            event_rows = list(dictionary["event"].iter_rows(values_only=True))
            self.assertEqual(event_rows[0][0], "IRB")
            self.assertEqual(event_rows[1][0], "111; 222")
            instrument_rows = list(dictionary["instrument"].iter_rows(values_only=True))
            self.assertEqual(instrument_rows[1][0], "111; 222")
            subject_timepoints_path = output / "data" / "cleaned" / "subjects" / "subject_timepoints.xlsx"
            self.assertTrue(subject_timepoints_path.exists())
            subject_timepoints = load_workbook(subject_timepoints_path, data_only=True)
            subject_timepoint_rows = list(subject_timepoints["subject_timepoints"].iter_rows(values_only=True))
            self.assertEqual(subject_timepoint_rows[0][:5], ("IRB", "subid", "arm", "visit", "earliest_entry_date"))
            self.assertIn(("111; 222", "s001", "1", "V1", "2026-01-01"), [row[:5] for row in subject_timepoint_rows])
            self.assertTrue((output / "studies" / "111-BRAINS" / "data" / "cleaned" / "dictionary.xlsx").exists())
            self.assertTrue((output / "studies" / "222-BRAINS" / "data" / "cleaned" / "dictionary.xlsx").exists())

    def test_merge_preserves_index_visit_when_cleaned_sheet_has_duplicate_raw_visit_column(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            study1 = root / "111-BRAINS"
            study2 = root / "222-BRAINS"
            write_dictionary(study1 / "data" / "cleaned" / "dictionary.xlsx", "111", [("eeg", "EEG")])
            write_dictionary(study2 / "data" / "cleaned" / "dictionary.xlsx", "222", [("eeg", "EEG")])
            write_duplicate_visit_workbook(study1 / "data" / "cleaned" / "neuroimaging" / "111-eeg.xlsx", "111")
            write_duplicate_visit_workbook(study2 / "data" / "cleaned" / "neuroimaging" / "222-eeg.xlsx", "222")

            result = merge_studies(study1, study2)
            merged_path = result.output_folder / "data" / "cleaned" / "neuroimaging" / "111-222-eeg.xlsx"
            workbook = load_workbook(merged_path, data_only=True)
            rows = list(workbook["cleaned"].iter_rows(values_only=True))

            self.assertEqual(
                rows[0][:9],
                ("IRB", "subid", "arm", "visit", "date", "record_id", "redcap_event_name", "visit.1", "score"),
            )
            self.assertEqual(rows[1][3], "V2")
            self.assertEqual(rows[1][7], "Screening (Visit 1)")

    def test_copies_non_overlapping_instruments_to_their_category(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            study1 = root / "111-BRAINS"
            study2 = root / "222-BRAINS"
            write_dictionary(study1 / "data" / "cleaned" / "dictionary.xlsx", "111", [("left_form", "Left Form")])
            write_dictionary(study2 / "data" / "cleaned" / "dictionary.xlsx", "222", [("right_form", "Right Form")])
            write_instrument_workbook(
                study1 / "data" / "cleaned" / "subjects" / "111-left_form.xlsx",
                "111",
                "left_form",
                ["value"],
                [["111", "s001", "1", "V1", "2026-01-01", "111_s001", "baseline_arm_1", "left"]],
            )
            write_instrument_workbook(
                study2 / "data" / "cleaned" / "treatments" / "222-right_form.xlsx",
                "222",
                "right_form",
                ["value"],
                [["222", "s001", "1", "V1", "2026-01-01", "222_s001", "baseline_arm_1", "right"]],
            )

            result = merge_studies(study1, study2)

            self.assertTrue((result.output_folder / "data" / "cleaned" / "subjects" / "111-left_form.xlsx").exists())
            self.assertTrue((result.output_folder / "data" / "cleaned" / "treatments" / "222-right_form.xlsx").exists())


if __name__ == "__main__":
    unittest.main()
