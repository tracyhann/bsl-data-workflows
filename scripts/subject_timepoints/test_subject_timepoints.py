import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.subject_timepoints.run import build_subject_timepoints


def write_instrument(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "cleaned"
    worksheet.append(["IRB", "subid", "arm", "visit", "date", "record_id", "redcap_event_name", "score"])
    for row in rows:
        worksheet.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_instrument_with_duplicate_visit(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "cleaned"
    worksheet.append(
        [
            "IRB",
            "subid",
            "arm",
            "visit",
            "date",
            "record_id",
            "redcap_event_name",
            "visit",
            "score",
        ]
    )
    worksheet.append(
        [
            "111",
            "s001",
            "1",
            "V1",
            "2026-01-01",
            "111_s001",
            "screening_visit_1_arm_1",
            "Screening (Visit 1)",
            10,
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


class SubjectTimepointsTests(unittest.TestCase):
    def test_builds_subject_timepoints_from_all_cleaned_instrument_workbooks(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study = Path(tmpdir) / "study"
            cleaned = study / "data" / "cleaned"
            write_instrument(
                cleaned / "assessments" / "111-scale.xlsx",
                [
                    ["111", "s001", "1", "V1", "2026-01-02", "111_s001", "baseline_arm_1", 10],
                    ["111", "s001", "1", "V1", "01/01/2026", "111_s001", "baseline_arm_1", 11],
                    ["111", "s002", "1", "V1", "", "111_s002", "baseline_arm_1", 12],
                ],
            )
            write_instrument(
                cleaned / "treatments" / "111-treatment.xlsx",
                [
                    ["111", "s001", "1", "V1", "2026-01-03 09:30", "111_s001", "baseline_arm_1", 20],
                    ["111", "s001", "1", "V2", "not a date", "111_s001", "visit_2_arm_1", 21],
                ],
            )
            dictionary = Workbook()
            dictionary.active.title = "subject_id"
            dictionary.save(cleaned / "dictionary.xlsx")

            result = build_subject_timepoints(study)

            self.assertEqual(result.output_path, cleaned / "subjects" / "subject_timepoints.xlsx")
            self.assertEqual(result.timepoint_count, 3)
            workbook = load_workbook(result.output_path, data_only=True)
            rows = list(workbook["subject_timepoints"].iter_rows(values_only=True))
            self.assertEqual(
                rows[0],
                (
                    "IRB",
                    "subid",
                    "arm",
                    "visit",
                    "earliest_entry_date",
                    "earliest_date_source",
                    "latest_entry_date",
                    "latest_date_source",
                    "span",
                    "values",
                ),
            )
            self.assertEqual(
                rows[1],
                (
                    "111",
                    "s001",
                    "1",
                    "V1",
                    "2026-01-01",
                    str(cleaned / "assessments" / "111-scale.xlsx"),
                    "2026-01-03",
                    str(cleaned / "treatments" / "111-treatment.xlsx"),
                    2,
                    "2026-01-01; 2026-01-02; 2026-01-03",
                ),
            )
            self.assertEqual(rows[2], ("111", "s001", "1", "V2", None, None, None, None, None, None))
            self.assertEqual(rows[3], ("111", "s002", "1", "V1", None, None, None, None, None, None))

    def test_ignores_nested_archived_study_folders(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study = Path(tmpdir) / "study"
            cleaned = study / "data" / "cleaned"
            write_instrument(
                cleaned / "assessments" / "111-scale.xlsx",
                [["111", "s001", "1", "V1", "2026-01-01", "111_s001", "baseline_arm_1", 10]],
            )
            write_instrument(
                study / "studies" / "111-old" / "data" / "cleaned" / "assessments" / "111-old.xlsx",
                [["111", "s999", "1", "V1", "2020-01-01", "111_s999", "baseline_arm_1", 10]],
            )

            result = build_subject_timepoints(study)
            workbook = load_workbook(result.output_path, data_only=True)
            rows = list(workbook["subject_timepoints"].iter_rows(values_only=True))

            self.assertEqual(len(rows), 2)
            self.assertEqual(rows[1][1], "s001")

    def test_uses_protected_index_visit_when_raw_instrument_has_duplicate_visit_column(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study = Path(tmpdir) / "study"
            cleaned = study / "data" / "cleaned"
            write_instrument_with_duplicate_visit(cleaned / "neuroimaging" / "111-eeg.xlsx")

            result = build_subject_timepoints(study)
            workbook = load_workbook(result.output_path, data_only=True)
            rows = list(workbook["subject_timepoints"].iter_rows(values_only=True))

            self.assertEqual(rows[1][3], "V1")

    def test_omits_date_sources_when_timepoint_span_is_zero(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study = Path(tmpdir) / "study"
            cleaned = study / "data" / "cleaned"
            write_instrument(
                cleaned / "assessments" / "111-scale.xlsx",
                [["111", "s001", "1", "V1", "2026-01-01", "111_s001", "baseline_arm_1", 10]],
            )
            write_instrument(
                cleaned / "treatments" / "111-treatment.xlsx",
                [["111", "s001", "1", "V1", "2026-01-01", "111_s001", "baseline_arm_1", 20]],
            )

            result = build_subject_timepoints(study)
            workbook = load_workbook(result.output_path, data_only=True)
            rows = list(workbook["subject_timepoints"].iter_rows(values_only=True))

            self.assertEqual(rows[1][4], "2026-01-01")
            self.assertIsNone(rows[1][5])
            self.assertEqual(rows[1][6], "2026-01-01")
            self.assertIsNone(rows[1][7])
            self.assertEqual(rows[1][8], 0)

    def test_combines_irb_variants_for_same_subject_arm_and_visit(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study = Path(tmpdir) / "study"
            cleaned = study / "data" / "cleaned"
            write_instrument(
                cleaned / "assessments" / "merged-scale.xlsx",
                [
                    ["54909", "s001", "1", "V1", "2026-01-01", "54909_s001", "baseline_arm_1", 10],
                    ["58807", "s001", "1", "V1", "2026-01-03", "58807_s001", "baseline_arm_1", 11],
                    ["54909; 58807", "s001", "1", "V1", "2026-01-02", "54909_s001; 58807_s001", "baseline_arm_1", 12],
                ],
            )

            result = build_subject_timepoints(study)
            workbook = load_workbook(result.output_path, data_only=True)
            rows = list(workbook["subject_timepoints"].iter_rows(values_only=True))

            self.assertEqual(result.timepoint_count, 1)
            self.assertEqual(rows[1][0], "54909; 58807")
            self.assertEqual(rows[1][1:5], ("s001", "1", "V1", "2026-01-01"))
            self.assertEqual(rows[1][6], "2026-01-03")
            self.assertEqual(rows[1][8], 2)
            self.assertEqual(rows[1][9], "2026-01-01; 2026-01-02; 2026-01-03")

    def test_sorts_by_subid_arm_earliest_entry_date_then_visit(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study = Path(tmpdir) / "study"
            cleaned = study / "data" / "cleaned"
            write_instrument(
                cleaned / "assessments" / "111-scale.xlsx",
                [
                    ["111", "s002", "1", "V1", "2026-01-01", "111_s002", "baseline_arm_1", 1],
                    ["111", "s001", "2", "V2", "2026-01-03", "111_s001", "visit_2_arm_2", 2],
                    ["111", "s001", "1", "V2", "2026-01-03", "111_s001", "visit_2_arm_1", 3],
                    ["111", "s001", "1", "V10", "2026-01-01", "111_s001", "visit_10_arm_1", 10],
                    ["111", "s001", "1", "V1", "", "111_s001", "visit_1_arm_1", 1],
                ],
            )

            result = build_subject_timepoints(study)
            workbook = load_workbook(result.output_path, data_only=True)
            rows = list(workbook["subject_timepoints"].iter_rows(min_row=2, values_only=True))

            self.assertEqual(
                [(row[1], row[2], row[3], row[4]) for row in rows],
                [
                    ("s001", "1", "V10", "2026-01-01"),
                    ("s001", "1", "V2", "2026-01-03"),
                    ("s001", "1", "V1", None),
                    ("s001", "2", "V2", "2026-01-03"),
                    ("s002", "1", "V1", "2026-01-01"),
                ],
            )

if __name__ == "__main__":
    unittest.main()
