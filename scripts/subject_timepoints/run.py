#!/usr/bin/env python3
"""Summarize subject timepoints and observed dates across cleaned instruments."""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook


if __package__ in {None, ""}:
    sys.path.append(str(Path(__file__).resolve().parents[2]))

from scripts.standardize_date.date_standardizer import standardize_date


OUTPUT_COLUMNS = [
    "IRB",
    "subid",
    "arm",
    "visit",
    "earliest_entry_date",
    "earliest_date_source",
    "latest_entry_date",
    "latest_date_source",
    "span",
    "values",
]
INDEX_COLUMNS = ["IRB", "subid", "arm", "visit"]
SKIP_DIRS = {"redcap", "raw_exports"}


@dataclass(frozen=True)
class SubjectTimepoint:
    irb: str
    subid: str
    arm: str
    visit: str
    dates: tuple[str, ...]
    date_sources: tuple[tuple[str, tuple[str, ...]], ...] = ()

    @property
    def earliest_entry_date(self) -> str | None:
        return self.dates[0] if self.dates else None

    @property
    def latest_entry_date(self) -> str | None:
        return self.dates[-1] if self.dates else None

    def sources_for_date(self, value: str | None) -> str | None:
        if not value:
            return None
        for observed_date, sources in self.date_sources:
            if observed_date == value:
                return "; ".join(sources) if sources else None
        return None

    @property
    def earliest_date_source(self) -> str | None:
        if self.span is None or self.span <= 0:
            return None
        return self.sources_for_date(self.earliest_entry_date)

    @property
    def latest_date_source(self) -> str | None:
        if self.span is None or self.span <= 0:
            return None
        return self.sources_for_date(self.latest_entry_date)

    @property
    def span(self) -> int | None:
        if len(self.dates) < 2:
            return 0 if self.dates else None
        return (date.fromisoformat(self.dates[-1]) - date.fromisoformat(self.dates[0])).days

    @property
    def values(self) -> str | None:
        return "; ".join(self.dates) if self.dates else None


@dataclass(frozen=True)
class SubjectTimepointsResult:
    study_folder: Path
    cleaned_dir: Path
    output_path: Path
    workbook_count: int
    row_count: int
    timepoint_count: int


def nonblank(value: object) -> bool:
    return value is not None and str(value).strip() != ""


def normalize_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def header_lookup(headers: list[object]) -> dict[str, int]:
    lookup: dict[str, int] = {}
    for index, header in enumerate(headers):
        normalized_header = normalize_header(header)
        if normalized_header and normalized_header not in lookup:
            lookup[normalized_header] = index
    return lookup


def required_index_positions(headers: list[object]) -> dict[str, int] | None:
    lookup = header_lookup(headers)
    positions: dict[str, int] = {}
    for column in INDEX_COLUMNS:
        position = lookup.get(normalize_header(column))
        if position is None:
            return None
        positions[column] = position
    return positions


def date_position(headers: list[object]) -> int | None:
    lookup = header_lookup(headers)
    return lookup.get("date")


def instrument_workbooks(cleaned_dir: Path) -> list[Path]:
    workbooks: list[Path] = []
    if not cleaned_dir.exists():
        raise FileNotFoundError(f"Missing cleaned directory: {cleaned_dir}")

    for category_dir in sorted(path for path in cleaned_dir.iterdir() if path.is_dir()):
        if category_dir.name in SKIP_DIRS:
            continue
        for path in sorted(category_dir.glob("*.xlsx")):
            if path.name.startswith("~$") or path.name == "dictionary.xlsx" or path.name == "subject_timepoints.xlsx":
                continue
            workbooks.append(path)
    return workbooks


def iter_cleaned_rows(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "cleaned" not in workbook.sheetnames:
        return
    worksheet = workbook["cleaned"]
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = list(next(rows))
    except StopIteration:
        return

    index_positions = required_index_positions(headers)
    if index_positions is None:
        return
    date_index = date_position(headers)

    for row in rows:
        if not any(nonblank(value) for value in row):
            continue
        key_values = {
            column: str(row[position] or "").strip()
            for column, position in index_positions.items()
        }
        if not all(key_values.values()):
            continue
        raw_date = row[date_index] if date_index is not None and date_index < len(row) else None
        yield key_values, standardize_date(raw_date)


def natural_subid(value: str) -> tuple[str, int, str]:
    match = re.search(r"(\d+)", value)
    if match:
        return (value[: match.start()], int(match.group(1)), value)
    return (value, -1, value)


def natural_visit(value: str) -> tuple[str, int, str]:
    match = re.search(r"(\d+)", value)
    if match:
        return (value[: match.start()], int(match.group(1)), value)
    return (value, -1, value)


def split_irb_values(value: object) -> list[str]:
    return [
        item.strip()
        for item in re.split(r"[;,]", str(value or ""))
        if item.strip()
    ]


def irb_sort_key(value: str) -> tuple[int, int, str]:
    if value.isdigit():
        return (0, int(value), value)
    return (1, 0, value)


def combine_irb_values(values: set[str]) -> str:
    expanded: set[str] = set()
    for value in values:
        expanded.update(split_irb_values(value))
    return "; ".join(sorted(expanded, key=irb_sort_key))


def timepoint_sort_key(timepoint: SubjectTimepoint) -> tuple[object, ...]:
    earliest = timepoint.earliest_entry_date
    return (
        natural_subid(timepoint.subid),
        timepoint.arm,
        (1, "") if earliest is None else (0, earliest),
        natural_visit(timepoint.visit),
        timepoint.irb,
    )


def source_path_for_workbook(cleaned_dir: Path, workbook_path: Path) -> str:
    study_folder = cleaned_dir.parent.parent
    try:
        return str(study_folder / workbook_path.relative_to(study_folder))
    except ValueError:
        return str(workbook_path)


def collect_subject_timepoints(cleaned_dir: Path) -> tuple[list[SubjectTimepoint], int, int]:
    date_sources_by_key: dict[tuple[str, str, str], dict[str, set[str]]] = {}
    irbs_by_key: dict[tuple[str, str, str], set[str]] = {}
    workbook_count = 0
    row_count = 0

    for workbook_path in instrument_workbooks(cleaned_dir):
        workbook_count += 1
        source_path = source_path_for_workbook(cleaned_dir, workbook_path)
        for key_values, canonical_date in iter_cleaned_rows(workbook_path):
            key = (
                key_values["subid"],
                key_values["arm"],
                key_values["visit"],
            )
            date_sources_by_key.setdefault(key, {})
            irbs_by_key.setdefault(key, set()).add(key_values["IRB"])
            if canonical_date:
                date_sources_by_key[key].setdefault(canonical_date, set()).add(source_path)
            row_count += 1

    timepoints = [
        SubjectTimepoint(
            irb=combine_irb_values(irbs_by_key.get(key, set())),
            subid=key[0],
            arm=key[1],
            visit=key[2],
            dates=tuple(sorted(date_sources)),
            date_sources=tuple(
                (observed_date, tuple(sorted(sources)))
                for observed_date, sources in sorted(date_sources.items())
            ),
        )
        for key, date_sources in date_sources_by_key.items()
    ]
    return sorted(timepoints, key=timepoint_sort_key), workbook_count, row_count


def write_subject_timepoints(output_path: Path, timepoints: list[SubjectTimepoint]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "subject_timepoints"
    worksheet.append(OUTPUT_COLUMNS)
    for timepoint in timepoints:
        worksheet.append(
            [
                timepoint.irb,
                timepoint.subid,
                timepoint.arm,
                timepoint.visit,
                timepoint.earliest_entry_date,
                timepoint.earliest_date_source,
                timepoint.latest_entry_date,
                timepoint.latest_date_source,
                timepoint.span,
                timepoint.values,
            ]
        )

    for column_cells in worksheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 80)
        worksheet.column_dimensions[column_cells[0].column_letter].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def build_subject_timepoints(study_folder: str | Path, output_path: str | Path | None = None) -> SubjectTimepointsResult:
    study_folder = Path(study_folder)
    cleaned_dir = study_folder / "data" / "cleaned"
    output_path = Path(output_path) if output_path is not None else cleaned_dir / "subjects" / "subject_timepoints.xlsx"

    timepoints, workbook_count, row_count = collect_subject_timepoints(cleaned_dir)
    write_subject_timepoints(output_path, timepoints)
    return SubjectTimepointsResult(
        study_folder=study_folder,
        cleaned_dir=cleaned_dir,
        output_path=output_path,
        workbook_count=workbook_count,
        row_count=row_count,
        timepoint_count=len(timepoints),
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--study-folder", required=True, type=Path, help="Study folder with data/cleaned outputs.")
    parser.add_argument("--out", type=Path, help="Optional output workbook path.")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    result = build_subject_timepoints(args.study_folder, output_path=args.out)
    print(result.output_path)
    print(f"workbooks={result.workbook_count}")
    print(f"cleaned_rows={result.row_count}")
    print(f"timepoints={result.timepoint_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
