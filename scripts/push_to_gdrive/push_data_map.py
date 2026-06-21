#!/usr/bin/env python3
"""Fill an existing Google Sheet from local data-map workbooks.

Each local data-map workbook maps to one Google Sheet tab by filename:

- ``platforms-data-map.xlsx`` -> ``platforms``
- ``assessments-data-map.xlsx`` -> ``assessments``
- ``biologics_biometrics-data-map.xlsx`` -> ``biologics_biometrics``

Within each matched tab, columns are matched by header name. Extra target
columns are preserved.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


if __package__ in {None, ""}:
    sys.path.append(str(Path(__file__).resolve().parents[2]))

from scripts.push_to_gdrive.fill_in_overview import (  # noqa: E402
    OverviewSheet,
    SheetsHttpClient,
    UrllibSheetsHttpClient,
    build_value_updates,
    ensure_row_capacity,
    is_blank,
    json_headers,
    load_sheet_properties,
    load_target_headers,
    normalize_key,
    sheets_values_batch_clear_url,
    sheets_values_batch_update_url,
)
from scripts.push_to_gdrive.push_instrument_workbook import (  # noqa: E402
    resolve_access_token,
    resolve_spreadsheet_id,
)


@dataclass(frozen=True)
class PushDataMapResult:
    spreadsheet_id: str
    data_map_files: list[Path]
    dry_run: bool
    updated_tabs: list[str] = field(default_factory=list)
    skipped_tabs: list[str] = field(default_factory=list)
    skipped_columns: dict[str, list[str]] = field(default_factory=dict)
    planned_ranges: list[str] = field(default_factory=list)
    updated_cell_count: int = 0

    @property
    def web_url(self) -> str:
        return f"https://docs.google.com/spreadsheets/d/{self.spreadsheet_id}/edit"


def data_map_tab_name(path: str | Path) -> str:
    stem = Path(path).stem.strip()
    return re.sub(r"[-_\s]*data[-_\s]*map$", "", stem, flags=re.IGNORECASE) or stem


def validate_data_map_file(path: str | Path) -> Path:
    data_map_file = Path(path).expanduser().resolve()
    if not data_map_file.exists():
        raise FileNotFoundError(data_map_file)
    if data_map_file.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError(f"Expected an Excel data-map workbook, got: {data_map_file}")
    return data_map_file


def discover_data_map_files(
    *,
    study_folder: str | Path | None = None,
    data_map_dir: str | Path | None = None,
    data_map_files: list[str | Path] | None = None,
) -> list[Path]:
    candidates: list[Path] = []
    if study_folder is not None:
        candidates.extend(sorted((Path(study_folder) / "data-map").glob("*-data-map.xlsx")))
    if data_map_dir is not None:
        candidates.extend(sorted(Path(data_map_dir).expanduser().glob("*-data-map.xlsx")))
    if data_map_files:
        candidates.extend(validate_data_map_file(path) for path in data_map_files)

    seen: set[Path] = set()
    output: list[Path] = []
    for candidate in candidates:
        path = validate_data_map_file(candidate)
        if path in seen:
            continue
        seen.add(path)
        output.append(path)

    if not output:
        raise ValueError("No data-map workbooks found. Pass --study-folder, --data-map-dir, or --data-map-file.")
    return output


def coerce_cell(value: Any) -> Any:
    if value is None:
        return ""
    return value


def read_data_map_workbook(path: str | Path) -> OverviewSheet:
    workbook_path = validate_data_map_file(path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return OverviewSheet(title=data_map_tab_name(workbook_path), headers=[], rows=[])
    headers = [str(value or "").strip() for value in rows[0]]
    data_rows = [
        [coerce_cell(value) for value in row[: len(headers)]]
        for row in rows[1:]
        if any(not is_blank(value) for value in row)
    ]
    return OverviewSheet(title=data_map_tab_name(workbook_path), headers=headers, rows=data_rows)


def push_data_map(
    *,
    target: str | Path,
    study_folder: str | Path | None = None,
    data_map_dir: str | Path | None = None,
    data_map_files: list[str | Path] | None = None,
    access_token: str | None = None,
    http_client: SheetsHttpClient | None = None,
    timeout: float = 120.0,
    dry_run: bool = False,
    clear_existing: bool = True,
) -> PushDataMapResult:
    spreadsheet_id = resolve_spreadsheet_id(target)
    files = discover_data_map_files(
        study_folder=study_folder,
        data_map_dir=data_map_dir,
        data_map_files=data_map_files,
    )
    source_sheets = [read_data_map_workbook(path) for path in files]
    token = resolve_access_token(access_token)
    headers = json_headers(token)
    client = http_client or UrllibSheetsHttpClient()

    target_sheets = load_sheet_properties(
        spreadsheet_id,
        http_client=client,
        headers=headers,
        timeout=timeout,
    )

    updated_tabs: list[str] = []
    skipped_tabs: list[str] = []
    skipped_columns: dict[str, list[str]] = {}
    planned_ranges: list[str] = []
    all_updates: list[dict[str, Any]] = []
    all_clear_ranges: list[str] = []
    updated_cell_count = 0

    for source in source_sheets:
        target_properties = target_sheets.get(normalize_key(source.title))
        if target_properties is None:
            skipped_tabs.append(source.title)
            continue
        target_headers = load_target_headers(
            spreadsheet_id,
            target_properties.title,
            http_client=client,
            headers=headers,
            timeout=timeout,
        )
        updates, clear_ranges, missing_columns, cell_count = build_value_updates(
            source,
            target_properties.title,
            target_headers,
        )
        if missing_columns:
            skipped_columns[target_properties.title] = missing_columns
        if not updates:
            continue
        updated_tabs.append(target_properties.title)
        planned_ranges.extend(update["range"] for update in updates)
        updated_cell_count += cell_count
        all_updates.extend(updates)
        all_clear_ranges.extend(clear_ranges)

        if not dry_run:
            ensure_row_capacity(
                spreadsheet_id,
                target_properties,
                len(source.rows) + 1,
                http_client=client,
                headers=headers,
                timeout=timeout,
            )

    if not dry_run and all_updates:
        if clear_existing and all_clear_ranges:
            client.post(
                sheets_values_batch_clear_url(spreadsheet_id),
                json.dumps({"ranges": all_clear_ranges}).encode("utf-8"),
                headers,
                timeout,
            )
        client.post(
            sheets_values_batch_update_url(spreadsheet_id),
            json.dumps({"valueInputOption": "RAW", "data": all_updates}).encode("utf-8"),
            headers,
            timeout,
        )

    return PushDataMapResult(
        spreadsheet_id=spreadsheet_id,
        data_map_files=files,
        dry_run=dry_run,
        updated_tabs=updated_tabs,
        skipped_tabs=skipped_tabs,
        skipped_columns=skipped_columns,
        planned_ranges=planned_ranges,
        updated_cell_count=updated_cell_count,
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--target",
        required=True,
        help="Target Google Sheet URL, raw spreadsheet id, or text file containing either.",
    )
    parser.add_argument(
        "--study-folder",
        type=Path,
        help="Study folder; data-map workbooks are read from STUDY/data-map/.",
    )
    parser.add_argument(
        "--data-map-dir",
        type=Path,
        help="Directory containing *-data-map.xlsx workbooks.",
    )
    parser.add_argument(
        "--data-map-file",
        type=Path,
        action="append",
        default=[],
        help="Explicit data-map workbook to push. May be passed more than once.",
    )
    parser.add_argument(
        "--access-token",
        default=None,
        help=(
            "Google OAuth access token with Sheets write access. If omitted, the "
            "script tries GOOGLE_OAUTH_ACCESS_TOKEN, then gcloud."
        ),
    )
    parser.add_argument("--timeout", type=float, default=120.0, help="API timeout in seconds.")
    parser.add_argument("--dry-run", action="store_true", help="Plan writes without modifying Google Sheets.")
    parser.add_argument(
        "--no-clear-existing",
        action="store_true",
        help="Do not clear matched target columns before writing new data-map values.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    result = push_data_map(
        target=args.target,
        study_folder=args.study_folder,
        data_map_dir=args.data_map_dir,
        data_map_files=args.data_map_file,
        access_token=args.access_token,
        timeout=args.timeout,
        dry_run=args.dry_run,
        clear_existing=not args.no_clear_existing,
    )
    print(
        json.dumps(
            {
                "dry_run": result.dry_run,
                "spreadsheet_id": result.spreadsheet_id,
                "data_map_files": [str(path) for path in result.data_map_files],
                "updated_tabs": result.updated_tabs,
                "skipped_tabs": result.skipped_tabs,
                "skipped_columns": result.skipped_columns,
                "planned_ranges": result.planned_ranges,
                "updated_cell_count": result.updated_cell_count,
                "url": result.web_url,
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
