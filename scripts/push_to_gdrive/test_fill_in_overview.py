import json
import tempfile
import unittest
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse

from openpyxl import Workbook

from scripts.push_to_gdrive.fill_in_overview import (
    SheetsApiResponse,
    fill_in_overview,
)


class FakeSheetsHttpClient:
    def __init__(self, metadata, headers_by_sheet):
        self.metadata = metadata
        self.headers_by_sheet = headers_by_sheet
        self.get_calls = []
        self.post_calls = []

    def get(self, url, headers, timeout):
        self.get_calls.append({"url": url, "headers": headers, "timeout": timeout})
        if "values/" in url:
            encoded_range = url.split("/values/", 1)[1].split("?", 1)[0]
            sheet_name = unquote(encoded_range).split("!", 1)[0].strip("'")
            return SheetsApiResponse(
                status_code=200,
                payload={"values": [self.headers_by_sheet.get(sheet_name, [])]},
                raw_text="{}",
            )
        return SheetsApiResponse(
            status_code=200,
            payload=self.metadata,
            raw_text=json.dumps(self.metadata),
        )

    def post(self, url, body, headers, timeout):
        payload = json.loads(body.decode("utf-8"))
        self.post_calls.append(
            {
                "url": url,
                "payload": payload,
                "headers": headers,
                "timeout": timeout,
            }
        )
        if url.endswith(":batchUpdate"):
            self.apply_batch_update(payload)
        return SheetsApiResponse(status_code=200, payload={}, raw_text="{}")

    def apply_batch_update(self, payload):
        for request in payload.get("requests", []):
            if "updateSheetProperties" in request:
                properties = request["updateSheetProperties"]["properties"]
                sheet = self.sheet_by_id(properties["sheetId"])
                old_title = sheet["properties"]["title"]
                new_title = properties["title"]
                sheet["properties"]["title"] = new_title
                self.headers_by_sheet[new_title] = self.headers_by_sheet.pop(old_title, [])
            elif "duplicateSheet" in request:
                duplicate = request["duplicateSheet"]
                source_sheet = self.sheet_by_id(duplicate["sourceSheetId"])
                source_props = source_sheet["properties"]
                new_sheet = {
                    "properties": {
                        "sheetId": duplicate["newSheetId"],
                        "title": duplicate["newSheetName"],
                        "gridProperties": dict(source_props.get("gridProperties", {})),
                    }
                }
                self.metadata["sheets"].append(new_sheet)
                self.headers_by_sheet[duplicate["newSheetName"]] = list(
                    self.headers_by_sheet.get(source_props["title"], [])
                )
            elif "appendDimension" in request:
                append = request["appendDimension"]
                sheet = self.sheet_by_id(append["sheetId"])
                grid = sheet["properties"].setdefault("gridProperties", {})
                if append["dimension"] == "ROWS":
                    grid["rowCount"] = grid.get("rowCount", 0) + append["length"]
                elif append["dimension"] == "COLUMNS":
                    grid["columnCount"] = grid.get("columnCount", 0) + append["length"]

    def sheet_by_id(self, sheet_id):
        for sheet in self.metadata.get("sheets", []):
            if sheet.get("properties", {}).get("sheetId") == sheet_id:
                return sheet
        raise KeyError(sheet_id)


def make_overview(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "assessments"
    worksheet.append(
        [
            "data",
            "description",
            "# of timepoints (collected)",
            "collection tool",
        ]
    )
    worksheet.append(["MADRS", "depression rating", 2, "REDCap (IRB: 58807)"])
    worksheet.append(["PHQ-8", "self-report depression scale", 1, "REDCap (IRB: 58807)"])
    workbook.create_sheet("neuroimaging").append(["data", "description"])
    workbook.save(path)


def make_dictionary(path: Path) -> None:
    workbook = Workbook()
    subject = workbook.active
    subject.title = "subject_id"
    subject.append(["IRB", "subid"])
    subject.append(["53879", "s001"])
    event = workbook.create_sheet("event")
    event.append(["arm", "visit"])
    event.append(["1", "V1"])
    instrument = workbook.create_sheet("instrument")
    instrument.append(["instrument", "instrument_label"])
    instrument.append(["ybocs2_7389", "YBOCS2"])
    workbook.save(path)


class FillInOverviewTests(unittest.TestCase):
    def test_fills_matching_google_tab_columns_by_header_name(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            overview = Path(tmpdir) / "overview.xlsx"
            make_overview(overview)
            fake_client = FakeSheetsHttpClient(
                metadata={
                    "sheets": [
                        {
                            "properties": {
                                "sheetId": 123,
                                "title": "Assessments",
                                "gridProperties": {"rowCount": 1, "columnCount": 6},
                            }
                        }
                    ]
                },
                headers_by_sheet={
                    "Assessments": [
                        "collection tool",
                        "Description",
                        "Data",
                        "# of timepoints (collected)",
                        "notes",
                    ]
                },
            )

            result = fill_in_overview(
                target="https://docs.google.com/spreadsheets/d/1MCjkVtR1lOJol8f95sK_W7bYdm8Bz1BkZbwJy410sYQ/edit",
                overview_file=overview,
                access_token="token",
                http_client=fake_client,
            )

            self.assertEqual(result.updated_tabs, ["Assessments"])
            self.assertEqual(result.skipped_tabs, ["neuroimaging"])
            self.assertEqual(result.updated_cell_count, 8)

            update_calls = [
                call for call in fake_client.post_calls if call["url"].endswith("/values:batchUpdate")
            ]
            self.assertEqual(len(update_calls), 1)
            ranges = {
                update["range"]: update["values"]
                for update in update_calls[0]["payload"]["data"]
            }
            self.assertEqual(
                ranges["'Assessments'!C2:C3"],
                [["MADRS"], ["PHQ-8"]],
            )
            self.assertEqual(
                ranges["'Assessments'!B2:B3"],
                [["depression rating"], ["self-report depression scale"]],
            )
            self.assertEqual(
                ranges["'Assessments'!D2:D3"],
                [[2], [1]],
            )
            self.assertEqual(
                ranges["'Assessments'!A2:A3"],
                [["REDCap (IRB: 58807)"], ["REDCap (IRB: 58807)"]],
            )

    def test_dry_run_reports_planned_updates_without_writes(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            overview = Path(tmpdir) / "overview.xlsx"
            make_overview(overview)
            fake_client = FakeSheetsHttpClient(
                metadata={
                    "sheets": [
                        {
                            "properties": {
                                "sheetId": 123,
                                "title": "Assessments",
                                "gridProperties": {"rowCount": 10, "columnCount": 6},
                            }
                        }
                    ]
                },
                headers_by_sheet={"Assessments": ["data", "description"]},
            )

            result = fill_in_overview(
                target="1MCjkVtR1lOJol8f95sK_W7bYdm8Bz1BkZbwJy410sYQ",
                overview_file=overview,
                access_token="token",
                http_client=fake_client,
                dry_run=True,
            )

            self.assertTrue(result.dry_run)
            self.assertEqual(result.updated_tabs, ["Assessments"])
            self.assertEqual(result.updated_cell_count, 4)
            self.assertEqual(fake_client.post_calls, [])

    def test_empty_overview_tab_does_not_generate_invalid_write_ranges(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            overview = Path(tmpdir) / "overview.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "assessments"
            worksheet.append(["data", "description"])
            workbook.save(overview)
            fake_client = FakeSheetsHttpClient(
                metadata={
                    "sheets": [
                        {
                            "properties": {
                                "sheetId": 123,
                                "title": "Assessments",
                                "gridProperties": {"rowCount": 10, "columnCount": 2},
                            }
                        }
                    ]
                },
                headers_by_sheet={"Assessments": ["data", "description"]},
            )

            result = fill_in_overview(
                target="1MCjkVtR1lOJol8f95sK_W7bYdm8Bz1BkZbwJy410sYQ",
                overview_file=overview,
                access_token="token",
                http_client=fake_client,
            )

            self.assertEqual(result.updated_tabs, [])
            self.assertEqual(result.planned_ranges, [])
            self.assertEqual(result.updated_cell_count, 0)
            self.assertEqual(fake_client.post_calls, [])

    def test_can_write_full_sheet_when_target_tab_has_no_headers(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            overview = Path(tmpdir) / "overview.xlsx"
            make_overview(overview)
            fake_client = FakeSheetsHttpClient(
                metadata={
                    "sheets": [
                        {
                            "properties": {
                                "sheetId": 123,
                                "title": "Assessments",
                                "gridProperties": {"rowCount": 1, "columnCount": 1},
                            }
                        }
                    ]
                },
                headers_by_sheet={"Assessments": []},
            )

            result = fill_in_overview(
                target="1MCjkVtR1lOJol8f95sK_W7bYdm8Bz1BkZbwJy410sYQ",
                overview_file=overview,
                access_token="token",
                http_client=fake_client,
                write_full_sheet_when_no_headers=True,
            )

            self.assertEqual(result.updated_tabs, ["Assessments"])
            self.assertEqual(result.updated_cell_count, 12)
            batch_calls = [
                call for call in fake_client.post_calls if call["url"].endswith(":batchUpdate")
            ]
            self.assertEqual(
                batch_calls[0]["payload"],
                {
                    "requests": [
                        {"appendDimension": {"sheetId": 123, "dimension": "ROWS", "length": 2}},
                        {"appendDimension": {"sheetId": 123, "dimension": "COLUMNS", "length": 3}},
                    ]
                },
            )
            update_calls = [
                call for call in fake_client.post_calls if call["url"].endswith("/values:batchUpdate")
            ]
            self.assertEqual(len(update_calls), 1)
            update = update_calls[0]["payload"]["data"][0]
            self.assertEqual(update["range"], "'Assessments'!A1:D3")
            self.assertEqual(
                update["values"],
                [
                    ["data", "description", "# of timepoints (collected)", "collection tool"],
                    ["MADRS", "depression rating", 2, "REDCap (IRB: 58807)"],
                    ["PHQ-8", "self-report depression scale", 1, "REDCap (IRB: 58807)"],
                ],
            )

    def test_can_rename_and_duplicate_blank_template_tab_for_multisheet_workbook(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            dictionary = Path(tmpdir) / "dictionary.xlsx"
            make_dictionary(dictionary)
            fake_client = FakeSheetsHttpClient(
                metadata={
                    "sheets": [
                        {
                            "properties": {
                                "sheetId": 123,
                                "title": "template",
                                "gridProperties": {"rowCount": 1, "columnCount": 1},
                            }
                        }
                    ]
                },
                headers_by_sheet={"template": []},
            )

            result = fill_in_overview(
                target="1MCjkVtR1lOJol8f95sK_W7bYdm8Bz1BkZbwJy410sYQ",
                overview_file=dictionary,
                access_token="token",
                http_client=fake_client,
                write_full_sheet_when_no_headers=True,
                sync_template_tab_to_source_sheets=True,
            )

            self.assertEqual(result.updated_tabs, ["subject_id", "event", "instrument"])
            tab_prep = fake_client.post_calls[0]["payload"]["requests"]
            self.assertEqual(
                tab_prep,
                [
                    {
                        "updateSheetProperties": {
                            "properties": {"sheetId": 123, "title": "subject_id"},
                            "fields": "title",
                        }
                    },
                    {
                        "duplicateSheet": {
                            "sourceSheetId": 123,
                            "newSheetId": 124,
                            "newSheetName": "event",
                            "insertSheetIndex": 1,
                        }
                    },
                    {
                        "duplicateSheet": {
                            "sourceSheetId": 123,
                            "newSheetId": 125,
                            "newSheetName": "instrument",
                            "insertSheetIndex": 2,
                        }
                    },
                ],
            )
            update_calls = [
                call for call in fake_client.post_calls if call["url"].endswith("/values:batchUpdate")
            ]
            ranges = {
                update["range"]: update["values"]
                for update in update_calls[0]["payload"]["data"]
            }
            self.assertEqual(ranges["'subject_id'!A1:B2"], [["IRB", "subid"], ["53879", "s001"]])
            self.assertEqual(ranges["'event'!A1:B2"], [["arm", "visit"], ["1", "V1"]])
            self.assertEqual(
                ranges["'instrument'!A1:B2"],
                [["instrument", "instrument_label"], ["ybocs2_7389", "YBOCS2"]],
            )


if __name__ == "__main__":
    unittest.main()
