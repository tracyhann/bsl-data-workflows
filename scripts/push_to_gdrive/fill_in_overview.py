#!/usr/bin/env python3
"""Fill an existing Google Sheet from a local study overview workbook.

The script matches overview workbook sheet names to Google Sheet tab names and
matches columns by header name. Only matched columns are written, which lets a
prepared Google Sheet keep extra columns, ordering, filters, and table setup.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Mapping, Protocol

from openpyxl import load_workbook


if __package__ in {None, ""}:
    sys.path.append(str(Path(__file__).resolve().parents[2]))

from scripts.push_to_gdrive.push_instrument_workbook import (  # noqa: E402
    resolve_access_token,
    resolve_spreadsheet_id,
)


SHEETS_API_ROOT = "https://sheets.googleapis.com/v4/spreadsheets"


@dataclass(frozen=True)
class SheetsApiResponse:
    status_code: int
    payload: dict[str, Any]
    raw_text: str


@dataclass(frozen=True)
class SheetProperties:
    sheet_id: int
    title: str
    row_count: int
    column_count: int


@dataclass(frozen=True)
class OverviewSheet:
    title: str
    headers: list[str]
    rows: list[list[Any]]


@dataclass(frozen=True)
class FillInOverviewResult:
    spreadsheet_id: str
    overview_file: Path
    dry_run: bool
    updated_tabs: list[str] = field(default_factory=list)
    skipped_tabs: list[str] = field(default_factory=list)
    skipped_columns: dict[str, list[str]] = field(default_factory=dict)
    planned_ranges: list[str] = field(default_factory=list)
    updated_cell_count: int = 0

    @property
    def web_url(self) -> str:
        return f"https://docs.google.com/spreadsheets/d/{self.spreadsheet_id}/edit"


class SheetsHttpClient(Protocol):
    def get(
        self,
        url: str,
        headers: Mapping[str, str],
        timeout: float,
    ) -> SheetsApiResponse:
        ...

    def post(
        self,
        url: str,
        body: bytes,
        headers: Mapping[str, str],
        timeout: float,
    ) -> SheetsApiResponse:
        ...


class UrllibSheetsHttpClient:
    def get(
        self,
        url: str,
        headers: Mapping[str, str],
        timeout: float,
    ) -> SheetsApiResponse:
        request = urllib.request.Request(url, headers=dict(headers), method="GET")
        return _send_request(request, timeout)

    def post(
        self,
        url: str,
        body: bytes,
        headers: Mapping[str, str],
        timeout: float,
    ) -> SheetsApiResponse:
        request = urllib.request.Request(
            url,
            data=body,
            headers=dict(headers),
            method="POST",
        )
        return _send_request(request, timeout)


def _send_request(request: urllib.request.Request, timeout: float) -> SheetsApiResponse:
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            raw = response.read().decode("utf-8")
            payload = json.loads(raw) if raw.strip() else {}
            return SheetsApiResponse(response.status, payload, raw)
    except urllib.error.HTTPError as exc:
        raw = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(
            f"Google Sheets request failed with HTTP {exc.code}: {raw}"
        ) from exc


def normalize_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def is_blank(value: object) -> bool:
    return value is None or str(value).strip() == ""


def validate_overview_file(path: str | Path) -> Path:
    overview = Path(path).expanduser().resolve()
    if not overview.exists():
        raise FileNotFoundError(overview)
    if overview.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError(f"Expected an Excel overview workbook, got: {overview}")
    return overview


def read_overview_workbook(path: str | Path) -> list[OverviewSheet]:
    workbook_path = validate_overview_file(path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    overview_sheets: list[OverviewSheet] = []
    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(value or "").strip() for value in rows[0]]
        data_rows = [
            [coerce_cell(value) for value in row[: len(headers)]]
            for row in rows[1:]
            if any(not is_blank(value) for value in row)
        ]
        overview_sheets.append(
            OverviewSheet(title=worksheet.title, headers=headers, rows=data_rows)
        )
    return overview_sheets


def coerce_cell(value: Any) -> Any:
    if value is None:
        return ""
    return value


def sheet_name_a1(title: str) -> str:
    return "'" + title.replace("'", "''") + "'"


def column_letter(index: int) -> str:
    if index < 0:
        raise ValueError(index)
    letters = ""
    number = index + 1
    while number:
        number, remainder = divmod(number - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def sheets_get_url(spreadsheet_id: str, range_name: str) -> str:
    encoded_range = urllib.parse.quote(range_name, safe="")
    return f"{SHEETS_API_ROOT}/{spreadsheet_id}/values/{encoded_range}"


def sheets_values_batch_update_url(spreadsheet_id: str) -> str:
    return f"{SHEETS_API_ROOT}/{spreadsheet_id}/values:batchUpdate"


def sheets_values_batch_clear_url(spreadsheet_id: str) -> str:
    return f"{SHEETS_API_ROOT}/{spreadsheet_id}/values:batchClear"


def sheets_batch_update_url(spreadsheet_id: str) -> str:
    return f"{SHEETS_API_ROOT}/{spreadsheet_id}:batchUpdate"


def metadata_url(spreadsheet_id: str) -> str:
    fields = "sheets(properties(sheetId,title,gridProperties(rowCount,columnCount)))"
    return (
        f"{SHEETS_API_ROOT}/{spreadsheet_id}"
        f"?includeGridData=false&fields={urllib.parse.quote(fields, safe='(),')}"
    )


def json_headers(token: str) -> dict[str, str]:
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json; charset=UTF-8",
    }


def load_sheet_properties(
    spreadsheet_id: str,
    *,
    http_client: SheetsHttpClient,
    headers: Mapping[str, str],
    timeout: float,
) -> dict[str, SheetProperties]:
    response = http_client.get(metadata_url(spreadsheet_id), headers, timeout)
    properties_by_key: dict[str, SheetProperties] = {}
    for sheet in response.payload.get("sheets", []):
        properties = sheet.get("properties", {})
        grid = properties.get("gridProperties", {})
        title = str(properties.get("title", "")).strip()
        if not title:
            continue
        properties_by_key[normalize_key(title)] = SheetProperties(
            sheet_id=int(properties.get("sheetId", 0)),
            title=title,
            row_count=int(grid.get("rowCount", 0) or 0),
            column_count=int(grid.get("columnCount", 0) or 0),
        )
    return properties_by_key


def load_target_headers(
    spreadsheet_id: str,
    sheet_title: str,
    *,
    http_client: SheetsHttpClient,
    headers: Mapping[str, str],
    timeout: float,
) -> list[str]:
    range_name = f"{sheet_name_a1(sheet_title)}!1:1"
    response = http_client.get(sheets_get_url(spreadsheet_id, range_name), headers, timeout)
    values = response.payload.get("values", [])
    if not values:
        return []
    return [str(value or "").strip() for value in values[0]]


def source_sheet_titles(source_sheets: list[OverviewSheet]) -> list[str]:
    titles: list[str] = []
    seen: set[str] = set()
    for index, source in enumerate(source_sheets, start=1):
        title = str(source.title or "").strip() or f"sheet-{index}"
        if normalize_key(title) in seen:
            title = f"{title}-{index}"
        seen.add(normalize_key(title))
        titles.append(title)
    return titles


def sync_template_tab_to_source_sheets_for_workbook(
    spreadsheet_id: str,
    *,
    target_sheets: dict[str, SheetProperties],
    source_sheets: list[OverviewSheet],
    http_client: SheetsHttpClient,
    headers: Mapping[str, str],
    timeout: float,
    dry_run: bool,
    template_tab_name: str = "template",
) -> dict[str, SheetProperties]:
    if not source_sheets:
        return target_sheets
    template = target_sheets.get(normalize_key(template_tab_name))
    if template is None:
        return target_sheets

    titles = source_sheet_titles(source_sheets)
    existing_ids = {sheet.sheet_id for sheet in target_sheets.values()}
    next_sheet_id = max(existing_ids or {template.sheet_id}) + 1
    requests: list[dict[str, Any]] = []

    first_title = titles[0]
    if normalize_key(template.title) != normalize_key(first_title):
        requests.append(
            {
                "updateSheetProperties": {
                    "properties": {"sheetId": template.sheet_id, "title": first_title},
                    "fields": "title",
                }
            }
        )

    synthetic_sheets = dict(target_sheets)
    synthetic_sheets.pop(normalize_key(template.title), None)
    synthetic_sheets[normalize_key(first_title)] = SheetProperties(
        sheet_id=template.sheet_id,
        title=first_title,
        row_count=template.row_count,
        column_count=template.column_count,
    )

    for insert_index, title in enumerate(titles[1:], start=1):
        while next_sheet_id in existing_ids:
            next_sheet_id += 1
        new_sheet_id = next_sheet_id
        existing_ids.add(new_sheet_id)
        next_sheet_id += 1
        requests.append(
            {
                "duplicateSheet": {
                    "sourceSheetId": template.sheet_id,
                    "newSheetId": new_sheet_id,
                    "newSheetName": title,
                    "insertSheetIndex": insert_index,
                }
            }
        )
        synthetic_sheets[normalize_key(title)] = SheetProperties(
            sheet_id=new_sheet_id,
            title=title,
            row_count=template.row_count,
            column_count=template.column_count,
        )

    if not requests:
        return target_sheets
    if dry_run:
        return synthetic_sheets

    http_client.post(
        sheets_batch_update_url(spreadsheet_id),
        json.dumps({"requests": requests}).encode("utf-8"),
        headers,
        timeout,
    )
    return load_sheet_properties(
        spreadsheet_id,
        http_client=http_client,
        headers=headers,
        timeout=timeout,
    )


def ensure_row_capacity(
    spreadsheet_id: str,
    sheet: SheetProperties,
    required_rows: int,
    *,
    http_client: SheetsHttpClient,
    headers: Mapping[str, str],
    timeout: float,
) -> None:
    if required_rows <= sheet.row_count:
        return
    body = {
        "requests": [
            {
                "appendDimension": {
                    "sheetId": sheet.sheet_id,
                    "dimension": "ROWS",
                    "length": required_rows - sheet.row_count,
                }
            }
        ]
    }
    http_client.post(
        sheets_batch_update_url(spreadsheet_id),
        json.dumps(body).encode("utf-8"),
        headers,
        timeout,
    )


def ensure_grid_capacity(
    spreadsheet_id: str,
    sheet: SheetProperties,
    *,
    required_rows: int,
    required_columns: int,
    http_client: SheetsHttpClient,
    headers: Mapping[str, str],
    timeout: float,
) -> None:
    requests: list[dict[str, Any]] = []
    if required_rows > sheet.row_count:
        requests.append(
            {
                "appendDimension": {
                    "sheetId": sheet.sheet_id,
                    "dimension": "ROWS",
                    "length": required_rows - sheet.row_count,
                }
            }
        )
    if required_columns > sheet.column_count:
        requests.append(
            {
                "appendDimension": {
                    "sheetId": sheet.sheet_id,
                    "dimension": "COLUMNS",
                    "length": required_columns - sheet.column_count,
                }
            }
        )
    if not requests:
        return
    http_client.post(
        sheets_batch_update_url(spreadsheet_id),
        json.dumps({"requests": requests}).encode("utf-8"),
        headers,
        timeout,
    )


def header_lookup(headers: list[str]) -> dict[str, int]:
    lookup: dict[str, int] = {}
    for index, header in enumerate(headers):
        key = normalize_key(header)
        if key and key not in lookup:
            lookup[key] = index
    return lookup


def build_value_updates(
    source: OverviewSheet,
    target_sheet: str,
    target_headers: list[str],
) -> tuple[list[dict[str, Any]], list[str], list[str], int]:
    if not source.rows:
        return [], [], [], 0

    target_lookup = header_lookup(target_headers)
    updates: list[dict[str, Any]] = []
    clear_ranges: list[str] = []
    skipped_columns: list[str] = []
    cell_count = 0

    for source_index, source_header in enumerate(source.headers):
        source_key = normalize_key(source_header)
        if not source_key:
            continue
        target_index = target_lookup.get(source_key)
        if target_index is None:
            skipped_columns.append(source_header)
            continue
        letter = column_letter(target_index)
        range_name = (
            f"{sheet_name_a1(target_sheet)}!{letter}2:{letter}{len(source.rows) + 1}"
        )
        clear_ranges.append(f"{sheet_name_a1(target_sheet)}!{letter}2:{letter}")
        values = [
            [row[source_index] if source_index < len(row) else ""]
            for row in source.rows
        ]
        updates.append({"range": range_name, "values": values})
        cell_count += len(values)
    planned_ranges = [update["range"] for update in updates]
    return updates, clear_ranges, skipped_columns, cell_count if updates else 0


def build_full_sheet_update(source: OverviewSheet, target_sheet: str) -> tuple[dict[str, Any] | None, str | None, int, int, int]:
    values = [source.headers, *source.rows]
    values = [row for row in values if any(not is_blank(value) for value in row)]
    if not values:
        return None, None, 0, 0, 0
    max_columns = max(len(row) for row in values)
    padded_values = [list(row) + [""] * (max_columns - len(row)) for row in values]
    row_count = len(padded_values)
    range_name = f"{sheet_name_a1(target_sheet)}!A1:{column_letter(max_columns - 1)}{row_count}"
    clear_range = f"{sheet_name_a1(target_sheet)}!A1:{column_letter(max_columns - 1)}"
    cell_count = row_count * max_columns
    return {"range": range_name, "values": padded_values}, clear_range, cell_count, row_count, max_columns


def fill_in_overview(
    *,
    target: str | Path,
    overview_file: str | Path,
    access_token: str | None = None,
    http_client: SheetsHttpClient | None = None,
    timeout: float = 120.0,
    dry_run: bool = False,
    clear_existing: bool = True,
    write_full_sheet_when_no_headers: bool = False,
    sync_template_tab_to_source_sheets: bool = False,
) -> FillInOverviewResult:
    spreadsheet_id = resolve_spreadsheet_id(target)
    overview_path = validate_overview_file(overview_file)
    source_sheets = read_overview_workbook(overview_path)
    token = resolve_access_token(access_token)
    headers = json_headers(token)
    client = http_client or UrllibSheetsHttpClient()

    target_sheets = load_sheet_properties(
        spreadsheet_id,
        http_client=client,
        headers=headers,
        timeout=timeout,
    )
    if sync_template_tab_to_source_sheets:
        target_sheets = sync_template_tab_to_source_sheets_for_workbook(
            spreadsheet_id,
            target_sheets=target_sheets,
            source_sheets=source_sheets,
            http_client=client,
            headers=headers,
            timeout=timeout,
            dry_run=dry_run,
        )

    updated_tabs: list[str] = []
    skipped_tabs: list[str] = []
    skipped_columns: dict[str, list[str]] = {}
    planned_ranges: list[str] = []
    all_updates: list[dict[str, Any]] = []
    all_clear_ranges: list[str] = []
    updated_cell_count = 0

    for source in source_sheets:
        target_properties = target_sheets.get(normalize_key(source.title))
        if target_properties is None:
            skipped_tabs.append(source.title)
            continue
        target_headers = load_target_headers(
            spreadsheet_id,
            target_properties.title,
            http_client=client,
            headers=headers,
            timeout=timeout,
        )
        if write_full_sheet_when_no_headers and not any(normalize_key(header) for header in target_headers):
            update, clear_range, cell_count, row_count, column_count = build_full_sheet_update(
                source,
                target_properties.title,
            )
            if update is None:
                continue
            updated_tabs.append(target_properties.title)
            planned_ranges.append(update["range"])
            updated_cell_count += cell_count
            all_updates.append(update)
            if clear_range:
                all_clear_ranges.append(clear_range)
            if not dry_run:
                ensure_grid_capacity(
                    spreadsheet_id,
                    target_properties,
                    required_rows=row_count,
                    required_columns=column_count,
                    http_client=client,
                    headers=headers,
                    timeout=timeout,
                )
            continue
        updates, clear_ranges, missing_columns, cell_count = build_value_updates(
            source,
            target_properties.title,
            target_headers,
        )
        if missing_columns:
            skipped_columns[target_properties.title] = missing_columns
        if not updates:
            continue
        updated_tabs.append(target_properties.title)
        planned_ranges.extend(update["range"] for update in updates)
        updated_cell_count += cell_count
        all_updates.extend(updates)
        all_clear_ranges.extend(clear_ranges)

        if not dry_run:
            ensure_row_capacity(
                spreadsheet_id,
                target_properties,
                len(source.rows) + 1,
                http_client=client,
                headers=headers,
                timeout=timeout,
            )

    if not dry_run and all_updates:
        if clear_existing and all_clear_ranges:
            clear_body = {"ranges": all_clear_ranges}
            client.post(
                sheets_values_batch_clear_url(spreadsheet_id),
                json.dumps(clear_body).encode("utf-8"),
                headers,
                timeout,
            )
        update_body = {"valueInputOption": "RAW", "data": all_updates}
        client.post(
            sheets_values_batch_update_url(spreadsheet_id),
            json.dumps(update_body).encode("utf-8"),
            headers,
            timeout,
        )

    return FillInOverviewResult(
        spreadsheet_id=spreadsheet_id,
        overview_file=overview_path,
        dry_run=dry_run,
        updated_tabs=updated_tabs,
        skipped_tabs=skipped_tabs,
        skipped_columns=skipped_columns,
        planned_ranges=planned_ranges,
        updated_cell_count=updated_cell_count,
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--target",
        required=True,
        help="Target Google Sheet URL, raw spreadsheet id, or text file containing either.",
    )
    parser.add_argument(
        "--overview-file",
        required=True,
        type=Path,
        help="Local study overview workbook to write into the target Google Sheet.",
    )
    parser.add_argument(
        "--access-token",
        default=None,
        help=(
            "Google OAuth access token with Sheets write access. If omitted, the "
            "script tries GOOGLE_OAUTH_ACCESS_TOKEN, then gcloud."
        ),
    )
    parser.add_argument("--timeout", type=float, default=120.0, help="API timeout in seconds.")
    parser.add_argument("--dry-run", action="store_true", help="Plan writes without modifying Google Sheets.")
    parser.add_argument(
        "--no-clear-existing",
        action="store_true",
        help="Do not clear matched target columns before writing new overview values.",
    )
    parser.add_argument(
        "--write-full-sheet-when-no-headers",
        action="store_true",
        help="If a target tab has no headers, write the complete local sheet from A1.",
    )
    parser.add_argument(
        "--sync-template-tab-to-source-sheets",
        action="store_true",
        help=(
            "If the target has one BLANK-template tab named template, rename/duplicate "
            "it to match the local workbook sheet names before writing."
        ),
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    result = fill_in_overview(
        target=args.target,
        overview_file=args.overview_file,
        access_token=args.access_token,
        timeout=args.timeout,
        dry_run=args.dry_run,
        clear_existing=not args.no_clear_existing,
        write_full_sheet_when_no_headers=args.write_full_sheet_when_no_headers,
        sync_template_tab_to_source_sheets=args.sync_template_tab_to_source_sheets,
    )
    print(
        json.dumps(
            {
                "dry_run": result.dry_run,
                "spreadsheet_id": result.spreadsheet_id,
                "overview_file": str(result.overview_file),
                "updated_tabs": result.updated_tabs,
                "skipped_tabs": result.skipped_tabs,
                "skipped_columns": result.skipped_columns,
                "planned_ranges": result.planned_ranges,
                "updated_cell_count": result.updated_cell_count,
                "url": result.web_url,
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
