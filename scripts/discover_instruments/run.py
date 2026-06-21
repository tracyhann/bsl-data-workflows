#!/usr/bin/env python3
"""Discover REDCap instrument blocks from CSV/Excel column headers."""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path
from typing import Iterable


if __package__ in {None, ""}:
    sys.path.append(str(Path(__file__).resolve().parents[2]))

from scripts.discover_instruments.instrument_discovery import discover_instruments, is_complete_column


def _dedupe_headers(headers: Iterable[str]) -> list[str]:
    seen: dict[str, int] = {}
    result = []
    for header in headers:
        name = str(header)
        if name not in seen:
            seen[name] = 0
            result.append(name)
            continue
        seen[name] += 1
        result.append(f"{name}.{seen[name]}")
    return result


def read_delimited_headers(path: Path, delimiter: str) -> list[str]:
    with path.open(newline="", encoding="utf-8-sig") as file:
        reader = csv.reader(file, delimiter=delimiter)
        return _dedupe_headers(next(reader))


def read_excel_headers(path: Path, sheet_name: str | int | None = None) -> list[str]:
    try:
        import pandas as pd
    except ImportError as exc:
        raise RuntimeError("Reading Excel files requires pandas/openpyxl in this environment.") from exc

    selected_sheet = 0 if sheet_name is None else sheet_name
    dataframe = pd.read_excel(path, sheet_name=selected_sheet, nrows=0)
    return _dedupe_headers([str(column) for column in dataframe.columns])


def read_excel_table(path: Path, sheet_name: str | int) -> tuple[list[str], list[list[str]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} was not found in {path}.")
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Sheet {sheet_name!r} in {path} is empty.")
    headers = _dedupe_headers([str(value or "").strip() for value in rows[0]])
    data_rows = [
        ["" if value is None else str(value).strip() for value in row]
        for row in rows[1:]
    ]
    return headers, data_rows


def read_headers(path: Path, sheet_name: str | int | None = None) -> list[str]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_delimited_headers(path, ",")
    if suffix == ".tsv":
        return read_delimited_headers(path, "\t")
    if suffix in {".xlsx", ".xls"}:
        return read_excel_headers(path, sheet_name)
    raise ValueError(f"Unsupported input extension {path.suffix!r}; expected CSV, TSV, XLSX, or XLS.")


def output_paths(input_path: Path, out_dir: Path) -> dict[str, Path]:
    stem = input_path.stem
    return {
        "instrument_summary": out_dir / f"{stem}_instrument_summary.csv",
        "instrument_columns": out_dir / f"{stem}_instrument_columns.csv",
    }


def _normalized_header(header: str) -> str:
    return str(header).strip().lower().replace("_", " ")


def _find_header(headers: list[str], names: set[str]) -> int:
    normalized_names = {_normalized_header(name) for name in names}
    for index, header in enumerate(headers):
        if _normalized_header(header) in normalized_names:
            return index
    raise ValueError(f"Could not find any of {sorted(names)} in headers {headers!r}.")


def _cell(row: list[str], index: int) -> str:
    return row[index].strip() if index < len(row) else ""


def expected_instruments_from_codebook(
    codebook_path: str | Path,
    codebook_sheet: str | int,
    discovered_names: Iterable[str],
) -> list[str]:
    """Read expected REDCap Form Name values from an instrument codebook sheet."""
    headers, rows = read_excel_table(Path(codebook_path), codebook_sheet)
    form_name_index = _find_header(headers, {"Form Name", "form_name"})
    instrument_index = _find_header(headers, {"Instrument", "instrument"})
    discovered_set = {str(name).strip() for name in discovered_names}
    expected: list[str] = []
    seen: set[str] = set()

    for row in rows:
        form_name = _cell(row, form_name_index)
        instrument_name = _cell(row, instrument_index)
        candidate = form_name or (instrument_name if instrument_name in discovered_set else "")
        if candidate and candidate not in seen:
            expected.append(candidate)
            seen.add(candidate)

    return expected


def codebook_verification_rows(discovered_names: list[str], expected_names: list[str]) -> list[dict[str, str]]:
    discovered_order = {name: str(index + 1) for index, name in enumerate(discovered_names)}
    expected_order = {name: str(index + 1) for index, name in enumerate(expected_names)}
    rows: list[dict[str, str]] = []

    for name in expected_names:
        if name not in discovered_order:
            rows.append(
                {
                    "status": "missing_from_discovery",
                    "expected_order": expected_order[name],
                    "discovered_order": "",
                    "expected_form_name": name,
                    "discovered_instrument_key": "",
                    "message": "Codebook form name was not found in discovered instrument keys.",
                }
            )

    for name in discovered_names:
        if name not in expected_order:
            rows.append(
                {
                    "status": "extra_in_discovery",
                    "expected_order": "",
                    "discovered_order": discovered_order[name],
                    "expected_form_name": "",
                    "discovered_instrument_key": name,
                    "message": "Discovered instrument key was not found in codebook form names.",
                }
            )

    if not rows:
        for index, (expected, discovered) in enumerate(zip(expected_names, discovered_names), start=1):
            if expected != discovered:
                rows.append(
                    {
                        "status": "order_mismatch",
                        "expected_order": str(index),
                        "discovered_order": str(index),
                        "expected_form_name": expected,
                        "discovered_instrument_key": discovered,
                        "message": "Expected and discovered instruments match as a set but differ in order.",
                    }
                )

    return rows


def write_codebook_verification(path: Path, rows: list[dict[str, str]]) -> None:
    fieldnames = [
        "status",
        "expected_order",
        "discovered_order",
        "expected_form_name",
        "discovered_instrument_key",
        "message",
    ]
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def process_file(
    input_path: str | Path,
    out_dir: str | Path,
    sheet_name: str | int | None = None,
    codebook_path: str | Path | None = None,
    codebook_sheet: str | int | None = None,
) -> dict[str, Path]:
    input_path = Path(input_path)
    out_dir = Path(out_dir)
    headers = read_headers(input_path, sheet_name)
    blocks = discover_instruments(headers)
    paths = output_paths(input_path, out_dir)
    if codebook_path:
        if codebook_sheet is None:
            raise ValueError("--codebook-sheet is required when --codebook is provided.")
        paths["instrument_codebook_verification"] = (
            out_dir / f"{input_path.stem}_instrument_codebook_verification.csv"
        )
    out_dir.mkdir(parents=True, exist_ok=True)

    with paths["instrument_summary"].open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=[
                "instrument_order",
                "instrument_key",
                "start_index",
                "end_index",
                "column_count",
                "start_column",
                "stop_column",
                "dominant_prefix",
                "prefix_evidence",
                "suffix_evidence",
                "has_timestamp_start",
            ],
        )
        writer.writeheader()
        for block in blocks:
            writer.writerow(
                {
                    "instrument_order": block.instrument_order,
                    "instrument_key": block.instrument_key,
                    "start_index": block.start_index,
                    "end_index": block.end_index,
                    "column_count": block.column_count,
                    "start_column": block.start_column,
                    "stop_column": block.stop_column,
                    "dominant_prefix": block.dominant_prefix,
                    "prefix_evidence": block.prefix_evidence,
                    "suffix_evidence": block.suffix_evidence,
                    "has_timestamp_start": str(block.has_timestamp_start),
                }
            )

    with paths["instrument_columns"].open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=[
                "instrument_order",
                "instrument_key",
                "column_index",
                "column_name",
                "is_stop_signal",
            ],
        )
        writer.writeheader()
        for block in blocks:
            for column_index in range(block.start_index, block.end_index + 1):
                column_name = headers[column_index]
                writer.writerow(
                    {
                        "instrument_order": block.instrument_order,
                        "instrument_key": block.instrument_key,
                        "column_index": column_index,
                        "column_name": column_name,
                        "is_stop_signal": str(is_complete_column(column_name)),
                    }
                )

    if codebook_path:
        discovered_names = [block.instrument_key for block in blocks]
        expected_names = expected_instruments_from_codebook(codebook_path, codebook_sheet, discovered_names)
        verification_rows = codebook_verification_rows(discovered_names, expected_names)
        write_codebook_verification(paths["instrument_codebook_verification"], verification_rows)

    return paths


def default_out_dir(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_instrument_discovery")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Discover REDCap instrument blocks from column headers.")
    parser.add_argument("input", type=Path, help="Input CSV, TSV, XLSX, or XLS file")
    parser.add_argument("--out-dir", type=Path, help="Output directory")
    parser.add_argument("--sheet", help="Excel sheet name; defaults to the first sheet")
    parser.add_argument("--codebook", type=Path, help="Optional REDCap codebook workbook for verification")
    parser.add_argument("--codebook-sheet", help="Codebook instrument sheet name for verification")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    out_dir = args.out_dir or default_out_dir(args.input)
    paths = process_file(
        args.input,
        out_dir,
        sheet_name=args.sheet,
        codebook_path=args.codebook,
        codebook_sheet=args.codebook_sheet,
    )
    for path in paths.values():
        print(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
