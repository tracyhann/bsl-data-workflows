#!/usr/bin/env python3
"""Build a reusable static clinical trial hub from an Excel index."""

from __future__ import annotations

import argparse
import base64
import html
import json
import re
import shutil
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable

import markdown
from openpyxl import load_workbook


REQUIRED_COLUMNS = [
    "data",
    "location (link to source)",
    "PHI",
    "access level",
    "stage",
    "permission",
]

REQUIRED_FACETS = ["access level", "PHI", "stage", "permission"]
OPTIONAL_FACETS = ["domain", "platform", "owner", "tags", "last reviewed"]
RENDERABLE_MARKDOWN_SUFFIXES = {".md", ".markdown"}
RENDERABLE_WORKBOOK_SUFFIXES = {".xlsx", ".xlsm"}
URL_RE = re.compile(r"https?://[^\s)>\"]+")
VENDORED_XLSX = Path("vendor/xlsx.full.min.js")


def load_index(index_path: Path) -> tuple[list[dict[str, str]], list[str]]:
    """Load the first worksheet of an Excel index into normalized row dicts."""
    workbook = load_workbook(index_path, data_only=True)
    worksheet = workbook["Index"] if "Index" in workbook.sheetnames else workbook.active
    header_cells = next(worksheet.iter_rows(min_row=1, max_row=1))
    headers = [str(cell.value or "").strip() for cell in header_cells]
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise ValueError(f"Missing required index columns: {', '.join(missing)}")

    rows: list[dict[str, str]] = []
    for excel_row in worksheet.iter_rows(min_row=2):
        row: dict[str, str] = {}
        has_content = False
        for header, cell in zip(headers, excel_row):
            value = "" if cell.value is None else str(cell.value).strip()
            row[header] = value
            if value:
                has_content = True
            if header == "location (link to source)" and cell.hyperlink:
                row["_location_href"] = str(cell.hyperlink.target)
        if has_content:
            rows.append(row)
    return rows, headers


def discover_facets(rows: list[dict[str, str]], headers: Iterable[str]) -> dict[str, list[str]]:
    """Return navigation facets discovered from required and optional columns."""
    available = set(headers)
    facet_names = [facet for facet in REQUIRED_FACETS if facet in available]
    facet_names.extend(facet for facet in OPTIONAL_FACETS if facet in available)
    facets: dict[str, list[str]] = {}
    for facet in facet_names:
        values = {
            value
            for row in rows
            for value in split_facet_values(row.get(facet, ""))
            if value
        }
        if values:
            facets[facet] = sorted(values, key=str.lower)
    return facets


def split_facet_values(value: str) -> list[str]:
    """Split semicolon/comma tag-style values while preserving single labels."""
    if not value:
        return []
    if ";" in value:
        return [part.strip() for part in value.split(";") if part.strip()]
    if "," in value and len(value) < 120:
        return [part.strip() for part in value.split(",") if part.strip()]
    return [value.strip()]


def build_site(index_path: Path, out_dir: Path) -> None:
    """Generate the static hub site."""
    rows, headers = load_index(index_path)
    facets = discover_facets(rows, headers)
    decorated = decorate_rows(rows, index_path)

    reset_output_dir(out_dir)
    for relative in ["assets", "data", "reference", "navigation", "facets", "workbooks"]:
        (out_dir / relative).mkdir(parents=True, exist_ok=True)

    public_pages = render_public_reference_pages(decorated, out_dir)
    write_manifest(decorated, facets, public_pages, out_dir)
    write_assets(out_dir)
    write_index_page(decorated, facets, public_pages, out_dir)
    write_access_navigation_pages(decorated, facets, public_pages, out_dir)
    write_facet_pages(decorated, facets, public_pages, out_dir)


def decorate_rows(rows: list[dict[str, str]], index_path: Path) -> list[dict[str, str]]:
    decorated = []
    for position, row in enumerate(rows, start=1):
        item = {key: value for key, value in row.items() if not key.startswith("_")}
        item["id"] = f"r{position:04d}"
        item["href"] = source_href(row)
        item["local_path"] = str(resolve_local_path(row, index_path) or "")
        item["slug"] = unique_slug(item.get("data", f"resource-{position}"), position)
        item["access_key"] = access_key(item.get("access level", ""))
        item["is_public_reference"] = item["access_key"] == "level-1"
        item["reference_page"] = ""
        item["workbook_asset"] = ""
        item["workbook_base64"] = ""
        decorated.append(item)
    return decorated


def source_href(row: dict[str, str]) -> str:
    hyperlink = row.get("_location_href", "").strip()
    if hyperlink:
        return hyperlink
    location = row.get("location (link to source)", "").strip()
    match = URL_RE.search(location)
    if match:
        return match.group(0)
    return location


def resolve_local_path(row: dict[str, str], index_path: Path) -> Path | None:
    location = row.get("location (link to source)", "").strip()
    if not location or URL_RE.search(location):
        return None
    candidate = Path(location)
    candidates = [
        candidate,
        Path.cwd() / candidate,
        index_path.parent / candidate,
    ]
    for possible in candidates:
        if possible.exists():
            return possible.resolve()
    return None


def access_key(access_level: str) -> str:
    if access_level.startswith("1 -"):
        return "level-1"
    if access_level.startswith("2 -"):
        return "level-2"
    if access_level.startswith("3 -"):
        return "level-3"
    return "unknown"


def reset_output_dir(out_dir: Path) -> None:
    if out_dir.exists():
        shutil.rmtree(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)


def render_public_reference_pages(
    rows: list[dict[str, str]], out_dir: Path
) -> dict[str, str]:
    public_pages: dict[str, str] = {}
    for row in rows:
        local_path = Path(row["local_path"]) if row.get("local_path") else None
        if not row["is_public_reference"] or not local_path:
            continue
        suffix = local_path.suffix.lower()
        if suffix in RENDERABLE_MARKDOWN_SUFFIXES:
            source_text = local_path.read_text(encoding="utf-8")
            body = markdown.markdown(
                source_text,
                extensions=["tables", "fenced_code", "toc"],
                output_format="html5",
            )
        elif suffix in RENDERABLE_WORKBOOK_SUFFIXES:
            row["workbook_asset"] = copy_workbook_asset(local_path, row, out_dir)
            row["workbook_base64"] = base64.b64encode(local_path.read_bytes()).decode("ascii")
            body = workbook_viewer_html(row["workbook_asset"], row["workbook_base64"], depth=1)
        else:
            continue
        filename = f"{row['slug']}.html"
        row["reference_page"] = f"reference/{filename}"
        public_pages[row["id"]] = row["reference_page"]
        page_html = page_shell(
            title=row["data"],
            body=f"""
            <nav class="crumbs"><a href="../index.html">Hub</a><span>{escape(row['data'])}</span></nav>
            <article class="markdown-body">{body}</article>
            <section class="resource-meta">
              {metadata_list(row)}
            </section>
            """,
            depth=1,
        )
        (out_dir / "reference" / filename).write_text(page_html, encoding="utf-8")
    return public_pages


def copy_workbook_asset(local_path: Path, row: dict[str, str], out_dir: Path) -> str:
    """Copy a level-1 workbook into the static site for runtime browser loading."""
    filename = f"{row['slug']}{local_path.suffix.lower()}"
    destination = out_dir / "workbooks" / filename
    shutil.copy2(local_path, destination)
    return f"workbooks/{filename}"


def workbook_viewer_html(workbook_asset: str, workbook_base64: str, depth: int = 0) -> str:
    src = "../" * depth + workbook_asset
    return f"""
    <section class="workbook-viewer" data-workbook-src="{escape_attr(src)}" data-workbook-base64="{escape_attr(workbook_base64)}">
      <div class="viewer-status">Loading workbook...</div>
      <div class="sheet-tabs" aria-label="Workbook sheets"></div>
      <div class="sheet-view"></div>
      <noscript>Enable JavaScript to inspect this workbook in the browser.</noscript>
    </section>
    """


def write_manifest(
    rows: list[dict[str, str]],
    facets: dict[str, list[str]],
    public_pages: dict[str, str],
    out_dir: Path,
) -> None:
    manifest_rows = []
    for row in rows:
        public_row = {
            key: value
            for key, value in row.items()
            if key not in {"local_path", "workbook_base64"} and not key.startswith("_")
        }
        public_row["reference_page"] = public_pages.get(row["id"], "")
        manifest_rows.append(public_row)
    manifest = {
        "title": "Clinical Trial Directory Hub",
        "facets": facets,
        "counts": {
            "resources": len(rows),
            "access level": dict(Counter(row.get("access level", "") for row in rows)),
            "stage": dict(Counter(row.get("stage", "") for row in rows)),
            "PHI": dict(Counter(row.get("PHI", "") for row in rows)),
        },
        "rows": manifest_rows,
    }
    (out_dir / "data" / "manifest.json").write_text(
        json.dumps(manifest, indent=2), encoding="utf-8"
    )


def write_assets(out_dir: Path) -> None:
    (out_dir / "assets" / "styles.css").write_text(STYLES, encoding="utf-8")
    (out_dir / "assets" / "app.js").write_text(APP_JS, encoding="utf-8")
    if VENDORED_XLSX.exists():
        shutil.copy2(VENDORED_XLSX, out_dir / "assets" / "xlsx.full.min.js")


def write_index_page(
    rows: list[dict[str, str]],
    facets: dict[str, list[str]],
    public_pages: dict[str, str],
    out_dir: Path,
) -> None:
    stats = {
        "Resources": len(rows),
        "Public Reference": sum(1 for row in rows if row["access_key"] == "level-1"),
        "Non-PHI": sum(1 for row in rows if row["access_key"] == "level-2"),
        "PHI": sum(1 for row in rows if row["access_key"] == "level-3"),
    }
    body = f"""
    <section class="topbar">
      <div>
        <p class="eyebrow">Index-driven clinical trial hub</p>
        <h1>Clinical Trial Directory Hub</h1>
      </div>
    </section>
    {stat_grid(stats)}
    <section class="panel">
      <div class="section-heading">
        <h2>Navigation</h2>
      </div>
      {navigation_links(facets)}
    </section>
    <section class="panel">
      <div class="section-heading">
        <h2>Resources</h2>
      </div>
      <div id="filters" class="filters"></div>
      <div class="search-row">
        <input id="resourceSearch" type="search" placeholder="Search resources" />
      </div>
      {resource_table(rows, public_pages)}
    </section>
    """
    (out_dir / "index.html").write_text(
        page_shell("Clinical Trial Directory Hub", body), encoding="utf-8"
    )


def write_access_navigation_pages(
    rows: list[dict[str, str]],
    facets: dict[str, list[str]],
    public_pages: dict[str, str],
    out_dir: Path,
) -> None:
    access_pages = [
        ("level-1", "1 - MOP derived facts", "Public References"),
        ("level-2", "2 - non-PHI study data", "Non-PHI Study Data"),
        ("level-3", "3 - PHI study data", "PHI Study Data"),
    ]
    for key, access_level, title in access_pages:
        subset = [row for row in rows if row["access_key"] == key]
        notice = ""
        if key in {"level-2", "level-3"}:
            notice = """
            <section class="notice restricted">
              <strong>Navigation only.</strong>
              Source-system authentication controls access to the linked material.
            </section>
            """
        body = f"""
        <nav class="crumbs"><a href="../index.html">Hub</a><span>{escape(title)}</span></nav>
        <section class="topbar compact"><h1>{escape(title)}</h1></section>
        {notice}
        {grouped_resources(subset, public_pages, depth=1)}
        """
        (out_dir / "navigation" / f"{key}.html").write_text(
            page_shell(title, body, depth=1), encoding="utf-8"
        )


def write_facet_pages(
    rows: list[dict[str, str]],
    facets: dict[str, list[str]],
    public_pages: dict[str, str],
    out_dir: Path,
) -> None:
    for facet, values in facets.items():
        for value in values:
            subset = [row for row in rows if value in split_facet_values(row.get(facet, ""))]
            filename = f"{slugify(facet)}-{slugify(value)}.html"
            body = f"""
            <nav class="crumbs"><a href="../index.html">Hub</a><span>{escape(facet)}: {escape(value)}</span></nav>
            <section class="topbar compact">
              <div>
                <p class="eyebrow">{escape(facet)}</p>
                <h1>{escape(value)}</h1>
              </div>
            </section>
            {grouped_resources(subset, public_pages, depth=1)}
            """
            (out_dir / "facets" / filename).write_text(
                page_shell(f"{facet}: {value}", body, depth=1), encoding="utf-8"
            )


def navigation_links(facets: dict[str, list[str]]) -> str:
    groups = [
        """
        <div class="nav-group">
          <h3>Access</h3>
          <a href="navigation/level-1.html">1 - MOP derived facts</a>
          <a href="navigation/level-2.html">2 - non-PHI study data</a>
          <a href="navigation/level-3.html">3 - PHI study data</a>
        </div>
        """
    ]
    for facet, values in facets.items():
        links = "\n".join(
            f'<a href="facets/{slugify(facet)}-{slugify(value)}.html">{escape(value)}</a>'
            for value in values
        )
        groups.append(
            f"""
            <div class="nav-group">
              <h3>{escape(titleize(facet))}</h3>
              {links}
            </div>
            """
        )
    return f'<div class="nav-grid">{"".join(groups)}</div>'


def grouped_resources(
    rows: list[dict[str, str]], public_pages: dict[str, str], depth: int = 0
) -> str:
    if not rows:
        return '<section class="empty">No resources in this section.</section>'
    groups: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        key_parts = [
            row.get("domain", ""),
            row.get("stage", ""),
            row.get("platform", ""),
        ]
        key = " / ".join(part for part in key_parts if part) or "Resources"
        groups[key].append(row)
    sections = []
    for group_name in sorted(groups, key=str.lower):
        sections.append(
            f"""
            <section class="panel">
              <div class="section-heading"><h2>{escape(titleize(group_name))}</h2></div>
              {resource_table(groups[group_name], public_pages, depth=depth)}
            </section>
            """
        )
    return "\n".join(sections)


def resource_table(
    rows: list[dict[str, str]], public_pages: dict[str, str], depth: int = 0
) -> str:
    body = []
    for row in rows:
        actions_html = resource_actions(row, public_pages, depth=depth)
        classes = f"access-{escape_attr(row['access_key'])}"
        searchable = escape_attr(" ".join(str(value) for value in row.values()))
        body.append(
            f"""
            <tr class="{classes}" data-search="{searchable}">
              <td><strong>{escape(row.get('data', ''))}</strong></td>
              <td>{badge(row.get('access level', ''))}</td>
              <td>{escape(row.get('stage', ''))}</td>
              <td>{escape(row.get('PHI', ''))}</td>
              <td>{escape(row.get('permission', ''))}</td>
              <td>{escape(row.get('domain', ''))}</td>
              <td>{escape(row.get('platform', ''))}</td>
              <td>{actions_html}</td>
            </tr>
            """
        )
    return f"""
    <div class="table-wrap">
      <table class="resource-table">
        <thead>
          <tr>
            <th>Data</th>
            <th>Access</th>
            <th>Stage</th>
            <th>PHI</th>
            <th>Permission</th>
            <th>Domain</th>
            <th>Platform</th>
            <th>Source</th>
          </tr>
        </thead>
        <tbody>{''.join(body)}</tbody>
      </table>
    </div>
    """


def resource_actions(
    row: dict[str, str], public_pages: dict[str, str], depth: int = 0
) -> str:
    href = public_pages.get(row["id"]) or row.get("href", "")
    if href and not URL_RE.search(href):
        href = "../" * depth + href
    actions = []
    if href:
        actions.append(
            f'<a class="button small" href="{escape_attr(href)}">Open</a>'
        )
    inline = inline_workbook_inspector(row, depth=depth)
    if inline:
        actions.append(inline)
    return f'<div class="source-actions">{"".join(actions)}</div>'


def inline_workbook_inspector(row: dict[str, str], depth: int = 0) -> str:
    if row.get("access_key") != "level-1":
        return ""
    local_path = Path(row["local_path"]) if row.get("local_path") else None
    if not local_path or local_path.suffix.lower() not in RENDERABLE_WORKBOOK_SUFFIXES:
        return ""
    workbook_asset = row.get("workbook_asset")
    if not workbook_asset:
        return ""
    return f"""
    <details class="inline-workbook">
      <summary>Inspect workbook</summary>
      {workbook_viewer_html(workbook_asset, row.get("workbook_base64", ""), depth=depth)}
    </details>
    """


def metadata_list(row: dict[str, str]) -> str:
    fields = ["access level", "PHI", "stage", "permission", "domain", "platform"]
    items = "\n".join(
        f"<dt>{escape(titleize(field))}</dt><dd>{escape(row.get(field, ''))}</dd>"
        for field in fields
        if row.get(field, "")
    )
    source = row.get("href", "")
    if source:
        items += (
            f'<dt>Source</dt><dd><a href="{escape_attr(source)}">{escape(source)}</a></dd>'
        )
    return f"<dl>{items}</dl>"


def stat_grid(stats: dict[str, int]) -> str:
    cards = "\n".join(
        f"""
        <div class="stat">
          <span>{escape(label)}</span>
          <strong>{count}</strong>
        </div>
        """
        for label, count in stats.items()
    )
    return f'<section class="stats">{cards}</section>'


def badge(value: str) -> str:
    label = escape(value)
    if value.startswith("3 -"):
        return f'<span class="badge danger">{label}</span>'
    if value.startswith("2 -"):
        return f'<span class="badge caution">{label}</span>'
    return f'<span class="badge safe">{label}</span>'


def page_shell(title: str, body: str, depth: int = 0) -> str:
    prefix = "../" * depth
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{escape(title)}</title>
  <link rel="stylesheet" href="{prefix}assets/styles.css" />
</head>
<body>
  <main class="page">
    {body}
  </main>
  <script src="{prefix}assets/xlsx.full.min.js"></script>
  <script src="{prefix}assets/app.js"></script>
</body>
</html>
"""


def unique_slug(value: str, position: int) -> str:
    return f"{slugify(value) or 'resource'}-{position:04d}"


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug[:80]


def titleize(value: str) -> str:
    return value.replace("_", " ").replace("-", " ").title()


def escape(value: object) -> str:
    return html.escape(str(value or ""))


def escape_attr(value: object) -> str:
    return html.escape(str(value or ""), quote=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--index", type=Path, required=True, help="Excel index path")
    parser.add_argument("--out", type=Path, required=True, help="Output site directory")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    build_site(args.index, args.out)
    print(f"Wrote static trial hub to {args.out}")


STYLES = """
:root {
  color-scheme: light;
  --ink: #17202a;
  --muted: #61707f;
  --line: #d8dee6;
  --paper: #f6f8fb;
  --surface: #ffffff;
  --blue: #2457a6;
  --teal: #0f766e;
  --amber: #a85f00;
  --red: #b42318;
}
* { box-sizing: border-box; }
body {
  margin: 0;
  font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
  color: var(--ink);
  background: var(--paper);
}
a { color: var(--blue); }
.page {
  width: min(1180px, calc(100vw - 32px));
  margin: 0 auto;
  padding: 28px 0 48px;
}
.topbar {
  display: flex;
  align-items: flex-start;
  justify-content: space-between;
  gap: 20px;
  margin-bottom: 22px;
}
.topbar.compact { margin-bottom: 16px; }
.eyebrow {
  margin: 0 0 6px;
  color: var(--teal);
  font-size: 0.78rem;
  font-weight: 800;
  text-transform: uppercase;
}
h1, h2, h3 { margin: 0; line-height: 1.2; letter-spacing: 0; }
h1 { font-size: 2rem; }
h2 { font-size: 1.05rem; }
h3 { font-size: 0.92rem; }
.stats {
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(170px, 1fr));
  gap: 12px;
  margin-bottom: 18px;
}
.stat, .panel, .notice, .empty, .markdown-body, .resource-meta {
  background: var(--surface);
  border: 1px solid var(--line);
  border-radius: 8px;
}
.stat { padding: 16px; }
.stat span { display: block; color: var(--muted); font-size: 0.85rem; }
.stat strong { display: block; margin-top: 6px; font-size: 1.7rem; }
.panel { padding: 16px; margin-top: 14px; }
.section-heading {
  display: flex;
  align-items: center;
  justify-content: space-between;
  margin-bottom: 12px;
}
.nav-grid {
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
  gap: 12px;
}
.nav-group {
  border-left: 3px solid var(--teal);
  padding: 2px 0 2px 12px;
}
.nav-group a {
  display: block;
  margin-top: 8px;
  overflow-wrap: anywhere;
}
.filters {
  display: flex;
  flex-wrap: wrap;
  gap: 10px;
  margin-bottom: 10px;
}
.filters select, .search-row input {
  border: 1px solid var(--line);
  border-radius: 6px;
  padding: 9px 10px;
  font: inherit;
  background: #fff;
}
.search-row { margin-bottom: 12px; }
.search-row input { width: min(420px, 100%); }
.table-wrap { overflow-x: auto; }
.resource-table {
  width: 100%;
  border-collapse: collapse;
  min-width: 860px;
}
.resource-table th, .resource-table td {
  border-bottom: 1px solid var(--line);
  padding: 10px 8px;
  text-align: left;
  vertical-align: top;
  font-size: 0.9rem;
}
.resource-table th {
  color: var(--muted);
  font-size: 0.78rem;
  text-transform: uppercase;
}
.badge {
  display: inline-block;
  border-radius: 999px;
  padding: 4px 8px;
  font-size: 0.78rem;
  font-weight: 700;
  white-space: nowrap;
}
.badge.safe { color: #075b4c; background: #dff8ef; }
.badge.caution { color: #7a4700; background: #fff1d8; }
.badge.danger { color: #8f1b13; background: #ffe3df; }
.button {
  display: inline-flex;
  align-items: center;
  justify-content: center;
  min-height: 36px;
  border-radius: 6px;
  border: 1px solid var(--blue);
  padding: 8px 12px;
  color: #fff;
  background: var(--blue);
  text-decoration: none;
  font-weight: 700;
}
.button.small { min-height: 30px; padding: 5px 9px; font-size: 0.82rem; }
.source-actions {
  display: flex;
  flex-direction: column;
  align-items: flex-start;
  gap: 8px;
}
.inline-workbook {
  width: min(720px, 70vw);
}
.inline-workbook summary {
  cursor: pointer;
  color: var(--blue);
  font-weight: 800;
}
.inline-workbook .workbook-inspector {
  margin-top: 10px;
}
.inline-workbook .workbook-viewer {
  margin-top: 10px;
}
.crumbs {
  display: flex;
  flex-wrap: wrap;
  gap: 8px;
  margin-bottom: 16px;
  color: var(--muted);
}
.crumbs span::before { content: "/ "; }
.notice { padding: 12px 14px; margin-bottom: 14px; }
.notice.restricted { border-left: 4px solid var(--amber); }
.empty { padding: 16px; color: var(--muted); }
.markdown-body { padding: 22px; }
.markdown-body h1 { font-size: 1.7rem; margin-bottom: 14px; }
.markdown-body h2 { margin-top: 26px; font-size: 1.25rem; }
.markdown-body table {
  border-collapse: collapse;
  width: 100%;
  margin: 14px 0;
}
.markdown-body th, .markdown-body td {
  border: 1px solid var(--line);
  padding: 8px;
  text-align: left;
  vertical-align: top;
}
.workbook-viewer {
  display: grid;
  gap: 12px;
}
.viewer-status {
  color: var(--muted);
  font-weight: 700;
}
.sheet-tabs {
  display: flex;
  flex-wrap: wrap;
  gap: 8px;
}
.sheet-tab {
  border: 1px solid var(--line);
  border-radius: 999px;
  background: #fff;
  color: var(--ink);
  cursor: pointer;
  padding: 7px 11px;
  font-weight: 800;
}
.sheet-tab.active {
  border-color: var(--teal);
  background: #dff8ef;
  color: #075b4c;
}
.workbook-table-wrap {
  max-height: 70vh;
}
.workbook-table {
  width: 100%;
  border-collapse: collapse;
  min-width: 720px;
}
.workbook-table th, .workbook-table td {
  border-bottom: 1px solid var(--line);
  border-right: 1px solid var(--line);
  padding: 8px;
  text-align: left;
  vertical-align: top;
  font-size: 0.88rem;
  overflow-wrap: anywhere;
}
.workbook-table th {
  position: sticky;
  top: 0;
  z-index: 1;
  background: #f9fbfd;
  color: var(--ink);
}
.resource-meta { padding: 16px; margin-top: 14px; }
.resource-meta dl {
  display: grid;
  grid-template-columns: minmax(120px, 180px) 1fr;
  gap: 8px 14px;
  margin: 0;
}
.resource-meta dt { color: var(--muted); font-weight: 800; }
.resource-meta dd { margin: 0; overflow-wrap: anywhere; }
@media (max-width: 700px) {
  .topbar { display: block; }
  .topbar .button { margin-top: 12px; }
  .page { width: min(100vw - 20px, 1180px); padding-top: 18px; }
  h1 { font-size: 1.55rem; }
  .resource-meta dl { grid-template-columns: 1fr; }
}
"""


APP_JS = """
(function () {
  const table = document.querySelector('.resource-table');
  const filterRoot = document.getElementById('filters');
  const search = document.getElementById('resourceSearch');

  if (table && filterRoot) {
    fetch('data/manifest.json')
      .then((response) => response.json())
      .then((manifest) => {
        Object.entries(manifest.facets || {}).forEach(([facet, values]) => {
          const select = document.createElement('select');
          select.dataset.facet = facet;
          const label = facet.replace(/[-_]/g, ' ');
          select.innerHTML = `<option value="">${label}</option>` + values.map((value) => {
            return `<option value="${escapeHtml(String(value).toLowerCase())}">${escapeHtml(value)}</option>`;
          }).join('');
          select.addEventListener('change', applyFilters);
          filterRoot.appendChild(select);
        });
        if (search) search.addEventListener('input', applyFilters);
      })
      .catch(() => {});
  }

  document.querySelectorAll('.workbook-viewer').forEach((viewer) => {
    loadWorkbookViewer(viewer);
  });

  function applyFilters() {
    const q = search ? search.value.trim().toLowerCase() : '';
    const filters = Array.from(filterRoot.querySelectorAll('select'))
      .map((select) => select.value)
      .filter(Boolean);
    table.querySelectorAll('tbody tr').forEach((row) => {
      const haystack = row.dataset.search.toLowerCase();
      const matchesSearch = !q || haystack.includes(q);
      const matchesFilters = filters.every((value) => haystack.includes(value));
      row.style.display = matchesSearch && matchesFilters ? '' : 'none';
    });
  }

  async function loadWorkbookViewer(viewer) {
    const status = viewer.querySelector('.viewer-status');
    const tabs = viewer.querySelector('.sheet-tabs');
    const sheetView = viewer.querySelector('.sheet-view');
    const src = viewer.dataset.workbookSrc;
    const embedded = viewer.dataset.workbookBase64;
    if (!src) {
      status.textContent = 'No workbook source configured.';
      return;
    }
    if (!window.XLSX) {
      status.innerHTML = 'Excel viewer library did not load. <a href="' + escapeAttr(src) + '">Open workbook file</a>.';
      return;
    }
    try {
      const buffer = embedded ? base64ToArrayBuffer(embedded) : await fetchWorkbook(src);
      const workbook = window.XLSX.read(buffer, { type: 'array' });
      status.textContent = workbook.SheetNames.length + ' sheet' + (workbook.SheetNames.length === 1 ? '' : 's') + ' loaded from workbook.';
      workbook.SheetNames.forEach((sheetName, index) => {
        const button = document.createElement('button');
        button.type = 'button';
        button.className = 'sheet-tab' + (index === 0 ? ' active' : '');
        button.textContent = sheetName;
        button.addEventListener('click', () => {
          tabs.querySelectorAll('.sheet-tab').forEach((tab) => tab.classList.remove('active'));
          button.classList.add('active');
          renderSheet(sheetView, workbook.Sheets[sheetName]);
        });
        tabs.appendChild(button);
      });
      if (workbook.SheetNames[0]) {
        renderSheet(sheetView, workbook.Sheets[workbook.SheetNames[0]]);
      }
    } catch (error) {
      status.innerHTML = 'Could not load workbook in browser. <a href="' + escapeAttr(src) + '">Open the workbook file</a>.';
    }
  }

  async function fetchWorkbook(src) {
    const response = await fetch(src);
    if (!response.ok) throw new Error('HTTP ' + response.status);
    return response.arrayBuffer();
  }

  function base64ToArrayBuffer(base64) {
    const binary = atob(base64);
    const bytes = new Uint8Array(binary.length);
    for (let i = 0; i < binary.length; i += 1) {
      bytes[i] = binary.charCodeAt(i);
    }
    return bytes.buffer;
  }

  function renderSheet(container, sheet) {
    const rows = window.XLSX.utils.sheet_to_json(sheet, { header: 1, defval: '' });
    if (!rows.length) {
      container.innerHTML = '<div class="empty">This sheet has no visible values.</div>';
      return;
    }
    const width = rows.reduce((max, row) => Math.max(max, row.length), 0);
    const padded = rows.map((row) => {
      const copy = row.slice();
      while (copy.length < width) copy.push('');
      return copy;
    });
    const header = padded[0].map((cell) => '<th>' + escapeHtml(String(cell)) + '</th>').join('');
    const body = padded.slice(1).map((row) => {
      return '<tr>' + row.map((cell) => '<td>' + escapeHtml(String(cell)) + '</td>').join('') + '</tr>';
    }).join('');
    container.innerHTML = '<div class="table-wrap workbook-table-wrap"><table class="workbook-table"><thead><tr>' + header + '</tr></thead><tbody>' + body + '</tbody></table></div>';
  }

  function escapeHtml(value) {
    return value.replace(/[&<>"']/g, (char) => ({
      '&': '&amp;',
      '<': '&lt;',
      '>': '&gt;',
      '"': '&quot;',
      "'": '&#39;'
    }[char]));
  }

  function escapeAttr(value) {
    return escapeHtml(value).replace(/`/g, '&#96;');
  }
}());
"""


if __name__ == "__main__":
    main()
