#!/usr/bin/env python3
"""Classify REDCap instruments into broad study-data categories."""

from __future__ import annotations

import argparse
import json
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


CONFIG_DIR = Path(__file__).resolve().parent
DEFAULT_KEYWORDS = CONFIG_DIR / "keywords.json"
DEFAULT_SEMANTICS = CONFIG_DIR / "semantics.json"
DEFAULT_MODEL = "sentence-transformers/all-MiniLM-L6-v2"
CATEGORY_ORDER = [
    "subjects",
    "assessments",
    "treatments",
    "neuroimaging",
    "biologics_biometrics",
    "safety_regulatory",
    "admin",
]


@dataclass(frozen=True)
class InstrumentInput:
    instrument_name: str
    instrument_label: str = ""


@dataclass(frozen=True)
class Classification:
    instrument_name: str
    instrument_label: str
    category: str
    confidence: float
    keyword_category: str
    semantic_category: str
    keyword_score: float
    semantic_score: float
    evidence: list[str]


@dataclass(frozen=True)
class ClassificationResult:
    output_path: Path
    instrument_count: int
    classifications: list[Classification]


def load_json(path: str | Path) -> dict:
    with Path(path).open(encoding="utf-8") as file:
        return json.load(file)


def normalize_text(value: object) -> str:
    text = str(value or "").lower()
    text = re.sub(r"(?<=[a-z])(?=[0-9])|(?<=[0-9])(?=[a-z])", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def tokens(value: object) -> set[str]:
    return set(normalize_text(value).split())


def pattern_matches(pattern: str, normalized_text: str, token_set: set[str]) -> bool:
    normalized_pattern = normalize_text(pattern)
    if not normalized_pattern:
        return False
    pattern_tokens = normalized_pattern.split()
    if len(pattern_tokens) == 1:
        token = pattern_tokens[0]
        return token in token_set or (token.endswith("s") and token[:-1] in token_set)

    text_tokens = normalized_text.split()
    for start in range(0, len(text_tokens) - len(pattern_tokens) + 1):
        window = text_tokens[start : start + len(pattern_tokens)]
        if all(tokens_equivalent(pattern_token, text_token) for pattern_token, text_token in zip(pattern_tokens, window)):
            return True
    return False


def tokens_equivalent(pattern_token: str, text_token: str) -> bool:
    if pattern_token == text_token:
        return True
    if text_token.endswith("s") and pattern_token == text_token[:-1]:
        return True
    if pattern_token.endswith("s") and pattern_token[:-1] == text_token:
        return True
    return False


def keyword_scores(
    instrument_name: str,
    instrument_label: str,
    keyword_rules: dict,
) -> tuple[dict[str, float], dict[str, list[str]]]:
    text = normalize_text(f"{instrument_name} {instrument_label}")
    token_set = tokens(text)
    scores = {category: 0.0 for category in CATEGORY_ORDER}
    evidence = {category: [] for category in CATEGORY_ORDER}

    for category in CATEGORY_ORDER:
        rules = keyword_rules.get(category, {})
        for pattern in rules.get("observed_patterns", []):
            if pattern_matches(pattern, text, token_set):
                scores[category] += 3.0
                evidence[category].append(f"observed:{pattern}")
        for pattern in rules.get("related_patterns", []):
            if pattern_matches(pattern, text, token_set):
                scores[category] += 1.5
                evidence[category].append(f"related:{pattern}")

    return scores, evidence


def semantic_documents(semantics: dict) -> dict[str, str]:
    documents: dict[str, str] = {}
    for category in CATEGORY_ORDER:
        entry = semantics.get(category, {})
        documents[category] = " ".join([entry.get("caption", ""), *entry.get("examples", [])])
    return documents


def lexical_semantic_scores(instrument_name: str, instrument_label: str, semantics: dict) -> dict[str, float]:
    text_tokens = tokens(f"{instrument_name} {instrument_label}")
    scores: dict[str, float] = {}
    for category, document in semantic_documents(semantics).items():
        category_tokens = tokens(document)
        if not text_tokens or not category_tokens:
            scores[category] = 0.0
            continue
        overlap = len(text_tokens & category_tokens)
        scores[category] = overlap / math.sqrt(len(text_tokens) * len(category_tokens))
    return scores


def cosine_similarity(left: list[float], right: list[float]) -> float:
    numerator = sum(a * b for a, b in zip(left, right))
    left_norm = math.sqrt(sum(a * a for a in left))
    right_norm = math.sqrt(sum(b * b for b in right))
    if left_norm == 0 or right_norm == 0:
        return 0.0
    return numerator / (left_norm * right_norm)


def transformer_semantic_scores(
    instrument_name: str,
    instrument_label: str,
    semantics: dict,
    model_name: str = DEFAULT_MODEL,
) -> dict[str, float]:
    from sentence_transformers import SentenceTransformer

    model = SentenceTransformer(model_name)
    text = f"{instrument_name} {instrument_label}"
    documents = semantic_documents(semantics)
    labels = list(documents)
    vectors = model.encode([text, *[documents[label] for label in labels]], normalize_embeddings=False)
    text_vector = [float(value) for value in vectors[0]]
    return {
        label: cosine_similarity(text_vector, [float(value) for value in vectors[index + 1]])
        for index, label in enumerate(labels)
    }


def top_category(scores: dict[str, float]) -> tuple[str, float, float]:
    ordered = sorted(scores.items(), key=lambda item: (-item[1], CATEGORY_ORDER.index(item[0])))
    top, top_score = ordered[0]
    second_score = ordered[1][1] if len(ordered) > 1 else 0.0
    return top, top_score, second_score


def confidence_from_agreement(
    keyword_top_score: float,
    keyword_second_score: float,
    semantic_top_score: float,
    agrees: bool,
    used_semantic_only: bool = False,
) -> float:
    if keyword_top_score <= 0 and semantic_top_score <= 0:
        return 0.0

    keyword_strength = min(keyword_top_score / 6.0, 1.0)
    margin = 0.0
    if keyword_top_score > 0:
        margin = max((keyword_top_score - keyword_second_score) / keyword_top_score, 0.0)

    if used_semantic_only:
        confidence = 0.25 + 0.55 * min(max(semantic_top_score, 0.0), 1.0)
    else:
        confidence = 0.20 + 0.45 * keyword_strength + 0.20 * margin + 0.15 * min(max(semantic_top_score, 0.0), 1.0)
        if semantic_top_score > 0:
            confidence += 0.10 if agrees else -0.15
    return round(max(0.0, min(confidence, 0.99)), 3)


def classify_one(
    instrument_name: str,
    instrument_label: str = "",
    keyword_rules: dict | None = None,
    semantics: dict | None = None,
    use_sentence_transformer: bool = False,
    model_name: str = DEFAULT_MODEL,
) -> Classification:
    keyword_rules = keyword_rules or load_json(DEFAULT_KEYWORDS)
    semantics = semantics or load_json(DEFAULT_SEMANTICS)

    key_scores, key_evidence = keyword_scores(instrument_name, instrument_label, keyword_rules)
    keyword_category, keyword_top_score, keyword_second_score = top_category(key_scores)

    if use_sentence_transformer:
        try:
            sem_scores = transformer_semantic_scores(instrument_name, instrument_label, semantics, model_name=model_name)
        except Exception:
            sem_scores = lexical_semantic_scores(instrument_name, instrument_label, semantics)
    else:
        sem_scores = lexical_semantic_scores(instrument_name, instrument_label, semantics)

    semantic_category, semantic_top_score, _ = top_category(sem_scores)
    if keyword_top_score > 0:
        category = keyword_category
        used_semantic_only = False
    elif semantic_top_score > 0:
        category = semantic_category
        used_semantic_only = True
    else:
        category = "unknown"
        used_semantic_only = False

    agrees = category == semantic_category
    confidence = confidence_from_agreement(
        keyword_top_score,
        keyword_second_score,
        semantic_top_score,
        agrees,
        used_semantic_only=used_semantic_only,
    )
    return Classification(
        instrument_name=instrument_name,
        instrument_label=instrument_label,
        category=category,
        confidence=confidence,
        keyword_category=keyword_category if keyword_top_score > 0 else "",
        semantic_category=semantic_category if semantic_top_score > 0 else "",
        keyword_score=round(keyword_top_score, 3),
        semantic_score=round(semantic_top_score, 3),
        evidence=key_evidence.get(category, []),
    )


def dictionary_path_from_study_folder(study_folder: str | Path) -> Path:
    return Path(study_folder) / "data" / "cleaned" / "redcap" / "dictionary.xlsx"


def read_dictionary_instruments(dictionary_path: str | Path) -> list[InstrumentInput]:
    dictionary_path = Path(dictionary_path)
    workbook = load_workbook(dictionary_path, read_only=True, data_only=True)
    if "instrument" not in workbook.sheetnames:
        raise ValueError(f"Workbook {dictionary_path} has no 'instrument' sheet.")
    worksheet = workbook["instrument"]
    headers = [str(cell.value or "") for cell in worksheet[1]]
    instrument_index = headers.index("instrument")
    label_index = headers.index("instrument_label") if "instrument_label" in headers else None

    instruments: list[InstrumentInput] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        name = str(row[instrument_index] or "").strip()
        label = str(row[label_index] or "").strip() if label_index is not None else ""
        if name:
            instruments.append(InstrumentInput(name, label))
    return instruments


def write_classification_workbook(output_path: str | Path, classifications: Iterable[Classification]) -> None:
    output_path = Path(output_path)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "instrument_classification"
    worksheet.append(["instrument_name", "instrument_label", "class", "confidence"])
    for classification in classifications:
        worksheet.append(
            [
                classification.instrument_name,
                classification.instrument_label,
                classification.category,
                classification.confidence,
            ]
        )
    for column_cells in worksheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 80)
        worksheet.column_dimensions[column_cells[0].column_letter].width = width
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_instrument_arg(value: str) -> InstrumentInput:
    if "::" in value:
        name, label = value.split("::", 1)
        return InstrumentInput(name.strip(), label.strip())
    return InstrumentInput(value.strip(), "")


def classify_instruments(
    study_folder: str | Path | None = None,
    dictionary_path: str | Path | None = None,
    instruments: Iterable[InstrumentInput] | None = None,
    output_path: str | Path | None = None,
    keywords_path: str | Path = DEFAULT_KEYWORDS,
    semantics_path: str | Path = DEFAULT_SEMANTICS,
    use_sentence_transformer: bool = False,
    model_name: str = DEFAULT_MODEL,
) -> ClassificationResult:
    if instruments is None:
        if dictionary_path is None:
            if study_folder is None:
                raise ValueError("Either study_folder, dictionary_path, or instruments must be supplied.")
            dictionary_path = dictionary_path_from_study_folder(study_folder)
        instrument_inputs = read_dictionary_instruments(dictionary_path)
    else:
        instrument_inputs = list(instruments)

    if output_path is None:
        if dictionary_path is not None:
            output_path = Path(dictionary_path).parent / "instrument_classification.xlsx"
        elif study_folder is not None:
            output_path = Path(study_folder) / "data" / "cleaned" / "redcap" / "instrument_classification.xlsx"
        else:
            output_path = Path("instrument_classification.xlsx")

    keyword_rules = load_json(keywords_path)
    semantics = load_json(semantics_path)
    classifications = [
        classify_one(
            instrument.instrument_name,
            instrument.instrument_label,
            keyword_rules=keyword_rules,
            semantics=semantics,
            use_sentence_transformer=use_sentence_transformer,
            model_name=model_name,
        )
        for instrument in instrument_inputs
    ]
    write_classification_workbook(output_path, classifications)
    return ClassificationResult(Path(output_path), len(classifications), classifications)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Classify REDCap instruments into broad study-data categories.")
    parser.add_argument("--study-folder", type=Path, help="Study folder with data/cleaned/redcap/dictionary.xlsx")
    parser.add_argument("--dictionary", type=Path, help="Dictionary workbook path")
    parser.add_argument("--instrument", action="append", default=[], help="Instrument as name or name::label; repeatable")
    parser.add_argument("--out", type=Path, help="Output instrument_classification.xlsx path")
    parser.add_argument("--keywords", type=Path, default=DEFAULT_KEYWORDS, help="Keyword rule JSON")
    parser.add_argument("--semantics", type=Path, default=DEFAULT_SEMANTICS, help="Semantic caption JSON")
    parser.add_argument("--use-sentence-transformer", action="store_true", help="Use sentence-transformers if installed")
    parser.add_argument("--model-name", default=DEFAULT_MODEL, help="SentenceTransformer model name/path")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    instruments = [parse_instrument_arg(value) for value in args.instrument] if args.instrument else None
    result = classify_instruments(
        study_folder=args.study_folder,
        dictionary_path=args.dictionary,
        instruments=instruments,
        output_path=args.out,
        keywords_path=args.keywords,
        semantics_path=args.semantics,
        use_sentence_transformer=args.use_sentence_transformer,
        model_name=args.model_name,
    )
    print(result.output_path)
    print(f"instruments={result.instrument_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
