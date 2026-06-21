import tempfile
import unittest
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.classify_instruments.run import classify_instruments, classify_one


def write_dictionary(path: Path, rows: list[tuple[str, str]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "instrument"
    worksheet.append(["instrument", "instrument_label", "abbreviation", "number_of_events", "events", "number_of_columns"])
    for instrument, label in rows:
        worksheet.append([instrument, label, "", "1", "Arm 1: baseline", "1"])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


class ClassifyInstrumentsTests(unittest.TestCase):
    def test_classifies_observed_instrument_patterns(self):
        examples = {
            "patient_email": ("Patient Email", "subjects"),
            "post_screening_yesno_enrollment_verification_54909": ("Post Screening Yes/no Enrollment Verification", "subjects"),
            "madrsc_mentor_rater": ("MADRS-C (Mentor Rater)", "assessments"),
            "holmes_and_rahe_lsss": ("Holmes And Rahe Lsss (SCREENING ONLY)", "assessments"),
            "ims12_survey": ("IMS-12 (survey)", "assessments"),
            "mssi": ("MSSI", "assessments"),
            "siat_confirmationnotes": ("S-IAT Confirmation/Notes", "assessments"),
            "athf_current": ("ATHF Current", "treatments"),
            "daily_tms_checklist": ("Daily TMS Checklist", "treatments"),
            "log_of_motor_thresholds": ("Log Of Motor Thresholds", "treatments"),
            "mt_session_log": ("MT Session Log", "treatments"),
            "participant_ap_rating_first_day_of_treatment": (
                "Participant A/P Rating (First Day Of Treatment)",
                "treatments",
            ),
            "fmri_targeting": ("fMRI Targeting", "neuroimaging"),
            "urine_toxicity_screen": ("Urine Toxicity Screen", "biologics_biometrics"),
            "subjectspecific_aesae_log": ("Subject-specific AE/SAE Log", "safety_regulatory"),
            "ae_daily_screening_nih": ("AE Daily Screening NIH", "safety_regulatory"),
            "tass_tms_safety_form": ("TASS TMS Safety Form", "safety_regulatory"),
            "progress_note": ("Progress Note", "safety_regulatory"),
            "survey_trigger_one_month": ("Survey Trigger (one month)", "admin"),
        }

        for instrument, (label, expected_category) in examples.items():
            with self.subTest(instrument=instrument):
                result = classify_one(instrument, label)
                self.assertEqual(result.category, expected_category)
                self.assertGreaterEqual(result.confidence, 0.55)

    def test_treatment_abbreviations_have_semantic_confidence(self):
        examples = {
            "participant_ap_rating_first_day_of_treatment": "Participant A/P Rating (First Day Of Treatment)",
            "participant_ap_rating_end_of_treatment": "Participant A/P Rating (End Of Treatment)",
            "log_of_motor_thresholds": "Log of Motor Thresholds; Log Of Motor Thresholds",
            "medication_changes": "Medication Changes",
        }

        for instrument, label in examples.items():
            with self.subTest(instrument=instrument):
                result = classify_one(instrument, label)
                self.assertEqual(result.category, "treatments")
                self.assertGreaterEqual(result.confidence, 0.70)

    def test_psychiatry_assessment_abbreviations_have_semantic_confidence(self):
        examples = {
            "mssi": "MSSI",
            "ymrs": "YMRS",
            "sigha": "SIGH-A",
            "sigh_a_mentor_rater": "SIGH-A (Mentor Rater)",
            "qids": "Quick Inventory Depressive Symptomatology (QIDS)",
            "gad_7": "GAD-7",
            "hamd6": "HAMD6",
            "madrss_survey": "MADRS-S Survey",
            "madrss_24_hour_version_survey": "MADRS-S 24 Hour Version (Survey)",
            "phq8": "PHQ-8",
            "ims12_survey": "IMS-12 (survey)",
            "bhs": "BHS",
            "bssssic": "BSS/SSI-C",
            "bdi_q9_si": "BDI (Q9, SI)",
            "panas": "PANAS",
            "panas_survey": "PANAS (Survey)",
            "ygtss_survey": "YGTSS (Survey)",
            "meq_survey": "MEQ (Survey)",
            "maudsley_thase_rush": "Maudsley Thase Rush",
            "cage": "CAGE",
            "psqi": "PSQI",
            "whoqolbref": "WHOQOL-BREF",
            "ybocs2_7389": "YBOCS2",
            "ybocs2_mentor_rater": "YBOCS2 (Mentor Rater)",
            "ybocs_2": "YBOCS-2",
            "visual_analog_scale_vas": "Visual Analog Scale (VAS)",
            "edinburgh_handedness_inventory_short_form": "Edinburgh Handedness Inventory - Short Form",
            "repetitive_behavior_questions": "Repetitive Behavior Questions",
            "pvss": "PVSS",
            "dars": "DARS",
            "wsas": "WSAS",
            "snapiv": "SNAP-IV",
            "bests_survey": "BEST-S (Survey)",
            "hacks_impairment_index": "Hacks Impairment Index",
            "mini_dsm5_certified_assessor": "MINI DSM5 (Certified Assessor)",
            "figs_lac_only": "FIGS Family Interview For Genetic Studies (FIGS) - LAC only",
            "ctq_lac_only": "CTQ Childhood Trauma Questionnaire (CTQ) - LAC only",
            "gad7": "Gad7 Generalized Anxiety Disorder 7item Scale Gad7",
            "cgi_i_mentor_rater": "CGI-I (Mentor Rater)",
            "cgi_s": "CGI-S",
            "dsii_s": "DSII-S Demoralization Scale-II: Self-report DSII-S",
            "shaps": "SHAPS",
            "pirs20_not_used": "PIRS-20 Insomnia Rating Scale (NOT USED)",
            "rrs22": "RRS-22 Ruminative Response Scale",
        }

        for instrument, label in examples.items():
            with self.subTest(instrument=instrument):
                result = classify_one(instrument, label)
                self.assertEqual(result.category, "assessments")
                self.assertGreaterEqual(result.confidence, 0.70)

    def test_tms_device_code_classifies_as_treatment(self):
        result = classify_one("magventure_code", "MagVenture Code")

        self.assertEqual(result.category, "treatments")
        self.assertGreaterEqual(result.confidence, 0.70)

    def test_geneactiv_checklist_classifies_as_biologics_biometrics(self):
        result = classify_one("geneactiv_checklist", "GENEActiv Checklist")

        self.assertEqual(result.category, "biologics_biometrics")
        self.assertGreaterEqual(result.confidence, 0.70)

    def test_safety_regulatory_abbreviations_have_semantic_confidence(self):
        examples = {
            "subjectspecific_pd_log": "Subject-specific PD Log",
            "safer_criteria": "SAFER Criteria",
        }

        for instrument, label in examples.items():
            with self.subTest(instrument=instrument):
                result = classify_one(instrument, label)
                self.assertEqual(result.category, "safety_regulatory")
                self.assertGreaterEqual(result.confidence, 0.70)

    def test_semantic_examples_expand_clinical_abbreviations(self):
        semantics_path = Path("scripts/classify_instruments/semantics.json")
        semantics = json.loads(semantics_path.read_text(encoding="utf-8"))
        assessment_examples = "\n".join(semantics["assessments"]["examples"])

        bad_fragments = [
            "MADRS rating scale",
            "GRID HAMD17 mentor rater",
            "PHQ depression questionnaire",
            "SCID diagnostic interview",
            "C-SSRS suicide severity scale",
            "YMRS mania rating",
            "IAT implicit association task",
            "Quick Inventory Depressive Symptomatology (QIDS)",
            "SNAP-IV (Survey)",
            "SIGH-A (Mentor Rater)",
            "YBOCS2 (Mentor Rater)",
        ]
        for fragment in bad_fragments:
            with self.subTest(fragment=fragment):
                self.assertNotIn(fragment, assessment_examples)

        expected_expansions = [
            "MADRS Montgomery Asberg Depression Rating Scale clinician-rated depression severity assessment",
            "GRID HAMD17 GRID Hamilton Depression Rating Scale 17-item clinician-rated depression severity assessment",
            "PHQ Patient Health Questionnaire depression symptom questionnaire",
            "SCID Structured Clinical Interview for DSM diagnostic interview",
            "C-SSRS Columbia Suicide Severity Rating Scale suicide risk assessment",
            "YMRS Young Mania Rating Scale clinician-rated mania severity assessment",
            "IAT Implicit Association Test behavioral association task",
            "SIGH-A Structured Interview Guide for the Hamilton Anxiety Rating Scale clinician-rated anxiety assessment",
            "QIDS Quick Inventory of Depressive Symptomatology depression symptom severity assessment",
            "HAMD6 Hamilton Depression Rating Scale 6-item clinician-rated depression severity assessment",
            "Holmes-Rahe LSSS Holmes and Rahe Life Stress Scale stressful life events and social readjustment stress burden assessment",
            "PANAS Positive and Negative Affect Schedule affect questionnaire",
            "YGTSS Yale Global Tic Severity Scale tic symptom severity assessment",
            "MEQ Morningness-Eveningness Questionnaire circadian preference and chronotype assessment",
            "PSQI Pittsburgh Sleep Quality Index sleep quality questionnaire",
            "WHOQOL-BREF World Health Organization Quality of Life Brief quality of life assessment",
            "Y-BOCS-II Yale-Brown Obsessive Compulsive Scale Second Edition obsessive-compulsive symptom severity assessment",
            "Repetitive Behavior Questions repetitive behavior symptom questionnaire",
            "PVSS Positive Valence Systems Scale positive valence reward and motivation assessment",
            "DARS Dimensional Anhedonia Rating Scale anhedonia symptom assessment",
            "WSAS Work and Social Adjustment Scale functional impairment assessment",
            "SNAP-IV Swanson Nolan and Pelham Rating Scale Fourth Edition ADHD and oppositional behavior symptom assessment",
            "BEST-S Borderline Evaluation of Severity over Time Self-report borderline personality symptom severity assessment",
            "Hacks Impairment Index neurologic psychomotor impairment assessment of speech cognition motor coordination and nystagmus",
            "MINI DSM5 Mini International Neuropsychiatric Interview for DSM-5 structured diagnostic interview with certified assessor administration",
            "FIGS Family Interview for Genetic Studies structured family psychiatric history interview",
            "CTQ Childhood Trauma Questionnaire retrospective self-report childhood trauma and maltreatment assessment",
            "GAD-7 Generalized Anxiety Disorder 7-item Scale anxiety symptom severity questionnaire",
            "CGI-I Clinical Global Impressions Improvement clinician-rated global improvement assessment",
            "CGI-S Clinical Global Impressions Severity clinician-rated global illness severity assessment",
            "VAS Visual Analog Scale subjective symptom intensity rating assessment",
            "YBOCS2 Yale-Brown Obsessive Compulsive Scale Second Edition obsessive-compulsive symptom severity assessment",
            "YBOCS-2 Yale-Brown Obsessive Compulsive Scale Second Edition obsessive-compulsive symptom severity assessment",
            "Edinburgh Handedness Inventory Short Form handedness laterality assessment",
            "DSII-S Demoralization Scale-II Self-report demoralization symptom assessment",
            "SHAPS Snaith-Hamilton Pleasure Scale hedonic capacity and anhedonia assessment",
            "PIRS-20 Insomnia Rating Scale 20-item insomnia symptom severity assessment",
            "RRS-22 Ruminative Responses Scale 22-item repetitive negative thinking and rumination assessment",
        ]
        for expansion in expected_expansions:
            with self.subTest(expansion=expansion):
                self.assertIn(expansion, assessment_examples)

    def test_semantic_examples_are_definition_like_across_categories(self):
        semantics_path = Path("scripts/classify_instruments/semantics.json")
        semantics = json.loads(semantics_path.read_text(encoding="utf-8"))
        all_examples = "\n".join(
            example
            for category in semantics.values()
            for example in category["examples"]
        )

        bad_fragments = [
            "patient contact log",
            "patient email",
            "link between screening id and study id",
            "remote econsent",
            "capacity to consent",
            "enrollment verification",
            "withdrawal status",
            "subject-specific adverse event log",
            "protocol deviation log",
            "note to file",
            "clinical free text narrative",
            "COVID19 screener",
            "Current Medication Form active medication exposure tracking",
        ]
        for fragment in bad_fragments:
            with self.subTest(fragment=fragment):
                self.assertNotIn(fragment, all_examples)

        expected_expansions = [
            "Patient Contact Log participant contact detail tracking form",
            "eConsent Electronic Consent remote consent documentation form",
            "PD Protocol Deviation regulatory protocol exception tracking log",
            "NTF Note To File regulatory documentation and study issue memo",
            "COVID19 Coronavirus Disease 2019 safety screener",
            "Current Medication Form current medication exposure and treatment tracking form",
            "Log of Motor Thresholds TMS motor threshold measurement tracking form",
            "Medication Changes medication adjustment and treatment change tracking form",
        ]
        for expansion in expected_expansions:
            with self.subTest(expansion=expansion):
                self.assertIn(expansion, all_examples)

    def test_outputs_classification_workbook_from_study_dictionary(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study_folder = Path(tmpdir) / "study"
            dictionary = study_folder / "data" / "cleaned" / "redcap" / "dictionary.xlsx"
            write_dictionary(
                dictionary,
                [
                    ("patient_contact_log", "Patient Contact Log"),
                    ("grid_hamd17_mentor_rater", "GRID HAMD17 (Mentor Rater)"),
                    ("outpatient_mri_checklist_concat_ag_5_13_2021", "Outpatient MRI Checklist"),
                    ("note_to_file", "Note To File"),
                ],
            )

            result = classify_instruments(study_folder=study_folder)
            workbook = load_workbook(result.output_path, data_only=True)
            rows = list(workbook["instrument_classification"].iter_rows(values_only=True))

        self.assertEqual(result.instrument_count, 4)
        self.assertEqual(rows[0], ("instrument_name", "instrument_label", "class", "confidence"))
        classes = {row[0]: row[2] for row in rows[1:]}
        self.assertEqual(classes["patient_contact_log"], "subjects")
        self.assertEqual(classes["grid_hamd17_mentor_rater"], "assessments")
        self.assertEqual(classes["outpatient_mri_checklist_concat_ag_5_13_2021"], "neuroimaging")
        self.assertEqual(classes["note_to_file"], "safety_regulatory")
        self.assertEqual(result.output_path, dictionary.parent / "instrument_classification.xlsx")


if __name__ == "__main__":
    unittest.main()
