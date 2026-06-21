import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.create_study_overview.run import create_study_overview, match_description


def write_workbook(path: Path, sheets: dict[str, list[list[object]]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in sheets.items():
        worksheet = workbook.create_sheet(sheet_name)
        for row in rows:
            worksheet.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


class CreateStudyOverviewTests(unittest.TestCase):
    def test_matches_hyphenated_clinical_abbreviations_before_delivery_wrappers(self):
        semantic_examples = {
            "assessments": [
                "SIGH-A Structured Interview Guide for the Hamilton Anxiety Rating Scale clinician-rated anxiety assessment",
                "YBOCS2 Yale-Brown Obsessive Compulsive Scale Second Edition obsessive-compulsive symptom severity assessment",
                "WHOQOL-BREF World Health Organization Quality of Life Brief quality of life assessment",
                "SNAP-IV Swanson Nolan and Pelham Rating Scale Fourth Edition ADHD and oppositional behavior symptom assessment",
                "QIDS Quick Inventory of Depressive Symptomatology depression symptom severity assessment",
                "GAD-7 Generalized Anxiety Disorder 7-item Scale anxiety symptom severity questionnaire",
                "HAMD6 Hamilton Depression Rating Scale 6-item clinician-rated depression severity assessment",
            ],
            "treatments": [
                "Medication Changes medication adjustment and treatment change tracking form",
                "Past Medication Form lifetime medication exposure history",
            ],
            "admin": [
                "Survey Trigger automated participant survey scheduling form",
            ],
        }

        expected = {
            "SIGH-A (Mentor Rater)": semantic_examples["assessments"][0],
            "YBOCS2 (Mentor Rater)": semantic_examples["assessments"][1],
            "WHOQOL-BREF": semantic_examples["assessments"][2],
            "SNAP-IV (Survey)": semantic_examples["assessments"][3],
            "Quick Inventory Depressive Symptomatology (QIDS)": semantic_examples["assessments"][4],
            "GAD-7": semantic_examples["assessments"][5],
            "HAMD6": semantic_examples["assessments"][6],
            "Past Medication Form": semantic_examples["treatments"][1],
        }

        for label, description in expected.items():
            with self.subTest(label=label):
                category = "treatments" if label == "Past Medication Form" else "assessments"
                self.assertEqual(match_description(category, label, label, semantic_examples), description)

    def test_creates_class_sheets_with_labels_descriptions_timepoints_and_tools(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            study_folder = Path(tmpdir) / "58807-BRAINS"
            cleaned_dir = study_folder / "data" / "cleaned"
            write_workbook(
                cleaned_dir / "dictionary.xlsx",
                {
                    "instrument": [
                        ["instrument", "instrument_label"],
                        ["madrs", "MADRS Montgomery Asberg Depression Rating Scale"],
                        ["panas_survey", "PANAS Survey"],
                        ["mini_certified", "MINI DSM5 Certified Assessor"],
                        ["phq_self_report", "PHQ-8 Self-report"],
                        ["cgi_trainee", "CGI-I Trainee Rater"],
                        ["ctq_do_not_use", "CTQ Childhood Trauma Questionnaire DO NOT USE"],
                        ["pirs_not_used", "PIRS-20 Insomnia Rating Scale (NOT USED)"],
                        ["old_figs", "FIGS Family Interview deprecated"],
                    ]
                },
            )
            write_workbook(
                cleaned_dir / "assessments" / "58807-madrs.xlsx",
                {
                    "cleaned": [
                        ["IRB", "subid", "arm", "visit", "date", "record_id", "redcap_event_name"],
                        ["58807", "s001", "1", "V1", "2024-01-01", "58807_s001", "baseline_arm_1"],
                        ["58807", "s001", "1", "V2", "2024-01-15", "58807_s001", "visit_2_arm_1"],
                        ["58807", "s002", "1", "V2", "2024-01-17", "58807_s002", "visit_2_arm_1"],
                    ]
                },
            )
            for instrument in ["panas_survey", "mini_certified", "phq_self_report"]:
                write_workbook(
                    cleaned_dir / "assessments" / f"58807-{instrument}.xlsx",
                    {
                        "cleaned": [
                            ["IRB", "subid", "arm", "visit"],
                            ["58807", "s001", "1", "V1"],
                        ]
                    },
                )
            write_workbook(
                cleaned_dir / "assessments" / "58807-unlabeled_assessment.xlsx",
                {
                    "cleaned": [
                        ["IRB", "subid", "arm", "visit"],
                        ["58807", "s001", "1", "V1"],
                    ]
                },
            )
            for instrument in ["cgi_trainee", "ctq_do_not_use", "pirs_not_used", "old_figs"]:
                write_workbook(
                    cleaned_dir / "assessments" / f"58807-{instrument}.xlsx",
                    {
                        "cleaned": [
                            ["IRB", "subid", "arm", "visit"],
                            ["58807", "s001", "1", "V1"],
                        ]
                    },
                )
            write_workbook(
                cleaned_dir / "subjects" / "subject_timepoints.xlsx",
                {
                    "subject_timepoints": [
                        ["IRB", "subid", "arm", "visit"],
                        ["58807", "s001", "1", "V1"],
                    ]
                },
            )
            write_workbook(
                cleaned_dir / "unknown" / "58807-mystery.xlsx",
                {
                    "cleaned": [
                        ["IRB", "subid", "arm", "visit"],
                        ["58807", "s001", "1", "V1"],
                    ]
                },
            )
            semantics_path = Path(tmpdir) / "semantics.json"
            semantics_path.write_text(
                json.dumps(
                    {
                        "assessments": {
                            "caption": "Clinical assessments.",
                            "examples": [
                                "MADRS Montgomery Asberg Depression Rating Scale clinician-rated depression severity assessment"
                            ],
                        }
                    }
                ),
                encoding="utf-8",
            )

            result = create_study_overview(study_folder, semantics_path=semantics_path)

            self.assertEqual(result.output_path, study_folder / "overview" / "58807-BRAINS.xlsx")
            workbook = load_workbook(result.output_path, data_only=True)
            self.assertIn("assessments", workbook.sheetnames)
            self.assertIn("subjects", workbook.sheetnames)
            self.assertNotIn("unknown", workbook.sheetnames)

            assessment_rows = list(workbook["assessments"].iter_rows(values_only=True))
            self.assertEqual(
                assessment_rows[0],
                (
                    "data",
                    "description",
                    "# of timepoints (collected)",
                    "timepoints (collected)",
                    "collection method",
                    "collection tool",
                ),
            )
            self.assertEqual(
                assessment_rows[1],
                (
                    "MADRS Montgomery Asberg Depression Rating Scale",
                    "MADRS Montgomery Asberg Depression Rating Scale clinician-rated depression severity assessment",
                    2,
                    "V1; V2",
                    None,
                    "REDCap (IRB: 58807)",
                ),
            )
            rows_by_data = {row[0]: row for row in assessment_rows[1:]}
            self.assertEqual(rows_by_data["PANAS Survey"][4], "Survey")
            self.assertEqual(rows_by_data["MINI DSM5 Certified Assessor"][4], "Clinician administered")
            self.assertEqual(rows_by_data["PHQ-8 Self-report"][4], "Self-report")
            self.assertEqual(len(assessment_rows), 5)

            subject_rows = list(workbook["subjects"].iter_rows(values_only=True))
            self.assertEqual(len(subject_rows), 1)


if __name__ == "__main__":
    unittest.main()
