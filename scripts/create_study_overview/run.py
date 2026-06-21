#!/usr/bin/env python3
"""Create a study overview workbook from cleaned study outputs."""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


if __package__ in {None, ""}:
    sys.path.append(str(Path(__file__).resolve().parents[2]))


DATA_CLASSES = [
    "assessments",
    "neuroimaging",
    "treatments",
    "biologics_biometrics",
    "subjects",
    "safety_regulatory",
]
OVERVIEW_COLUMNS = [
    "data",
    "description",
    "# of timepoints (collected)",
    "timepoints (collected)",
    "collection method",
    "collection tool",
]
DEFAULT_SEMANTICS = Path(__file__).resolve().parents[1] / "classify_instruments" / "semantics.json"
EXCLUDED_OVERVIEW_TERMS = [
    "trainee",
    "do not use",
    "not used",
    "deprecated",
    "obsolete",
    "retired",
    "inactive",
    "not in use",
    "no longer used",
    "unused",
]


@dataclass(frozen=True)
class StudyOverviewResult:
    study_folder: Path
    output_path: Path
    row_count: int


def normalize_text(value: object) -> str:
    text = str(value or "").lower()
    text = re.sub(r"(?<=[a-z])(?=[0-9])|(?<=[0-9])(?=[a-z])", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def tokens(value: object) -> set[str]:
    return set(normalize_text(value).split())


def is_blank(value: object) -> bool:
    return value is None or str(value).strip() == ""


def instrument_key_from_workbook(path: Path) -> str:
    stem = path.stem
    stripped = re.sub(r"^(?:\d+[-_])+", "", stem)
    return stripped or stem


def overview_label_for_workbook(path: Path, dictionary_labels: dict[str, str]) -> str | None:
    instrument_key = instrument_key_from_workbook(path)
    return dictionary_labels.get(instrument_key) or dictionary_labels.get(normalize_text(instrument_key))


def has_excluded_overview_term(*values: object) -> bool:
    text = normalize_text(" ".join(str(value or "") for value in values))
    compact = text.replace(" ", "")
    if "donotuse" in compact:
        return True
    if "notused" in compact:
        return True
    return any(term in text for term in EXCLUDED_OVERVIEW_TERMS)


def collection_method_from_label(label: object) -> str | None:
    text = normalize_text(label)
    if "self report" in text or "selfreport" in text:
        return "Self-report"
    if "mentor rater" in text or "certified assessor" in text or "clinician" in text:
        return "Clinician administered"
    if "survey" in text:
        return "Survey"
    return None


def dictionary_path_for_study(study_folder: Path) -> Path:
    candidates = [
        study_folder / "data" / "cleaned" / "dictionary.xlsx",
        study_folder / "data" / "cleaned" / "redcap" / "dictionary.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def load_dictionary_labels(study_folder: Path) -> dict[str, str]:
    dictionary_path = dictionary_path_for_study(study_folder)
    if not dictionary_path.exists():
        return {}

    workbook = load_workbook(dictionary_path, read_only=True, data_only=True)
    if "instrument" not in workbook.sheetnames:
        return {}

    rows = list(workbook["instrument"].iter_rows(values_only=True))
    if not rows:
        return {}

    headers = [str(value or "").strip() for value in rows[0]]
    lookup = {normalize_header(header): index for index, header in enumerate(headers)}
    instrument_index = lookup.get("instrument")
    label_index = lookup.get("instrument_label")
    if instrument_index is None or label_index is None:
        return {}

    labels: dict[str, str] = {}
    for row in rows[1:]:
        instrument = str(row[instrument_index] or "").strip() if instrument_index < len(row) else ""
        label = str(row[label_index] or "").strip() if label_index < len(row) else ""
        if not instrument or not label:
            continue
        labels[instrument] = label
        labels[normalize_text(instrument)] = label
    return labels


def load_semantic_examples(path: str | Path) -> dict[str, list[str]]:
    path = Path(path)
    if not path.exists():
        return {}
    with path.open(encoding="utf-8") as file:
        semantics = json.load(file)
    return {
        category: [str(example) for example in entry.get("examples", []) if str(example).strip()]
        for category, entry in semantics.items()
        if isinstance(entry, dict)
    }


def example_matches(example: str, text: str) -> bool:
    normalized_example = normalize_text(example)
    normalized_text = normalize_text(text)
    if not normalized_example or not normalized_text:
        return False
    if normalized_example in normalized_text or normalized_text in normalized_example:
        return True

    text_tokens = tokens(text)
    example_words = str(example).strip().split()
    if not example_words:
        return False
    lead = normalize_text(example_words[0])
    original_lead = example_words[0]
    is_acronym_lead = any(character.isupper() for character in original_lead) and not any(
        character.islower() for character in original_lead
    )
    if len(lead) >= 2 and ((" " in lead and f" {lead} " in f" {normalized_text} ") or (is_acronym_lead and lead in text_tokens)):
        return True

    first_two = normalize_text(" ".join(example_words[:2]))
    return len(first_two) >= 5 and first_two in normalized_text


def match_description(category: str, data_label: str, instrument_key: str, semantic_examples: dict[str, list[str]]) -> str:
    text = f"{data_label} {instrument_key}"
    candidate_examples = [
        *semantic_examples.get(category, []),
        *[
            example
            for other_category, examples in semantic_examples.items()
            if other_category != category
            for example in examples
        ],
    ]
    for example in candidate_examples:
        if example_matches(example, text):
            return example
    return ""


def workbook_paths_for_class(cleaned_dir: Path, data_class: str) -> list[Path]:
    class_dir = cleaned_dir / data_class
    if not class_dir.exists():
        return []
    return sorted(
        path
        for path in class_dir.glob("*.xlsx")
        if not path.name.startswith("~$") and path.name != "dictionary.xlsx"
    )


def read_cleaned_sheet(path: Path) -> tuple[list[object], list[tuple[object, ...]]] | None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "cleaned" not in workbook.sheetnames:
        return None
    rows = list(workbook["cleaned"].iter_rows(values_only=True))
    if not rows:
        return [], []
    return list(rows[0]), [tuple(row) for row in rows[1:] if any(not is_blank(value) for value in row)]


def natural_visit_key(value: object) -> tuple[object, ...]:
    text = str(value or "").strip()
    match = re.search(r"(\d+)", text)
    if match:
        return (text[: match.start()], int(match.group(1)), text)
    return (text, -1, text)


def unique_values(values: Iterable[object]) -> list[str]:
    seen: set[str] = set()
    output: list[str] = []
    for value in values:
        if is_blank(value):
            continue
        text = str(value).strip()
        if text in seen:
            continue
        seen.add(text)
        output.append(text)
    return output


def cleaned_summary(path: Path) -> tuple[int | None, str | None, str | None]:
    cleaned = read_cleaned_sheet(path)
    if cleaned is None:
        return None, None, None
    _, rows = cleaned
    visits = unique_values(row[3] for row in rows if len(row) > 3)
    irbs = unique_values(row[0] for row in rows if len(row) > 0)
    visits = sorted(visits, key=natural_visit_key)
    irbs = sorted(irbs, key=lambda value: (0, int(value), value) if value.isdigit() else (1, 0, value))
    visit_count = len(visits)
    visit_text = "; ".join(visits) if visits else None
    collection_tool = f"REDCap (IRB: {', '.join(irbs)})" if irbs else None
    return visit_count, visit_text, collection_tool


def overview_row(
    path: Path,
    data_class: str,
    data_label: str,
    semantic_examples: dict[str, list[str]],
) -> list[object]:
    instrument_key = instrument_key_from_workbook(path)
    description = match_description(data_class, data_label, instrument_key, semantic_examples) or None
    visit_count, visit_text, collection_tool = cleaned_summary(path)
    return [
        data_label,
        description,
        visit_count,
        visit_text,
        collection_method_from_label(data_label),
        collection_tool,
    ]


def autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 80)
        worksheet.column_dimensions[column_cells[0].column_letter].width = width


def create_study_overview(
    study_folder: str | Path,
    output_path: str | Path | None = None,
    semantics_path: str | Path = DEFAULT_SEMANTICS,
) -> StudyOverviewResult:
    study_folder = Path(study_folder)
    cleaned_dir = study_folder / "data" / "cleaned"
    output_path = (
        Path(output_path)
        if output_path is not None
        else study_folder / "overview" / f"{study_folder.name}.xlsx"
    )

    dictionary_labels = load_dictionary_labels(study_folder)
    semantic_examples = load_semantic_examples(semantics_path)

    workbook = Workbook()
    workbook.remove(workbook.active)
    row_count = 0
    for data_class in DATA_CLASSES:
        worksheet = workbook.create_sheet(data_class)
        worksheet.append(OVERVIEW_COLUMNS)
        for path in workbook_paths_for_class(cleaned_dir, data_class):
            instrument_key = instrument_key_from_workbook(path)
            data_label = overview_label_for_workbook(path, dictionary_labels)
            if not data_label:
                continue
            if has_excluded_overview_term(instrument_key, data_label, path.stem):
                continue
            worksheet.append(overview_row(path, data_class, data_label, semantic_examples))
            row_count += 1
        autosize_columns(worksheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return StudyOverviewResult(study_folder=study_folder, output_path=output_path, row_count=row_count)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--study-folder", required=True, type=Path, help="Cleaned study folder.")
    parser.add_argument("--out", type=Path, help="Optional overview workbook output path.")
    parser.add_argument(
        "--semantics",
        type=Path,
        default=DEFAULT_SEMANTICS,
        help="Instrument semantics JSON used to fill description values.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    result = create_study_overview(args.study_folder, output_path=args.out, semantics_path=args.semantics)
    print(result.output_path)
    print(f"rows={result.row_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
